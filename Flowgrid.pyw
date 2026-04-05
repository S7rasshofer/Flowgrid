#!/usr/bin/env python3
"""Flowgrid - PySide6 production/part tracking desktop utility.
"""

from __future__ import annotations

import csv
import ctypes
import getpass
import importlib
import json
import math
import os
import re
import shutil
import socket
import struct
import subprocess
import sys
import time
import traceback
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable

import sqlite3
from datetime import date, datetime, timedelta


DATA_ROOT_ENV_VAR = "FLOWGRID_DATA_ROOT"
DEFAULT_SHARED_DATA_ROOT = Path(r"Z:\DATA\Flowgrid")
LAUNCH_LOG_FILENAME = "Flowgrid_launch_errors.log"
RUNTIME_LOG_FILENAME_PREFIX = "Flowgrid_runtime"
RUNTIME_LOG_FILENAME_SUFFIX = ".log.jsonl"
RUNTIME_LOG_MAX_BYTES = 10 * 1024 * 1024
RUNTIME_LOG_MAX_BACKUPS = 20
RUNTIME_PROMPT_TITLE = "Flowgrid Runtime Issue"
_RESOLVED_DATA_ROOT: Path | None = None
_DATA_ROOT_FALLBACK_DETAILS = ""
_DATA_ROOT_FALLBACK_NOTIFIED = False
_RUNTIME_ESCALATED_EVENTS: set[str] = set()
_RUNTIME_LOG_WRITE_IN_PROGRESS = False
_LAUNCH_LOGGED_ONCE_EVENTS: set[str] = set()
ASSETS_DIR_NAME = "Assets"
FLOWGRID_ICON_PACK_DIR_NAME = "Flowgrid Icons"
ASSET_AGENT_ICON_DIR_NAME = "agent_icons"
ASSET_ADMIN_ICON_DIR_NAME = "admin_icons"
ASSET_QA_FLAG_ICON_DIR_NAME = "qa_flag_icons"
ASSET_PART_FLAG_IMAGE_DIR_NAME = "part_flag_images"
ASSET_UI_ICON_COMPAT_DIR_NAME = "ui_icons"
_CLI_FLAGS = {str(arg or "").strip().lower() for arg in sys.argv[1:] if str(arg or "").strip()}
_INSTALLER_FLAGS_ACTIVE = bool({"--install", "--create-shortcut"} & _CLI_FLAGS)

# ============================================================================
# CENTRALIZED PATH CONFIGURATION SYSTEM
# ============================================================================
# Flowgrid_paths.json lives on the shared drive and defines all path locations.
# User only edits this ONE file if they change shared drive locations.
# Both installer and app read from this config.

_FLOWGRID_PATHS_CONFIG: dict[str, Any] | None = None
_FLOWGRID_PATHS_CONFIG_ERROR: str = ""


def _find_paths_config_on_shared_drive() -> Path | None:
    """Locate Flowgrid_paths.json starting from the script location or env override."""
    candidates: list[Path] = []
    
    # Try environment variable override first
    env_override = str(os.environ.get("FLOWGRID_PATHS_CONFIG", "") or "").strip()
    if env_override:
        candidates.append(Path(env_override))
    
    # Try script parent directory (common when running from shared drive)
    try:
        script_dir = Path(__file__).resolve().parent
        candidates.append(script_dir / "Flowgrid_paths.json")
    except Exception:
        pass
    
    # Try current working directory
    candidates.append(Path.cwd() / "Flowgrid_paths.json")
    
    # Try Documents\Flowgrid (after local install)
    try:
        local_app = Path.home() / "Documents" / "Flowgrid"
        candidates.append(local_app / "Flowgrid_paths.json")
    except Exception:
        pass
    
    for candidate in candidates:
        try:
            if candidate.exists() and candidate.is_file():
                return candidate.resolve()
        except Exception:
            pass
    
    return None


def _load_paths_config() -> dict[str, Any]:
    """Load and cache Flowgrid_paths.json configuration."""
    global _FLOWGRID_PATHS_CONFIG, _FLOWGRID_PATHS_CONFIG_ERROR
    
    if _FLOWGRID_PATHS_CONFIG is not None:
        return _FLOWGRID_PATHS_CONFIG
    
    config_path = _find_paths_config_on_shared_drive()
    if config_path is None:
        _FLOWGRID_PATHS_CONFIG_ERROR = (
            "Flowgrid_paths.json not found. "
            "Searched: script dir, working dir, Documents\Flowgrid. "
            "Set FLOWGRID_PATHS_CONFIG env var or place file on shared drive."
        )
        _FLOWGRID_PATHS_CONFIG = {}
        return _FLOWGRID_PATHS_CONFIG
    
    try:
        with config_path.open("r", encoding="utf-8") as handle:
            _FLOWGRID_PATHS_CONFIG = json.load(handle)
        return _FLOWGRID_PATHS_CONFIG
    except Exception as exc:
        _FLOWGRID_PATHS_CONFIG_ERROR = f"Failed to parse {config_path}: {type(exc).__name__}: {exc}"
        _FLOWGRID_PATHS_CONFIG = {}
        return _FLOWGRID_PATHS_CONFIG


def _substitute_path_variables(template: str, shared_root: Path | None = None) -> str:
    """Substitute {DOCUMENTS}, {SHARED_ROOT}, etc. in path strings."""
    if not template:
        return ""
    
    result = str(template)
    
    # Substitute {DOCUMENTS}
    if "{DOCUMENTS}" in result:
        try:
            documents = Path.home() / "Documents"
            result = result.replace("{DOCUMENTS}", str(documents))
        except Exception:
            pass
    
    # Substitute {SHARED_ROOT}
    if "{SHARED_ROOT}" in result:
        root = shared_root or Path(r"Z:\DATA\Tracker - Test")
        result = result.replace("{SHARED_ROOT}", str(root))
    
    return result


def _resolve_path_from_config(config_key: str, default: str | Path | None = None, shared_root: Path | None = None) -> Path:
    """Retrieve a path from the config, with substitution and fallback."""
    config = _load_paths_config()
    
    # Navigate nested keys (e.g., "local_paths.database_folder")
    parts = str(config_key).split(".")
    value = config
    for part in parts:
        if isinstance(value, dict):
            value = value.get(part)
        else:
            value = None
            break
    
    if value is None:
        if default is None:
            return Path.cwd()
        return Path(_substitute_path_variables(str(default), shared_root))
    
    return Path(_substitute_path_variables(str(value), shared_root))


def _get_shared_root_from_config() -> Path:
    """Get the configured shared drive root."""
    config = _load_paths_config()
    shared_root_str = config.get("shared_drive_root", "Z:\\DATA\\Tracker - Test")
    return Path(shared_root_str)


# ============================================================================
# PATH GETTERS FOR SPECIFIC DIRECTORIES (using centralized config)
# ============================================================================

def _get_local_app_folder() -> Path:
    """Get the local app installation folder (e.g., Documents\Flowgrid)."""
    shared_root = _get_shared_root_from_config()
    return _resolve_path_from_config("local_paths.app_folder", "{DOCUMENTS}\\Flowgrid", shared_root)


def _get_local_database_folder() -> Path:
    """Get the local database folder (e.g., Documents\Flowgrid\Data)."""
    shared_root = _get_shared_root_from_config()
    db_folder = _resolve_path_from_config("local_paths.database_folder", "{DOCUMENTS}\\Flowgrid\\Data", shared_root)
    db_folder.mkdir(parents=True, exist_ok=True)
    return db_folder


def _get_local_queue_folder() -> Path:
    """Get the local sync queue folder (e.g., Documents\Flowgrid\Queue)."""
    shared_root = _get_shared_root_from_config()
    queue_folder = _resolve_path_from_config("local_paths.queue_folder", "{DOCUMENTS}\\Flowgrid\\Queue", shared_root)
    queue_folder.mkdir(parents=True, exist_ok=True)
    return queue_folder


def _get_local_config_folder() -> Path:
    """Get the local config folder (e.g., Documents\Flowgrid\Config)."""
    shared_root = _get_shared_root_from_config()
    config_folder = _resolve_path_from_config("local_paths.config_folder", "{DOCUMENTS}\\Flowgrid\\Config", shared_root)
    config_folder.mkdir(parents=True, exist_ok=True)
    return config_folder


def _get_local_assets_folder() -> Path:
    """Get the local assets folder (e.g., Documents\Flowgrid\Assets)."""
    shared_root = _get_shared_root_from_config()
    assets_folder = _resolve_path_from_config("local_paths.assets_folder", "{DOCUMENTS}\\Flowgrid\\Assets", shared_root)
    assets_folder.mkdir(parents=True, exist_ok=True)
    return assets_folder


def _get_shared_archive_folder() -> Path:
    """Get the shared drive archive folder for backups/exports."""
    shared_root = _get_shared_root_from_config()
    archive_folder = _resolve_path_from_config("shared_paths.archive_folder", "{SHARED_ROOT}\\archive", shared_root)
    archive_folder.mkdir(parents=True, exist_ok=True)
    return archive_folder


def _local_data_root() -> Path:
    try:
        return Path(__file__).resolve().parent
    except Exception as exc:
        fallback = Path.cwd()
        _runtime_log_event(
            "bootstrap.local_data_root_fallback",
            severity="warning",
            summary="Fell back to current working directory for local data root.",
            exc=exc,
            context={"fallback_path": str(fallback)},
        )
        return fallback


def _configured_data_root() -> Path:
    """Get configured data root from paths config, env var, or fallback to local."""
    # First, try reading from centralized config
    try:
        config = _load_paths_config()
        shared_root = _get_shared_root_from_config()
        if config and shared_root:
            return shared_root
    except Exception:
        pass
    
    # Fall back to env var
    override = str(os.environ.get(DATA_ROOT_ENV_VAR, "") or "").strip()
    if override:
        return Path(override)
    
    # Final fallback: local app folder
    return _local_data_root()


def _paths_equal(left: Path, right: Path) -> bool:
    try:
        return left.resolve() == right.resolve()
    except Exception as exc:
        _runtime_log_event(
            "bootstrap.paths_equal_resolve_failed",
            severity="warning",
            summary="Path resolution failed while comparing paths; using string comparison fallback.",
            exc=exc,
            context={"left": str(left), "right": str(right)},
        )
        return str(left) == str(right)


def _legacy_data_candidates(filename: str) -> list[Path]:
    candidates: list[Path] = []
    try:
        candidates.append(Path(__file__).with_name(filename))
    except Exception as exc:
        _runtime_log_event(
            "bootstrap.legacy_candidate_file_dir_failed",
            severity="warning",
            summary="Unable to derive __file__-relative legacy data candidate.",
            exc=exc,
            context={"filename": filename},
        )
    candidates.append(Path.cwd() / filename)

    unique: list[Path] = []
    for path in candidates:
        if any(_paths_equal(path, existing) for existing in unique):
            continue
        unique.append(path)
    return unique


def _resolve_data_root() -> Path:
    global _RESOLVED_DATA_ROOT, _DATA_ROOT_FALLBACK_DETAILS
    if _RESOLVED_DATA_ROOT is not None:
        return _RESOLVED_DATA_ROOT

    target = _configured_data_root()
    try:
        target.mkdir(parents=True, exist_ok=True)
        _RESOLVED_DATA_ROOT = target
        return _RESOLVED_DATA_ROOT
    except Exception as exc:
        fallback = _local_data_root()
        try:
            fallback.mkdir(parents=True, exist_ok=True)
        except Exception as fallback_exc:
            _runtime_log_event(
                "bootstrap.data_root_fallback_mkdir_failed",
                severity="error",
                summary="Failed to create fallback data root directory.",
                exc=fallback_exc,
                context={"fallback_path": str(fallback)},
            )
        _RESOLVED_DATA_ROOT = fallback
        _DATA_ROOT_FALLBACK_DETAILS = (
            f"Configured data root: {target}\n"
            f"Reason: {type(exc).__name__}: {exc}\n"
            f"Fallback data root: {fallback}\n"
            f"Set {DATA_ROOT_ENV_VAR} to override the data path."
        )
        _runtime_log_event(
            "bootstrap.data_root_fallback_used",
            severity="warning",
            summary="Configured data root failed; using fallback data root.",
            exc=exc,
            context={"configured_path": str(target), "fallback_path": str(fallback)},
        )
        return _RESOLVED_DATA_ROOT


def _data_file_path(filename: str, migrate_legacy: bool = True) -> Path:
    """
    Resolve path for a data file.
    
    All files use shared data root for centralized access.
    Local folders are only for temporary queues and user-specific data.
    """
    # All data files go to shared root for centralized reading
    target = _resolve_data_root() / filename
    
    if not migrate_legacy or target.exists():
        return target

    for legacy in _legacy_data_candidates(filename):
        if _paths_equal(legacy, target):
            continue
        if not legacy.exists() or not legacy.is_file():
            continue
        try:
            target.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(legacy, target)
            break
        except Exception as exc:
            _runtime_log_event(
                "bootstrap.data_file_legacy_copy_failed",
                severity="warning",
                summary="Legacy data file copy failed; continuing without migration for this source.",
                exc=exc,
                context={
                    "filename": filename,
                    "legacy_path": str(legacy),
                    "target_path": str(target),
                },
            )
            continue
    return target


def _migrate_legacy_agent_icons(target_db_path: Path) -> None:
    data_root = target_db_path.parent
    assets_root = data_root / ASSETS_DIR_NAME
    managed_folders = (
        ASSET_AGENT_ICON_DIR_NAME,
        ASSET_ADMIN_ICON_DIR_NAME,
        ASSET_QA_FLAG_ICON_DIR_NAME,
        ASSET_PART_FLAG_IMAGE_DIR_NAME,
        ASSET_UI_ICON_COMPAT_DIR_NAME,
        FLOWGRID_ICON_PACK_DIR_NAME,
    )
    try:
        assets_root.mkdir(parents=True, exist_ok=True)
        for folder in managed_folders:
            (assets_root / folder).mkdir(parents=True, exist_ok=True)
    except Exception as exc:
        _runtime_log_event(
            "bootstrap.assets_dir_create_failed",
            severity="warning",
            summary="Unable to create Assets folders; skipping legacy asset migration.",
            exc=exc,
            context={"assets_root": str(assets_root)},
        )
        return

    candidate_roots: list[Path] = []
    try:
        candidate_roots.append(Path(__file__).resolve().parent)
    except Exception as exc:
        _runtime_log_event(
            "bootstrap.legacy_asset_root_resolve_failed",
            severity="warning",
            summary="Unable to resolve module directory while gathering legacy asset roots.",
            exc=exc,
            context={"target_db_path": str(target_db_path)},
        )
    candidate_roots.append(Path.cwd())
    candidate_roots.append(data_root)

    unique_roots: list[Path] = []
    for root in candidate_roots:
        if any(_paths_equal(root, existing) for existing in unique_roots):
            continue
        unique_roots.append(root)

    legacy_to_assets = {
        ASSET_AGENT_ICON_DIR_NAME: ASSET_AGENT_ICON_DIR_NAME,
        ASSET_ADMIN_ICON_DIR_NAME: ASSET_ADMIN_ICON_DIR_NAME,
        ASSET_QA_FLAG_ICON_DIR_NAME: ASSET_QA_FLAG_ICON_DIR_NAME,
        ASSET_PART_FLAG_IMAGE_DIR_NAME: ASSET_PART_FLAG_IMAGE_DIR_NAME,
        "ui_icons": ASSET_UI_ICON_COMPAT_DIR_NAME,
    }

    for root in unique_roots:
        for legacy_folder, asset_folder in legacy_to_assets.items():
            source_dir = root / legacy_folder
            target_dir = assets_root / asset_folder
            if not source_dir.exists() or not source_dir.is_dir():
                continue
            try:
                if _paths_equal(source_dir, target_dir):
                    continue
            except Exception as exc:
                _runtime_log_event(
                    "bootstrap.legacy_asset_path_compare_failed",
                    severity="warning",
                    summary="Failed comparing legacy and target asset directories; continuing migration scan.",
                    exc=exc,
                    context={"source_dir": str(source_dir), "target_dir": str(target_dir)},
                )
            try:
                for source_file in source_dir.rglob("*"):
                    if not source_file.is_file():
                        continue
                    rel = source_file.relative_to(source_dir)
                    target_file = target_dir / rel
                    if target_file.exists():
                        continue
                    target_file.parent.mkdir(parents=True, exist_ok=True)
                    try:
                        shutil.copy2(source_file, target_file)
                    except Exception as exc:
                        _runtime_log_event(
                            "bootstrap.legacy_asset_copy_failed",
                            severity="warning",
                            summary="Failed copying a legacy asset file into Assets.",
                            exc=exc,
                            context={
                                "source_file": str(source_file),
                                "target_file": str(target_file),
                            },
                        )
                        continue
            except Exception as exc:
                _runtime_log_event(
                    "bootstrap.legacy_asset_root_scan_failed",
                    severity="warning",
                    summary="Failed scanning a legacy asset source directory.",
                    exc=exc,
                    context={"source_dir": str(source_dir), "target_dir": str(target_dir)},
                )
                continue


def _error_log_path() -> Path:
    try:
        return _data_file_path(LAUNCH_LOG_FILENAME, migrate_legacy=False)
    except Exception as exc:
        fallback = Path.cwd() / LAUNCH_LOG_FILENAME
        _runtime_log_event(
            "bootstrap.launch_log_path_fallback",
            severity="warning",
            summary="Failed to resolve launch log path from data root; using current working directory.",
            exc=exc,
            context={"fallback_path": str(fallback)},
        )
        return fallback


def _log_launch_error(code: str, summary: str, details: str = "") -> None:
    try:
        now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
        lines = [f"[{now}] [{code}] {summary}"]
        if details:
            lines.append(details)
        lines.append("-" * 72)
        with _error_log_path().open("a", encoding="utf-8") as handle:
            handle.write("\n".join(lines) + "\n")
    except Exception as exc:
        _runtime_log_event(
            "bootstrap.launch_log_write_failed",
            severity="error",
            summary="Failed to append launch error log entry.",
            exc=exc,
            context={"code": code, "summary": summary},
        )


def _log_launch_error_once(event_key: str, code: str, summary: str, details: str = "") -> None:
    key = str(event_key or "").strip().lower()
    if not key:
        key = f"{code}:{summary}".strip().lower()
    if key in _LAUNCH_LOGGED_ONCE_EVENTS:
        return
    _LAUNCH_LOGGED_ONCE_EVENTS.add(key)
    _log_launch_error(code, summary, details)


def _show_windows_toast(title: str, message: str) -> bool:
    # Intentionally disabled to avoid spawning shell-based toast commands.
    return False


def _show_error_dialog(title: str, message: str) -> None:
    if os.name == "nt":
        try:
            ctypes.windll.user32.MessageBoxW(None, str(message), str(title), 0x10 | 0x1000)
            return
        except Exception as exc:
            _runtime_log_event(
                "bootstrap.error_dialog_native_failed",
                severity="warning",
                summary="Native error dialog failed; falling back to console print.",
                exc=exc,
                context={"title": str(title)},
            )
    try:
        print(f"{title}\n{message}")
    except Exception as exc:
        _runtime_log_event(
            "bootstrap.error_dialog_console_print_failed",
            severity="error",
            summary="Fallback console error dialog print failed.",
            exc=exc,
            context={"title": str(title)},
        )


def _notify_launch_error(code: str, summary: str, details: str = "") -> None:
    _log_launch_error(code, summary, details)
    short_msg = f"[{code}] {summary}"[:240]
    if _show_windows_toast("Flowgrid Launch Error", short_msg):
        return
    log_path = _error_log_path()
    body = short_msg
    if details:
        body = f"{body}\n\nDetails:\n{details}"
    body = f"{body}\n\nLog: {log_path}"
    _show_error_dialog("Flowgrid Launch Error", body)


def _fatal_launch_error(code: str, summary: str, details: str = "") -> None:
    _notify_launch_error(code, summary, details)
    raise SystemExit(f"[{code}] {summary}")


def _notify_data_root_fallback_once() -> None:
    global _DATA_ROOT_FALLBACK_NOTIFIED
    if _DATA_ROOT_FALLBACK_NOTIFIED or not _DATA_ROOT_FALLBACK_DETAILS:
        return
    _DATA_ROOT_FALLBACK_NOTIFIED = True
    _notify_launch_error(
        "TH-1201",
        "Shared data path unavailable. Using a local fallback path.",
        _DATA_ROOT_FALLBACK_DETAILS,
    )


def _safe_print(message: str = "", end: str = "\n") -> None:
    try:
        print(message, end=end)
    except Exception as exc:
        _runtime_log_event(
            "runtime.safe_print_failed",
            severity="warning",
            summary="Console print failed in safe print helper.",
            exc=exc,
            context={"message_preview": str(message)[:240]},
        )


def detect_current_user_id() -> str:
    candidates = [
        os.environ.get("USERNAME", ""),
        os.environ.get("USER", ""),
        os.environ.get("LOGNAME", ""),
    ]
    try:
        candidates.append(getpass.getuser() or "")
    except Exception as exc:
        _runtime_log_event(
            "bootstrap.detect_user_getpass_failed",
            severity="warning",
            summary="getpass.getuser() failed while detecting current user ID.",
            exc=exc,
        )
    for raw in candidates:
        value = str(raw or "").strip()
        if value:
            return value.upper()
    return "UNKNOWN"


def _sanitize_log_filename_component(raw: str) -> str:
    text = str(raw or "").strip()
    if not text:
        return "UNKNOWN"
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", text)
    cleaned = cleaned.strip("._-")
    return cleaned or "UNKNOWN"


def _runtime_log_dir() -> Path:
    """Get runtime log directory. Logs stay LOCAL for reliability."""
    if _RESOLVED_DATA_ROOT is not None:
        return _RESOLVED_DATA_ROOT
    try:
        # Logs go to local folder for reliability and privacy
        target = _get_local_config_folder()
        target.mkdir(parents=True, exist_ok=True)
        return target
    except Exception as exc:
        _log_launch_error_once(
            "runtime.log_dir_primary_failed",
            "TH-9802",
            "Primary runtime log directory unavailable.",
            f"Path: {_get_local_config_folder()}\nReason: {type(exc).__name__}: {exc}",
        )
        try:
            fallback = _local_data_root()
            fallback.mkdir(parents=True, exist_ok=True)
            return fallback
        except Exception as fallback_exc:
            _log_launch_error_once(
                "runtime.log_dir_fallback_failed",
                "TH-9803",
                "Fallback runtime log directory unavailable.",
                f"Fallback path: {_local_data_root()}\nReason: {type(fallback_exc).__name__}: {fallback_exc}",
            )
            return Path.cwd()


def _runtime_log_path(user_id: str | None = None) -> Path:
    normalized_user = _sanitize_log_filename_component(user_id or detect_current_user_id())
    filename = f"{RUNTIME_LOG_FILENAME_PREFIX}_{normalized_user}{RUNTIME_LOG_FILENAME_SUFFIX}"
    return _runtime_log_dir() / filename


def _json_safe(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, dict):
        return {str(k): _json_safe(v) for k, v in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [_json_safe(v) for v in value]
    return str(value)


def _brief_runtime_context(context: dict[str, Any] | None) -> str:
    if not context:
        return ""
    parts: list[str] = []
    for key, value in list(context.items())[:6]:
        rendered = str(_json_safe(value))
        if len(rendered) > 120:
            rendered = rendered[:117] + "..."
        parts.append(f"{key}={rendered}")
    return "; ".join(parts)


def _rotate_runtime_log(path: Path) -> None:
    if not path.exists():
        return
    try:
        if path.stat().st_size < RUNTIME_LOG_MAX_BYTES:
            return
    except Exception as exc:
        _log_launch_error_once(
            "runtime.log_rotate_stat_failed",
            "TH-9804",
            "Runtime log rotation stat check failed.",
            f"Path: {path}\nReason: {type(exc).__name__}: {exc}",
        )
        return

    oldest = path.with_name(f"{path.name}.{RUNTIME_LOG_MAX_BACKUPS}")
    if oldest.exists():
        try:
            oldest.unlink()
        except Exception as exc:
            _log_launch_error_once(
                "runtime.log_rotate_delete_oldest_failed",
                "TH-9805",
                "Runtime log rotation failed deleting oldest backup.",
                f"Path: {oldest}\nReason: {type(exc).__name__}: {exc}",
            )

    for idx in range(RUNTIME_LOG_MAX_BACKUPS - 1, 0, -1):
        src = path.with_name(f"{path.name}.{idx}")
        if not src.exists():
            continue
        dst = path.with_name(f"{path.name}.{idx + 1}")
        try:
            src.replace(dst)
        except Exception as exc:
            _log_launch_error_once(
                "runtime.log_rotate_shift_failed",
                "TH-9806",
                "Runtime log rotation failed while shifting backups.",
                f"Source: {src}\nTarget: {dst}\nReason: {type(exc).__name__}: {exc}",
            )

    first = path.with_name(f"{path.name}.1")
    try:
        path.replace(first)
    except Exception as exc:
        _log_launch_error_once(
            "runtime.log_rotate_final_move_failed",
            "TH-9807",
            "Runtime log rotation failed moving active log to first backup.",
            f"Source: {path}\nTarget: {first}\nReason: {type(exc).__name__}: {exc}",
        )


def _write_runtime_log_entry(path: Path, entry: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    _rotate_runtime_log(path)
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(entry, ensure_ascii=False) + "\n")


def _runtime_log_event(
    event_key: str,
    severity: str = "warning",
    *,
    summary: str = "",
    exc: BaseException | None = None,
    context: dict[str, Any] | None = None,
) -> Path | None:
    global _RUNTIME_LOG_WRITE_IN_PROGRESS
    if _RUNTIME_LOG_WRITE_IN_PROGRESS:
        return None

    _RUNTIME_LOG_WRITE_IN_PROGRESS = True
    try:
        user_id = detect_current_user_id()
        payload: dict[str, Any] = {
            "timestamp_utc": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.%fZ"),
            "event_key": str(event_key),
            "severity": str(severity or "warning"),
            "user_id": str(user_id),
            "pid": int(os.getpid()),
            "host": str(socket.gethostname() or ""),
            "exception_type": "",
            "exception_message": "",
            "traceback": "",
        }
        if summary:
            payload["summary"] = str(summary)
        if exc is not None:
            payload["exception_type"] = type(exc).__name__
            payload["exception_message"] = str(exc)
            payload["traceback"] = "".join(traceback.format_exception(type(exc), exc, exc.__traceback__))
        if context:
            payload["context"] = _json_safe(context)
        log_path = _runtime_log_path(user_id)
        _write_runtime_log_entry(log_path, payload)
        return log_path
    except Exception as log_exc:
        try:
            _log_launch_error(
                "TH-9801",
                "Runtime log write failed.",
                f"Event key: {event_key}\nReason: {type(log_exc).__name__}: {log_exc}",
            )
        except Exception:
            try:
                _safe_print(f"[Flowgrid] Runtime log write failed and launch log fallback also failed: {log_exc}")
            except Exception as print_exc:
                try:
                    sys.stderr.write(
                        "[Flowgrid] Runtime logging failure fallback also failed: "
                        f"{type(print_exc).__name__}: {print_exc}\n"
                    )
                except Exception:
                    return None
        return None
    finally:
        _RUNTIME_LOG_WRITE_IN_PROGRESS = False


def _escalate_runtime_issue_once(
    event_key: str,
    summary: str,
    *,
    details: str = "",
    context: dict[str, Any] | None = None,
) -> None:
    if event_key in _RUNTIME_ESCALATED_EVENTS:
        return
    _RUNTIME_ESCALATED_EVENTS.add(event_key)

    clipped_details = str(details or "").strip()
    if len(clipped_details) > 1200:
        clipped_details = clipped_details[:1197] + "..."
    context_line = _brief_runtime_context(context)
    message_lines: list[str] = [str(summary)]
    if clipped_details:
        message_lines.extend(["", f"Details: {clipped_details}"])
    if context_line:
        message_lines.extend(["", f"Context: {context_line}"])
    message_lines.extend(["", f"Runtime log: {_runtime_log_path()}"])
    _show_error_dialog(RUNTIME_PROMPT_TITLE, "\n".join(message_lines))


def _unhandled_exception_hook(exc_type, exc_value, exc_tb) -> None:
    details = "".join(traceback.format_exception(exc_type, exc_value, exc_tb))
    _runtime_log_event(
        "bootstrap.unhandled_exception",
        severity="error",
        summary=f"Unhandled exception: {getattr(exc_type, '__name__', 'Exception')}",
        exc=exc_value,
        context={},
    )
    _notify_launch_error(
        "TH-9001",
        f"Unhandled exception: {getattr(exc_type, '__name__', 'Exception')}",
        details,
    )
    try:
        sys.__excepthook__(exc_type, exc_value, exc_tb)
    except Exception as exc:
        _runtime_log_event(
            "bootstrap.sys_excepthook_forward_failed",
            severity="warning",
            summary="Forwarding unhandled exception to sys.__excepthook__ failed.",
            exc=exc,
            context={"original_exception_type": str(getattr(exc_type, "__name__", "Exception"))},
        )


sys.excepthook = _unhandled_exception_hook


def _check_python_version() -> None:
    """Check if Python version meets minimum requirements."""
    MIN_PYTHON_VERSION = (3, 8, 0)
    current_version = sys.version_info[:3]

    if current_version < MIN_PYTHON_VERSION:
        version_str = ".".join(map(str, current_version))
        min_version_str = ".".join(map(str, MIN_PYTHON_VERSION))
        _fatal_launch_error(
            "TH-1001",
            f"Python {min_version_str}+ is required.",
            f"Python {min_version_str} or higher is required.\n"
            f"Current version: {version_str}\n"
            "Please upgrade Python from https://python.org",
        )


DEPENDENCY_SPECS: tuple[tuple[str, str, str, bool], ...] = (
    ("PySide6", "PySide6", "Qt GUI framework", True),
    ("openpyxl", "openpyxl", "Excel workbook import support", False),
)


def _dependency_specs_from_env() -> list[tuple[str, str, str, bool]]:
    specs: list[tuple[str, str, str, bool]] = []

    def parse_list(raw: str, required: bool) -> None:
        for token in str(raw or "").replace(";", ",").split(","):
            name = token.strip()
            if not name:
                continue
            module_name = name.replace("-", "_")
            desc = f"Extra {'required' if required else 'optional'} dependency"
            specs.append((name, module_name, desc, required))

    parse_list(os.environ.get("QI_EXTRA_PACKAGES", ""), False)
    parse_list(os.environ.get("QI_EXTRA_REQUIRED_PACKAGES", ""), True)
    return specs


def _module_import_status(module_name: str) -> tuple[bool, str]:
    try:
        importlib.import_module(module_name)
        return True, ""
    except Exception as exc:
        return False, f"{type(exc).__name__}: {exc}"


def _format_pip_failure(stderr_text: str, stdout_text: str) -> str:
    stderr_clean = (stderr_text or "").strip()
    stdout_clean = (stdout_text or "").strip()
    if stderr_clean:
        return stderr_clean[-2000:]
    if stdout_clean:
        return stdout_clean[-2000:]
    return "No pip output captured."


def _install_package(package_name: str, description: str = "") -> tuple[bool, str]:
    """Attempt to install a Python package using pip."""
    try:
        _safe_print(f"Installing {package_name}...{f' ({description})' if description else ''}")
        result = subprocess.run(
            [sys.executable, "-m", "pip", "--disable-pip-version-check", "install", package_name],
            capture_output=True,
            text=True,
            timeout=300  # 5 minute timeout
        )
        if result.returncode == 0:
            _safe_print(f"[OK] Successfully installed {package_name}")
            return True, ""
        else:
            _safe_print(f"[FAIL] Failed to install {package_name}")
            detail = _format_pip_failure(result.stderr, result.stdout)
            _safe_print(f"Error: {detail}")
            return False, detail
    except subprocess.TimeoutExpired:
        _safe_print(f"[TIMEOUT] Installation of {package_name} timed out")
        return False, "pip install timed out after 300 seconds."
    except Exception as e:
        _safe_print(f"[ERROR] Error installing {package_name}: {e}")
        return False, f"{type(e).__name__}: {e}"


def _ensure_dependencies() -> None:
    """Check dependencies and auto-install missing ones when possible."""
    missing_specs: list[dict[str, Any]] = []
    all_specs = list(DEPENDENCY_SPECS) + _dependency_specs_from_env()
    for package_name, module_name, description, required in all_specs:
        ok, reason = _module_import_status(module_name)
        if ok:
            continue
        missing_specs.append(
            {
                "package_name": package_name,
                "module_name": module_name,
                "description": description,
                "required": required,
                "reason": reason,
            }
        )

    if not missing_specs:
        return

    _safe_print("Missing required packages detected:")
    for item in missing_specs:
        _safe_print(f"  - {item['package_name']}: {item['description']} ({item['reason']})")

    auto_install_disabled = os.environ.get("QI_AUTO_INSTALL") == "0"
    if auto_install_disabled:
        required_missing = [item for item in missing_specs if bool(item["required"])]
        if required_missing:
            details = [
                f"Interpreter: {sys.executable}",
                "Required dependencies missing and QI_AUTO_INSTALL=0.",
            ]
            for item in required_missing:
                details.append(f"- {item['package_name']} ({item['module_name']}): {item['reason']}")
            _fatal_launch_error(
                "TH-1102",
                "Required dependencies missing and automatic installation is disabled.",
                "\n".join(details),
            )
        return

    _safe_print("\nAttempting automatic installation...")
    for item in missing_specs:
        package_name = str(item["package_name"])
        module_name = str(item["module_name"])
        description = str(item["description"])
        required = bool(item["required"])

        installed, install_detail = _install_package(package_name, description)
        if not installed:
            if required:
                details = (
                    f"Interpreter: {sys.executable}\n"
                    f"Package: {package_name}\n"
                    f"Module: {module_name}\n"
                    f"Failure: {install_detail}\n"
                    "See the launch/runtime log files for diagnostics."
                )
                _fatal_launch_error(
                    "TH-1101",
                    f"Required dependency installation failed: {package_name}.",
                    details,
                )
            _safe_print(f"Warning: {package_name} installation failed. Some features may be disabled.")
            continue

        import_ok, import_reason = _module_import_status(module_name)
        if import_ok:
            continue
        if required:
            details = (
                f"Interpreter: {sys.executable}\n"
                f"Package installed but module import still fails.\n"
                f"Package: {package_name}\n"
                f"Module: {module_name}\n"
                f"Import error: {import_reason}"
            )
            _fatal_launch_error(
                "TH-1107",
                f"Required dependency import failed after install: {module_name}.",
                details,
            )
        _safe_print(f"Warning: {module_name} still cannot be imported after install: {import_reason}")


try:
    import openpyxl
except Exception as exc:
    openpyxl = None  # Workbook import will be disabled if not installed
    _runtime_log_event(
        "bootstrap.openpyxl_import_failed",
        severity="warning",
        summary="openpyxl import failed; workbook import will remain disabled.",
        exc=exc,
    )


def _hide_console_window() -> None:
    if os.name != "nt" or os.environ.get("QI_KEEP_CONSOLE") == "1":
        return
    try:
        kernel32 = ctypes.WinDLL("kernel32", use_last_error=True)
        user32_local = ctypes.WinDLL("user32", use_last_error=True)
        hwnd = kernel32.GetConsoleWindow()
        if hwnd:
            user32_local.ShowWindow(hwnd, 0)  # SW_HIDE
            kernel32.FreeConsole()
    except Exception as exc:
        _runtime_log_event(
            "bootstrap.hide_console_failed",
            severity="warning",
            summary="Unable to hide or detach console window.",
            exc=exc,
        )


def _ensure_pyside6() -> None:
    try:
        importlib.import_module("PySide6")
        return
    except Exception as exc:
        _fatal_launch_error(
            "TH-1106",
            "PySide6 import failed.",
            f"Interpreter: {sys.executable}\n"
            f"Import error: {type(exc).__name__}: {exc}\n"
            "See the launch/runtime log files for diagnostics.",
        )


# Run checks in order
try:
    _resolve_data_root()
    _check_python_version()
    if not _INSTALLER_FLAGS_ACTIVE:
        _hide_console_window()
    _ensure_dependencies()
    _ensure_pyside6()
    _notify_data_root_fallback_once()
except SystemExit:
    raise
except Exception as exc:
    _fatal_launch_error("TH-1900", "Unexpected startup initialization failure.", repr(exc))

from PySide6.QtCore import (
    QByteArray,
    QBuffer,
    QDate,
    QEvent,
    QEasingCurve,
    QIODevice,
    QPoint,
    QPointF,
    QRect,
    QRectF,
    QSize,
    Qt,
    QProcess,
    QTimer,
    QUrl,
    Signal,
    QVariantAnimation,
)
from PySide6.QtGui import (
    QColor,
    QCursor,
    QDesktopServices,
    QFont,
    QGuiApplication,
    QIcon,
    QImage,
    QImageReader,
    QMouseEvent,
    QPainter,
    QPainterPath,
    QPen,
    QPolygon,
    QPixmap,
    QPalette,
    QRegion,
    QTextCursor,
)
from PySide6.QtWidgets import (
    QAbstractButton,
    QAbstractItemView,
    QApplication,
    QCheckBox,
    QColorDialog,
    QComboBox,
    QDateEdit,
    QDialog,
    QFileDialog,
    QFontComboBox,
    QFormLayout,
    QFrame,
    QGraphicsDropShadowEffect,
    QGraphicsOpacityEffect,
    QGridLayout,
    QHeaderView,
    QHBoxLayout,
    QInputDialog,
    QLabel,
    QLineEdit,
    QListWidget,
    QTabWidget,
    QMessageBox,
    QListWidgetItem,
    QMainWindow,
    QPushButton,
    QScrollArea,
    QScrollBar,
    QSizePolicy,
    QSlider,
    QSpinBox,
    QStackedWidget,
    QStyle,
    QStyleOptionTab,
    QStylePainter,
    QTableWidget,
    QTableWidgetItem,
    QTabBar,
    QTextEdit,
    QToolButton,
    QVBoxLayout,
    QWidget,
)

APP_TITLE = "Flowgrid"
CONFIG_FILENAME = "Flowgrid_config.json"
DEPOT_DB_FILENAME = "Flowgrid_depot.db"
QA_FLAG_OPTIONS: tuple[str, ...] = (
    "None",
    "Follow Up",
    "Need Parts",
    "Escalation",
    "Client Callback",
    "Return Visit",
    "Safety",
    "Other",
)
QA_FLAG_SEVERITY_OPTIONS: tuple[str, ...] = ("Low", "Medium", "High", "Critical")
LAUNCH_WIDTH = 430
LAUNCH_HEIGHT = 485
TITLEBAR_HEIGHT = 34
SIDEBAR_WIDTH = 52
SHIFT_CONTEXT_SCRIPT_LAUNCHERS: dict[str, str] = {
}

WORKBOOK_IMPORT_SPECS: tuple[tuple[str, str, str], ...] = (
    ("submissions", "Table1", "Submissions"),
    ("parts", "tblParts", "Parts"),
    ("rtvs", "RTVs", "RTVs"),
    ("client_jo", "Client_JO", "Client JO"),
    ("client_parts", "client_parts", "Client Parts"),
    ("agents", "agents", "Agents"),
    ("qa_flags", "qa_flags", "QA Flags"),
    ("admin_users", "admin", "Admin Users"),
)

TRACKER_DASHBOARD_TABLES: tuple[tuple[str, str], ...] = (
    ("submissions", "Submissions"),
    ("parts", "Parts"),
    ("rtvs", "RTVs"),
    ("client_jo", "Client JO"),
    ("client_parts", "Client Parts"),
)


DEFAULT_THEME_PRESETS: dict[str, dict[str, str]] = {
    "Default": {"primary": "#C35A00", "accent": "#FF9A1F", "surface": "#090A0F"},
    "Classic": {"primary": "#0A246A", "accent": "#C0C0C0", "surface": "#D4D0C8"},
    "Slate": {"primary": "#3A4A6A", "accent": "#D97706", "surface": "#E8ECF3"},
    "Forest": {"primary": "#205E55", "accent": "#D66A1A", "surface": "#E9F1ED"},
    "Ocean": {"primary": "#15D3E3", "accent": "#D1A91F", "surface": "#70D7E9"},
    "Midnight": {"primary": "#1E2B3A", "accent": "#4DA3FF", "surface": "#0F141C"},
    "Desert": {"primary": "#7A4A2A", "accent": "#D9A25E", "surface": "#F1E5D6"},
    "Sage": {"primary": "#2F5D50", "accent": "#9DC66B", "surface": "#E8F0E6"},
    "Crimson": {"primary": "#6A1E1E", "accent": "#D85A5A", "surface": "#F1E6E6"},
    "Steel": {"primary": "#3C4B5C", "accent": "#8FA7BF", "surface": "#E3EAF2"},
    "Amber": {"primary": "#70420C", "accent": "#F3B33E", "surface": "#F6EDD9"},
}

DEFAULT_THEME_PRIMARY = "#C35A00"
DEFAULT_THEME_ACCENT = "#FF9A1F"
DEFAULT_THEME_SURFACE = "#090A0F"
LEGACY_DEFAULT_THEME_PRIMARY = "#2F6FED"
LEGACY_DEFAULT_THEME_ACCENT = "#16A085"
LEGACY_DEFAULT_THEME_SURFACE = "#E9EEF5"

DEFAULT_CONFIG: dict[str, Any] = {
    "grid_columns": 3,
    "always_on_top": False,
    "agent_window_always_on_top": True,
    "qa_window_always_on_top": True,
    "sidebar_on_right": False,
    "auto_minimize_after_insert": False,
    "compact_mode": True,
    "background_tint_enabled": True,
    "window_opacity": 1.0,
    "hover_reveal_delay_s": 5,
    "hover_fade_in_s": 5,
    "hover_fade_out_s": 5,
    "popup_control_style": "Fade Left to Right",
    "popup_control_opacity": 82,
    "popup_control_tail_opacity": 0,
    "popup_control_fade_enabled": True,
    "popup_control_fade_strength": 65,
    "popup_header_color": "",
    "popup_row_hover_color": "",
    "popup_row_selected_color": "",
    "popup_auto_reinherit_enabled": True,
    "quick_button_opacity": 1.0,
    "window_position": None,
    "popup_positions": {"image_layers": None, "quick_layout": None, "depot_dashboard": None},
    "theme": {"primary": DEFAULT_THEME_PRIMARY, "accent": DEFAULT_THEME_ACCENT, "surface": DEFAULT_THEME_SURFACE},
    "theme_presets": DEFAULT_THEME_PRESETS,
    "selected_theme_preset": "Default",
    "theme_image_layers": [],
    "quick_button_width": 140,
    "quick_button_height": 40,
    "quick_button_font_size": 11,
    "quick_button_font_family": "Segoe UI",
    "quick_button_shape": "Soft",
    "active_quick_tab": 0,
    "current_user": "",
    "agent_theme": {
        "background": "#FFFFFF",
        "text": "#000000",
        "field_bg": "#FFFFFF",
        "transparent": False,
        "inherit_main_theme": True,
        "image_layers": [],
        "control_style": "Fade Left to Right",
        "control_opacity": 82,
        "control_tail_opacity": 0,
        "control_fade_strength": 65,
        "header_color": "",
        "row_hover_color": "",
        "row_selected_color": "",
    },
    "qa_theme": {
        "background": "#FFFFFF",
        "text": "#000000",
        "field_bg": "#FFFFFF",
        "transparent": False,
        "inherit_main_theme": True,
        "image_layers": [],
        "control_style": "Fade Left to Right",
        "control_opacity": 82,
        "control_tail_opacity": 0,
        "control_fade_strength": 65,
        "header_color": "",
        "row_hover_color": "",
        "row_selected_color": "",
    },
    "admin_theme": {
        "background": "#FFFFFF",
        "text": "#000000",
        "field_bg": "#FFFFFF",
        "transparent": False,
        "inherit_main_theme": True,
        "image_layers": [],
        "control_style": "Fade Left to Right",
        "control_opacity": 82,
        "control_tail_opacity": 0,
        "control_fade_strength": 65,
        "header_color": "",
        "row_hover_color": "",
        "row_selected_color": "",
    },
    "dashboard_theme": {
        "background": "#FFFFFF",
        "text": "#000000",
        "field_bg": "#FFFFFF",
        "transparent": False,
        "inherit_main_theme": True,
        "image_layers": [],
        "control_style": "Fade Left to Right",
        "control_opacity": 82,
        "control_tail_opacity": 0,
        "control_fade_strength": 65,
        "header_color": "",
        "row_hover_color": "",
        "row_selected_color": "",
    },
    "app_icon_path": "",
    "quick_texts": [
        {
            "title": "Greeting",
            "tooltip": "Quick opening line",
            "text": "Hi there,",
            "action": "paste_text",
            "open_target": "",
            "app_targets": "",
            "urls": "",
            "browser_path": "",
        },
        {
            "title": "Follow-up",
            "tooltip": "Ask for updates",
            "text": "Checking in on this when you have a moment.",
            "action": "paste_text",
            "open_target": "",
            "app_targets": "",
            "urls": "",
            "browser_path": "",
        },
    ],
}

DEFAULT_WINDOW_ICON_FILENAME = "wrench.png"
MANAGED_SHORTCUT_ICON_FILENAME = "Flowgrid_shortcut.ico"
DESKTOP_SHORTCUT_FILENAME = f"{APP_TITLE}.lnk"
WINDOWS_SHORTCUT_DESCRIPTION = "Launch Flowgrid"


def _flowgrid_script_path() -> Path:
    try:
        return Path(__file__).resolve()
    except Exception as exc:
        fallback = Path.cwd() / "Flowgrid.pyw"
        _runtime_log_event(
            "installer.script_path_resolve_failed",
            severity="warning",
            summary="Failed resolving Flowgrid script path; using current working directory fallback.",
            exc=exc,
            context={"fallback_path": str(fallback)},
        )
        return fallback


def _load_installer_config_snapshot() -> dict[str, Any]:
    config_path = _data_file_path(CONFIG_FILENAME)
    if not config_path.exists():
        return {}
    try:
        data = json.loads(config_path.read_text(encoding="utf-8"))
    except Exception as exc:
        _runtime_log_event(
            "installer.config_snapshot_parse_failed",
            severity="warning",
            summary="Failed parsing config while preparing installer icon state; default icon will be used.",
            exc=exc,
            context={"config_path": str(config_path)},
        )
        return {}
    if isinstance(data, dict):
        return data
    _runtime_log_event(
        "installer.config_snapshot_invalid",
        severity="warning",
        summary="Config snapshot was not a JSON object; default icon will be used for shortcut sync.",
        context={"config_path": str(config_path), "value_type": type(data).__name__},
    )
    return {}


def _resolve_existing_file_path(raw_path: str) -> Path | None:
    expanded = os.path.expandvars(os.path.expanduser(str(raw_path or "").strip()))
    if not expanded:
        return None

    base_candidate = Path(expanded)
    candidates: list[Path] = [base_candidate]
    if not base_candidate.is_absolute():
        candidates.extend((_resolve_data_root() / base_candidate, _local_data_root() / base_candidate))

    unique: list[Path] = []
    for candidate in candidates:
        if any(_paths_equal(candidate, existing) for existing in unique):
            continue
        unique.append(candidate)

    for candidate in unique:
        try:
            if candidate.exists() and candidate.is_file():
                return candidate
        except Exception as exc:
            _runtime_log_event(
                "installer.icon_candidate_stat_failed",
                severity="warning",
                summary="Failed checking an icon path candidate while resolving installer icon state.",
                exc=exc,
                context={"candidate": str(candidate)},
            )
    return None


def _flowgrid_icon_pack_dir() -> Path:
    return _resolve_data_root() / ASSETS_DIR_NAME / FLOWGRID_ICON_PACK_DIR_NAME


def _ensure_flowgrid_icon_pack_dir() -> Path:
    icon_dir = _flowgrid_icon_pack_dir()
    try:
        icon_dir.mkdir(parents=True, exist_ok=True)
    except Exception as exc:
        _runtime_log_event(
            "installer.icon_pack_dir_create_failed",
            severity="warning",
            summary="Failed creating Flowgrid icon pack directory.",
            exc=exc,
            context={"icon_dir": str(icon_dir)},
        )
    return icon_dir


def _default_wrench_icon_source_path() -> Path | None:
    target = _ensure_flowgrid_icon_pack_dir() / DEFAULT_WINDOW_ICON_FILENAME
    if target.exists() and target.is_file():
        return target

    candidate_dirs = [
        _local_data_root() / ASSETS_DIR_NAME / FLOWGRID_ICON_PACK_DIR_NAME,
        _resolve_data_root() / ASSETS_DIR_NAME / FLOWGRID_ICON_PACK_DIR_NAME,
        _local_data_root() / "ui_icons",
        _resolve_data_root() / "ui_icons",
    ]

    seen: list[Path] = []
    for directory in candidate_dirs:
        if any(_paths_equal(directory, existing) for existing in seen):
            continue
        seen.append(directory)
        candidate = directory / DEFAULT_WINDOW_ICON_FILENAME
        if not candidate.exists() or not candidate.is_file():
            continue
        if _paths_equal(candidate, target):
            return candidate
        try:
            target.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(candidate, target)
            return target
        except Exception as exc:
            _runtime_log_event(
                "installer.default_icon_copy_failed",
                severity="warning",
                summary="Failed copying the default wrench icon into the managed icon pack directory.",
                exc=exc,
                context={"source_path": str(candidate), "target_path": str(target)},
            )
            return candidate
    return None


def _resolve_active_app_icon_path(config: dict[str, Any] | None = None) -> Path | None:
    config_data = config if isinstance(config, dict) else _load_installer_config_snapshot()
    stored = str(config_data.get("app_icon_path", "") or "").strip() if isinstance(config_data, dict) else ""
    custom_icon = _resolve_existing_file_path(stored)
    if custom_icon is not None:
        return custom_icon
    return _default_wrench_icon_source_path()


def _load_icon_image_file(icon_path: str | Path) -> QImage:
    resolved = _resolve_existing_file_path(str(icon_path))
    if resolved is None:
        return QImage()

    reader = QImageReader(str(resolved))
    reader.setAutoTransform(True)
    image = reader.read()
    if image.isNull():
        return QImage()

    max_dim = max(image.width(), image.height())
    if max_dim > 512:
        image = image.scaled(
            512,
            512,
            Qt.AspectRatioMode.KeepAspectRatio,
            Qt.TransformationMode.SmoothTransformation,
        )
    return image.convertToFormat(QImage.Format.Format_ARGB32)


def _is_image_mostly_opaque(image: QImage) -> bool:
    width = image.width()
    height = image.height()
    if width <= 0 or height <= 0:
        return False

    step = max(1, min(width, height) // 64)
    total = 0
    opaque = 0
    for y in range(0, height, step):
        for x in range(0, width, step):
            total += 1
            if image.pixelColor(x, y).alpha() >= 250:
                opaque += 1

    if total == 0:
        return False
    return (opaque / total) >= 0.96


def _estimate_icon_corner_matte(image: QImage) -> QColor:
    width = image.width()
    height = image.height()
    points = [
        (0, 0),
        (min(width - 1, 1), 0),
        (0, min(height - 1, 1)),
        (width - 1, 0),
        (width - 1, min(height - 1, 1)),
        (max(0, width - 2), 0),
        (0, height - 1),
        (min(width - 1, 1), height - 1),
        (0, max(0, height - 2)),
        (width - 1, height - 1),
        (max(0, width - 2), height - 1),
        (width - 1, max(0, height - 2)),
    ]
    rs = 0
    gs = 0
    bs = 0
    count = 0
    for x, y in points:
        color = image.pixelColor(x, y)
        rs += color.red()
        gs += color.green()
        bs += color.blue()
        count += 1
    if count == 0:
        return QColor(0, 0, 0)
    return QColor(rs // count, gs // count, bs // count)


def _cleanup_icon_transparency_image(image: QImage) -> QImage:
    if image.isNull() or not _is_image_mostly_opaque(image):
        return image

    matte = _estimate_icon_corner_matte(image)
    hard = 24
    soft = 72
    cleaned = QImage(image)

    for y in range(cleaned.height()):
        for x in range(cleaned.width()):
            color = cleaned.pixelColor(x, y)
            dist = (
                abs(color.red() - matte.red())
                + abs(color.green() - matte.green())
                + abs(color.blue() - matte.blue())
            )
            alpha = color.alpha()
            if dist <= hard:
                cleaned.setPixelColor(x, y, QColor(color.red(), color.green(), color.blue(), 0))
            elif dist < soft:
                ratio = (dist - hard) / float(soft - hard)
                cleaned.setPixelColor(
                    x,
                    y,
                    QColor(color.red(), color.green(), color.blue(), int(alpha * ratio)),
                )
    return cleaned


def _build_smoothed_qicon(icon_path: str | Path) -> QIcon:
    image = _load_icon_image_file(icon_path)
    if image.isNull():
        return QIcon()

    cleaned = _cleanup_icon_transparency_image(image)
    icon = QIcon()
    for size in (16, 20, 24, 32, 40, 48, 64, 96, 128, 256):
        scaled = cleaned.scaled(
            size,
            size,
            Qt.AspectRatioMode.KeepAspectRatio,
            Qt.TransformationMode.SmoothTransformation,
        )
        canvas = QPixmap(size, size)
        canvas.fill(Qt.GlobalColor.transparent)
        painter = QPainter(canvas)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)
        painter.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform, True)
        x = (size - scaled.width()) // 2
        y = (size - scaled.height()) // 2
        painter.drawImage(x, y, scaled)
        painter.end()
        icon.addPixmap(canvas)
    return icon


def _normalized_icon_export_image(image: QImage, size: int = 256) -> QImage:
    canvas = QImage(size, size, QImage.Format.Format_ARGB32)
    canvas.fill(Qt.GlobalColor.transparent)
    if image.isNull():
        return canvas
    scaled = image.scaled(
        size,
        size,
        Qt.AspectRatioMode.KeepAspectRatio,
        Qt.TransformationMode.SmoothTransformation,
    )
    painter = QPainter(canvas)
    painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)
    painter.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform, True)
    x = (size - scaled.width()) // 2
    y = (size - scaled.height()) // 2
    painter.drawImage(x, y, scaled)
    painter.end()
    return canvas


def _qimage_to_png_bytes(image: QImage) -> bytes:
    buffer_bytes = QByteArray()
    buffer = QBuffer(buffer_bytes)
    if not buffer.open(QIODevice.OpenModeFlag.WriteOnly):
        return b""
    try:
        if not image.save(buffer, "PNG"):
            return b""
    finally:
        buffer.close()
    return bytes(buffer_bytes)


def _png_dimensions(png_bytes: bytes) -> tuple[int, int]:
    if len(png_bytes) < 24 or png_bytes[:8] != b"\x89PNG\r\n\x1a\n" or png_bytes[12:16] != b"IHDR":
        raise ValueError("PNG byte stream missing a valid IHDR header.")
    width, height = struct.unpack(">II", png_bytes[16:24])
    return int(width), int(height)


def _write_png_bytes_as_ico(png_bytes: bytes, target_path: Path) -> None:
    width, height = _png_dimensions(png_bytes)
    directory_entry = struct.pack(
        "<BBBBHHII",
        0 if width >= 256 else width,
        0 if height >= 256 else height,
        0,
        0,
        1,
        32,
        len(png_bytes),
        6 + 16,
    )
    payload = struct.pack("<HHH", 0, 1, 1) + directory_entry + png_bytes

    target_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = target_path.with_name(f"{target_path.name}.tmp")
    temp_path.write_bytes(payload)
    os.replace(temp_path, target_path)


def _write_managed_shortcut_icon(source_path: str | Path, target_path: Path) -> Path:
    image = _load_icon_image_file(source_path)
    if image.isNull():
        raise ValueError(f"Unable to decode icon source: {source_path}")

    cleaned = _cleanup_icon_transparency_image(image)
    export_image = _normalized_icon_export_image(cleaned, 256)
    png_bytes = _qimage_to_png_bytes(export_image)
    if not png_bytes:
        raise ValueError(f"Unable to encode icon source as PNG: {source_path}")

    _write_png_bytes_as_ico(png_bytes, target_path)
    return target_path


def _preferred_gui_python_executable() -> Path:
    candidates: list[Path] = []
    for raw in (getattr(sys, "_base_executable", ""), sys.executable):
        text = str(raw or "").strip()
        if not text:
            continue
        path = Path(text)
        candidates.append(path)
        candidates.append(path.parent / "pythonw.exe")
        if path.name.lower() == "python.exe":
            candidates.append(path.with_name("pythonw.exe"))

    unique: list[Path] = []
    for candidate in candidates:
        if any(_paths_equal(candidate, existing) for existing in unique):
            continue
        unique.append(candidate)

    for candidate in unique:
        if candidate.name.lower() == "pythonw.exe" and candidate.exists() and candidate.is_file():
            return candidate
    for candidate in unique:
        if candidate.exists() and candidate.is_file():
            return candidate
    return Path(sys.executable)


def _resolve_windows_desktop_directory() -> Path | None:
    if os.name != "nt":
        return None

    try:
        buffer = ctypes.create_unicode_buffer(260)
        result = ctypes.windll.shell32.SHGetFolderPathW(None, 0x0010, None, 0, buffer)
        if result == 0 and str(buffer.value).strip():
            path = Path(str(buffer.value).strip())
            if path.exists() and path.is_dir():
                return path
    except Exception as exc:
        _runtime_log_event(
            "installer.desktop_path_native_resolve_failed",
            severity="warning",
            summary="Native desktop path lookup failed; falling back to environment-based detection.",
            exc=exc,
        )

    fallback_candidates = []
    onedrive = str(os.environ.get("OneDrive", "") or "").strip()
    if onedrive:
        fallback_candidates.append(Path(onedrive) / "Desktop")
    fallback_candidates.append(Path.home() / "Desktop")

    for candidate in fallback_candidates:
        try:
            if candidate.exists() and candidate.is_dir():
                return candidate
        except Exception as exc:
            _runtime_log_event(
                "installer.desktop_path_fallback_stat_failed",
                severity="warning",
                summary="Desktop fallback path check failed during installer preparation.",
                exc=exc,
                context={"candidate": str(candidate)},
            )
    return None


def _powershell_single_quote(value: str) -> str:
    return "'" + str(value).replace("'", "''") + "'"


def _create_or_update_windows_shortcut(
    shortcut_path: Path,
    target_path: Path,
    arguments: str,
    working_directory: Path,
    icon_path: Path,
    description: str,
) -> tuple[bool, str]:
    script = "\n".join(
        [
            "$ErrorActionPreference = 'Stop'",
            "$shell = New-Object -ComObject WScript.Shell",
            f"$shortcut = $shell.CreateShortcut({_powershell_single_quote(str(shortcut_path))})",
            f"$shortcut.TargetPath = {_powershell_single_quote(str(target_path))}",
            f"$shortcut.Arguments = {_powershell_single_quote(arguments)}",
            f"$shortcut.WorkingDirectory = {_powershell_single_quote(str(working_directory))}",
            f"$shortcut.IconLocation = {_powershell_single_quote(str(icon_path) + ',0')}",
            f"$shortcut.Description = {_powershell_single_quote(description)}",
            "$shortcut.WindowStyle = 1",
            "$shortcut.Save()",
        ]
    )

    try:
        result = subprocess.run(
            ["powershell", "-NoProfile", "-NonInteractive", "-ExecutionPolicy", "Bypass", "-Command", script],
            capture_output=True,
            text=True,
            timeout=60,
        )
    except Exception as exc:
        return False, f"{type(exc).__name__}: {exc}"

    if result.returncode == 0:
        return True, ""

    detail = _format_pip_failure(result.stderr, result.stdout)
    return False, detail


def _sync_desktop_shortcut(
    config: dict[str, Any] | None = None,
    *,
    create_if_missing: bool,
) -> tuple[str, str]:
    if os.name != "nt":
        return "skipped", "Desktop shortcut creation is only supported on Windows."

    desktop_dir = _resolve_windows_desktop_directory()
    if desktop_dir is None:
        detail = "Unable to locate the current user's Desktop folder."
        _runtime_log_event(
            "installer.desktop_path_unavailable",
            severity="error",
            summary="Desktop shortcut sync failed because the Desktop folder could not be located.",
            context={"app_title": APP_TITLE},
        )
        return "failed", detail

    shortcut_path = desktop_dir / DESKTOP_SHORTCUT_FILENAME
    already_exists = shortcut_path.exists()
    if not create_if_missing and not already_exists:
        return "missing", ""

    icon_source = _resolve_active_app_icon_path(config)
    if icon_source is None:
        detail = "Unable to find the default wrench icon or the configured custom icon."
        _runtime_log_event(
            "installer.icon_source_unavailable",
            severity="error",
            summary="Desktop shortcut sync failed because no usable icon source was available.",
            context={"shortcut_path": str(shortcut_path)},
        )
        return "failed", detail

    managed_icon_path = _ensure_flowgrid_icon_pack_dir() / MANAGED_SHORTCUT_ICON_FILENAME
    try:
        _write_managed_shortcut_icon(icon_source, managed_icon_path)
    except Exception as exc:
        detail = f"Failed preparing shortcut icon: {type(exc).__name__}: {exc}"
        _runtime_log_event(
            "installer.shortcut_icon_write_failed",
            severity="error",
            summary="Desktop shortcut sync failed while preparing the managed shortcut icon file.",
            exc=exc,
            context={
                "icon_source": str(icon_source),
                "managed_icon_path": str(managed_icon_path),
                "shortcut_path": str(shortcut_path),
            },
        )
        return "failed", detail

    launcher_path = _preferred_gui_python_executable()
    script_path = _flowgrid_script_path()
    if not launcher_path.exists() or not launcher_path.is_file():
        detail = f"Python launcher not found: {launcher_path}"
        _runtime_log_event(
            "installer.launcher_not_found",
            severity="error",
            summary="Desktop shortcut sync failed because the Python launcher executable could not be found.",
            context={"launcher_path": str(launcher_path), "shortcut_path": str(shortcut_path)},
        )
        return "failed", detail
    if not script_path.exists() or not script_path.is_file():
        detail = f"Flowgrid script not found: {script_path}"
        _runtime_log_event(
            "installer.script_not_found",
            severity="error",
            summary="Desktop shortcut sync failed because the Flowgrid script file could not be found.",
            context={"script_path": str(script_path), "shortcut_path": str(shortcut_path)},
        )
        return "failed", detail

    arguments = f'"{script_path}"'
    ok, detail = _create_or_update_windows_shortcut(
        shortcut_path,
        launcher_path,
        arguments,
        script_path.parent,
        managed_icon_path,
        WINDOWS_SHORTCUT_DESCRIPTION,
    )
    if not ok:
        _runtime_log_event(
            "installer.shortcut_save_failed",
            severity="error",
            summary="Desktop shortcut save failed.",
            context={
                "shortcut_path": str(shortcut_path),
                "launcher_path": str(launcher_path),
                "managed_icon_path": str(managed_icon_path),
                "detail": detail,
            },
        )
        return "failed", f"Failed saving desktop shortcut: {detail}"

    status = "updated" if already_exists else "created"
    return status, f"Desktop shortcut {status} at {shortcut_path}"


def _launch_flowgrid_detached() -> tuple[bool, str]:
    launcher_path = _preferred_gui_python_executable()
    script_path = _flowgrid_script_path()
    if not launcher_path.exists() or not launcher_path.is_file():
        return False, f"Python launcher not found: {launcher_path}"
    if not script_path.exists() or not script_path.is_file():
        return False, f"Flowgrid script not found: {script_path}"

    try:
        subprocess.Popen([str(launcher_path), str(script_path)], cwd=str(script_path.parent))
        return True, ""
    except Exception as exc:
        _runtime_log_event(
            "installer.launch_subprocess_failed",
            severity="error",
            summary="Installer failed to launch Flowgrid after creating the desktop shortcut.",
            exc=exc,
            context={"launcher_path": str(launcher_path), "script_path": str(script_path)},
        )
        return False, f"{type(exc).__name__}: {exc}"


def _run_installer_mode(*, launch_after_install: bool) -> int:
    _safe_print(f"{APP_TITLE} installer")
    _safe_print(f"Interpreter: {sys.executable}")
    _safe_print("Checking dependencies: complete")

    status, detail = _sync_desktop_shortcut(create_if_missing=True)
    if status == "failed":
        _notify_launch_error("TH-1301", "Flowgrid installer could not create the desktop shortcut.", detail)
        return 1

    if detail:
        _safe_print(detail)

    if not launch_after_install:
        return 0

    launched, launch_detail = _launch_flowgrid_detached()
    if not launched:
        _notify_launch_error("TH-1302", "Flowgrid installer could not launch the app.", launch_detail)
        return 1

    _safe_print("Flowgrid launched.")
    return 0


def _can_start_window_drag(root: QWidget, local_pos: QPoint) -> bool:
    blocked_types = (
        QLineEdit,
        QTextEdit,
        QAbstractButton,
        QComboBox,
        QSpinBox,
        QSlider,
        QListWidget,
        QAbstractItemView,
        QScrollArea,
        QScrollBar,
        QTabBar,
    )
    child = root.childAt(local_pos)
    while child is not None and child is not root:
        if isinstance(child, blocked_types):
            return False
        child = child.parentWidget()
    return True


def configure_standard_table(
    table: QTableWidget,
    headers: list[str] | tuple[str, ...],
    *,
    resize_modes: dict[int, QHeaderView.ResizeMode] | None = None,
    stretch_last: bool = True,
) -> None:
    table.setColumnCount(len(headers))
    table.setHorizontalHeaderLabels([str(h) for h in headers])
    table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
    table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
    table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
    table.verticalHeader().setVisible(False)
    table.setAlternatingRowColors(True)
    table.setShowGrid(True)
    table.setGridStyle(Qt.PenStyle.SolidLine)
    table.setWordWrap(False)
    table.setTextElideMode(Qt.TextElideMode.ElideRight)
    table.horizontalHeader().setStretchLastSection(bool(stretch_last))
    table.horizontalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
    if resize_modes:
        for col_idx, mode in resize_modes.items():
            if 0 <= int(col_idx) < table.columnCount():
                table.horizontalHeader().setSectionResizeMode(int(col_idx), mode)


# --- Shared UI layout (frameless depot tools vs standard framed dialogs) ---
UI_MARGIN_FRAMELESS_TOOL = 6
UI_SPACING_FRAMELESS_TOOL = 6
UI_MARGIN_STANDARD_DIALOG = 10
UI_SPACING_STANDARD_DIALOG = 8
HEADER_CLOSE_BUTTON_WIDTH = 26
HEADER_CLOSE_BUTTON_HEIGHT = 22


def show_flowgrid_themed_message(
    parent: QWidget,
    app_window: "QuickInputsWindow" | None,
    theme_kind: str,
    icon: QMessageBox.Icon,
    title: str,
    text: str,
) -> None:
    """Themed validation/confirmation popup that follows the active popup theme."""
    dialog = FlowgridThemedMessageDialog(parent, app_window, theme_kind, icon, title, text)
    dialog.exec()


def _resolve_flowgrid_popup_app_window(
    parent: QWidget | None,
    app_window: "QuickInputsWindow" | None,
) -> "QuickInputsWindow" | None:
    if app_window is not None:
        return app_window
    probe = parent
    while probe is not None:
        candidate = getattr(probe, "app_window", None)
        if candidate is not None:
            return candidate
        probe = probe.parentWidget()
    return None


def _apply_flowgrid_popup_stylesheet(
    widget: QWidget,
    app_window: "QuickInputsWindow" | None,
    theme_kind: str,
    *,
    force_opaque_root: bool = True,
    event_key: str = "ui.themed_popup_stylesheet_failed",
    title: str = "",
) -> None:
    if app_window is None:
        return
    try:
        widget.setStyleSheet(app_window._popup_theme_stylesheet(theme_kind, force_opaque_root=force_opaque_root))
    except Exception as exc:
        _runtime_log_event(
            event_key,
            severity="warning",
            summary="Failed applying themed popup stylesheet.",
            exc=exc,
            context={"theme_kind": str(theme_kind), "title": str(title or widget.windowTitle())},
        )


def _paint_flowgrid_popup_background(
    widget: QWidget,
    painter: QPainter,
    app_window: "QuickInputsWindow" | None,
    theme_kind: str,
) -> None:
    if app_window is None:
        return

    target_rect = widget.rect()
    if target_rect.width() <= 0 or target_rect.height() <= 0:
        return

    target_size = target_rect.size()
    parent_widget = widget.parentWidget()
    reference_size = target_size
    if parent_widget is not None:
        parent_size = parent_widget.size()
        if parent_size.width() > 0 and parent_size.height() > 0:
            reference_size = parent_size

    resolved_kind = str(theme_kind or "").strip() or "main"
    if resolved_kind == "main":
        base_bg = normalize_hex(
            app_window.palette_data.get("control_bg", app_window.palette_data.get("surface", DEFAULT_THEME_SURFACE)),
            DEFAULT_THEME_SURFACE,
        )
        transparent = bool(app_window.config.get("theme_page_transparent_primary_bg", False))
        bg = app_window.render_background_pixmap(reference_size, kind="main")
    else:
        resolved = app_window._resolved_popup_theme(resolved_kind)
        base_bg = normalize_hex(
            resolved.get("background", app_window.palette_data.get("control_bg", DEFAULT_THEME_SURFACE)),
            app_window.palette_data.get("control_bg", DEFAULT_THEME_SURFACE),
        )
        transparent = bool(resolved.get("transparent", False))
        bg = app_window.render_background_pixmap(reference_size, kind=resolved_kind)

    bg_color = QColor(base_bg)
    bg_color.setAlpha(220 if transparent else 244)
    painter.fillRect(target_rect, bg_color)

    if not bg.isNull():
        src = QRect(0, 0, bg.width(), bg.height())
        if bg.width() >= target_size.width() and bg.height() >= target_size.height():
            src = QRect(
                max(0, (bg.width() - target_size.width()) // 2),
                max(0, (bg.height() - target_size.height()) // 2),
                int(target_size.width()),
                int(target_size.height()),
            )
        painter.setOpacity(0.84 if transparent else 0.94)
        painter.drawPixmap(target_rect, bg, src)
        painter.setOpacity(1.0)

    overlay = QColor(app_window.palette_data.get("shell_overlay", base_bg))
    overlay.setAlpha(56 if transparent else 28)
    painter.fillRect(target_rect, overlay)


class FlowgridThemedDialog(QDialog):
    def __init__(
        self,
        parent: QWidget | None,
        app_window: "QuickInputsWindow" | None,
        theme_kind: str,
    ) -> None:
        super().__init__(parent)
        self._app_window = _resolve_flowgrid_popup_app_window(parent, app_window)
        self._theme_kind = str(theme_kind or "").strip() or "main"
        self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)

    def apply_theme_styles(self, *, force_opaque_root: bool = True) -> None:
        _apply_flowgrid_popup_stylesheet(
            self,
            self._app_window,
            self._theme_kind,
            force_opaque_root=force_opaque_root,
            event_key="ui.themed_dialog_stylesheet_failed",
        )

    def paintEvent(self, event) -> None:  # noqa: N802
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform)
        _paint_flowgrid_popup_background(self, painter, self._app_window, self._theme_kind)
        super().paintEvent(event)


class FlowgridThemedFileDialog(QFileDialog):
    def __init__(
        self,
        parent: QWidget | None,
        app_window: "QuickInputsWindow" | None,
        theme_kind: str,
    ) -> None:
        super().__init__(parent)
        self._app_window = _resolve_flowgrid_popup_app_window(parent, app_window)
        self._theme_kind = str(theme_kind or "").strip() or "main"
        self.setOption(QFileDialog.Option.DontUseNativeDialog, True)
        self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        _apply_flowgrid_popup_stylesheet(
            self,
            self._app_window,
            self._theme_kind,
            force_opaque_root=True,
            event_key="ui.themed_file_dialog_stylesheet_failed",
        )

    def paintEvent(self, event) -> None:  # noqa: N802
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform)
        _paint_flowgrid_popup_background(self, painter, self._app_window, self._theme_kind)
        super().paintEvent(event)


class FlowgridThemedColorDialog(QColorDialog):
    def __init__(
        self,
        parent: QWidget | None,
        app_window: "QuickInputsWindow" | None,
        theme_kind: str,
    ) -> None:
        super().__init__(parent)
        self._app_window = _resolve_flowgrid_popup_app_window(parent, app_window)
        self._theme_kind = str(theme_kind or "").strip() or "main"
        self.setOption(QColorDialog.ColorDialogOption.DontUseNativeDialog, True)
        self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        _apply_flowgrid_popup_stylesheet(
            self,
            self._app_window,
            self._theme_kind,
            force_opaque_root=True,
            event_key="ui.themed_color_dialog_stylesheet_failed",
        )

    def paintEvent(self, event) -> None:  # noqa: N802
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform)
        _paint_flowgrid_popup_background(self, painter, self._app_window, self._theme_kind)
        super().paintEvent(event)


class FlowgridThemedMessageDialog(FlowgridThemedDialog):
    def __init__(
        self,
        parent: QWidget | None,
        app_window: "QuickInputsWindow" | None,
        theme_kind: str,
        icon: QMessageBox.Icon,
        title: str,
        text: str,
    ) -> None:
        super().__init__(parent, app_window, theme_kind)
        self.setWindowTitle(str(title or "").strip() or "Message")
        self.setModal(True)
        self.setMinimumWidth(360)
        self.setMaximumWidth(620)
        self.apply_theme_styles(force_opaque_root=True)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(UI_MARGIN_STANDARD_DIALOG, UI_MARGIN_STANDARD_DIALOG, UI_MARGIN_STANDARD_DIALOG, UI_MARGIN_STANDARD_DIALOG)
        layout.setSpacing(UI_SPACING_STANDARD_DIALOG)

        content_row = QHBoxLayout()
        content_row.setContentsMargins(0, 0, 0, 0)
        content_row.setSpacing(UI_SPACING_STANDARD_DIALOG)

        icon_label = QLabel()
        icon_label.setAlignment(Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignHCenter)
        icon_label.setPixmap(self._message_icon_pixmap(icon))
        icon_label.setMinimumWidth(40)
        content_row.addWidget(icon_label, 0)

        text_label = QLabel(str(text or ""))
        text_label.setWordWrap(True)
        text_label.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        content_row.addWidget(text_label, 1)
        layout.addLayout(content_row)

        button_row = QHBoxLayout()
        button_row.setContentsMargins(0, 0, 0, 0)
        button_row.setSpacing(UI_SPACING_STANDARD_DIALOG)
        button_row.addStretch(1)
        ok_button = QPushButton("OK")
        ok_button.setProperty("actionRole", "pick")
        ok_button.clicked.connect(self.accept)
        ok_button.setDefault(True)
        ok_button.setAutoDefault(True)
        button_row.addWidget(ok_button, 0)
        layout.addLayout(button_row)

    def _message_icon_pixmap(self, icon: QMessageBox.Icon) -> QPixmap:
        style = self.style() if self.style() is not None else QApplication.style()
        if style is None:
            return QPixmap()
        mapping = {
            QMessageBox.Icon.Information: QStyle.StandardPixmap.SP_MessageBoxInformation,
            QMessageBox.Icon.Warning: QStyle.StandardPixmap.SP_MessageBoxWarning,
            QMessageBox.Icon.Critical: QStyle.StandardPixmap.SP_MessageBoxCritical,
            QMessageBox.Icon.Question: QStyle.StandardPixmap.SP_MessageBoxQuestion,
        }
        standard = mapping.get(icon, QStyle.StandardPixmap.SP_MessageBoxInformation)
        return style.standardIcon(standard).pixmap(32, 32)


class FlowgridThemedInputDialog(QInputDialog):
    def __init__(
        self,
        parent: QWidget | None,
        app_window: "QuickInputsWindow" | None,
        theme_kind: str,
    ) -> None:
        super().__init__(parent)
        self._app_window = _resolve_flowgrid_popup_app_window(parent, app_window)
        self._theme_kind = str(theme_kind or "").strip() or "main"
        self.setObjectName("FlowgridThemedInputDialog")
        self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)

    def paintEvent(self, event) -> None:  # noqa: N802
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform)
        _paint_flowgrid_popup_background(self, painter, self._app_window, self._theme_kind)
        super().paintEvent(event)


def _configure_flowgrid_file_dialog(
    dialog: FlowgridThemedFileDialog,
    title: str,
    directory: str,
    file_filter: str,
) -> None:
    dialog.setWindowTitle(str(title or "").strip() or "Select File")
    initial_path = str(directory or "").strip()
    if initial_path:
        candidate = Path(initial_path)
        if candidate.exists() and candidate.is_dir():
            dialog.setDirectory(str(candidate))
        else:
            parent_dir = candidate.parent if str(candidate.parent) else Path.home()
            if parent_dir.exists():
                dialog.setDirectory(str(parent_dir))
            if candidate.name:
                dialog.selectFile(candidate.name)
    filters = [item for item in str(file_filter or "").split(";;") if str(item or "").strip()]
    if filters:
        dialog.setNameFilters(filters)
        dialog.selectNameFilter(filters[0])


def show_flowgrid_themed_open_file_name(
    parent: QWidget | None,
    app_window: "QuickInputsWindow" | None,
    theme_kind: str,
    title: str,
    directory: str,
    file_filter: str,
) -> tuple[str, str]:
    dialog = FlowgridThemedFileDialog(parent, app_window, theme_kind)
    _configure_flowgrid_file_dialog(dialog, title, directory, file_filter)
    dialog.setFileMode(QFileDialog.FileMode.ExistingFile)
    dialog.setAcceptMode(QFileDialog.AcceptMode.AcceptOpen)
    if dialog.exec() != QDialog.DialogCode.Accepted:
        return "", dialog.selectedNameFilter()
    files = dialog.selectedFiles()
    return (str(files[0]), dialog.selectedNameFilter()) if files else ("", dialog.selectedNameFilter())


def show_flowgrid_themed_open_file_names(
    parent: QWidget | None,
    app_window: "QuickInputsWindow" | None,
    theme_kind: str,
    title: str,
    directory: str,
    file_filter: str,
) -> tuple[list[str], str]:
    dialog = FlowgridThemedFileDialog(parent, app_window, theme_kind)
    _configure_flowgrid_file_dialog(dialog, title, directory, file_filter)
    dialog.setFileMode(QFileDialog.FileMode.ExistingFiles)
    dialog.setAcceptMode(QFileDialog.AcceptMode.AcceptOpen)
    if dialog.exec() != QDialog.DialogCode.Accepted:
        return [], dialog.selectedNameFilter()
    return [str(path) for path in dialog.selectedFiles()], dialog.selectedNameFilter()


def show_flowgrid_themed_save_file_name(
    parent: QWidget | None,
    app_window: "QuickInputsWindow" | None,
    theme_kind: str,
    title: str,
    directory: str,
    file_filter: str,
) -> tuple[str, str]:
    dialog = FlowgridThemedFileDialog(parent, app_window, theme_kind)
    _configure_flowgrid_file_dialog(dialog, title, directory, file_filter)
    dialog.setAcceptMode(QFileDialog.AcceptMode.AcceptSave)
    dialog.setFileMode(QFileDialog.FileMode.AnyFile)
    if dialog.exec() != QDialog.DialogCode.Accepted:
        return "", dialog.selectedNameFilter()
    files = dialog.selectedFiles()
    return (str(files[0]), dialog.selectedNameFilter()) if files else ("", dialog.selectedNameFilter())


def show_flowgrid_themed_existing_directory(
    parent: QWidget | None,
    app_window: "QuickInputsWindow" | None,
    theme_kind: str,
    title: str,
    directory: str,
) -> str:
    dialog = FlowgridThemedFileDialog(parent, app_window, theme_kind)
    _configure_flowgrid_file_dialog(dialog, title, directory, "")
    dialog.setFileMode(QFileDialog.FileMode.Directory)
    dialog.setOption(QFileDialog.Option.ShowDirsOnly, True)
    dialog.setAcceptMode(QFileDialog.AcceptMode.AcceptOpen)
    if dialog.exec() != QDialog.DialogCode.Accepted:
        return ""
    files = dialog.selectedFiles()
    return str(files[0]) if files else ""


def show_flowgrid_themed_color(
    parent: QWidget | None,
    app_window: "QuickInputsWindow" | None,
    theme_kind: str,
    title: str,
    current: QColor,
) -> QColor:
    dialog = FlowgridThemedColorDialog(parent, app_window, theme_kind)
    dialog.setWindowTitle(str(title or "").strip() or "Pick Color")
    dialog.setCurrentColor(current)
    if dialog.exec() != QDialog.DialogCode.Accepted:
        return QColor()
    return dialog.currentColor()


def _build_flowgrid_themed_input_dialog(
    parent: QWidget | None,
    app_window: "QuickInputsWindow" | None,
    theme_kind: str,
    title: str,
) -> QInputDialog:
    if app_window is None and parent is not None:
        app_window = getattr(parent, "app_window", None)
    dialog = FlowgridThemedInputDialog(parent, app_window, theme_kind)
    dialog.setWindowTitle(str(title or "").strip() or "Input")
    dialog.setModal(True)
    dialog.setOption(QInputDialog.InputDialogOption.UseListViewForComboBoxItems, False)
    dialog.setMinimumWidth(340)
    dialog.setMaximumWidth(520)
    resolved_theme = str(theme_kind or "").strip() or "main"
    if app_window is not None:
        try:
            dialog.setStyleSheet(app_window._popup_theme_stylesheet(resolved_theme, force_opaque_root=True))
        except Exception as exc:
            _runtime_log_event(
                "ui.themed_input_dialog_stylesheet_failed",
                severity="warning",
                summary="Failed applying themed stylesheet to input dialog.",
                exc=exc,
                context={"theme_kind": str(resolved_theme), "title": str(title)},
            )
    return dialog


def show_flowgrid_themed_input_text(
    parent: QWidget | None,
    app_window: "QuickInputsWindow" | None,
    theme_kind: str,
    title: str,
    label: str,
    default_text: str = "",
) -> tuple[str, bool]:
    dialog = _build_flowgrid_themed_input_dialog(parent, app_window, theme_kind, title)
    dialog.setInputMode(QInputDialog.InputMode.TextInput)
    dialog.setLabelText(str(label or ""))
    dialog.setTextValue(str(default_text or ""))
    accepted = dialog.exec() == QDialog.DialogCode.Accepted
    return str(dialog.textValue() or ""), bool(accepted)


def show_flowgrid_themed_input_int(
    parent: QWidget | None,
    app_window: "QuickInputsWindow" | None,
    theme_kind: str,
    title: str,
    label: str,
    value: int,
    min_value: int,
    max_value: int,
    step: int = 1,
) -> tuple[int, bool]:
    dialog = _build_flowgrid_themed_input_dialog(parent, app_window, theme_kind, title)
    dialog.setInputMode(QInputDialog.InputMode.IntInput)
    dialog.setLabelText(str(label or ""))
    dialog.setIntRange(int(min_value), int(max_value))
    dialog.setIntStep(max(1, int(step)))
    dialog.setIntValue(int(value))
    accepted = dialog.exec() == QDialog.DialogCode.Accepted
    return int(dialog.intValue()), bool(accepted)


def show_flowgrid_themed_input_item(
    parent: QWidget | None,
    app_window: "QuickInputsWindow" | None,
    theme_kind: str,
    title: str,
    label: str,
    items: list[str] | tuple[str, ...],
    current_index: int = 0,
    editable: bool = False,
) -> tuple[str, bool]:
    dialog = _build_flowgrid_themed_input_dialog(parent, app_window, theme_kind, title)
    values = [str(item or "").strip() for item in items if str(item or "").strip()]
    if not values:
        return "", False
    index = int(clamp(int(current_index), 0, len(values) - 1))
    dialog.setInputMode(QInputDialog.InputMode.TextInput)
    dialog.setLabelText(str(label or ""))
    dialog.setComboBoxItems(values)
    dialog.setComboBoxEditable(bool(editable))
    dialog.setTextValue(values[index])
    dialog.setMinimumHeight(148)
    dialog.setMaximumHeight(230)
    accepted = dialog.exec() == QDialog.DialogCode.Accepted
    return str(dialog.textValue() or "").strip(), bool(accepted)


def _center_table_item(item: QTableWidgetItem) -> QTableWidgetItem:
    item.setTextAlignment(int(Qt.AlignmentFlag.AlignCenter))
    return item


def _table_column_index_by_header(table: QTableWidget, header_text: str) -> int:
    for col in range(int(table.columnCount())):
        hdr = table.horizontalHeaderItem(col)
        if hdr is not None and str(hdr.text() or "").strip() == header_text:
            return col
    return -1


def _selected_part_id_from_table(table: QTableWidget) -> int | None:
    row = table.currentRow()
    if row < 0:
        return None
    for col in range(int(table.columnCount())):
        cell_item = table.item(row, col)
        if cell_item is None:
            continue
        raw_id = cell_item.data(Qt.ItemDataRole.UserRole)
        try:
            return int(raw_id)
        except Exception:
            continue
    return None


def _copy_work_order_from_table_item(
    item: QTableWidgetItem | None,
    *,
    header_text: str = "Work Order",
) -> tuple[QTableWidget | None, str]:
    if item is None:
        return None, ""
    table = item.tableWidget()
    if table is None:
        return None, ""
    work_col = _table_column_index_by_header(table, header_text)
    if work_col < 0:
        return table, ""
    work_item = table.item(item.row(), work_col)
    if work_item is None:
        return table, ""
    work_order = str(work_item.text() or "").strip()
    if work_order:
        QApplication.clipboard().setText(work_order)
    return table, work_order


def _select_table_row_by_context_pos(
    table: QTableWidget | None,
    pos: QPoint,
    *,
    header_text: str = "Work Order",
) -> bool:
    if table is None:
        return False
    row = int(table.rowAt(int(pos.y())))
    if row < 0:
        return False
    work_col = _table_column_index_by_header(table, header_text)
    if work_col < 0:
        work_col = 0
    table.setCurrentCell(row, work_col)
    table.selectRow(row)
    return True


class DepotFramelessToolWindow(QDialog):
    """Frameless depot panels: shared chrome, drag-to-move, themed background, themed message boxes."""

    def __init__(
        self,
        app_window: "QuickInputsWindow" | None,
        *,
        window_title: str,
        theme_kind: str,
        size: tuple[int, int] = (780, 560),
        minimum_size: tuple[int, int] | None = None,
    ) -> None:
        super().__init__()
        self.app_window = app_window
        self._theme_kind = theme_kind
        self._drag_offset: QPoint | None = None

        self.setWindowTitle(window_title)
        self.setModal(False)
        self.setWindowModality(Qt.WindowModality.NonModal)
        self.setWindowFlags(Qt.WindowType.Window | Qt.WindowType.FramelessWindowHint | Qt.WindowType.WindowStaysOnTopHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground, True)
        self.resize(int(size[0]), int(size[1]))
        if minimum_size is not None:
            self.setMinimumSize(int(minimum_size[0]), int(minimum_size[1]))
        self._copy_notice_label: QLabel | None = None
        self._copy_notice_host: QWidget | None = None

        self.root_layout = QVBoxLayout(self)
        self.root_layout.setContentsMargins(
            UI_MARGIN_FRAMELESS_TOOL,
            UI_MARGIN_FRAMELESS_TOOL,
            UI_MARGIN_FRAMELESS_TOOL,
            UI_MARGIN_FRAMELESS_TOOL,
        )
        self.root_layout.setSpacing(UI_SPACING_FRAMELESS_TOOL)

        header_row = QHBoxLayout()
        header_row.setContentsMargins(0, 0, 0, 0)
        header_row.setSpacing(UI_SPACING_FRAMELESS_TOOL)
        self.header_title = QLabel(window_title)
        self.header_title.setProperty("section", True)
        self.header_close_btn = QPushButton("x")
        self.header_close_btn.setFixedSize(HEADER_CLOSE_BUTTON_WIDTH, HEADER_CLOSE_BUTTON_HEIGHT)
        self.header_close_btn.setObjectName("DepotFramelessCloseButton")
        self.header_close_btn.setAutoDefault(False)
        self.header_close_btn.setDefault(False)
        self.header_close_btn.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.header_close_btn.clicked.connect(self.close)
        header_row.addWidget(self.header_title)
        header_row.addStretch(1)
        header_row.addWidget(self.header_close_btn)
        self.root_layout.addLayout(header_row)

    def apply_theme_styles(self) -> None:
        if self.app_window is None:
            return
        self.setStyleSheet(self.app_window._popup_theme_stylesheet(self._theme_kind))

    def set_window_always_on_top(self, enabled: bool) -> None:
        keep_on_top = bool(enabled)
        was_visible = self.isVisible()
        self.setWindowFlag(Qt.WindowType.WindowStaysOnTopHint, keep_on_top)
        if was_visible:
            self.show()
            if keep_on_top:
                self.raise_()

    def _load_window_always_on_top_preference(self, config_key: str, default: bool = True) -> bool:
        if self.app_window is None:
            return bool(default)
        return bool(self.app_window.config.get(config_key, default))

    def _apply_window_always_on_top_preference(self, config_key: str, enabled: bool) -> bool:
        keep_on_top = bool(enabled)
        self.set_window_always_on_top(keep_on_top)
        if self.app_window is not None:
            self.app_window.config[config_key] = keep_on_top
            self.app_window.queue_save_config()
        return keep_on_top

    def _show_themed_message(self, icon: QMessageBox.Icon, title: str, text: str) -> None:
        show_flowgrid_themed_message(self, self.app_window, self._theme_kind, icon, title, text)

    def _clear_copy_notice(self) -> None:
        label = self._copy_notice_label
        self._copy_notice_label = None
        self._copy_notice_host = None
        if label is None:
            return
        try:
            label.hide()
            label.deleteLater()
        except Exception as exc:
            _runtime_log_event(
                "ui.copy_notice_clear_failed",
                severity="warning",
                summary="Failed clearing copy-notice overlay.",
                exc=exc,
                context={"window_title": str(self.windowTitle())},
            )

    def _show_copy_notice(self, anchor: QWidget | None, text: str, *, duration_ms: int = 4200) -> None:
        self._clear_copy_notice()
        host = anchor
        if host is not None and hasattr(host, "viewport"):
            try:
                vp = host.viewport()  # type: ignore[attr-defined]
                if isinstance(vp, QWidget):
                    host = vp
            except Exception:
                host = anchor
        if not isinstance(host, QWidget):
            host = self
        message = str(text or "").strip()
        if not message:
            return
        notice = QLabel(message, host)
        notice.setObjectName("FlowgridCopyNotice")
        notice.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents, True)
        notice.setAlignment(Qt.AlignmentFlag.AlignCenter)
        notice.setWordWrap(False)
        notice.setStyleSheet(
            "QLabel#FlowgridCopyNotice {"
            "background-color: rgba(8, 16, 28, 228);"
            "color: #FFFFFF;"
            "border: 1px solid rgba(74, 196, 186, 230);"
            "border-radius: 8px;"
            "padding: 4px 10px;"
            "font-weight: 800;"
            "}"
        )
        hint = notice.sizeHint()
        max_w = max(170, int(host.width()) - 16)
        width = int(clamp(int(hint.width()) + 20, 180, max_w))
        height = int(max(28, int(hint.height()) + 10))
        x = int(max(8, int(host.width()) - width - 8))
        y = int(max(8, int(host.height()) - height - 8))
        notice.setGeometry(x, y, width, height)
        notice.show()
        notice.raise_()
        self._copy_notice_label = notice
        self._copy_notice_host = host
        QTimer.singleShot(max(1200, int(duration_ms)), self._clear_copy_notice)

    def mousePressEvent(self, event: QMouseEvent) -> None:  # noqa: N802
        if event.button() == Qt.MouseButton.LeftButton and _can_start_window_drag(self, event.position().toPoint()):
            self._drag_offset = event.globalPosition().toPoint() - self.frameGeometry().topLeft()
            event.accept()
            return
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event: QMouseEvent) -> None:  # noqa: N802
        if self._drag_offset and event.buttons() & Qt.MouseButton.LeftButton:
            self.move(event.globalPosition().toPoint() - self._drag_offset)
            event.accept()
            return
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event: QMouseEvent) -> None:  # noqa: N802
        self._drag_offset = None
        super().mouseReleaseEvent(event)

    def paintEvent(self, event) -> None:  # noqa: N802
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform)
        if self.app_window is not None:
            theme = self.app_window._resolved_popup_theme(self._theme_kind)
            bg_color = QColor(theme.get("background", "#FFFFFF"))
            if not bool(theme.get("transparent", False)):
                painter.fillRect(self.rect(), bg_color)
            bg = self.app_window.render_background_pixmap(self.rect().size(), kind=self._theme_kind)
            if not bg.isNull():
                painter.drawPixmap(self.rect(), bg)
        super().paintEvent(event)

    def keyPressEvent(self, event) -> None:  # noqa: N802
        key = event.key()
        if key in (int(Qt.Key.Key_Return), int(Qt.Key.Key_Enter)):
            event.accept()
            return
        super().keyPressEvent(event)


if os.name == "nt":
    user32 = ctypes.WinDLL("user32", use_last_error=True)
    # Standard keys
    VK_CONTROL = 0x11
    VK_V = 0x56
    KEYEVENTF_KEYUP = 0x0002
    SW_RESTORE = 9
    # Extended keycodes for macro sequences
    VK_TAB = 0x09
    VK_ENTER = 0x0D
    VK_ESCAPE = 0x1B
    VK_SPACE = 0x20
    VK_SHIFT = 0x10
    VK_ALT = 0x12
    VK_BACKSPACE = 0x08
    VK_DELETE = 0x2E
    VK_HOME = 0x24
    VK_END = 0x23
    VK_PAGE_UP = 0x21
    VK_PAGE_DOWN = 0x22
    VK_LEFT = 0x25
    VK_RIGHT = 0x27
    VK_UP = 0x26
    VK_DOWN = 0x28
    VK_RETURN = 0x0D
    VK_ENTER = 0x0D
    
    # Key code mapping for macro parsing
    KEY_CODES: dict[str, int] = {
        "tab": VK_TAB,
        "enter": VK_RETURN,
        "return": VK_RETURN,
        "escape": VK_ESCAPE,
        "esc": VK_ESCAPE,
        "space": VK_SPACE,
        "shift": VK_SHIFT,
        "alt": VK_ALT,
        "backspace": VK_BACKSPACE,
        "delete": VK_DELETE,
        "del": VK_DELETE,
        "home": VK_HOME,
        "end": VK_END,
        "pageup": VK_PAGE_UP,
        "pagedown": VK_PAGE_DOWN,
        "left": VK_LEFT,
        "right": VK_RIGHT,
        "up": VK_UP,
        "down": VK_DOWN,
    }


@dataclass
class LayerRenderInfo:
    layer: dict[str, Any]
    rect: QRectF
    pixmap: QPixmap


def deep_clone(value: Any) -> Any:
    return json.loads(json.dumps(value))


def deep_merge(defaults: Any, incoming: Any) -> Any:
    if isinstance(defaults, dict):
        out: dict[str, Any] = {}
        incoming_dict = incoming if isinstance(incoming, dict) else {}
        for key, default_value in defaults.items():
            out[key] = deep_merge(default_value, incoming_dict.get(key))
        for key, value in incoming_dict.items():
            if key not in out:
                out[key] = value
        return out
    if isinstance(defaults, list):
        if isinstance(incoming, list):
            return incoming
        return deep_clone(defaults)
    return defaults if incoming is None else incoming


def clamp(value: float, lo: float, hi: float) -> float:
    return max(lo, min(hi, value))


def safe_int(value: Any, fallback: int = 0) -> int:
    try:
        return int(value)
    except Exception:
        return int(fallback)


def normalize_hex(color: str, fallback: str = "#FFFFFF") -> str:
    if not isinstance(color, str):
        return fallback
    value = color.strip().upper()
    if len(value) == 7 and value.startswith("#"):
        try:
            int(value[1:], 16)
            return value
        except ValueError:
            return fallback
    return fallback


def hex_to_rgb(color: str) -> tuple[int, int, int]:
    color = normalize_hex(color)
    return int(color[1:3], 16), int(color[3:5], 16), int(color[5:7], 16)


def rgb_to_hex(r: int, g: int, b: int) -> str:
    return f"#{int(clamp(r, 0, 255)):02X}{int(clamp(g, 0, 255)):02X}{int(clamp(b, 0, 255)):02X}"


def blend(color_a: str, color_b: str, ratio: float) -> str:
    ratio = clamp(ratio, 0.0, 1.0)
    ar, ag, ab = hex_to_rgb(color_a)
    br, bg, bb = hex_to_rgb(color_b)
    return rgb_to_hex(ar + (br - ar) * ratio, ag + (bg - ag) * ratio, ab + (bb - ab) * ratio)


def luminance(color: str) -> float:
    r, g, b = hex_to_rgb(color)

    def channel(v: int) -> float:
        x = v / 255.0
        return x / 12.92 if x <= 0.03928 else ((x + 0.055) / 1.055) ** 2.4

    return 0.2126 * channel(r) + 0.7152 * channel(g) + 0.0722 * channel(b)


def contrast_ratio(color_a: str, color_b: str) -> float:
    l1, l2 = luminance(color_a), luminance(color_b)
    hi, lo = (l1, l2) if l1 > l2 else (l2, l1)
    return (hi + 0.05) / (lo + 0.05)


def readable_text(background: str) -> str:
    white_ratio = contrast_ratio("#FFFFFF", background)
    black_ratio = contrast_ratio("#101418", background)
    return "#FFFFFF" if white_ratio >= black_ratio else "#101418"


def shift(color: str, amount: float) -> str:
    target = "#FFFFFF" if amount >= 0 else "#000000"
    return blend(color, target, abs(amount))


def rgba_css(color: str, alpha: float) -> str:
    r, g, b = hex_to_rgb(color)
    a = int(clamp(alpha, 0.0, 1.0) * 255)
    return f"rgba({r}, {g}, {b}, {a})"


def compute_palette(theme: dict[str, str]) -> dict[str, str]:
    primary = normalize_hex(theme.get("primary", DEFAULT_THEME_PRIMARY), DEFAULT_THEME_PRIMARY)
    accent = normalize_hex(theme.get("accent", DEFAULT_THEME_ACCENT), DEFAULT_THEME_ACCENT)
    surface = normalize_hex(theme.get("surface", DEFAULT_THEME_SURFACE), DEFAULT_THEME_SURFACE)

    shell_overlay = shift(primary, -0.60)
    sidebar_overlay = shift(primary, -0.70)
    nav_active = blend(accent, primary, 0.35)
    text_color = readable_text(shift(surface, -0.55))
    control_bg = blend(surface, "#1E2A34", 0.22)
    input_bg = blend(surface, "#FFFFFF", 0.08)
    button_bg = blend(primary, accent, 0.30)

    return {
        "primary": primary,
        "accent": accent,
        "surface": surface,
        "shell_overlay": shell_overlay,
        "sidebar_overlay": sidebar_overlay,
        "label_text": text_color,
        "muted_text": blend(text_color, "#AAB7C2", 0.35),
        "control_bg": control_bg,
        "input_bg": input_bg,
        "button_bg": button_bg,
        "button_text": readable_text(button_bg),
        "nav_active": nav_active,
    }


def safe_layer_defaults(layer: dict[str, Any]) -> dict[str, Any]:
    visible_raw = layer.get("visible", True)
    if isinstance(visible_raw, str):
        visible_text = visible_raw.strip().lower()
        if visible_text in {"0", "false", "no", "off"}:
            visible_value = False
        elif visible_text in {"1", "true", "yes", "on"}:
            visible_value = True
        else:
            visible_value = True
    else:
        visible_value = bool(visible_raw)
    return {
        "image_path": layer.get("image_path", ""),
        "image_x": int(layer.get("image_x", 0)),
        "image_y": int(layer.get("image_y", 0)),
        "image_scale_mode": layer.get("image_scale_mode", "Fill"),
        "image_anchor": layer.get("image_anchor", "Center"),
        "image_scale_percent": int(layer.get("image_scale_percent", 100)),
        "image_opacity": float(clamp(float(layer.get("image_opacity", 1.0)), 0.0, 1.0)),
        "visible": visible_value,
        "name": layer.get("name") or Path(layer.get("image_path", "")).name or "Layer",
    }


def build_quick_shape_polygon(shape: str, w: int, h: int) -> QPolygon | None:
    if w <= 8 or h <= 8:
        return None

    if shape == "Diamond":
        return QPolygon(
            [
                QPoint(w // 2, 0),
                QPoint(w - 1, h // 2),
                QPoint(w // 2, h - 1),
                QPoint(0, h // 2),
            ]
        )

    if shape == "Hex":
        dx = max(8, int(w * 0.14))
        return QPolygon(
            [
                QPoint(dx, 0),
                QPoint(w - dx - 1, 0),
                QPoint(w - 1, h // 2),
                QPoint(w - dx - 1, h - 1),
                QPoint(dx, h - 1),
                QPoint(0, h // 2),
            ]
        )

    if shape == "Slant":
        dx = max(8, int(w * 0.12))
        return QPolygon(
            [
                QPoint(dx, 0),
                QPoint(w - 1, 0),
                QPoint(w - dx - 1, h - 1),
                QPoint(0, h - 1),
            ]
        )

    if shape == "CutCorner":
        cut = max(6, int(min(w, h) * 0.22))
        return QPolygon(
            [
                QPoint(cut, 0),
                QPoint(w - cut - 1, 0),
                QPoint(w - 1, cut),
                QPoint(w - 1, h - cut - 1),
                QPoint(w - cut - 1, h - 1),
                QPoint(cut, h - 1),
                QPoint(0, h - cut - 1),
                QPoint(0, cut),
            ]
        )

    if shape == "Trapezoid":
        top_inset = max(8, int(w * 0.18))
        return QPolygon(
            [
                QPoint(top_inset, 0),
                QPoint(w - top_inset - 1, 0),
                QPoint(w - 1, h - 1),
                QPoint(0, h - 1),
            ]
        )

    return None


class BackgroundCanvas(QWidget):
    def __init__(self, app_window: "QuickInputsWindow") -> None:
        super().__init__()
        self.app_window = app_window
        self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)

    def paintEvent(self, event) -> None:  # noqa: N802
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform)
        self.app_window.paint_background(painter, self.rect())
        super().paintEvent(event)


class TitleBar(QWidget):
    def __init__(self, app_window: "QuickInputsWindow") -> None:
        super().__init__(app_window)
        self.app_window = app_window
        self._drag_offset: QPoint | None = None
        self.setFixedHeight(TITLEBAR_HEIGHT)

        layout = QHBoxLayout(self)
        layout.setContentsMargins(10, 4, 4, 4)
        layout.setSpacing(6)

        self.icon_label = QLabel()
        self.icon_label.setFixedSize(16, 16)

        self.title_label = QLabel(APP_TITLE)
        self.title_label.setObjectName("TitleText")

        self.min_button = QToolButton()
        self.min_button.setObjectName("TitleMinButton")
        self.min_button.setText("-")
        self.min_button.setFixedSize(26, 22)
        self.min_button.clicked.connect(self.app_window.showMinimized)

        self.close_button = QToolButton()
        self.close_button.setObjectName("TitleCloseButton")
        self.close_button.setText("x")
        self.close_button.setFixedSize(26, 22)
        self.close_button.clicked.connect(self.app_window.close)

        layout.addWidget(self.icon_label)
        layout.addWidget(self.title_label)
        layout.addStretch(1)
        layout.addWidget(self.min_button)
        layout.addWidget(self.close_button)

    def update_icon(self, icon: QIcon | None) -> None:
        if icon and not icon.isNull():
            source = icon.pixmap(64, 64)
            if source.isNull():
                source = icon.pixmap(16, 16)
        else:
            source = QIcon.fromTheme("applications-utilities").pixmap(64, 64)

        if source.isNull():
            self.icon_label.clear()
            return

        scaled = source.scaled(16, 16, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
        canvas = QPixmap(16, 16)
        canvas.fill(Qt.GlobalColor.transparent)
        painter = QPainter(canvas)
        painter.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform, True)
        x = (16 - scaled.width()) // 2
        y = (16 - scaled.height()) // 2
        painter.drawPixmap(x, y, scaled)
        painter.end()
        self.icon_label.setPixmap(canvas)

    def mousePressEvent(self, event: QMouseEvent) -> None:  # noqa: N802
        if event.button() == Qt.MouseButton.LeftButton:
            self._drag_offset = event.globalPosition().toPoint() - self.app_window.frameGeometry().topLeft()
            event.accept()
            return
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event: QMouseEvent) -> None:  # noqa: N802
        if self._drag_offset and event.buttons() & Qt.MouseButton.LeftButton:
            self.app_window.move(event.globalPosition().toPoint() - self._drag_offset)
            event.accept()
            return
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event: QMouseEvent) -> None:  # noqa: N802
        self._drag_offset = None
        super().mouseReleaseEvent(event)


class QuickButtonCard(QWidget):
    edit_requested = Signal(int)
    insert_requested = Signal(int)
    move_requested = Signal(int, int, int, bool)
    _icon_cache: dict[tuple[str, int], QIcon] = {}

    def __init__(self, index: int, title: str, tooltip: str, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.index = index
        self._shape_for_mask = "Soft"
        self._layout_mode = False
        self._drag_anchor: QPoint | None = None
        self._drag_origin = QPoint(0, 0)
        self._drag_active = False
        self._suppress_click_once = False
        self.main_button = QPushButton(title)
        self.main_button.setToolTip(tooltip or "")
        self.main_button.clicked.connect(self._on_main_button_clicked)
        self.main_button.installEventFilter(self)
        self.main_button.setCursor(Qt.CursorShape.PointingHandCursor)

        self.edit_button = QToolButton(self.main_button)
        self.edit_button.setText("")
        self.edit_button.setToolTip("Edit")
        self.edit_button.setCursor(Qt.CursorShape.PointingHandCursor)
        self.edit_button.clicked.connect(lambda: self.edit_requested.emit(self.index))

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        layout.addWidget(self.main_button)

    def _on_main_button_clicked(self) -> None:
        if self._layout_mode:
            self._suppress_click_once = False
            return
        if self._suppress_click_once:
            self._suppress_click_once = False
            return
        self.insert_requested.emit(self.index)

    def set_layout_mode(self, enabled: bool) -> None:
        self._layout_mode = bool(enabled)
        if self._layout_mode:
            self.main_button.setCursor(Qt.CursorShape.OpenHandCursor)
        else:
            self.main_button.setCursor(Qt.CursorShape.PointingHandCursor)
            self._drag_anchor = None
            self._drag_active = False
            self._suppress_click_once = False

    def eventFilter(self, watched, event) -> bool:  # noqa: N802
        if watched is self.main_button and self._layout_mode:
            et = event.type()
            if et == QEvent.Type.MouseButtonPress and event.button() == Qt.MouseButton.LeftButton:
                self._drag_anchor = event.globalPosition().toPoint()
                self._drag_origin = self.pos()
                self._drag_active = False
                self._suppress_click_once = False
                self.main_button.setCursor(Qt.CursorShape.ClosedHandCursor)
                return True

            if et == QEvent.Type.MouseMove and event.buttons() & Qt.MouseButton.LeftButton and self._drag_anchor is not None:
                delta = event.globalPosition().toPoint() - self._drag_anchor
                if not self._drag_active and delta.manhattanLength() < 3:
                    return True
                self._drag_active = True
                self._suppress_click_once = True
                new_pos = self._drag_origin + delta
                self.move_requested.emit(self.index, int(new_pos.x()), int(new_pos.y()), False)
                return True

            if et == QEvent.Type.MouseButtonRelease and event.button() == Qt.MouseButton.LeftButton and self._drag_anchor is not None:
                if self._drag_active:
                    delta = event.globalPosition().toPoint() - self._drag_anchor
                    new_pos = self._drag_origin + delta
                    self.move_requested.emit(self.index, int(new_pos.x()), int(new_pos.y()), True)
                self._drag_anchor = None
                self._drag_active = False
                self.main_button.setCursor(Qt.CursorShape.OpenHandCursor)
                return True

        return super().eventFilter(watched, event)

    def apply_visual_style(
        self,
        width: int,
        height: int,
        font_size: int,
        font_family: str,
        shape: str,
        button_opacity: float,
        palette: dict[str, str],
        action_type: str = "paste_text",
    ) -> None:
        width = int(clamp(width, 90, 220))
        height = int(clamp(height, 35, 100))
        font_size = int(clamp(font_size, 8, 20))
        button_opacity = float(clamp(button_opacity, 0.15, 1.0))

        # Adjust base color based on action type for subtle distinction
        base_bg = palette['button_bg']
        if action_type == "open_url":
            # Shift toward accent color for web links
            base_bg = blend(palette['button_bg'], palette['accent'], 0.25)
        elif action_type == "open_app":
            # Shift toward primary color for apps
            base_bg = blend(palette['button_bg'], palette['primary'], 0.25)
        elif action_type in {"input_sequence", "macro_sequence"}:
            # Shift toward surface color for input sequences.
            base_bg = blend(palette['button_bg'], palette['surface'], 0.35)

        self.main_button.setMinimumSize(width, height)
        self.main_button.setMaximumSize(width, height)
        self.main_button.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
        self.setFixedSize(width, height)
        self._shape_for_mask = shape

        temp_window = self.window()
        style_provider = None
        if temp_window is not None and hasattr(temp_window, "_quick_button_stylesheet"):
            style_provider = temp_window
        elif temp_window is not None and hasattr(temp_window, "app_window") and hasattr(temp_window.app_window, "_quick_button_stylesheet"):
            style_provider = temp_window.app_window
        style_css = None
        if style_provider is not None:
            style_css = style_provider._quick_button_stylesheet(
                font_size,
                button_opacity,
                shape,
                font_family=font_family,
                padding="2px 20px 2px 8px",
                action_type=action_type,
            )
        if style_css is None:
            safe_family = str(font_family or "Segoe UI").replace("'", "\\'")
            style_css = (
                "QPushButton {"
                f"background-color: {rgba_css(base_bg, button_opacity)};"
                f"color: {palette['button_text']};"
                f"border: 1px solid {shift(base_bg, -0.35)};"
                "border-radius: 11px;"
                f"font-size: {font_size}px;"
                f"font-family: '{safe_family}';"
                "font-weight: 700;"
                "padding: 2px 20px 2px 8px;"
                "text-align: center;"
                "}"
            )
        self.main_button.setStyleSheet(style_css)
        # User-created quick buttons stay text-only; no action icons on button faces.
        self.main_button.setIcon(QIcon())

        icon_color = shift(palette["button_text"], 0.03)
        self.edit_button.setIcon(self._build_pencil_icon(icon_color, 14))
        self.edit_button.setIconSize(QSize(12, 12))
        self.edit_button.setStyleSheet(
            "QToolButton {"
            "background: rgba(0, 0, 0, 28);"
            "border: none;"
            "border-radius: 8px;"
            "padding: 0px;"
            "}"
            "QToolButton:hover { background: rgba(255, 255, 255, 52); }"
        )
        self.edit_button.setFixedSize(16, 16)
        self._apply_shape_edge_effect(shape, palette, width, height)
        self._position_edit_button()
        self._apply_shape_mask()

    def _apply_shape_edge_effect(self, shape: str, palette: dict[str, str], width: int, height: int) -> None:
        if self._build_shape_polygon(shape, width, height) is None:
            self.main_button.setGraphicsEffect(None)
            return
        effect = self.main_button.graphicsEffect()
        if not isinstance(effect, QGraphicsDropShadowEffect):
            effect = QGraphicsDropShadowEffect(self.main_button)
            self.main_button.setGraphicsEffect(effect)
        edge = QColor(shift(palette["button_bg"], 0.28))
        edge.setAlpha(110)
        effect.setColor(edge)
        effect.setOffset(0, 0)
        effect.setBlurRadius(5.0)

    @classmethod
    def _build_pencil_icon(cls, color_hex: str, size: int) -> QIcon:
        key = (color_hex, size)
        cached = cls._icon_cache.get(key)
        if cached is not None:
            return cached

        canvas_size = 64
        pixmap = QPixmap(canvas_size, canvas_size)
        pixmap.fill(Qt.GlobalColor.transparent)

        painter = QPainter(pixmap)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)
        painter.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform, True)

        body_color = QColor(color_hex)
        body_pen = QPen(body_color)
        body_pen.setWidth(8)
        body_pen.setCapStyle(Qt.PenCapStyle.RoundCap)
        body_pen.setJoinStyle(Qt.PenJoinStyle.RoundJoin)
        painter.setPen(body_pen)
        painter.drawLine(16, 48, 44, 20)

        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(body_color)
        painter.drawPolygon(
            QPolygon(
                [
                    QPoint(44, 20),
                    QPoint(54, 10),
                    QPoint(52, 24),
                ]
            )
        )

        eraser_color = QColor("#F2F5F8")
        eraser_color.setAlpha(230)
        painter.setBrush(eraser_color)
        painter.drawPolygon(
            QPolygon(
                [
                    QPoint(12, 52),
                    QPoint(18, 58),
                    QPoint(24, 52),
                    QPoint(18, 46),
                ]
            )
        )
        painter.end()

        icon = QIcon(pixmap)
        cls._icon_cache[key] = icon
        return icon

    def _position_edit_button(self) -> None:
        pad = 4
        x = max(0, self.main_button.width() - self.edit_button.width() - pad)
        self.edit_button.move(x, pad)
        self.edit_button.raise_()

    def _build_shape_polygon(self, shape: str, w: int, h: int) -> QPolygon | None:
        return build_quick_shape_polygon(shape, w, h)

    def _apply_shape_mask(self) -> None:
        w = self.main_button.width()
        h = self.main_button.height()
        polygon = self._build_shape_polygon(self._shape_for_mask, w, h)
        if polygon is None:
            self.main_button.clearMask()
            return
        self.main_button.setMask(QRegion(polygon))

    def resizeEvent(self, event) -> None:  # noqa: N802
        self._position_edit_button()
        self._apply_shape_mask()
        super().resizeEvent(event)


class QuickButtonCanvas(QWidget):
    def __init__(self) -> None:
        super().__init__()
        self._snap_x = 1
        self._snap_y = 1
        self._snap_enabled = False
        self._show_grid = False
        self._viewport_width = 300
        self._guide_v_lines: list[int] = []
        self._guide_h_lines: list[int] = []
        self._background_drawer: Callable[[QPainter, QRect], None] | None = None
        self._cards: dict[int, QuickButtonCard] = {}
        self._placeholder = QLabel("", self)
        self._placeholder.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop)
        self._placeholder.hide()
        self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        self.setStyleSheet("background: transparent;")

    def set_background_drawer(self, drawer: Callable[[QPainter, QRect], None] | None) -> None:
        self._background_drawer = drawer
        self.update()

    def configure_grid(self, snap_x: int = 1, snap_y: int = 1, show_grid: bool = False, snap_enabled: bool = False) -> None:
        self._snap_x = max(1, int(snap_x))
        self._snap_y = max(1, int(snap_y))
        self._show_grid = bool(show_grid)
        self._snap_enabled = bool(snap_enabled)
        self.update()

    def set_alignment_guides(self, vertical_lines: list[int], horizontal_lines: list[int]) -> None:
        self._guide_v_lines = [int(v) for v in vertical_lines]
        self._guide_h_lines = [int(h) for h in horizontal_lines]
        self.update()

    def clear_alignment_guides(self) -> None:
        self._guide_v_lines = []
        self._guide_h_lines = []
        self.update()

    def set_viewport_width(self, width: int) -> None:
        self._viewport_width = max(120, int(width))
        self._refresh_canvas_size()

    def clear_cards(self) -> None:
        for card in list(self._cards.values()):
            card.setParent(None)
            card.deleteLater()
        self._cards.clear()
        self._placeholder.hide()
        self._refresh_canvas_size()

    def set_placeholder(self, text: str, color: str) -> None:
        self._placeholder.setText(text)
        self._placeholder.setStyleSheet(
            "QLabel {"
            f"color: {color};"
            "background: transparent;"
            "font-weight: 700;"
            "}"
        )
        self._placeholder.move(8, 8)
        self._placeholder.adjustSize()
        self._placeholder.show()
        self._refresh_canvas_size()

    def place_card(self, card: QuickButtonCard, x: int, y: int, snap: bool = False) -> tuple[int, int]:
        card.setParent(self)
        card.show()
        self._cards[card.index] = card
        snapped_x, snapped_y = self.snap_position(x, y, card.width(), card.height(), snap=snap)
        card.move(snapped_x, snapped_y)
        self._refresh_canvas_size()
        return snapped_x, snapped_y

    def snap_position(self, x: int, y: int, width: int, height: int, snap: bool = False) -> tuple[int, int]:
        snapped_x = int(x)
        snapped_y = int(y)
        do_snap = bool(snap or self._snap_enabled)
        if do_snap:
            sx = max(1, self._snap_x)
            sy = max(1, self._snap_y)
            snapped_x = int(round(float(snapped_x) / float(sx)) * sx)
            snapped_y = int(round(float(snapped_y) / float(sy)) * sy)
        max_x = max(0, self._viewport_width - max(1, int(width)))
        snapped_x = int(clamp(snapped_x, 0, max_x))
        snapped_y = max(0, snapped_y)
        return snapped_x, snapped_y

    def move_card(self, index: int, x: int, y: int, snap: bool = False) -> tuple[int, int]:
        card = self._cards.get(index)
        if card is None:
            return 0, 0
        snapped_x, snapped_y = self.snap_position(x, y, card.width(), card.height(), snap=snap)
        card.move(snapped_x, snapped_y)
        self._refresh_canvas_size()
        return snapped_x, snapped_y

    def card_geometry(self, index: int) -> QRect | None:
        card = self._cards.get(index)
        if card is None:
            return None
        return QRect(card.x(), card.y(), card.width(), card.height())

    def iter_card_geometries(self, exclude_index: int | None = None) -> list[QRect]:
        rects: list[QRect] = []
        for idx, card in self._cards.items():
            if exclude_index is not None and idx == exclude_index:
                continue
            rects.append(QRect(card.x(), card.y(), card.width(), card.height()))
        return rects

    def _refresh_canvas_size(self) -> None:
        width = max(120, self._viewport_width)
        max_bottom = 0
        for card in self._cards.values():
            max_bottom = max(max_bottom, card.y() + card.height())
        if self._placeholder.isVisible():
            max_bottom = max(max_bottom, self._placeholder.y() + self._placeholder.height())
        height = max(220, max_bottom + 12)
        self.resize(width, height)
        self.setMinimumSize(width, height)
        self.update()

    def paintEvent(self, event) -> None:  # noqa: N802
        painter = QPainter(self)
        if self._background_drawer is not None:
            self._background_drawer(painter, self.rect())
        if self._show_grid:
            grid_pen = QPen(QColor(255, 255, 255, 28))
            grid_pen.setWidth(1)
            painter.setPen(grid_pen)
            for x in range(0, self.width(), max(1, self._snap_x)):
                painter.drawLine(x, 0, x, self.height())
            for y in range(0, self.height(), max(1, self._snap_y)):
                painter.drawLine(0, y, self.width(), y)
        if self._guide_v_lines or self._guide_h_lines:
            guide_pen = QPen(QColor(255, 216, 76, 220))
            guide_pen.setWidth(1)
            painter.setPen(guide_pen)
            for x in self._guide_v_lines:
                painter.drawLine(int(x), 0, int(x), self.height())
            for y in self._guide_h_lines:
                painter.drawLine(0, int(y), self.width(), int(y))


class QuickRadialMenu(QDialog):
    action_requested = Signal(str)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setWindowFlags(Qt.WindowType.Popup | Qt.WindowType.FramelessWindowHint | Qt.WindowType.NoDropShadowWindowHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground, True)
        self.setModal(False)
        self.setFixedSize(290, 290)

        self._center = QPoint(self.width() // 2, self.height() // 2)
        self._radius = 90
        self._button_size = QSize(92, 36)
        self._buttons: dict[str, QPushButton] = {}
        self._action_meta: list[tuple[str, str, str, int]] = [
            ("add", "Add", "add", 142),
            ("layout", "Layout", "pick", 94),
            ("new_tab", "New Tab", "new", 46),
            ("rename", "Rename", "pick", -2),
            ("remove", "Remove", "reset", -50),
        ]

        self.center_chip = QLabel("+", self)
        self.center_chip.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.center_chip.setGeometry(self._center.x() - 22, self._center.y() - 22, 44, 44)

        for action_key, label, action_role, _angle in self._action_meta:
            button = QPushButton(label, self)
            button.setProperty("actionRole", action_role)
            button.setProperty("radialAction", action_key)
            button.setCursor(Qt.CursorShape.PointingHandCursor)
            button.setFixedSize(self._button_size)
            button.clicked.connect(lambda _checked=False, key=action_key: self._emit_action(key))
            self._buttons[action_key] = button

        self._layout_buttons()
        self.apply_theme_styles(compute_palette(DEFAULT_CONFIG.get("theme", {})))

    def _layout_buttons(self) -> None:
        for action_key, _label, _role, angle_deg in self._action_meta:
            button = self._buttons[action_key]
            radians = math.radians(float(angle_deg))
            cx = self._center.x() + int(math.cos(radians) * self._radius)
            cy = self._center.y() - int(math.sin(radians) * self._radius)
            button.move(cx - (button.width() // 2), cy - (button.height() // 2))

    def _emit_action(self, action_key: str) -> None:
        self.hide()
        self.action_requested.emit(action_key)

    def set_action_enabled(self, action_key: str, enabled: bool) -> None:
        button = self._buttons.get(action_key)
        if button is not None:
            button.setEnabled(bool(enabled))

    def open_anchored_to(self, anchor_widget: QWidget) -> None:
        anchor_center = anchor_widget.mapToGlobal(anchor_widget.rect().center())
        top_left = QPoint(anchor_center.x() - (self.width() // 2), anchor_center.y() - (self.height() // 2))
        screen = QGuiApplication.screenAt(anchor_center)
        if screen is not None:
            avail = screen.availableGeometry()
            top_left.setX(int(clamp(top_left.x(), avail.left(), avail.right() - self.width())))
            top_left.setY(int(clamp(top_left.y(), avail.top(), avail.bottom() - self.height())))
        self.move(top_left)
        self.show()
        self.raise_()
        self.activateWindow()

    def apply_theme_styles(self, palette_data: dict[str, str]) -> None:
        accent = palette_data["accent"]
        control_bg = palette_data["control_bg"]
        label_text = palette_data["label_text"]
        base_btn = palette_data["button_bg"]
        reset_base = shift(palette_data["surface"], -0.45)
        self.setStyleSheet(
            "QPushButton {"
            f"background-color: {rgba_css(base_btn, 0.92)};"
            f"color: {readable_text(base_btn)};"
            f"border: 2px solid {shift(base_btn, -0.62)};"
            "border-radius: 18px;"
            "font-size: 11px;"
            "font-weight: 800;"
            "padding: 2px 8px;"
            "}"
            "QPushButton:hover {"
            f"background-color: {rgba_css(shift(base_btn, 0.08), 0.96)};"
            f"border: 2px solid {shift(base_btn, -0.72)};"
            "}"
            "QPushButton:pressed {"
            f"background-color: {rgba_css(shift(base_btn, -0.06), 0.98)};"
            "}"
            "QPushButton[actionRole='add'] {"
            f"background-color: {rgba_css(palette_data['primary'], 0.94)};"
            f"color: {readable_text(palette_data['primary'])};"
            f"border: 2px solid {shift(palette_data['primary'], -0.56)};"
            "}"
            "QPushButton[actionRole='pick'] {"
            f"background-color: {rgba_css(blend(palette_data['primary'], palette_data['surface'], 0.45), 0.94)};"
            f"color: {readable_text(blend(palette_data['primary'], palette_data['surface'], 0.45))};"
            f"border: 2px solid {shift(blend(palette_data['primary'], palette_data['surface'], 0.45), -0.56)};"
            "}"
            "QPushButton[actionRole='new'] {"
            f"background-color: {rgba_css(blend(palette_data['surface'], palette_data['primary'], 0.35), 0.94)};"
            f"color: {readable_text(blend(palette_data['surface'], palette_data['primary'], 0.35))};"
            f"border: 2px solid {shift(blend(palette_data['surface'], palette_data['primary'], 0.35), -0.56)};"
            "}"
            "QPushButton[actionRole='reset'] {"
            f"background-color: {rgba_css(reset_base, 0.94)};"
            f"color: {readable_text(reset_base)};"
            f"border: 2px solid {shift(reset_base, -0.56)};"
            "}"
            "QPushButton:disabled {"
            "background-color: rgba(80,80,80,160);"
            "color: rgba(230,230,230,140);"
            "border: 2px solid rgba(60,60,60,180);"
            "}"
            "QLabel {"
            f"background: {rgba_css(control_bg, 0.92)};"
            f"color: {readable_text(control_bg)};"
            f"border: 2px solid {shift(control_bg, -0.58)};"
            "border-radius: 22px;"
            "font-size: 20px;"
            "font-weight: 900;"
            "}"
        )
        self.center_chip.setText("+")
        self.center_chip.setToolTip("Quick actions")
        self.center_chip.setStyleSheet(
            "QLabel {"
            f"background: {rgba_css(accent, 0.95)};"
            f"color: {readable_text(accent)};"
            f"border: 2px solid {shift(accent, -0.52)};"
            "border-radius: 22px;"
            "font-size: 20px;"
            "font-weight: 900;"
            "}"
        )
        self.update()

    def paintEvent(self, event) -> None:  # noqa: N802
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QColor(0, 0, 0, 95))
        painter.drawEllipse(self._center, 106, 106)

        guide_pen = QPen(QColor(255, 255, 255, 55))
        guide_pen.setWidth(2)
        painter.setPen(guide_pen)
        for action_key, _label, _role, _angle in self._action_meta:
            button = self._buttons[action_key]
            button_center = button.geometry().center()
            painter.drawLine(self._center, button_center)
        super().paintEvent(event)


class QuickLayoutDialog(QDialog):
    def __init__(self, app_window: "QuickInputsWindow") -> None:
        super().__init__(app_window)
        self.app_window = app_window
        self.setWindowTitle("Arrange Quick Buttons")
        self.setMinimumSize(LAUNCH_WIDTH, LAUNCH_HEIGHT)
        self.setMaximumWidth(LAUNCH_WIDTH)
        self.resize(LAUNCH_WIDTH, LAUNCH_HEIGHT)
        self.setModal(False)
        self.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose, False)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(8)

        self.hint_label = QLabel("Drag freely. Hold Shift while dragging to snap to center/alignment guides.")
        self.hint_label.setWordWrap(True)
        layout.addWidget(self.hint_label)

        style_form = QFormLayout()
        style_form.setContentsMargins(0, 0, 0, 0)
        style_form.setSpacing(6)

        self.style_width_slider = QSlider(Qt.Orientation.Horizontal)
        self.style_width_slider.setRange(90, 220)
        self.style_width_value = QLabel("140")
        width_row = QHBoxLayout()
        width_row.addWidget(self.style_width_slider, 1)
        width_row.addWidget(self.style_width_value)
        width_wrap = QWidget()
        width_wrap.setLayout(width_row)

        self.style_height_slider = QSlider(Qt.Orientation.Horizontal)
        self.style_height_slider.setRange(35, 100)
        self.style_height_value = QLabel("40")
        height_row = QHBoxLayout()
        height_row.addWidget(self.style_height_slider, 1)
        height_row.addWidget(self.style_height_value)
        height_wrap = QWidget()
        height_wrap.setLayout(height_row)

        self.style_font_slider = QSlider(Qt.Orientation.Horizontal)
        self.style_font_slider.setRange(8, 20)
        self.style_font_value = QLabel("11")
        font_row = QHBoxLayout()
        font_row.addWidget(self.style_font_slider, 1)
        font_row.addWidget(self.style_font_value)
        font_wrap = QWidget()
        font_wrap.setLayout(font_row)

        self.style_font_family_combo = QFontComboBox()
        self.style_font_family_combo.setEditable(True)
        self.style_font_family_combo.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.style_font_family_combo.setMaxVisibleItems(16)
        if self.style_font_family_combo.lineEdit() is not None:
            self.style_font_family_combo.lineEdit().setPlaceholderText("Search fonts...")

        self.style_shape_combo = QComboBox()
        self.style_shape_combo.addItems(
            [
                "Soft",
                "Bordered",
                "Block",
                "Pill",
                "Outline",
                "Glass",
                "Diamond",
                "Hex",
                "Slant",
                "Raised3D",
                "Bevel3D",
                "Ridge3D",
                "Neumorph",
                "Retro3D",
                "Neon3D",
            ]
        )

        self.style_opacity_slider = QSlider(Qt.Orientation.Horizontal)
        self.style_opacity_slider.setRange(20, 100)
        self.style_opacity_value = QLabel("1.00")
        opacity_row = QHBoxLayout()
        opacity_row.addWidget(self.style_opacity_slider, 1)
        opacity_row.addWidget(self.style_opacity_value)
        opacity_wrap = QWidget()
        opacity_wrap.setLayout(opacity_row)

        style_form.addRow("Width", width_wrap)
        style_form.addRow("Height", height_wrap)
        style_form.addRow("Font", font_wrap)
        style_form.addRow("Font Family", self.style_font_family_combo)
        style_form.addRow("Shape", self.style_shape_combo)
        style_form.addRow("Opacity", opacity_wrap)
        layout.addLayout(style_form)

        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(False)
        self.scroll.setFrameShape(QFrame.Shape.NoFrame)
        self.scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.scroll.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop)
        self.canvas = QuickButtonCanvas()
        self.canvas.setObjectName("QuickLayoutCanvas")
        self.canvas.configure_grid(show_grid=False, snap_enabled=False)
        self.canvas.set_background_drawer(self._paint_canvas_background)
        self.scroll.setWidget(self.canvas)
        layout.addWidget(self.scroll, 1)

        actions = QHBoxLayout()
        self.reset_button = QPushButton("Auto Layout")
        self.done_button = QPushButton("Done")
        self.reset_button.setProperty("actionRole", "pick")
        self.done_button.setProperty("actionRole", "save")
        actions.addWidget(self.reset_button)
        actions.addStretch(1)
        actions.addWidget(self.done_button)
        layout.addLayout(actions)

        self.scroll.viewport().installEventFilter(self)
        self.style_width_slider.valueChanged.connect(self._on_style_controls_changed)
        self.style_height_slider.valueChanged.connect(self._on_style_controls_changed)
        self.style_font_slider.valueChanged.connect(self._on_style_controls_changed)
        self.style_font_family_combo.currentFontChanged.connect(lambda _font: self._on_style_controls_changed())
        self.style_shape_combo.currentTextChanged.connect(self._on_style_controls_changed)
        self.style_opacity_slider.valueChanged.connect(self._on_style_controls_changed)
        self.reset_button.clicked.connect(self.reset_auto_layout)
        self.done_button.clicked.connect(self.close)
        self._sync_style_controls_from_config()
        self.apply_theme_styles()
        self.refresh_cards()

    def closeEvent(self, event) -> None:  # noqa: N802
        self.app_window.config.setdefault("popup_positions", {})["quick_layout"] = {
            "x": self.x(),
            "y": self.y(),
        }
        self.app_window.queue_save_config()
        self.app_window.refresh_quick_grid()
        super().closeEvent(event)

    def eventFilter(self, watched, event) -> bool:  # noqa: N802
        if watched is self.scroll.viewport() and event.type() == QEvent.Type.Resize:
            self._sync_canvas_viewport_width()
            return False
        return super().eventFilter(watched, event)

    def _reference_main_canvas_width(self) -> int:
        try:
            return max(120, int(self.app_window.quick_scroll.viewport().width()))
        except Exception:
            return max(120, int(self.scroll.viewport().width()))

    def _sync_canvas_viewport_width(self) -> None:
        popup_viewport_w = max(120, int(self.scroll.viewport().width()))
        main_w = self._reference_main_canvas_width()
        target_w = min(popup_viewport_w, main_w)
        self.canvas.set_viewport_width(target_w)

    def apply_theme_styles(self) -> None:
        p = self.app_window.palette_data
        bg = rgba_css(p["shell_overlay"], 0.92)
        border = shift(p["control_bg"], -0.30)
        self.setStyleSheet(
            "QDialog {"
            f"background: {bg};"
            f"color: {p['label_text']};"
            "font-family: 'Segoe UI';"
            "font-size: 13px;"
            "}"
            "QLabel { background: transparent; color: inherit; font-weight: 700; }"
            "QScrollArea { background: transparent; border: none; }"
            "QComboBox, QLineEdit, QSpinBox {"
            f"background: {rgba_css(p['input_bg'], 0.86)};"
            f"border: 1px solid {shift(p['input_bg'], -0.38)};"
            "border-radius: 3px;"
            "padding: 2px 6px;"
            "}"
            "QSlider::groove:horizontal {"
            "background: rgba(0,0,0,90);"
            "height: 6px;"
            "border-radius: 3px;"
            "}"
            "QSlider::handle:horizontal {"
            f"background: {p['accent']};"
            "width: 12px;"
            "margin: -4px 0px;"
            "border-radius: 3px;"
            "}"
            "QPushButton {"
            f"background-color: {rgba_css(p['button_bg'], 0.78)};"
            f"color: {p['button_text']};"
            f"border: 1px solid {shift(p['button_bg'], -0.40)};"
            "border-radius: 4px;"
            "padding: 4px 10px;"
            "min-height: 30px;"
            "font-size: 11px;"
            "font-weight: 700;"
            "}"
            "QPushButton[actionRole='save'] {"
            f"background-color: {rgba_css(p['accent'], 0.78)};"
            f"color: {readable_text(p['accent'])};"
            f"border: 1px solid {shift(p['accent'], -0.42)};"
            "}"
        )
        self.canvas.setStyleSheet(
            "QWidget#QuickLayoutCanvas {"
            "background: transparent;"
            f"border: 1px solid {border};"
            "border-radius: 4px;"
            "}"
        )
        self.canvas.update()

    def _sync_style_controls_from_config(self) -> None:
        self.style_width_slider.blockSignals(True)
        self.style_height_slider.blockSignals(True)
        self.style_font_slider.blockSignals(True)
        self.style_font_family_combo.blockSignals(True)
        self.style_shape_combo.blockSignals(True)
        self.style_opacity_slider.blockSignals(True)

        width = int(clamp(int(self.app_window.config.get("quick_button_width", 140)), 90, 220))
        height = int(clamp(int(self.app_window.config.get("quick_button_height", 40)), 35, 100))
        font_size = int(clamp(int(self.app_window.config.get("quick_button_font_size", 11)), 8, 20))
        font_family = str(self.app_window.config.get("quick_button_font_family", "Segoe UI"))
        shape = self.app_window.config.get("quick_button_shape", "Soft")
        if self.style_shape_combo.findText(shape) == -1:
            shape = "Soft"
        opacity = float(clamp(float(self.app_window.config.get("quick_button_opacity", 1.0)), 0.2, 1.0))

        self.style_width_slider.setValue(width)
        self.style_height_slider.setValue(height)
        self.style_font_slider.setValue(font_size)
        self.style_font_family_combo.setCurrentFont(QFont(font_family))
        self.style_shape_combo.setCurrentText(shape)
        self.style_opacity_slider.setValue(int(opacity * 100))

        self.style_width_slider.blockSignals(False)
        self.style_height_slider.blockSignals(False)
        self.style_font_slider.blockSignals(False)
        self.style_font_family_combo.blockSignals(False)
        self.style_shape_combo.blockSignals(False)
        self.style_opacity_slider.blockSignals(False)

        self.style_width_value.setText(str(width))
        self.style_height_value.setText(str(height))
        self.style_font_value.setText(str(font_size))
        self.style_opacity_value.setText(f"{opacity:.2f}")

    def _on_style_controls_changed(self) -> None:
        width = int(self.style_width_slider.value())
        height = int(self.style_height_slider.value())
        font_size = int(self.style_font_slider.value())
        font_family = str(self.style_font_family_combo.currentFont().family() or "Segoe UI")
        shape = self.style_shape_combo.currentText()
        opacity = float(clamp(self.style_opacity_slider.value() / 100.0, 0.2, 1.0))

        self.app_window.config["quick_button_width"] = width
        self.app_window.config["quick_button_height"] = height
        self.app_window.config["quick_button_font_size"] = font_size
        self.app_window.config["quick_button_font_family"] = font_family
        self.app_window.config["quick_button_shape"] = shape
        self.app_window.config["quick_button_opacity"] = opacity

        self.style_width_value.setText(str(width))
        self.style_height_value.setText(str(height))
        self.style_font_value.setText(str(font_size))
        self.style_opacity_value.setText(f"{opacity:.2f}")

        self.app_window._refresh_theme_preview_buttons()
        self.app_window.refresh_quick_grid()
        self.refresh_cards()
        self.app_window.queue_save_config()

    def _default_position(self, index: int, width: int, height: int) -> tuple[int, int]:
        return self.app_window._default_quick_position(index, width, height)

    def _paint_canvas_background(self, painter: QPainter, rect: QRect) -> None:
        base = QColor(self.app_window.palette_data["surface"])
        base.setAlpha(110)
        painter.fillRect(rect, base)

        surface_size = self.app_window.surface.size()
        if surface_size.width() <= 0 or surface_size.height() <= 0:
            return
        bg = self.app_window.render_background_pixmap(surface_size)
        if bg.isNull():
            return

        source_offset = self.app_window.quick_scroll.viewport().mapTo(self.app_window.surface, QPoint(0, 0))
        painter.save()
        painter.setClipRect(rect)
        painter.drawPixmap(-source_offset.x(), -source_offset.y(), bg)
        overlay = QColor(self.app_window.palette_data["shell_overlay"])
        overlay.setAlpha(24)
        painter.fillRect(rect, overlay)
        painter.restore()

    def refresh_cards(self) -> None:
        self._sync_style_controls_from_config()
        self.canvas.clear_cards()
        self.canvas.clear_alignment_guides()
        self.canvas.configure_grid(show_grid=False, snap_enabled=False)
        self._sync_canvas_viewport_width()
        can_persist_positions = bool(
            self.isVisible() and self.scroll.viewport().width() > 120 and self._reference_main_canvas_width() > 120
        )

        quick_texts = self.app_window._active_quick_texts()
        if not quick_texts:
            self.canvas.set_placeholder("No quick text buttons yet.", self.app_window.palette_data["muted_text"])
            return

        width = int(self.app_window.config.get("quick_button_width", 140))
        height = int(self.app_window.config.get("quick_button_height", 40))
        font_size = int(self.app_window.config.get("quick_button_font_size", 11))
        font_family = str(self.app_window.config.get("quick_button_font_family", "Segoe UI"))
        shape = self.app_window.config.get("quick_button_shape", "Soft")
        button_opacity = float(clamp(float(self.app_window.config.get("quick_button_opacity", 1.0)), 0.15, 1.0))
        updated_positions = False

        for idx, item in enumerate(quick_texts):
            card = QuickButtonCard(
                idx,
                str(item.get("title", "Untitled"))[:28],
                str(item.get("tooltip", "")),
                self.canvas,
            )
            action_type = self.app_window._quick_action_kind(item)
            card.apply_visual_style(width, height, font_size, font_family, shape, button_opacity, self.app_window.palette_data, action_type)
            card.set_layout_mode(True)
            card.edit_button.hide()
            card.move_requested.connect(self.on_card_move)
            raw_x = item.get("x")
            raw_y = item.get("y")
            if isinstance(raw_x, (int, float)) and isinstance(raw_y, (int, float)):
                pos_x, pos_y = int(raw_x), int(raw_y)
            else:
                pos_x, pos_y = self._default_position(idx, width, height)
                if can_persist_positions:
                    item["x"] = int(pos_x)
                    item["y"] = int(pos_y)
                    updated_positions = True

            px, py = self.canvas.place_card(card, pos_x, pos_y, snap=False)
            if (
                can_persist_positions
                and (
                    safe_int(item.get("x", -99999), -99999) != px
                    or safe_int(item.get("y", -99999), -99999) != py
                )
            ):
                item["x"] = int(px)
                item["y"] = int(py)
                updated_positions = True

        if updated_positions:
            self.app_window.queue_save_config()

    def _alignment_guides_for(self, moving_index: int, x: int, y: int, w: int, h: int) -> tuple[list[int], list[int]]:
        tolerance = 8
        moving_left = x
        moving_right = x + w - 1
        moving_cx = x + (w // 2)
        moving_top = y
        moving_bottom = y + h - 1
        moving_cy = y + (h // 2)

        v_lines: set[int] = set()
        h_lines: set[int] = set()

        center_x = self.canvas.width() // 2
        center_y = self.canvas.height() // 2
        if abs(moving_cx - center_x) <= tolerance:
            v_lines.add(center_x)
        if abs(moving_cy - center_y) <= tolerance:
            h_lines.add(center_y)

        for rect in self.canvas.iter_card_geometries(exclude_index=moving_index):
            other_left = rect.left()
            other_right = rect.right()
            other_cx = rect.left() + (rect.width() // 2)
            other_top = rect.top()
            other_bottom = rect.bottom()
            other_cy = rect.top() + (rect.height() // 2)

            for moving_anchor in (moving_left, moving_cx, moving_right):
                for other_anchor in (other_left, other_cx, other_right):
                    if abs(moving_anchor - other_anchor) <= tolerance:
                        v_lines.add(int(other_anchor))
            for moving_anchor in (moving_top, moving_cy, moving_bottom):
                for other_anchor in (other_top, other_cy, other_bottom):
                    if abs(moving_anchor - other_anchor) <= tolerance:
                        h_lines.add(int(other_anchor))

        return sorted(v_lines), sorted(h_lines)

    @staticmethod
    def _snap_axis_to_lines(start: int, size: int, guide_lines: list[int], tolerance: int = 12) -> int:
        if not guide_lines:
            return start
        anchors = [start, start + (size // 2), start + size - 1]
        best_delta: int | None = None
        for anchor in anchors:
            for line in guide_lines:
                delta = int(line - anchor)
                if abs(delta) > tolerance:
                    continue
                if best_delta is None or abs(delta) < abs(best_delta):
                    best_delta = delta
        if best_delta is None:
            return start
        return start + best_delta

    def on_card_move(self, index: int, x: int, y: int, finished: bool) -> None:
        gx, gy = self.canvas.move_card(index, x, y, snap=False)
        card_rect = self.canvas.card_geometry(index)
        if card_rect is not None:
            v_lines, h_lines = self._alignment_guides_for(index, gx, gy, card_rect.width(), card_rect.height())
            if bool(QApplication.keyboardModifiers() & Qt.KeyboardModifier.ShiftModifier):
                snapped_x = self._snap_axis_to_lines(gx, card_rect.width(), v_lines)
                snapped_y = self._snap_axis_to_lines(gy, card_rect.height(), h_lines)
                if snapped_x != gx or snapped_y != gy:
                    gx, gy = self.canvas.move_card(index, snapped_x, snapped_y, snap=False)
                    card_rect = self.canvas.card_geometry(index)
                    if card_rect is not None:
                        v_lines, h_lines = self._alignment_guides_for(index, gx, gy, card_rect.width(), card_rect.height())
            self.canvas.set_alignment_guides(v_lines, h_lines)
        if finished:
            self.canvas.clear_alignment_guides()

        quick_texts = self.app_window._active_quick_texts()
        if 0 <= index < len(quick_texts):
            entry = quick_texts[index]
            entry["x"] = int(gx)
            entry["y"] = int(gy)
            if finished:
                self.app_window.queue_save_config()
                self.app_window.refresh_quick_grid()

    def reset_auto_layout(self) -> None:
        quick_texts = self.app_window._active_quick_texts()
        width = int(self.app_window.config.get("quick_button_width", 140))
        height = int(self.app_window.config.get("quick_button_height", 40))
        for idx, entry in enumerate(quick_texts):
            x, y = self._default_position(idx, width, height)
            entry["x"] = int(x)
            entry["y"] = int(y)
        self.app_window.queue_save_config()
        self.refresh_cards()
        self.app_window.refresh_quick_grid()


class ImageLayerPreview(QWidget):
    layer_changed = Signal(dict)

    def __init__(self, app_window: "QuickInputsWindow", get_layer: Callable[[], dict[str, Any] | None], kind: str = "main") -> None:
        super().__init__()
        self.app_window = app_window
        self.get_layer = get_layer
        self.kind = kind
        self._dragging = False
        self._drag_start = QPointF(0.0, 0.0)
        self._layer_start = QPointF(0.0, 0.0)
        self._drag_scale = 1.0
        self.setMinimumSize(260, 190)

    def _virtual_size(self) -> QSize:
        if self.kind == "main":
            size = self.app_window.surface.size()
            if size.width() <= 0 or size.height() <= 0:
                return QSize(LAUNCH_WIDTH, LAUNCH_HEIGHT)
            return size

        if self.kind == "agent":
            active_agent = getattr(self.app_window, "active_agent_window", None)
            if active_agent is not None and active_agent.isVisible():
                size = active_agent.size()
                if size.width() > 0 and size.height() > 0:
                    return size
            return QSize(460, 380)

        if self.kind == "qa":
            active_qa = getattr(self.app_window, "active_qa_window", None)
            if active_qa is not None and active_qa.isVisible():
                size = active_qa.size()
                if size.width() > 0 and size.height() > 0:
                    return size
            return QSize(500, 420)
        
        if self.kind == "dashboard":
            active_dashboard = getattr(self.app_window, "depot_dashboard_dialog", None)
            if active_dashboard is not None and active_dashboard.isVisible():
                size = active_dashboard.size()
                if size.width() > 0 and size.height() > 0:
                    return size
            return QSize(780, 420)

        if self.kind == "admin":
            active_admin = getattr(self.app_window, "admin_dialog", None)
            if active_admin is not None and active_admin.isVisible():
                size = active_admin.size()
                if size.width() > 0 and size.height() > 0:
                    return size
            return QSize(620, 500)

        return QSize(LAUNCH_WIDTH, LAUNCH_HEIGHT)

    def _mapping(self) -> tuple[QSize, QRectF, float]:
        virtual_size = self._virtual_size()
        outer = QRectF(self.rect())
        if virtual_size.width() <= 0 or virtual_size.height() <= 0 or outer.width() <= 0 or outer.height() <= 0:
            return virtual_size, outer, 1.0

        scale = min(outer.width() / virtual_size.width(), outer.height() / virtual_size.height())
        draw_w = virtual_size.width() * scale
        draw_h = virtual_size.height() * scale
        draw_rect = QRectF(
            outer.left() + (outer.width() - draw_w) / 2.0,
            outer.top() + (outer.height() - draw_h) / 2.0,
            draw_w,
            draw_h,
        )
        return virtual_size, draw_rect, scale

    @staticmethod
    def _virtual_rect_to_preview(rect: QRectF, preview_rect: QRectF, scale: float) -> QRectF:
        if scale <= 0:
            return QRectF()
        return QRectF(
            preview_rect.left() + rect.left() * scale,
            preview_rect.top() + rect.top() * scale,
            rect.width() * scale,
            rect.height() * scale,
        )

    def paintEvent(self, event) -> None:  # noqa: N802
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform)

        virtual_size, draw_rect, scale = self._mapping()
        frame_tint = QColor(self.app_window.palette_data["shell_overlay"])
        frame_tint.setAlpha(72)
        painter.fillRect(self.rect(), frame_tint)
        bg = self.app_window.render_background_pixmap(virtual_size, kind=self.kind)
        if not bg.isNull():
            painter.drawPixmap(draw_rect.toRect(), bg)

        layer = self.get_layer()
        if not layer:
            return

        render_info = self.app_window.compute_layer_render(layer, virtual_size)
        if not render_info:
            return
        layer_rect = self._virtual_rect_to_preview(render_info.rect, draw_rect, scale)

        pen = QPen(QColor("#FFFFFF"))
        pen.setWidth(2)
        painter.setPen(pen)
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.drawRect(layer_rect)

    def mousePressEvent(self, event: QMouseEvent) -> None:  # noqa: N802
        layer = self.get_layer()
        if not layer:
            return

        virtual_size, draw_rect, scale = self._mapping()
        info = self.app_window.compute_layer_render(layer, virtual_size)
        if not info:
            return
        layer_rect = self._virtual_rect_to_preview(info.rect, draw_rect, scale)

        if event.button() == Qt.MouseButton.LeftButton and layer_rect.contains(event.position()):
            self._dragging = True
            self._drag_start = event.position()
            self._layer_start = QPointF(float(layer.get("image_x", 0)), float(layer.get("image_y", 0)))
            self._drag_scale = scale if scale > 0 else 1.0
            self.setCursor(Qt.CursorShape.ClosedHandCursor)

    def mouseMoveEvent(self, event: QMouseEvent) -> None:  # noqa: N802
        if not self._dragging:
            return

        layer = self.get_layer()
        if not layer:
            return

        delta = event.position() - self._drag_start
        dx = delta.x() / self._drag_scale
        dy = delta.y() / self._drag_scale
        layer["image_x"] = int(round(self._layer_start.x() + dx))
        layer["image_y"] = int(round(self._layer_start.y() + dy))
        self.layer_changed.emit(layer)
        self.update()

    def mouseReleaseEvent(self, event: QMouseEvent) -> None:  # noqa: N802
        if event.button() == Qt.MouseButton.LeftButton:
            self._dragging = False
            self.setCursor(Qt.CursorShape.ArrowCursor)


class ImageLayersDialog(QDialog):
    def __init__(self, app_window: "QuickInputsWindow", kind: str = "main") -> None:
        super().__init__(app_window)
        self.app_window = app_window
        self.kind = kind
        title = "Image Layers - Flowgrid" if kind == "main" else f"Image Layers - {kind.title()}"
        self.setWindowTitle(title)
        self.setMinimumSize(500, 380)
        self.setModal(False)
        self.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose, False)

        layout = QHBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)

        left = QVBoxLayout()
        self.layer_list = QListWidget()
        left.addWidget(self.layer_list, 1)

        row_buttons = QHBoxLayout()
        self.add_button = QPushButton("Add")
        self.remove_button = QPushButton("Remove")
        self.up_button = QPushButton("Up")
        self.down_button = QPushButton("Down")
        self.add_button.setProperty("actionRole", "add")
        self.remove_button.setProperty("actionRole", "reset")
        self.up_button.setProperty("actionRole", "pick")
        self.down_button.setProperty("actionRole", "pick")
        row_buttons.addWidget(self.add_button)
        row_buttons.addWidget(self.remove_button)
        row_buttons.addWidget(self.up_button)
        row_buttons.addWidget(self.down_button)
        left.addLayout(row_buttons)

        self.visible_check = QCheckBox("Visible")
        left.addWidget(self.visible_check)

        form = QFormLayout()
        form.setContentsMargins(0, 0, 0, 0)
        form.setSpacing(6)

        self.x_spin = QSpinBox()
        self.x_spin.setRange(-3000, 3000)
        self.y_spin = QSpinBox()
        self.y_spin.setRange(-3000, 3000)

        self.scale_mode = QComboBox()
        self.scale_mode.addItems(["Fill", "Fit", "Stretch", "Place"])

        self.anchor_combo = QComboBox()
        self.anchor_combo.addItems(["TopLeft", "Top", "TopRight", "Left", "Center", "Right", "BottomLeft", "Bottom", "BottomRight"])

        self.scale_spin = QSpinBox()
        self.scale_spin.setRange(10, 400)

        self.opacity_slider = QSlider(Qt.Orientation.Horizontal)
        self.opacity_slider.setRange(0, 100)

        form.addRow("X", self.x_spin)
        form.addRow("Y", self.y_spin)
        form.addRow("Scale", self.scale_mode)
        form.addRow("Anchor", self.anchor_combo)
        form.addRow("Scale %", self.scale_spin)
        form.addRow("Opacity", self.opacity_slider)

        left.addLayout(form)
        layout.addLayout(left, 0)

        right = QVBoxLayout()
        self.preview = ImageLayerPreview(app_window, self.current_layer, kind=self.kind)
        right.addWidget(self.preview, 1)

        hint = QLabel("Tip: drag the highlighted image in preview to reposition it.")
        hint.setWordWrap(True)
        right.addWidget(hint)

        layout.addLayout(right, 1)

        self.layer_list.currentRowChanged.connect(self._on_layer_selected)
        self.add_button.clicked.connect(self._add_image)
        self.remove_button.clicked.connect(self._remove_selected)
        self.up_button.clicked.connect(lambda: self._move_selected(-1))
        self.down_button.clicked.connect(lambda: self._move_selected(1))

        self.visible_check.toggled.connect(lambda value: self._update_layer_field("visible", value))
        self.x_spin.valueChanged.connect(lambda value: self._update_layer_field("image_x", value))
        self.y_spin.valueChanged.connect(lambda value: self._update_layer_field("image_y", value))
        self.scale_mode.currentTextChanged.connect(lambda value: self._update_layer_field("image_scale_mode", value))
        self.anchor_combo.currentTextChanged.connect(lambda value: self._update_layer_field("image_anchor", value))
        self.scale_spin.valueChanged.connect(lambda value: self._update_layer_field("image_scale_percent", value))
        self.opacity_slider.valueChanged.connect(lambda value: self._update_layer_field("image_opacity", value / 100.0))

        self.preview.layer_changed.connect(self._on_layer_dragged)
        self.apply_theme_styles()
        self.refresh_list()

    def apply_theme_styles(self) -> None:
        p = self.app_window.palette_data
        dialog_bg = rgba_css(p["shell_overlay"], 0.92)
        list_bg = rgba_css(p["input_bg"], 0.42)
        input_bg = rgba_css(p["input_bg"], 0.86)
        border = shift(p["control_bg"], -0.30)
        field_border = shift(p["input_bg"], -0.38)
        button_bg = rgba_css(p["button_bg"], 0.78)
        button_hover = rgba_css(shift(p["button_bg"], 0.08), 0.86)
        self.setStyleSheet(
            "QDialog {"
            f"background: {dialog_bg};"
            f"color: {p['label_text']};"
            "font-family: 'Segoe UI';"
            "font-size: 13px;"
            "}"
            "QLabel { background: transparent; color: inherit; font-weight: 700; }"
            "QListWidget {"
            f"background: {list_bg};"
            f"border: 1px solid {field_border};"
            "border-radius: 4px;"
            "padding: 4px;"
            "}"
            "QLineEdit, QTextEdit, QSpinBox, QComboBox {"
            f"background: {input_bg};"
            f"border: 1px solid {field_border};"
            "border-radius: 3px;"
            "padding: 2px 6px;"
            "}"
            "QCheckBox { background: transparent; spacing: 8px; font-weight: 700; }"
            "QPushButton {"
            f"background: {button_bg};"
            f"color: {p['button_text']};"
            f"border: 1px solid {shift(p['button_bg'], -0.40)};"
            "border-radius: 4px;"
            "padding: 4px 10px;"
            "min-height: 28px;"
            "font-size: 11px;"
            "font-weight: 700;"
            "}"
            f"QPushButton:hover {{ background: {button_hover}; }}"
            "QPushButton[actionRole='add'] {"
            f"background-color: {rgba_css(p['primary'], 0.78)};"
            f"color: {readable_text(p['primary'])};"
            f"border: 1px solid {shift(p['primary'], -0.42)};"
            "}"
            "QPushButton[actionRole='reset'] {"
            f"background-color: {rgba_css(p['accent'], 0.78)};"
            f"color: {readable_text(p['accent'])};"
            f"border: 1px solid {shift(p['accent'], -0.42)};"
            "}"
        )
        self.preview.setStyleSheet(
            "QWidget {"
            "background: transparent;"
            f"border: 1px solid {border};"
            "border-radius: 4px;"
            "}"
        )
        self.preview.update()

    def _layers_key(self) -> str:
        if self.kind == "main":
            return "theme_image_layers"
        else:
            return f"{self.kind}_theme"

    def _popup_uses_inherited_layers(self, theme: dict[str, Any]) -> bool:
        if self.kind == "main":
            return False
        if not isinstance(theme, dict):
            return True
        if bool(theme.get("inherit_main_theme", False)):
            return True
        if self.app_window._looks_like_unconfigured_popup_theme(theme):
            return True
        return False

    def _materialize_popup_layers_for_edit(self) -> None:
        if self.kind == "main":
            return
        key = self._layers_key()
        theme = self.app_window.config.setdefault(key, {})
        if not isinstance(theme, dict):
            theme = {}
            self.app_window.config[key] = theme
        if not self._popup_uses_inherited_layers(theme):
            return
        effective_layers = self._get_layers()
        theme["image_layers"] = [
            safe_layer_defaults(layer) for layer in effective_layers if isinstance(layer, dict)
        ]
        theme["inherit_main_theme"] = False

    def _get_layers(self) -> list[dict[str, Any]]:
        key = self._layers_key()
        if self.kind == "main":
            return self.app_window.config.get(key, [])
        else:
            theme = self.app_window.config.get(key, {})
            if self._popup_uses_inherited_layers(theme if isinstance(theme, dict) else {}):
                resolved = self.app_window._resolved_popup_theme(self.kind)
                inherited_layers = resolved.get("image_layers", [])
                if isinstance(inherited_layers, list):
                    return [
                        safe_layer_defaults(layer)
                        for layer in inherited_layers
                        if isinstance(layer, dict)
                    ]
                return []
            if isinstance(theme, dict):
                raw_layers = theme.get("image_layers", [])
                if isinstance(raw_layers, list):
                    return raw_layers
            return []

    def _set_layers(self, layers: list[dict[str, Any]]) -> None:
        key = self._layers_key()
        if self.kind == "main":
            self.app_window.config[key] = layers
        else:
            theme = self.app_window.config.setdefault(key, {})
            if not isinstance(theme, dict):
                theme = {}
                self.app_window.config[key] = theme
            theme["image_layers"] = layers
            theme["inherit_main_theme"] = False

    def closeEvent(self, event) -> None:  # noqa: N802
        popup_positions = self.app_window.config.setdefault("popup_positions", {})
        if self.kind == "main":
            popup_positions["image_layers"] = {"x": self.x(), "y": self.y()}
        else:
            popup_positions[f"image_layers_{self.kind}"] = {"x": self.x(), "y": self.y()}
        self.app_window.queue_save_config()
        super().closeEvent(event)

    def refresh_list(self) -> None:
        current = self.layer_list.currentRow()
        self.layer_list.blockSignals(True)
        self.layer_list.clear()
        for layer in self._get_layers():
            item = QListWidgetItem(layer.get("name") or "Layer")
            self.layer_list.addItem(item)
        self.layer_list.blockSignals(False)

        if self.layer_list.count() == 0:
            self._load_layer_to_controls(None)
            return

        if current < 0:
            current = 0
        current = int(clamp(current, 0, self.layer_list.count() - 1))
        self.layer_list.setCurrentRow(current)

    def current_layer(self) -> dict[str, Any] | None:
        row = self.layer_list.currentRow()
        layers = self._get_layers()
        if row < 0 or row >= len(layers):
            return None
        return layers[row]

    def _on_layer_selected(self, _row: int) -> None:
        self._load_layer_to_controls(self.current_layer())

    def _load_layer_to_controls(self, layer: dict[str, Any] | None) -> None:
        controls: list[QWidget] = [
            self.visible_check,
            self.x_spin,
            self.y_spin,
            self.scale_mode,
            self.anchor_combo,
            self.scale_spin,
            self.opacity_slider,
            self.remove_button,
            self.up_button,
            self.down_button,
        ]
        enabled = layer is not None
        for control in controls:
            control.setEnabled(enabled)

        if not layer:
            self.preview.update()
            return

        self.visible_check.blockSignals(True)
        self.x_spin.blockSignals(True)
        self.y_spin.blockSignals(True)
        self.scale_mode.blockSignals(True)
        self.anchor_combo.blockSignals(True)
        self.scale_spin.blockSignals(True)
        self.opacity_slider.blockSignals(True)

        self.visible_check.setChecked(bool(layer.get("visible", True)))
        self.x_spin.setValue(int(layer.get("image_x", 0)))
        self.y_spin.setValue(int(layer.get("image_y", 0)))
        self.scale_mode.setCurrentText(layer.get("image_scale_mode", "Fill"))
        self.anchor_combo.setCurrentText(layer.get("image_anchor", "Center"))
        self.scale_spin.setValue(int(layer.get("image_scale_percent", 100)))
        self.opacity_slider.setValue(int(float(layer.get("image_opacity", 1.0)) * 100))

        self.visible_check.blockSignals(False)
        self.x_spin.blockSignals(False)
        self.y_spin.blockSignals(False)
        self.scale_mode.blockSignals(False)
        self.anchor_combo.blockSignals(False)
        self.scale_spin.blockSignals(False)
        self.opacity_slider.blockSignals(False)

        self.preview.update()

    def _add_image(self) -> None:
        files, _ = show_flowgrid_themed_open_file_names(
            self,
            self.app_window,
            self.kind,
            "Add Image Layers",
            str(Path.home()),
            "Images (*.png *.jpg *.jpeg *.bmp *.gif *.webp);;All Files (*.*)",
        )
        if not files:
            return

        for file_path in files:
            layer = safe_layer_defaults(
                {
                    "image_path": file_path,
                    "name": Path(file_path).name,
                }
            )
            layers = self._get_layers()
            layers.append(layer)
            self._set_layers(layers)

        self.app_window.mark_background_dirty()
        self.app_window.queue_save_config()
        self.refresh_list()
        self.layer_list.setCurrentRow(self.layer_list.count() - 1)
        self.app_window.refresh_all_views()

    def _remove_selected(self) -> None:
        row = self.layer_list.currentRow()
        if row < 0:
            return

        layers = self._get_layers()
        if 0 <= row < len(layers):
            layers.pop(row)
            self._set_layers(layers)
            self.app_window.mark_background_dirty()
            self.app_window.queue_save_config()
            self.refresh_list()
            self.app_window.refresh_all_views()

    def _move_selected(self, direction: int) -> None:
        row = self.layer_list.currentRow()
        layers = self._get_layers()
        new_index = row + direction
        if row < 0 or new_index < 0 or new_index >= len(layers):
            return

        layers[row], layers[new_index] = layers[new_index], layers[row]
        self._set_layers(layers)
        self.refresh_list()
        self.layer_list.setCurrentRow(new_index)
        self.app_window.mark_background_dirty()
        self.app_window.queue_save_config()
        self.app_window.refresh_all_views()

    def _update_layer_field(self, field: str, value: Any) -> None:
        self._materialize_popup_layers_for_edit()
        layer = self.current_layer()
        if not layer:
            return

        layer[field] = value
        self.app_window.mark_background_dirty()
        self.app_window.queue_save_config()
        self.preview.update()
        self.app_window.refresh_all_views()

    def _on_layer_dragged(self, layer: dict[str, Any]) -> None:
        self._materialize_popup_layers_for_edit()
        editable_layer = self.current_layer()
        if editable_layer is not None:
            editable_layer["image_x"] = int(layer.get("image_x", 0))
            editable_layer["image_y"] = int(layer.get("image_y", 0))
            layer = editable_layer
        self.x_spin.blockSignals(True)
        self.y_spin.blockSignals(True)
        self.x_spin.setValue(int(layer.get("image_x", 0)))
        self.y_spin.setValue(int(layer.get("image_y", 0)))
        self.x_spin.blockSignals(False)
        self.y_spin.blockSignals(False)

        self.app_window.mark_background_dirty()
        self.app_window.queue_save_config()
        self.app_window.refresh_all_views()

# -------------------------- Depot Tracker Data Layer --------------------------

class DepotDB:
    def __init__(self, db_path: Path):
        self.db_path = db_path
        self.conn = sqlite3.connect(str(db_path), check_same_thread=False)
        self.conn.row_factory = sqlite3.Row
        self._create_tables()

    def _create_tables(self) -> None:
        cursor = self.conn.cursor()
        cursor.execute("PRAGMA foreign_keys = ON")

        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS submissions (
                id INTEGER PRIMARY KEY,
                created_at TEXT NOT NULL,
                user_id TEXT NOT NULL,
                work_order TEXT NOT NULL,
                touch TEXT NOT NULL,
                client_unit INTEGER NOT NULL DEFAULT 0,
                entry_date TEXT NOT NULL,
                part_order_count INTEGER NOT NULL DEFAULT 0
            )
            """
        )
        cursor.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_submissions_user ON submissions(user_id)
            """
        )
        cursor.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_submissions_work_order ON submissions(work_order)
            """
        )
        cursor.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_submissions_touch ON submissions(touch)
            """
        )
        cursor.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_submissions_entry_date ON submissions(entry_date)
            """
        )

        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS parts (
                id INTEGER PRIMARY KEY,
                created_at TEXT NOT NULL,
                user_id TEXT NOT NULL,
                assigned_user_id TEXT NOT NULL,
                work_order TEXT NOT NULL,
                client_unit INTEGER NOT NULL DEFAULT 0,
                category TEXT NOT NULL,
                comments TEXT,
                is_active INTEGER NOT NULL DEFAULT 1,
                working_user_id TEXT NOT NULL DEFAULT '',
                working_updated_at TEXT NOT NULL DEFAULT '',
                parts_on_hand INTEGER NOT NULL DEFAULT 0,
                parts_installed INTEGER NOT NULL DEFAULT 0,
                parts_installed_by TEXT NOT NULL DEFAULT '',
                parts_installed_at TEXT NOT NULL DEFAULT ''
            )
            """
        )
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS part_details (
                id INTEGER PRIMARY KEY,
                part_id INTEGER NOT NULL UNIQUE,
                created_at TEXT NOT NULL,
                lpn TEXT NOT NULL,
                part_number TEXT NOT NULL DEFAULT '',
                part_description TEXT NOT NULL DEFAULT '',
                shipping_info TEXT NOT NULL DEFAULT '',
                installed_keys TEXT NOT NULL DEFAULT '',
                delivered INTEGER NOT NULL DEFAULT 0,
                FOREIGN KEY(part_id) REFERENCES parts(id) ON DELETE CASCADE
            )
            """
        )
        cursor.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_part_details_lpn ON part_details(lpn)
            """
        )
        cursor.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_parts_user ON parts(user_id)
            """
        )
        cursor.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_parts_work_order ON parts(work_order)
            """
        )

        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS rtvs (
                id INTEGER PRIMARY KEY,
                created_at TEXT NOT NULL,
                user_id TEXT NOT NULL,
                work_order TEXT NOT NULL,
                comments TEXT
            )
            """
        )

        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS client_jo (
                id INTEGER PRIMARY KEY,
                created_at TEXT NOT NULL,
                user_id TEXT NOT NULL,
                work_order TEXT NOT NULL,
                comments TEXT
            )
            """
        )

        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS client_parts (
                id INTEGER PRIMARY KEY,
                created_at TEXT NOT NULL,
                user_id TEXT NOT NULL,
                assigned_user_id TEXT,
                work_order TEXT NOT NULL UNIQUE,
                comments TEXT,
                followup_last_action TEXT NOT NULL DEFAULT '',
                followup_last_action_at TEXT NOT NULL DEFAULT '',
                followup_last_actor TEXT NOT NULL DEFAULT '',
                followup_no_contact_count INTEGER NOT NULL DEFAULT 0,
                followup_stage_logged INTEGER NOT NULL DEFAULT -1
            )
            """
        )

        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS agents (
                id INTEGER PRIMARY KEY,
                agent_name TEXT NOT NULL,
                user_id TEXT NOT NULL UNIQUE,
                tier INTEGER NOT NULL DEFAULT 1,
                location TEXT NOT NULL DEFAULT '',
                icon_path TEXT NOT NULL DEFAULT ''
            )
            """
        )

        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS admin_users (
                id INTEGER PRIMARY KEY,
                user_id TEXT NOT NULL UNIQUE,
                admin_name TEXT NOT NULL DEFAULT '',
                position TEXT NOT NULL DEFAULT '',
                location TEXT NOT NULL DEFAULT '',
                icon_path TEXT NOT NULL DEFAULT ''
            )
            """
        )

        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS qa_flags (
                id INTEGER PRIMARY KEY,
                flag_name TEXT NOT NULL UNIQUE,
                severity TEXT NOT NULL DEFAULT 'Medium',
                icon_path TEXT NOT NULL DEFAULT '',
                sort_order INTEGER NOT NULL DEFAULT 0
            )
            """
        )
        cursor.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_qa_flags_sort ON qa_flags(sort_order, severity, flag_name)
            """
        )

        self._ensure_column("agents", "tier", "INTEGER NOT NULL DEFAULT 1")
        self._ensure_column("agents", "location", "TEXT NOT NULL DEFAULT ''")
        self._ensure_column("agents", "icon_path", "TEXT NOT NULL DEFAULT ''")
        self._ensure_column("admin_users", "admin_name", "TEXT NOT NULL DEFAULT ''")
        self._ensure_column("admin_users", "position", "TEXT NOT NULL DEFAULT ''")
        self._ensure_column("admin_users", "location", "TEXT NOT NULL DEFAULT ''")
        self._ensure_column("admin_users", "icon_path", "TEXT NOT NULL DEFAULT ''")
        self._ensure_column("parts", "qa_comment", "TEXT NOT NULL DEFAULT ''")
        self._ensure_column("parts", "agent_comment", "TEXT NOT NULL DEFAULT ''")
        self._ensure_column("parts", "qa_flag", "TEXT NOT NULL DEFAULT ''")
        self._ensure_column("parts", "qa_flag_image_path", "TEXT NOT NULL DEFAULT ''")
        self._ensure_column("parts", "working_user_id", "TEXT NOT NULL DEFAULT ''")
        self._ensure_column("parts", "working_updated_at", "TEXT NOT NULL DEFAULT ''")
        self._ensure_column("parts", "parts_on_hand", "INTEGER NOT NULL DEFAULT 0")
        self._ensure_column("parts", "parts_installed", "INTEGER NOT NULL DEFAULT 0")
        self._ensure_column("parts", "parts_installed_by", "TEXT NOT NULL DEFAULT ''")
        self._ensure_column("parts", "parts_installed_at", "TEXT NOT NULL DEFAULT ''")
        self._ensure_column("part_details", "installed_keys", "TEXT NOT NULL DEFAULT ''")
        self._ensure_column("qa_flags", "severity", "TEXT NOT NULL DEFAULT 'Medium'")
        self._ensure_column("qa_flags", "icon_path", "TEXT NOT NULL DEFAULT ''")
        self._ensure_column("qa_flags", "sort_order", "INTEGER NOT NULL DEFAULT 0")
        self._ensure_column("client_parts", "followup_last_action", "TEXT NOT NULL DEFAULT ''")
        self._ensure_column("client_parts", "followup_last_action_at", "TEXT NOT NULL DEFAULT ''")
        self._ensure_column("client_parts", "followup_last_actor", "TEXT NOT NULL DEFAULT ''")
        self._ensure_column("client_parts", "followup_no_contact_count", "INTEGER NOT NULL DEFAULT 0")
        self._ensure_column("client_parts", "followup_stage_logged", "INTEGER NOT NULL DEFAULT -1")

        # Backfill legacy comments into qa_comment when first migrating older DBs.
        try:
            cursor.execute(
                "UPDATE parts SET qa_comment=COALESCE(comments, '') "
                "WHERE (qa_comment IS NULL OR qa_comment='') AND comments IS NOT NULL AND TRIM(comments) <> ''"
            )
            cursor.execute("UPDATE parts SET qa_comment='' WHERE qa_comment IS NULL")
            cursor.execute("UPDATE parts SET agent_comment='' WHERE agent_comment IS NULL")
        except Exception as exc:
            context = {"db_path": str(self.db_path)}
            _runtime_log_event(
                "depot.db.backfill_migration_failed",
                severity="critical",
                summary="DB schema backfill failed while creating or upgrading tables.",
                exc=exc,
                context=context,
            )
            _escalate_runtime_issue_once(
                "depot.db.backfill_migration_failed",
                "Flowgrid database migration encountered an error. Some historical fields may be incomplete.",
                details=f"{type(exc).__name__}: {exc}",
                context=context,
            )

        self.conn.commit()

    def _ensure_column(self, table_name: str, column_name: str, column_sql: str) -> None:
        cursor = self.conn.cursor()
        cursor.execute(f"PRAGMA table_info({table_name})")
        columns = {str(row[1]).lower() for row in cursor.fetchall()}
        if column_name.lower() in columns:
            return
        cursor.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_sql}")

    def execute(self, query: str, params: tuple = ()) -> sqlite3.Cursor:
        cursor = self.conn.cursor()
        cursor.execute(query, params)
        self.conn.commit()
        return cursor

    def fetchall(self, query: str, params: tuple = ()) -> list[sqlite3.Row]:
        cursor = self.conn.cursor()
        cursor.execute(query, params)
        return cursor.fetchall()

    def fetchone(self, query: str, params: tuple = ()) -> sqlite3.Row | None:
        cursor = self.conn.cursor()
        cursor.execute(query, params)
        return cursor.fetchone()


class DepotRules:
    TOUCH_RTV = "RTV"
    TOUCH_COMPLETE = "Complete"
    TOUCH_JUNK = "Junk Out"
    TOUCH_PART_ORDER = "Part Order"
    TOUCH_OTHER = "Other"
    CLOSING_TOUCHES: tuple[str, ...] = (
        TOUCH_COMPLETE,
        TOUCH_JUNK,
        TOUCH_RTV,
    )
    AGENT_TIER_LABELS: dict[int, str] = {
        1: "Tech 1",
        2: "Tech 2",
        3: "Tech 3",
        4: "MP",
    }
    TOUCH_CHART_LABELS: dict[str, str] = {
        TOUCH_COMPLETE: "Com.",
        TOUCH_PART_ORDER: "PO",
        TOUCH_JUNK: "JO",
        "Triaged": "Tri",
    }
    CLIENT_FOLLOWUP_WORK_APPROVED = "Work approved"
    CLIENT_FOLLOWUP_LEFT_MESSAGE = "Left message"
    CLIENT_FOLLOWUP_COULDNT_CONTACT = "Couldn't contact"
    CLIENT_FOLLOWUP_ACTIONS: tuple[str, ...] = (
        CLIENT_FOLLOWUP_WORK_APPROVED,
        CLIENT_FOLLOWUP_LEFT_MESSAGE,
        CLIENT_FOLLOWUP_COULDNT_CONTACT,
    )
    CLIENT_FOLLOWUP_NO_CONTACT_ACTIONS: tuple[str, ...] = (
        CLIENT_FOLLOWUP_LEFT_MESSAGE,
        CLIENT_FOLLOWUP_COULDNT_CONTACT,
    )
    CLIENT_FOLLOWUP_STAGE_LABELS: tuple[str, ...] = ("Day 1", "Day 2", "Day 3")

    @staticmethod
    def normalize_user_id(value: str) -> str:
        return str(value or "").strip().upper()

    @staticmethod
    def normalize_work_order(value: str) -> str:
        return str(value or "").strip().upper()

    @staticmethod
    def chart_touch_label(value: str) -> str:
        touch = str(value or "").strip()
        if not touch:
            return ""
        return DepotRules.TOUCH_CHART_LABELS.get(touch, touch)

    @staticmethod
    def normalize_followup_action(value: str) -> str:
        text = str(value or "").strip().lower()
        if not text:
            return ""
        canonical = {
            "work approved": DepotRules.CLIENT_FOLLOWUP_WORK_APPROVED,
            "approved": DepotRules.CLIENT_FOLLOWUP_WORK_APPROVED,
            "left message": DepotRules.CLIENT_FOLLOWUP_LEFT_MESSAGE,
            "message left": DepotRules.CLIENT_FOLLOWUP_LEFT_MESSAGE,
            "couldn't contact": DepotRules.CLIENT_FOLLOWUP_COULDNT_CONTACT,
            "couldnt contact": DepotRules.CLIENT_FOLLOWUP_COULDNT_CONTACT,
            "no contact": DepotRules.CLIENT_FOLLOWUP_COULDNT_CONTACT,
        }
        return canonical.get(text, "")

    @staticmethod
    def followup_stage_label(stage_index: int) -> str:
        idx = int(clamp(int(stage_index), 0, len(DepotRules.CLIENT_FOLLOWUP_STAGE_LABELS) - 1))
        return DepotRules.CLIENT_FOLLOWUP_STAGE_LABELS[idx]

    @staticmethod
    def normalize_agent_tier(value: Any, default: int = 1) -> int:
        raw = value
        if isinstance(raw, str):
            text = raw.strip().upper()
            if text in {"MP", "TECH MP"}:
                return 4
            if text.startswith("TECH "):
                text = text[5:].strip()
            elif text.startswith("TIER "):
                text = text[5:].strip()
            raw = text
        try:
            numeric = int(raw)
        except Exception:
            numeric = int(default)
        return int(clamp(numeric, 1, 4))

    @staticmethod
    def agent_tier_label(value: Any) -> str:
        tier = DepotRules.normalize_agent_tier(value)
        return DepotRules.AGENT_TIER_LABELS.get(tier, f"Tech {tier}")


# ---------------------------- Depot Tracker UI Layer ----------------------------

class DepotTracker:
    DASHBOARD_NOTE_TARGET_SPECS: dict[str, dict[str, Any]] = {
        "parts.qa_comment": {
            "label": "Parts - QA Note",
            "table": "parts",
            "column": "qa_comment",
            "order_by": "created_at DESC, id DESC",
            "sync_comments_with_column": True,
        },
        "parts.agent_comment": {
            "label": "Parts - Agent Note",
            "table": "parts",
            "column": "agent_comment",
            "order_by": "created_at DESC, id DESC",
            "sync_comments_with_column": False,
        },
        "client_parts.comments": {
            "label": "Client Parts - Comment",
            "table": "client_parts",
            "column": "comments",
            "order_by": "created_at DESC, id DESC",
            "sync_comments_with_column": False,
        },
        "rtvs.comments": {
            "label": "RTVs - Comment",
            "table": "rtvs",
            "column": "comments",
            "order_by": "created_at DESC, id DESC",
            "sync_comments_with_column": False,
        },
        "client_jo.comments": {
            "label": "Client JO - Comment",
            "table": "client_jo",
            "column": "comments",
            "order_by": "created_at DESC, id DESC",
            "sync_comments_with_column": False,
        },
    }

    def __init__(self, db: DepotDB):
        self.db = db
        self._ensure_default_qa_flags()
        self._repair_closed_workorder_queues()

    def _repair_closed_workorder_queues(self) -> None:
        try:
            close_statuses = (
                DepotRules.TOUCH_COMPLETE,
                DepotRules.TOUCH_JUNK,
                DepotRules.TOUCH_RTV,
            )
            # Do not auto-close active parts on startup based only on latest submission touch.
            # Reason: teams frequently reuse work-order IDs in testing/iteration cycles; applying
            # a historical "Complete/Junk Out/RTV" touch during app startup can incorrectly
            # deactivate newly queued QA parts and make lists appear empty after restart.
            #
            # Active-part closure remains enforced at submit time in submit_work().
            self.db.execute(
                "DELETE FROM client_parts "
                "WHERE COALESCE(("
                "SELECT s.touch FROM submissions s WHERE s.work_order=client_parts.work_order "
                "ORDER BY s.created_at DESC, s.id DESC LIMIT 1"
                "), '') IN (?, ?, ?)",
                close_statuses,
            )
        except Exception as exc:
            _runtime_log_event(
                "depot.closed_queue_repair_failed",
                severity="warning",
                summary="Failed repairing closed work orders in active queues during startup.",
                exc=exc,
            )

    def dashboard_note_target_options(self) -> list[tuple[str, str]]:
        return [
            (key, str(spec.get("label", key)))
            for key, spec in self.DASHBOARD_NOTE_TARGET_SPECS.items()
        ]

    def _dashboard_note_target_spec(self, target_key: str) -> dict[str, Any]:
        normalized_key = str(target_key or "").strip()
        spec = self.DASHBOARD_NOTE_TARGET_SPECS.get(normalized_key)
        if spec is None:
            raise ValueError("Invalid dashboard note target.")
        return spec

    def fetch_dashboard_note_rows(
        self,
        target_key: str,
        *,
        limit: int = 200,
        work_order_filter: str | None = None,
    ) -> list[dict[str, Any]]:
        spec = self._dashboard_note_target_spec(target_key)
        table_name = str(spec.get("table", "")).strip()
        column_name = str(spec.get("column", "")).strip()
        order_by = str(spec.get("order_by", "id DESC")).strip() or "id DESC"
        if not table_name or not column_name:
            raise ValueError("Dashboard note target configuration is incomplete.")

        where_parts: list[str] = []
        params: list[Any] = []
        normalized_work_order = DepotRules.normalize_work_order(str(work_order_filter or ""))
        if normalized_work_order:
            where_parts.append("UPPER(COALESCE(work_order, '')) LIKE ?")
            params.append(f"%{normalized_work_order}%")
        where_clause = f" WHERE {' AND '.join(where_parts)}" if where_parts else ""

        max_rows = int(clamp(safe_int(limit, 200), 1, 5000))
        query = (
            "SELECT id, COALESCE(created_at, '') AS created_at, COALESCE(user_id, '') AS user_id, "
            "COALESCE(work_order, '') AS work_order, COALESCE("
            f"{column_name}, '') AS note_text "
            f"FROM {table_name}{where_clause} ORDER BY {order_by} LIMIT ?"
        )
        params.append(max_rows)
        rows = self.db.fetchall(query, tuple(params))
        result: list[dict[str, Any]] = []
        for row in rows:
            result.append(
                {
                    "id": int(max(0, safe_int(row["id"], 0))),
                    "created_at": str(row["created_at"] or "").strip(),
                    "user_id": str(row["user_id"] or "").strip(),
                    "work_order": str(row["work_order"] or "").strip(),
                    "note_text": str(row["note_text"] or "").strip(),
                }
            )
        return result

    def update_dashboard_note_value(self, target_key: str, row_id: int, note_text: str) -> None:
        spec = self._dashboard_note_target_spec(target_key)
        table_name = str(spec.get("table", "")).strip()
        column_name = str(spec.get("column", "")).strip()
        if not table_name or not column_name:
            raise ValueError("Dashboard note target configuration is incomplete.")

        normalized_row_id = int(row_id)
        existing = self.db.fetchone(f"SELECT id FROM {table_name} WHERE id=? LIMIT 1", (normalized_row_id,))
        if existing is None:
            raise ValueError("Selected row no longer exists.")

        normalized_note = str(note_text or "").strip()
        if bool(spec.get("sync_comments_with_column", False)) and column_name != "comments":
            self.db.execute(
                f"UPDATE {table_name} SET {column_name}=?, comments=? WHERE id=?",
                (normalized_note, normalized_note, normalized_row_id),
            )
            return
        self.db.execute(
            f"UPDATE {table_name} SET {column_name}=? WHERE id=?",
            (normalized_note, normalized_row_id),
        )

    def ensure_admin_user(self, user_id: str) -> None:
        normalized = DepotRules.normalize_user_id(user_id)
        if normalized == 'KIDDS':                               
            return  # Skip adding hardcoded admin to table
        if not normalized:
            return
        existing = self.db.fetchone(
            "SELECT admin_name, position, location, icon_path FROM admin_users WHERE user_id=? LIMIT 1",
            (normalized,),
        )
        existing_name = str(existing["admin_name"] or "").strip() if existing is not None else ""
        existing_position = str(existing["position"] or "").strip() if existing is not None else ""
        existing_location = str(existing["location"] or "").strip() if existing is not None else ""
        existing_icon_path = str(existing["icon_path"] or "").strip() if existing is not None else ""
        self.db.execute(
            "INSERT OR REPLACE INTO admin_users (user_id, admin_name, position, location, icon_path) VALUES (?, ?, ?, ?, ?)",
            (
                normalized,
                existing_name or normalized,
                existing_position or "Admin",
                existing_location,
                existing_icon_path,
            ),
        )

    def is_admin_user(self, user_id: str) -> bool:
        normalized = DepotRules.normalize_user_id(user_id)
        if normalized == 'KIDDS':                                
            return True  # Hardcoded admin access
        if not normalized:
            return False
        row = self.db.fetchone("SELECT 1 FROM admin_users WHERE user_id=? LIMIT 1", (normalized,))
        return row is not None

    def _asset_subdir(self, folder_name: str) -> Path:
        folder = str(folder_name or "").strip()
        if not folder:
            return self.db.db_path.parent / ASSETS_DIR_NAME
        path = self.db.db_path.parent / ASSETS_DIR_NAME / folder
        path.mkdir(parents=True, exist_ok=True)
        return path

    def _resolve_stored_asset_path(self, stored_path: str, folder_name: str) -> Path | None:
        raw = str(stored_path or "").strip()
        if not raw:
            return None
        data_root = self.db.db_path.parent
        folder = str(folder_name or "").strip()
        asset_dir = self._asset_subdir(folder)
        raw_norm = raw.replace("\\", "/").lstrip("./")
        path_obj = Path(raw_norm)

        candidates: list[Path] = []
        if path_obj.is_absolute():
            candidates.append(path_obj)
            candidates.append(asset_dir / path_obj.name)
        else:
            candidates.append(data_root / path_obj)
            candidates.append(asset_dir / path_obj.name)
            parts = [p for p in path_obj.parts if p]
            if parts:
                if parts[0].lower() == folder.lower():
                    tail = Path(*parts[1:]) if len(parts) > 1 else Path(path_obj.name)
                    candidates.append(asset_dir / tail)
                elif parts[0].lower() == ASSETS_DIR_NAME.lower() and len(parts) > 1 and parts[1].lower() == folder.lower():
                    tail = Path(*parts[2:]) if len(parts) > 2 else Path(path_obj.name)
                    candidates.append(asset_dir / tail)

        seen: set[str] = set()
        for candidate in candidates:
            key = str(candidate).replace("\\", "/").lower()
            if key in seen:
                continue
            seen.add(key)
            if candidate.exists() and candidate.is_file():
                return candidate
        return None

    def _admin_icons_dir(self) -> Path:
        return self._asset_subdir(ASSET_ADMIN_ICON_DIR_NAME)

    def _stored_admin_icon_to_abs_path(self, stored_path: str) -> Path | None:
        return self._resolve_stored_asset_path(stored_path, ASSET_ADMIN_ICON_DIR_NAME)

    def _find_icon_for_admin_user(self, user_id: str) -> Path | None:
        normalized = DepotRules.normalize_user_id(user_id)
        if not normalized:
            return None
        icon_dir = self._admin_icons_dir()
        candidates = [p for p in icon_dir.glob(f"{normalized}.*") if p.is_file()]
        if not candidates:
            return None
        candidates.sort(key=lambda p: p.suffix.lower())
        return candidates[0]

    def _relative_admin_icon_store_path(self, abs_path: Path) -> str:
        try:
            rel = abs_path.relative_to(self.db.db_path.parent)
            return str(rel).replace("\\", "/")
        except Exception as exc:
            if not isinstance(exc, ValueError):
                _runtime_log_event(
                    "depot.admin_icon_relative_path_failed",
                    severity="warning",
                    summary="Failed converting admin icon path to relative path; storing absolute path.",
                    exc=exc,
                    context={"abs_path": str(abs_path), "data_root": str(self.db.db_path.parent)},
                )
            return str(abs_path)

    def _store_admin_icon(self, user_id: str, icon_path: str, existing_stored_path: str = "") -> str:
        normalized = DepotRules.normalize_user_id(user_id)
        if not normalized:
            return ""
        requested = str(icon_path or "").strip()
        icon_dir = self._admin_icons_dir()

        if not requested:
            for existing in icon_dir.glob(f"{normalized}.*"):
                try:
                    if existing.is_file():
                        existing.unlink()
                except Exception as exc:
                    _runtime_log_event(
                        "depot.admin_icon_cleanup_failed",
                        severity="warning",
                        summary="Failed deleting existing admin icon during clear operation.",
                        exc=exc,
                        context={"user_id": normalized, "path": str(existing)},
                    )
            return ""

        source = Path(requested)
        if not source.is_absolute():
            source = (self.db.db_path.parent / source).resolve()
        if not source.exists() or not source.is_file():
            existing_abs = self._stored_admin_icon_to_abs_path(existing_stored_path)
            if existing_abs is not None:
                return self._relative_admin_icon_store_path(existing_abs)
            fallback = self._find_icon_for_admin_user(normalized)
            return self._relative_admin_icon_store_path(fallback) if fallback is not None else ""

        suffix = source.suffix.lower() or ".img"
        target = icon_dir / f"{normalized}{suffix}"
        for existing in icon_dir.glob(f"{normalized}.*"):
            if existing.resolve() == target.resolve():
                continue
            try:
                if existing.is_file():
                    existing.unlink()
            except Exception as exc:
                _runtime_log_event(
                    "depot.admin_icon_replace_cleanup_failed",
                    severity="warning",
                    summary="Failed deleting stale admin icon during replacement.",
                    exc=exc,
                    context={"user_id": normalized, "path": str(existing), "target_path": str(target)},
                )
        try:
            if source.resolve() != target.resolve():
                shutil.copy2(source, target)
            elif not target.exists():
                shutil.copy2(source, target)
        except Exception as exc:
            _runtime_log_event(
                "depot.admin_icon_copy_failed",
                severity="warning",
                summary="Failed copying admin icon; keeping previous icon path when possible.",
                exc=exc,
                context={"user_id": normalized, "source_path": str(source), "target_path": str(target)},
            )
            existing_abs = self._stored_admin_icon_to_abs_path(existing_stored_path)
            if existing_abs is not None:
                return self._relative_admin_icon_store_path(existing_abs)
            return ""
        return self._relative_admin_icon_store_path(target)

    def list_admin_users(self) -> list[dict[str, Any]]:
        rows = self.db.fetchall(
            "SELECT user_id, admin_name, position, location, icon_path "
            "FROM admin_users ORDER BY user_id ASC"
        )
        out: list[dict[str, Any]] = []
        for row in rows:
            user_id = DepotRules.normalize_user_id(str(row["user_id"] or ""))
            stored_icon = str(row["icon_path"] or "").strip()
            abs_icon = self._stored_admin_icon_to_abs_path(stored_icon)
            if abs_icon is None:
                fallback = self._find_icon_for_admin_user(user_id)
                if fallback is not None:
                    fallback_stored = self._relative_admin_icon_store_path(fallback)
                    if fallback_stored != stored_icon:
                        self.db.execute("UPDATE admin_users SET icon_path=? WHERE user_id=?", (fallback_stored, user_id))
                    abs_icon = fallback
            out.append(
                {
                    "user_id": user_id,
                    "admin_name": str(row["admin_name"] or "").strip(),
                    "position": str(row["position"] or "").strip(),
                    "location": str(row["location"] or "").strip(),
                    "icon_path": str(abs_icon) if abs_icon is not None else "",
                }
            )
        return out

    def add_admin_user(
        self,
        user_id: str,
        admin_name: str = "",
        position: str = "",
        location: str = "",
        icon_path: str = "",
    ) -> str:
        normalized = DepotRules.normalize_user_id(user_id)
        if not normalized:
            return ""
        normalized_name = str(admin_name or "").strip()
        normalized_position = str(position or "").strip()
        normalized_location = str(location or "").strip()
        normalized_icon = str(icon_path or "").strip()

        existing = self.db.fetchone("SELECT icon_path FROM admin_users WHERE user_id=?", (normalized,))
        existing_stored = str(existing["icon_path"] or "").strip() if existing is not None else ""
        stored_icon = self._store_admin_icon(normalized, normalized_icon, existing_stored)

        self.db.execute(
            "INSERT INTO admin_users (user_id, admin_name, position, location, icon_path) VALUES (?, ?, ?, ?, ?) "
            "ON CONFLICT(user_id) DO UPDATE SET "
            "admin_name=excluded.admin_name, position=excluded.position, location=excluded.location, icon_path=excluded.icon_path",
            (
                normalized,
                normalized_name,
                normalized_position,
                normalized_location,
                stored_icon,
            ),
        )
        return stored_icon

    def remove_admin_user(self, user_id: str) -> None:
        normalized = DepotRules.normalize_user_id(user_id)
        if not normalized:
            return
        self.db.execute("DELETE FROM admin_users WHERE user_id=?", (normalized,))
        for existing in self._admin_icons_dir().glob(f"{normalized}.*"):
            try:
                if existing.is_file():
                    existing.unlink()
            except Exception as exc:
                _runtime_log_event(
                    "depot.admin_icon_delete_failed",
                    severity="warning",
                    summary="Failed deleting admin icon file during admin deletion.",
                    exc=exc,
                    context={"user_id": normalized, "path": str(existing)},
                )

    @staticmethod
    def _normalize_flag_name(value: str) -> str:
        return str(value or "").strip()

    @staticmethod
    def _normalize_flag_severity(value: str) -> str:
        raw = str(value or "").strip().title()
        return raw if raw in QA_FLAG_SEVERITY_OPTIONS else "Medium"

    def _qa_flag_icons_dir(self) -> Path:
        return self._asset_subdir(ASSET_QA_FLAG_ICON_DIR_NAME)

    def _stored_qa_flag_icon_to_abs_path(self, stored_path: str) -> Path | None:
        return self._resolve_stored_asset_path(stored_path, ASSET_QA_FLAG_ICON_DIR_NAME)

    def _find_icon_for_qa_flag_id(self, flag_id: int) -> Path | None:
        icon_dir = self._qa_flag_icons_dir()
        candidates = [p for p in icon_dir.glob(f"flag_{int(flag_id)}.*") if p.is_file()]
        if not candidates:
            return None
        candidates.sort(key=lambda p: p.suffix.lower())
        return candidates[0]

    def _relative_qa_flag_icon_store_path(self, abs_path: Path) -> str:
        try:
            rel = abs_path.relative_to(self.db.db_path.parent)
            return str(rel).replace("\\", "/")
        except Exception as exc:
            if not isinstance(exc, ValueError):
                _runtime_log_event(
                    "depot.qa_flag_relative_path_failed",
                    severity="warning",
                    summary="Failed to convert QA flag icon path to data-root-relative path; storing absolute path.",
                    exc=exc,
                    context={"abs_path": str(abs_path), "data_root": str(self.db.db_path.parent)},
                )
            return str(abs_path)

    def _store_qa_flag_icon(self, flag_id: int, icon_path: str, existing_stored_path: str = "") -> str:
        fid = int(flag_id)
        requested = str(icon_path or "").strip()
        icon_dir = self._qa_flag_icons_dir()

        if not requested:
            for existing in icon_dir.glob(f"flag_{fid}.*"):
                try:
                    if existing.is_file():
                        existing.unlink()
                except Exception as exc:
                    _runtime_log_event(
                        "depot.qa_flag_icon_cleanup_failed",
                        severity="warning",
                        summary="Failed deleting existing QA flag icon during clear operation.",
                        exc=exc,
                        context={"flag_id": fid, "path": str(existing)},
                    )
            return ""

        source = Path(requested)
        if not source.is_absolute():
            source = (self.db.db_path.parent / source).resolve()
        if not source.exists() or not source.is_file():
            existing_abs = self._stored_qa_flag_icon_to_abs_path(existing_stored_path)
            if existing_abs is not None:
                return self._relative_qa_flag_icon_store_path(existing_abs)
            fallback = self._find_icon_for_qa_flag_id(fid)
            return self._relative_qa_flag_icon_store_path(fallback) if fallback is not None else ""

        suffix = source.suffix.lower() or ".img"
        target = icon_dir / f"flag_{fid}{suffix}"
        for existing in icon_dir.glob(f"flag_{fid}.*"):
            if existing.resolve() == target.resolve():
                continue
            try:
                if existing.is_file():
                    existing.unlink()
            except Exception as exc:
                _runtime_log_event(
                    "depot.qa_flag_icon_replace_cleanup_failed",
                    severity="warning",
                    summary="Failed deleting stale QA flag icon while replacing icon.",
                    exc=exc,
                    context={"flag_id": fid, "path": str(existing), "target_path": str(target)},
                )
        try:
            if source.resolve() != target.resolve():
                shutil.copy2(source, target)
            elif not target.exists():
                shutil.copy2(source, target)
        except Exception as exc:
            _runtime_log_event(
                "depot.qa_flag_icon_copy_failed",
                severity="warning",
                summary="Failed to copy QA flag icon; keeping previous icon path when possible.",
                exc=exc,
                context={"flag_id": fid, "source_path": str(source), "target_path": str(target)},
            )
            existing_abs = self._stored_qa_flag_icon_to_abs_path(existing_stored_path)
            if existing_abs is not None:
                return self._relative_qa_flag_icon_store_path(existing_abs)
            return ""
        return self._relative_qa_flag_icon_store_path(target)

    def _ensure_default_qa_flags(self) -> None:
        row = self.db.fetchone("SELECT COUNT(*) AS c FROM qa_flags")
        if row is not None and int(row["c"] or 0) > 0:
            return
        default_names = [name for name in QA_FLAG_OPTIONS if name and name.lower() != "none"]
        for idx, name in enumerate(default_names, start=1):
            severity = "High" if name in {"Escalation", "Safety"} else "Medium"
            self.db.execute(
                "INSERT OR IGNORE INTO qa_flags (flag_name, severity, icon_path, sort_order) VALUES (?, ?, '', ?)",
                (name, severity, idx),
            )

    def list_qa_flags(self) -> list[dict[str, Any]]:
        rows = self.db.fetchall(
            "SELECT id, flag_name, severity, icon_path, sort_order "
            "FROM qa_flags ORDER BY sort_order ASC, flag_name ASC"
        )
        out: list[dict[str, Any]] = []
        for row in rows:
            flag_id = int(row["id"])
            stored_icon = str(row["icon_path"] or "").strip()
            abs_icon = self._stored_qa_flag_icon_to_abs_path(stored_icon)
            if abs_icon is None:
                fallback = self._find_icon_for_qa_flag_id(flag_id)
                if fallback is not None:
                    fallback_stored = self._relative_qa_flag_icon_store_path(fallback)
                    if fallback_stored != stored_icon:
                        self.db.execute("UPDATE qa_flags SET icon_path=? WHERE id=?", (fallback_stored, flag_id))
                    abs_icon = fallback
            out.append(
                {
                    "id": flag_id,
                    "flag_name": str(row["flag_name"] or "").strip(),
                    "severity": self._normalize_flag_severity(str(row["severity"] or "Medium")),
                    "icon_path": str(abs_icon) if abs_icon is not None else "",
                    "sort_order": int(row["sort_order"] or 0),
                }
            )
        return out

    def get_qa_flag_options(self, include_none: bool = True) -> list[str]:
        flags = [row["flag_name"] for row in self.list_qa_flags() if row.get("flag_name")]
        if include_none:
            return ["None", *flags]
        return flags

    def resolve_qa_flag_icon(self, qa_flag: str, legacy_part_image_path: str = "") -> str:
        flag_name = self._normalize_flag_name(qa_flag)
        if flag_name:
            row = self.db.fetchone(
                "SELECT icon_path FROM qa_flags WHERE flag_name=? LIMIT 1",
                (flag_name,),
            )
            if row is not None:
                stored = str(row["icon_path"] or "").strip()
                abs_icon = self._stored_qa_flag_icon_to_abs_path(stored)
                if abs_icon is not None:
                    return str(abs_icon)
        # Backward compatibility for old per-unit icons
        return self.resolve_part_flag_image_path(legacy_part_image_path)

    def upsert_qa_flag(self, flag_name: str, severity: str, icon_path: str = "") -> int:
        normalized_name = self._normalize_flag_name(flag_name)
        normalized_severity = self._normalize_flag_severity(severity)
        if not normalized_name:
            return 0
        existing = self.db.fetchone(
            "SELECT id, icon_path, sort_order FROM qa_flags WHERE flag_name=?",
            (normalized_name,),
        )
        if existing is None:
            order_row = self.db.fetchone("SELECT COALESCE(MAX(sort_order), 0) AS m FROM qa_flags")
            next_order = int(order_row["m"] or 0) + 1 if order_row is not None else 1
            cur = self.db.execute(
                "INSERT INTO qa_flags (flag_name, severity, icon_path, sort_order) VALUES (?, ?, '', ?)",
                (normalized_name, normalized_severity, next_order),
            )
            flag_id = int(cur.lastrowid or 0)
            existing_stored = ""
        else:
            flag_id = int(existing["id"] or 0)
            existing_stored = str(existing["icon_path"] or "").strip()
            self.db.execute(
                "UPDATE qa_flags SET severity=? WHERE id=?",
                (normalized_severity, flag_id),
            )
        if flag_id <= 0:
            return 0
        stored_icon = self._store_qa_flag_icon(flag_id, icon_path, existing_stored)
        self.db.execute(
            "UPDATE qa_flags SET flag_name=?, severity=?, icon_path=? WHERE id=?",
            (normalized_name, normalized_severity, stored_icon, flag_id),
        )
        return flag_id

    def delete_qa_flag(self, flag_name: str) -> None:
        normalized_name = self._normalize_flag_name(flag_name)
        if not normalized_name:
            return
        row = self.db.fetchone("SELECT id FROM qa_flags WHERE flag_name=?", (normalized_name,))
        if row is None:
            return
        flag_id = int(row["id"] or 0)
        self.db.execute("DELETE FROM qa_flags WHERE id=?", (flag_id,))
        if flag_id > 0:
            icon_dir = self._qa_flag_icons_dir()
            for existing in icon_dir.glob(f"flag_{flag_id}.*"):
                try:
                    if existing.is_file():
                        existing.unlink()
                except Exception as exc:
                    _runtime_log_event(
                        "depot.qa_flag_icon_delete_failed",
                        severity="warning",
                        summary="Failed deleting QA flag icon file during QA flag deletion.",
                        exc=exc,
                        context={"flag_id": flag_id, "path": str(existing)},
                    )

    def _agent_icons_dir(self) -> Path:
        return self._asset_subdir(ASSET_AGENT_ICON_DIR_NAME)

    def _stored_icon_to_abs_path(self, stored_path: str) -> Path | None:
        return self._resolve_stored_asset_path(stored_path, ASSET_AGENT_ICON_DIR_NAME)

    def _find_icon_for_user(self, user_id: str) -> Path | None:
        normalized = DepotRules.normalize_user_id(user_id)
        if not normalized:
            return None
        icon_dir = self._agent_icons_dir()
        candidates = [p for p in icon_dir.glob(f"{normalized}.*") if p.is_file()]
        if not candidates:
            return None
        candidates.sort(key=lambda p: p.suffix.lower())
        return candidates[0]

    def _relative_icon_store_path(self, abs_path: Path) -> str:
        try:
            rel = abs_path.relative_to(self.db.db_path.parent)
            return str(rel).replace("\\", "/")
        except Exception as exc:
            if not isinstance(exc, ValueError):
                _runtime_log_event(
                    "depot.agent_icon_relative_path_failed",
                    severity="warning",
                    summary="Failed to convert agent icon path to data-root-relative path; storing absolute path.",
                    exc=exc,
                    context={"abs_path": str(abs_path), "data_root": str(self.db.db_path.parent)},
                )
            return str(abs_path)

    def _part_flag_images_dir(self) -> Path:
        return self._asset_subdir(ASSET_PART_FLAG_IMAGE_DIR_NAME)

    def _stored_part_flag_image_to_abs_path(self, stored_path: str) -> Path | None:
        return self._resolve_stored_asset_path(stored_path, ASSET_PART_FLAG_IMAGE_DIR_NAME)

    def resolve_part_flag_image_path(self, stored_path: str) -> str:
        abs_path = self._stored_part_flag_image_to_abs_path(stored_path)
        return str(abs_path) if abs_path is not None else ""

    def _find_flag_image_for_part(self, part_id: int) -> Path | None:
        flag_dir = self._part_flag_images_dir()
        candidates = [p for p in flag_dir.glob(f"part_{int(part_id)}.*") if p.is_file()]
        if not candidates:
            return None
        candidates.sort(key=lambda p: p.suffix.lower())
        return candidates[0]

    def _relative_part_flag_image_store_path(self, abs_path: Path) -> str:
        try:
            rel = abs_path.relative_to(self.db.db_path.parent)
            return str(rel).replace("\\", "/")
        except Exception as exc:
            if not isinstance(exc, ValueError):
                _runtime_log_event(
                    "depot.part_flag_relative_path_failed",
                    severity="warning",
                    summary="Failed to convert part flag image path to data-root-relative path; storing absolute path.",
                    exc=exc,
                    context={"abs_path": str(abs_path), "data_root": str(self.db.db_path.parent)},
                )
            return str(abs_path)

    def _store_part_flag_image(self, part_id: int, image_path: str, existing_stored_path: str = "") -> str:
        pid = int(part_id)
        requested = str(image_path or "").strip()
        flag_dir = self._part_flag_images_dir()

        if not requested:
            for existing in flag_dir.glob(f"part_{pid}.*"):
                try:
                    if existing.is_file():
                        existing.unlink()
                except Exception as exc:
                    _runtime_log_event(
                        "depot.part_flag_cleanup_failed",
                        severity="warning",
                        summary="Failed deleting existing part-flag image during clear operation.",
                        exc=exc,
                        context={"part_id": pid, "path": str(existing)},
                    )
            return ""

        source = Path(requested)
        if not source.is_absolute():
            source = (self.db.db_path.parent / source).resolve()

        if not source.exists() or not source.is_file():
            existing_abs = self._stored_part_flag_image_to_abs_path(existing_stored_path)
            if existing_abs is not None:
                return self._relative_part_flag_image_store_path(existing_abs)
            fallback = self._find_flag_image_for_part(pid)
            return self._relative_part_flag_image_store_path(fallback) if fallback is not None else ""

        suffix = source.suffix.lower() or ".img"
        target = flag_dir / f"part_{pid}{suffix}"

        for existing in flag_dir.glob(f"part_{pid}.*"):
            if existing.resolve() == target.resolve():
                continue
            try:
                if existing.is_file():
                    existing.unlink()
            except Exception as exc:
                _runtime_log_event(
                    "depot.part_flag_replace_cleanup_failed",
                    severity="warning",
                    summary="Failed deleting stale part-flag image during replacement.",
                    exc=exc,
                    context={"part_id": pid, "path": str(existing), "target_path": str(target)},
                )

        try:
            if source.resolve() != target.resolve():
                shutil.copy2(source, target)
            elif not target.exists():
                shutil.copy2(source, target)
        except Exception as exc:
            _runtime_log_event(
                "depot.part_flag_copy_failed",
                severity="warning",
                summary="Failed to copy part-flag image; keeping previous image path when possible.",
                exc=exc,
                context={"part_id": pid, "source_path": str(source), "target_path": str(target)},
            )
            existing_abs = self._stored_part_flag_image_to_abs_path(existing_stored_path)
            if existing_abs is not None:
                return self._relative_part_flag_image_store_path(existing_abs)
            return ""

        return self._relative_part_flag_image_store_path(target)

    def _store_agent_icon(self, user_id: str, icon_path: str, existing_stored_path: str = "") -> str:
        normalized = DepotRules.normalize_user_id(user_id)
        if not normalized:
            return ""

        requested = str(icon_path or "").strip()
        icon_dir = self._agent_icons_dir()

        # Empty icon path means explicit clear.
        if not requested:
            for existing in icon_dir.glob(f"{normalized}.*"):
                try:
                    if existing.is_file():
                        existing.unlink()
                except Exception as exc:
                    _runtime_log_event(
                        "depot.agent_icon_cleanup_failed",
                        severity="warning",
                        summary="Failed deleting existing agent icon during clear operation.",
                        exc=exc,
                        context={"user_id": normalized, "path": str(existing)},
                    )
            return ""

        source = Path(requested)
        if not source.is_absolute():
            source = (self.db.db_path.parent / source).resolve()

        if not source.exists() or not source.is_file():
            existing_abs = self._stored_icon_to_abs_path(existing_stored_path)
            if existing_abs is not None:
                return self._relative_icon_store_path(existing_abs)
            fallback = self._find_icon_for_user(normalized)
            return self._relative_icon_store_path(fallback) if fallback is not None else ""

        suffix = source.suffix.lower() or ".img"
        target = icon_dir / f"{normalized}{suffix}"

        for existing in icon_dir.glob(f"{normalized}.*"):
            if existing.resolve() == target.resolve():
                continue
            try:
                if existing.is_file():
                    existing.unlink()
            except Exception as exc:
                _runtime_log_event(
                    "depot.agent_icon_replace_cleanup_failed",
                    severity="warning",
                    summary="Failed deleting stale agent icon during replacement.",
                    exc=exc,
                    context={"user_id": normalized, "path": str(existing), "target_path": str(target)},
                )

        try:
            if source.resolve() != target.resolve():
                shutil.copy2(source, target)
            elif not target.exists():
                shutil.copy2(source, target)
        except Exception as exc:
            _runtime_log_event(
                "depot.agent_icon_copy_failed",
                severity="warning",
                summary="Failed copying agent icon; keeping previous icon path when possible.",
                exc=exc,
                context={"user_id": normalized, "source_path": str(source), "target_path": str(target)},
            )
            existing_abs = self._stored_icon_to_abs_path(existing_stored_path)
            if existing_abs is not None:
                return self._relative_icon_store_path(existing_abs)
            return ""

        return self._relative_icon_store_path(target)

    def list_agents(self, tier_filter: int | None = None) -> list[dict[str, Any]]:
        params: list[Any] = []
        where = ""
        if tier_filter in (1, 2, 3, 4):
            where = "WHERE tier=?"
            params.append(int(tier_filter))
        rows = self.db.fetchall(
            f"SELECT agent_name, user_id, tier, location, icon_path FROM agents {where} ORDER BY tier ASC, agent_name ASC, user_id ASC",
            tuple(params),
        )
        result: list[dict[str, Any]] = []
        for row in rows:
            user_id = DepotRules.normalize_user_id(row["user_id"])
            stored_icon = str(row["icon_path"] or "").strip()
            abs_icon = self._stored_icon_to_abs_path(stored_icon)

            if abs_icon is None:
                fallback = self._find_icon_for_user(user_id)
                if fallback is not None:
                    fallback_stored = self._relative_icon_store_path(fallback)
                    if fallback_stored != stored_icon:
                        self.db.execute("UPDATE agents SET icon_path=? WHERE user_id=?", (fallback_stored, user_id))
                    abs_icon = fallback

            result.append(
                {
                    "agent_name": str(row["agent_name"] or ""),
                    "user_id": user_id,
                    "tier": DepotRules.normalize_agent_tier(row["tier"]),
                    "location": str(row["location"] or "").strip(),
                    "icon_path": str(abs_icon) if abs_icon is not None else "",
                }
            )
        return result

    def upsert_agent(self, user_id: str, agent_name: str, tier: int, icon_path: str = "", location: str = "") -> str:
        normalized_user = DepotRules.normalize_user_id(user_id)
        normalized_name = str(agent_name or "").strip()
        normalized_tier = DepotRules.normalize_agent_tier(tier)
        normalized_icon = str(icon_path or "").strip()
        normalized_location = str(location or "").strip()
        if not normalized_user or not normalized_name:
            return ""

        existing = self.db.fetchone("SELECT icon_path FROM agents WHERE user_id=?", (normalized_user,))
        existing_stored = str(existing["icon_path"] or "").strip() if existing else ""
        stored_icon = self._store_agent_icon(normalized_user, normalized_icon, existing_stored)
        self.db.execute(
            """
            INSERT INTO agents (agent_name, user_id, tier, location, icon_path)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(user_id) DO UPDATE SET
              agent_name=excluded.agent_name,
              tier=excluded.tier,
              location=excluded.location,
              icon_path=excluded.icon_path
            """,
            (normalized_name, normalized_user, normalized_tier, normalized_location, stored_icon),
        )
        return stored_icon

    def delete_agent(self, user_id: str) -> None:
        normalized_user = DepotRules.normalize_user_id(user_id)
        if not normalized_user:
            return
        self.db.execute("DELETE FROM agents WHERE user_id=?", (normalized_user,))
        for existing in self._agent_icons_dir().glob(f"{normalized_user}.*"):
            try:
                if existing.is_file():
                    existing.unlink()
            except Exception as exc:
                _runtime_log_event(
                    "depot.agent_icon_delete_failed",
                    severity="warning",
                    summary="Failed deleting agent icon file during agent deletion.",
                    exc=exc,
                    context={"user_id": normalized_user, "path": str(existing)},
                )

    def submit_work(self, user_id: str, work_order: str, touch: str, client_unit: bool, comments: str | None = None) -> None:
        now_dt = datetime.now()
        now = now_dt.isoformat(timespec="seconds")
        user_id = DepotRules.normalize_user_id(user_id)
        work_order = DepotRules.normalize_work_order(work_order)
        entry_date = now_dt.date().isoformat()
        client_unit_int = 1 if client_unit else 0

        # insert submission record
        self.db.execute(
            "INSERT INTO submissions (created_at, user_id, work_order, touch, client_unit, entry_date) VALUES (?, ?, ?, ?, ?, ?)",
            (now, user_id, work_order, touch, client_unit_int, entry_date),
        )

        # RTV
        if touch == DepotRules.TOUCH_RTV:
            self.db.execute(
                "INSERT INTO rtvs (created_at, user_id, work_order, comments) VALUES (?, ?, ?, ?)",
                (now, user_id, work_order, comments or ""),
            )

        # client jo
        if client_unit and touch == DepotRules.TOUCH_JUNK:
            self.db.execute(
                "INSERT INTO client_jo (created_at, user_id, work_order, comments) VALUES (?, ?, ?, ?)",
                (now, user_id, work_order, comments or ""),
            )

        # client_parts upsert
        if client_unit and touch in (DepotRules.TOUCH_PART_ORDER, DepotRules.TOUCH_OTHER):
            existing = self.db.fetchone(
                "SELECT id FROM client_parts WHERE work_order = ?", (work_order,)
            )
            if existing:
                self.db.execute(
                    "UPDATE client_parts SET user_id=?, comments=?, created_at=? WHERE work_order = ?",
                    (user_id, comments or "", now, work_order),
                )
            else:
                self.db.execute(
                    "INSERT INTO client_parts (created_at, user_id, work_order, comments) VALUES (?, ?, ?, ?)",
                    (now, user_id, work_order, comments or ""),
                )

        # removal rule
        if touch in DepotRules.CLOSING_TOUCHES:
            # Completed/junked/RTV work orders are no longer active in parts queues.
            self.db.execute(
                "DELETE FROM part_details WHERE part_id IN (SELECT id FROM parts WHERE work_order=?)",
                (work_order,),
            )
            self.db.execute(
                "UPDATE parts SET is_active=0, working_user_id='', working_updated_at='' WHERE work_order=?",
                (work_order,),
            )
            self.db.execute(
                "DELETE FROM client_parts WHERE work_order = ?", (work_order,)
            )

    def submit_part(
        self,
        user_id: str,
        assigned_user_id: str,
        work_order: str,
        category: str,
        client_unit: bool,
        qa_comment: str | None = "",
        qa_flag: str | None = "",
        parts_on_hand: bool = False,
    ) -> int:
        now = datetime.utcnow().isoformat()
        user_id = DepotRules.normalize_user_id(user_id)
        assigned_user_id = DepotRules.normalize_user_id(assigned_user_id)
        work_order = DepotRules.normalize_work_order(work_order)
        category = str(category or "").strip()
        client_unit_int = 1 if client_unit else 0
        parts_on_hand_int = 1 if bool(parts_on_hand) else 0
        qa_comment_text = str(qa_comment or "").strip()
        qa_flag_text = str(qa_flag or "").strip()
        if qa_flag_text.lower() == "none":
            qa_flag_text = ""
        existing = self.db.fetchone(
            "SELECT id FROM parts WHERE is_active=1 AND work_order=? ORDER BY id DESC LIMIT 1",
            (work_order,),
        )
        if existing is not None:
            existing_id = int(existing["id"])
            self.db.execute(
                "UPDATE parts SET user_id=?, assigned_user_id=?, client_unit=?, category=?, comments=?, qa_comment=?, "
                "qa_flag=?, qa_flag_image_path='', "
                "parts_on_hand=CASE WHEN ?=1 THEN 1 ELSE parts_on_hand END "
                "WHERE id=?",
                (
                    user_id,
                    assigned_user_id,
                    client_unit_int,
                    category,
                    qa_comment_text,
                    qa_comment_text,
                    qa_flag_text,
                    parts_on_hand_int,
                    existing_id,
                ),
            )
            return existing_id
        insert_cursor = self.db.execute(
            "INSERT INTO parts (created_at, user_id, assigned_user_id, work_order, client_unit, category, comments, qa_comment, "
            "agent_comment, qa_flag, qa_flag_image_path, is_active, parts_on_hand, parts_installed, parts_installed_by, parts_installed_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, 0, '', '')",
            (
                now,
                user_id,
                assigned_user_id,
                work_order,
                client_unit_int,
                category,
                qa_comment_text,
                qa_comment_text,
                "",
                qa_flag_text,
                "",
                parts_on_hand_int,
            ),
        )
        return int(insert_cursor.lastrowid or 0)

    def upsert_part_detail(
        self,
        part_id: int,
        lpn: str,
        part_number: str,
        part_description: str,
        shipping_info: str,
        delivered: bool,
    ) -> None:
        now = datetime.utcnow().isoformat()
        normalized_lpn = DepotRules.normalize_work_order(lpn)
        part_no_text = str(part_number or "").strip()
        part_desc_text = str(part_description or "").strip()
        shipping_text = str(shipping_info or "").strip()

        existing = self.db.fetchone(
            "SELECT COALESCE(lpn, '') AS lpn, COALESCE(part_number, '') AS part_number, "
            "COALESCE(part_description, '') AS part_description, COALESCE(shipping_info, '') AS shipping_info, "
            "COALESCE(installed_keys, '') AS installed_keys, COALESCE(delivered, 0) AS delivered "
            "FROM part_details WHERE part_id=?",
            (int(part_id),),
        )
        if existing is None:
            self.db.execute(
                "INSERT INTO part_details (part_id, created_at, lpn, part_number, part_description, shipping_info, delivered) "
                "VALUES (?, ?, ?, ?, ?, ?, ?)",
                (
                    int(part_id),
                    now,
                    normalized_lpn,
                    part_no_text,
                    part_desc_text,
                    shipping_text,
                    1 if bool(delivered) else 0,
                ),
            )
            return

        def split_piped(text: str) -> list[str]:
            raw = str(text or "")
            if raw == "":
                return []
            return [str(piece or "").strip() for piece in raw.split(" | ")]

        def line_key(lpn_value: str, part_value: str, desc_value: str, ship_value: str) -> str:
            return json.dumps(
                [
                    str(lpn_value or "").strip().casefold(),
                    str(part_value or "").strip().casefold(),
                    str(desc_value or "").strip().casefold(),
                    str(ship_value or "").strip().casefold(),
                ],
                ensure_ascii=True,
                separators=(",", ":"),
            )

        def value_for(values: list[str], idx: int, row_count: int) -> str:
            if idx < len(values):
                return str(values[idx] or "").strip()
            # If one value was historically de-duplicated, broadcast it to all line-items.
            if len(values) == 1 and row_count > 1:
                return str(values[0] or "").strip()
            return ""

        lpn_values = split_piped(str(existing["lpn"] or ""))
        part_no_values = split_piped(str(existing["part_number"] or ""))
        desc_values = split_piped(str(existing["part_description"] or ""))
        shipping_values = split_piped(str(existing["shipping_info"] or ""))
        installed_keys_raw = str(existing["installed_keys"] or "").strip()
        installed_key_set: set[str] = set()
        if installed_keys_raw:
            try:
                parsed_installed = json.loads(installed_keys_raw)
                if isinstance(parsed_installed, list):
                    for value in parsed_installed:
                        value_text = str(value or "").strip()
                        if value_text:
                            installed_key_set.add(value_text)
            except Exception:
                for value in installed_keys_raw.split(" | "):
                    value_text = str(value or "").strip()
                    if value_text:
                        installed_key_set.add(value_text)
        existing_line_count = max(len(lpn_values), len(part_no_values), len(desc_values), len(shipping_values), 0)

        line_rows: list[tuple[str, str, str, str]] = []
        for idx in range(existing_line_count):
            line_rows.append(
                (
                    value_for(lpn_values, idx, existing_line_count),
                    value_for(part_no_values, idx, existing_line_count),
                    value_for(desc_values, idx, existing_line_count),
                    value_for(shipping_values, idx, existing_line_count),
                )
            )

        # Repair previously misaligned rows where one shared shipping value was expected for all lines.
        non_empty_ship = [row[3] for row in line_rows if str(row[3] or "").strip()]
        ship_seen: set[str] = set()
        ship_unique: list[str] = []
        for ship in non_empty_ship:
            key = str(ship or "").strip().casefold()
            if key in ship_seen:
                continue
            ship_seen.add(key)
            ship_unique.append(str(ship or "").strip())
        if len(ship_unique) == 1:
            shared_ship = ship_unique[0]
            line_rows = [
                (row[0], row[1], row[2], shared_ship if not str(row[3] or "").strip() else str(row[3] or "").strip())
                for row in line_rows
            ]

        incoming_row = (
            normalized_lpn,
            part_no_text,
            part_desc_text,
            shipping_text,
        )
        if any(str(piece or "").strip() for piece in incoming_row):
            existing_keys: set[tuple[str, str, str, str]] = {
                tuple(str(piece or "").strip().casefold() for piece in row)
                for row in line_rows
                if any(str(piece or "").strip() for piece in row)
            }
            incoming_key = tuple(str(piece or "").strip().casefold() for piece in incoming_row)
            appended_new_row = False
            if incoming_key not in existing_keys:
                line_rows.append(incoming_row)
                appended_new_row = True
            else:
                appended_new_row = False
        else:
            appended_new_row = False

        # De-duplicate exact line repeats while preserving order.
        deduped_rows: list[tuple[str, str, str, str]] = []
        deduped_keys: set[tuple[str, str, str, str]] = set()
        for row in line_rows:
            if not any(str(piece or "").strip() for piece in row):
                continue
            key = tuple(str(piece or "").strip().casefold() for piece in row)
            if key in deduped_keys:
                continue
            deduped_keys.add(key)
            deduped_rows.append(tuple(str(piece or "").strip() for piece in row))
        line_rows = deduped_rows

        retained_installed_keys: list[str] = []
        for row in line_rows:
            key = line_key(row[0], row[1], row[2], row[3])
            if key in installed_key_set:
                retained_installed_keys.append(key)

        merged_lpn = " | ".join(str(row[0] or "").strip() for row in line_rows)
        merged_part_number = " | ".join(str(row[1] or "").strip() for row in line_rows)
        merged_part_desc = " | ".join(str(row[2] or "").strip() for row in line_rows)
        merged_shipping = " | ".join(str(row[3] or "").strip() for row in line_rows)
        merged_installed_keys = (
            json.dumps(retained_installed_keys, ensure_ascii=True, separators=(",", ":"))
            if retained_installed_keys
            else ""
        )
        delivered_value = 1 if (bool(delivered) or bool(int(existing["delivered"] or 0))) else 0
        self.db.execute(
            "UPDATE part_details SET lpn=?, part_number=?, part_description=?, shipping_info=?, installed_keys=?, delivered=? WHERE part_id=?",
            (
                merged_lpn,
                merged_part_number,
                merged_part_desc,
                merged_shipping,
                merged_installed_keys,
                delivered_value,
                int(part_id),
            ),
        )
        if appended_new_row:
            # New delivered rows become pending install by default.
            self.db.execute(
                "UPDATE parts SET parts_installed=0, parts_installed_by='', parts_installed_at='' WHERE id=?",
                (int(part_id),),
            )

    def find_active_part_work_orders(self, work_orders: list[str]) -> set[str]:
        normalized: list[str] = []
        seen: set[str] = set()
        for raw_value in work_orders:
            work_order = DepotRules.normalize_work_order(str(raw_value or ""))
            if not work_order or work_order in seen:
                continue
            seen.add(work_order)
            normalized.append(work_order)
        if not normalized:
            return set()

        active: set[str] = set()
        chunk_size = 400
        for start in range(0, len(normalized), chunk_size):
            chunk = normalized[start : start + chunk_size]
            placeholders = ",".join("?" for _ in chunk)
            rows = self.db.fetchall(
                f"SELECT DISTINCT work_order FROM parts WHERE is_active=1 AND work_order IN ({placeholders})",
                tuple(chunk),
            )
            for row in rows:
                existing_work_order = DepotRules.normalize_work_order(str(row["work_order"] or ""))
                if existing_work_order:
                    active.add(existing_work_order)
        return active

    def find_active_parts_by_work_orders(self, work_orders: list[str]) -> dict[str, int]:
        normalized: list[str] = []
        seen: set[str] = set()
        for raw_value in work_orders:
            work_order = DepotRules.normalize_work_order(str(raw_value or ""))
            if not work_order or work_order in seen:
                continue
            seen.add(work_order)
            normalized.append(work_order)
        if not normalized:
            return {}

        by_work_order: dict[str, int] = {}
        chunk_size = 400
        for start in range(0, len(normalized), chunk_size):
            chunk = normalized[start : start + chunk_size]
            placeholders = ",".join("?" for _ in chunk)
            rows = self.db.fetchall(
                f"SELECT id, work_order FROM parts WHERE is_active=1 AND work_order IN ({placeholders}) "
                "ORDER BY id DESC",
                tuple(chunk),
            )
            for row in rows:
                existing_work_order = DepotRules.normalize_work_order(str(row["work_order"] or ""))
                if not existing_work_order or existing_work_order in by_work_order:
                    continue
                by_work_order[existing_work_order] = int(row["id"])
        return by_work_order

    def update_part_comments(self, part_id: int, qa_comment: str, agent_comment: str) -> None:
        self.db.execute(
            "UPDATE parts SET qa_comment=?, agent_comment=?, comments=? WHERE id=?",
            (str(qa_comment or "").strip(), str(agent_comment or "").strip(), str(qa_comment or "").strip(), int(part_id)),
        )

    def update_part_agent_comment(self, part_id: int, agent_comment: str) -> None:
        self.db.execute(
            "UPDATE parts SET agent_comment=? WHERE id=?",
            (str(agent_comment or "").strip(), int(part_id)),
        )

    def set_part_working_user(self, part_id: int, working_user_id: str) -> None:
        normalized = DepotRules.normalize_user_id(working_user_id)
        stamp = datetime.now().isoformat(timespec="seconds") if normalized else ""
        # One agent may actively "work" only one part row at a time.
        if normalized:
            self.db.execute(
                "UPDATE parts SET working_user_id='', working_updated_at='' "
                "WHERE working_user_id=? AND id<>?",
                (normalized, int(part_id)),
            )
        self.db.execute(
            "UPDATE parts SET working_user_id=?, working_updated_at=? WHERE id=?",
            (normalized, stamp, int(part_id)),
        )

    def set_part_on_hand(self, part_id: int, on_hand: bool) -> None:
        if bool(on_hand):
            self.db.execute(
                "UPDATE parts SET parts_on_hand=1 WHERE id=?",
                (int(part_id),),
            )
            return
        self.db.execute(
            "UPDATE parts SET parts_on_hand=0, parts_installed=0, parts_installed_by='', parts_installed_at='' WHERE id=?",
            (int(part_id),),
        )

    def set_part_installed(self, part_id: int, installed: bool, actor_user_id: str = "") -> None:
        if bool(installed):
            actor = DepotRules.normalize_user_id(actor_user_id)
            stamp = datetime.now().isoformat(timespec="seconds")
            self.db.execute(
                "UPDATE parts SET parts_on_hand=1, parts_installed=1, parts_installed_by=?, parts_installed_at=?, "
                    "working_user_id='', working_updated_at='' WHERE id=?",
                (actor, stamp, int(part_id)),
            )
            detail = self.db.fetchone(
                "SELECT COALESCE(lpn, '') AS lpn, COALESCE(part_number, '') AS part_number, "
                "COALESCE(part_description, '') AS part_description, COALESCE(shipping_info, '') AS shipping_info "
                "FROM part_details WHERE part_id=?",
                (int(part_id),),
            )
            if detail is not None:
                def split_piped(text: str) -> list[str]:
                    raw = str(text or "")
                    if raw == "":
                        return []
                    return [str(piece or "").strip() for piece in raw.split(" | ")]

                def value_for(values: list[str], idx: int, row_count: int) -> str:
                    if idx < len(values):
                        return str(values[idx] or "").strip()
                    if len(values) == 1 and row_count > 1:
                        return str(values[0] or "").strip()
                    return ""

                def line_key(lpn_value: str, part_value: str, desc_value: str, ship_value: str) -> str:
                    return json.dumps(
                        [
                            str(lpn_value or "").strip().casefold(),
                            str(part_value or "").strip().casefold(),
                            str(desc_value or "").strip().casefold(),
                            str(ship_value or "").strip().casefold(),
                        ],
                        ensure_ascii=True,
                        separators=(",", ":"),
                    )

                lpn_values = split_piped(str(detail["lpn"] or ""))
                part_values = split_piped(str(detail["part_number"] or ""))
                desc_values = split_piped(str(detail["part_description"] or ""))
                ship_values = split_piped(str(detail["shipping_info"] or ""))
                row_count = max(len(lpn_values), len(part_values), len(desc_values), len(ship_values), 0)
                installed_line_keys: list[str] = []
                for idx in range(row_count):
                    installed_line_keys.append(
                        line_key(
                            value_for(lpn_values, idx, row_count),
                            value_for(part_values, idx, row_count),
                            value_for(desc_values, idx, row_count),
                            value_for(ship_values, idx, row_count),
                        )
                    )
                serialized = (
                    json.dumps(installed_line_keys, ensure_ascii=True, separators=(",", ":"))
                    if installed_line_keys
                    else ""
                )
                self.db.execute(
                    "UPDATE part_details SET installed_keys=? WHERE part_id=?",
                    (serialized, int(part_id)),
                )
            return
        self.db.execute(
            "UPDATE parts SET parts_installed=0, parts_installed_by='', parts_installed_at='' WHERE id=?",
            (int(part_id),),
        )
        self.db.execute(
            "UPDATE part_details SET installed_keys='' WHERE part_id=?",
            (int(part_id),),
        )

    def mark_client_followup_action(self, client_part_id: int, action: str, actor_user_id: str) -> int:
        normalized_action = DepotRules.normalize_followup_action(action)
        if not normalized_action:
            raise ValueError("Invalid follow-up action.")
        row = self.db.fetchone(
            "SELECT COALESCE(followup_no_contact_count, 0) AS followup_no_contact_count, "
            "COALESCE(followup_last_action, '') AS followup_last_action, "
            "COALESCE(followup_last_action_at, '') AS followup_last_action_at, "
            "COALESCE(comments, '') AS comments "
            "FROM client_parts WHERE id=?",
            (int(client_part_id),),
        )
        if row is None:
            raise ValueError("Client follow-up row no longer exists.")
        existing_count = int(max(0, safe_int(row["followup_no_contact_count"], 0)))
        previous_action = DepotRules.normalize_followup_action(str(row["followup_last_action"] or ""))
        previous_stamp = str(row["followup_last_action_at"] or "").strip()
        previous_day = previous_stamp[:10] if len(previous_stamp) >= 10 else ""
        stamp_dt = datetime.now()
        stamp = stamp_dt.isoformat(timespec="seconds")
        today_iso = stamp_dt.date().isoformat()
        same_day_replacement = bool(previous_day == today_iso)

        if normalized_action in DepotRules.CLIENT_FOLLOWUP_NO_CONTACT_ACTIONS:
            if same_day_replacement:
                if previous_action in DepotRules.CLIENT_FOLLOWUP_NO_CONTACT_ACTIONS:
                    existing_count = int(max(1, existing_count))
                else:
                    existing_count = 1
            else:
                existing_count += 1
        else:
            existing_count = 0

        actor = DepotRules.normalize_user_id(actor_user_id)
        action_tag = f"Follow-up {today_iso}: {normalized_action} - [{actor or 'UNKNOWN'}]"
        stage_logged = -1
        existing_comments = str(row["comments"] or "").strip()
        lines = [line.rstrip() for line in existing_comments.splitlines() if str(line).strip()]
        action_prefix = f"Follow-up {today_iso}:"
        stage_prefix = f"Waiting Stage {today_iso}:"
        lines = [line for line in lines if not line.startswith(action_prefix)]
        lines = [line for line in lines if not line.startswith(stage_prefix)]
        stage_lines: list[str] = [action_tag]
        if normalized_action in DepotRules.CLIENT_FOLLOWUP_NO_CONTACT_ACTIONS:
            stage_logged = 0
            stage_lines.append(f"{stage_prefix} {DepotRules.followup_stage_label(0)} - [SYSTEM]")
        lines.extend(stage_lines)
        updated_comments = "\n".join([line for line in lines if str(line).strip()])
        self.db.execute(
            "UPDATE client_parts SET followup_last_action=?, followup_last_action_at=?, "
            "followup_last_actor=?, followup_no_contact_count=?, followup_stage_logged=?, comments=? WHERE id=?",
            (normalized_action, stamp, actor, int(existing_count), int(stage_logged), updated_comments, int(client_part_id)),
        )
        return int(existing_count)

    def update_client_followup_stage(self, client_part_id: int, stage_index: int) -> str:
        next_stage = int(clamp(int(stage_index), 0, 2))
        row = self.db.fetchone(
            "SELECT COALESCE(comments, '') AS comments, "
            "COALESCE(followup_stage_logged, -1) AS followup_stage_logged "
            "FROM client_parts WHERE id=?",
            (int(client_part_id),),
        )
        if row is None:
            raise ValueError("Client follow-up row no longer exists.")
        current_stage = int(safe_int(row["followup_stage_logged"], -1))
        current_comments = str(row["comments"] or "").strip()
        if next_stage <= current_stage:
            return current_comments

        stamp_day = datetime.now().date().isoformat()
        stage_prefix = f"Waiting Stage {stamp_day}:"
        chunks: list[str] = [line.rstrip() for line in current_comments.splitlines() if str(line).strip()]
        chunks = [line for line in chunks if not line.startswith(stage_prefix)]
        for idx in range(max(0, current_stage + 1), next_stage + 1):
            chunks.append(f"{stage_prefix} {DepotRules.followup_stage_label(idx)} - [SYSTEM]")
        updated_comments = "\n".join([chunk for chunk in chunks if str(chunk).strip()])
        self.db.execute(
            "UPDATE client_parts SET followup_stage_logged=?, comments=? WHERE id=?",
            (int(next_stage), updated_comments, int(client_part_id)),
        )
        return updated_comments

    def update_part_qa_fields(self, part_id: int, qa_comment: str, qa_flag: str) -> None:
        qa_flag_text = str(qa_flag or "").strip()
        if qa_flag_text.lower() == "none":
            qa_flag_text = ""
        qa_comment_text = str(qa_comment or "").strip()
        self.db.execute(
            "UPDATE parts SET qa_comment=?, comments=?, qa_flag=?, qa_flag_image_path=? WHERE id=?",
            (qa_comment_text, qa_comment_text, qa_flag_text, "", int(part_id)),
        )

    def get_dashboard_metrics(self, start_date: str | None = None, end_date: str | None = None, user_id: str | None = None, touch: str | None = None, client_only: bool | None = None) -> dict[str, Any]:
        where = []
        params: list[Any] = []

        if start_date:
            where.append("entry_date >= ?")
            params.append(start_date)
        if end_date:
            where.append("entry_date <= ?")
            params.append(end_date)
        if user_id:
            where.append("user_id = ?")
            params.append(DepotRules.normalize_user_id(user_id))
        if touch:
            where.append("touch = ?")
            params.append(touch)
        if client_only is not None:
            where.append("client_unit = ?")
            params.append(1 if client_only else 0)

        where_clause = "WHERE " + " AND ".join(where) if where else ""

        total = self.db.fetchone(f"SELECT COUNT(*) AS c FROM submissions {where_clause}", tuple(params))
        by_touch = self.db.fetchall(f"SELECT touch, COUNT(*) AS c FROM submissions {where_clause} GROUP BY touch", tuple(params))
        by_user = self.db.fetchall(f"SELECT user_id, COUNT(*) AS c FROM submissions {where_clause} GROUP BY user_id", tuple(params))
        daily = self.db.fetchall(f"SELECT entry_date, COUNT(*) AS c FROM submissions {where_clause} GROUP BY entry_date ORDER BY entry_date DESC LIMIT 30", tuple(params))

        actividad = {
            "total_submissions": int(total["c"] if total else 0),
            "by_touch": {row["touch"]: row["c"] for row in by_touch},
            "by_user": {row["user_id"]: row["c"] for row in by_user},
            "daily": [{"entry_date": r["entry_date"], "count": r["c"]} for r in daily],
            "active_client_follow_up": self.db.fetchone("SELECT COUNT(*) AS c FROM client_parts", ())["c"],
            "active_parts": self.db.fetchone("SELECT COUNT(*) AS c FROM parts WHERE is_active=1", ())["c"],
            "rtv_count": self.db.fetchone("SELECT COUNT(*) AS c FROM rtvs", ())["c"],
            "client_jo_count": self.db.fetchone("SELECT COUNT(*) AS c FROM client_jo", ())["c"],
        }
        return actividad

    def import_workbook(self, xlsm_path: Path, import_tables: set[str] | None = None) -> tuple[bool, str]:
        if openpyxl is None:
            return False, "openpyxl not installed"
        if not xlsm_path.exists():
            return False, "workbook not found"
        try:
            wb = openpyxl.load_workbook(str(xlsm_path), keep_vba=True, data_only=True)
            selected: set[str] | None = None
            if import_tables is not None:
                selected = {str(name or "").strip().lower() for name in import_tables if str(name or "").strip()}
                if not selected:
                    return False, "No tables were selected for import."

            def nx(val):
                return DepotRules.normalize_user_id(str(val)) if val is not None else ""

            def include(table_key: str) -> bool:
                return selected is None or table_key in selected

            counts = {key: 0 for key, _sheet, _label in WORKBOOK_IMPORT_SPECS}
            missing_sheets: list[str] = []

            # Submissions
            if include("submissions") and "Table1" in wb.sheetnames:
                sh = wb["Table1"]
                for i, row in enumerate(sh.iter_rows(min_row=2), start=2):
                    date_time = row[0].value
                    user = nx(row[1].value)
                    work_order = DepotRules.normalize_work_order(row[2].value)
                    touch = str(row[3].value or "").strip()
                    client_unit = bool(row[4].value)
                    comments = str(row[6].value or "") if len(row) > 6 else ""
                    if not user or not work_order or not touch:
                        continue
                    self.submit_work(user, work_order, touch, client_unit, comments)
                    counts["submissions"] += 1
            elif include("submissions"):
                missing_sheets.append("Table1")

            # Parts
            if include("parts") and "tblParts" in wb.sheetnames:
                sh = wb["tblParts"]
                for row in sh.iter_rows(min_row=2):
                    date_time = row[0].value
                    user = nx(row[1].value)
                    work_order = DepotRules.normalize_work_order(row[2].value)
                    client_unit = bool(row[4].value) if len(row) > 4 else False
                    category = str(row[5].value or "").strip()
                    if not user or not work_order or not category:
                        continue
                    self.submit_part(user, user, work_order, category, client_unit, "")
                    counts["parts"] += 1
            elif include("parts"):
                missing_sheets.append("tblParts")

            # RTVs
            if include("rtvs") and "RTVs" in wb.sheetnames:
                sh = wb["RTVs"]
                for row in sh.iter_rows(min_row=2):
                    user = nx(row[1].value)
                    work_order = DepotRules.normalize_work_order(row[2].value)
                    comments = str(row[3].value or "")
                    if user and work_order:
                        now = datetime.utcnow().isoformat()
                        self.db.execute("INSERT OR IGNORE INTO rtvs (created_at, user_id, work_order, comments) VALUES (?, ?, ?, ?)", (now, user, work_order, comments))
                        counts["rtvs"] += 1
            elif include("rtvs"):
                missing_sheets.append("RTVs")

            # Client JO
            if include("client_jo") and "Client_JO" in wb.sheetnames:
                sh = wb["Client_JO"]
                for row in sh.iter_rows(min_row=2):
                    user = nx(row[1].value)
                    work_order = DepotRules.normalize_work_order(row[2].value)
                    comments = str(row[3].value or "")
                    if user and work_order:
                        now = datetime.utcnow().isoformat()
                        self.db.execute("INSERT OR IGNORE INTO client_jo (created_at, user_id, work_order, comments) VALUES (?, ?, ?, ?)", (now, user, work_order, comments))
                        counts["client_jo"] += 1
            elif include("client_jo"):
                missing_sheets.append("Client_JO")

            # client_parts
            if include("client_parts") and "client_parts" in wb.sheetnames:
                sh = wb["client_parts"]
                for row in sh.iter_rows(min_row=2):
                    user = nx(row[1].value)
                    work_order = DepotRules.normalize_work_order(row[2].value)
                    comments = str(row[4].value or "") if len(row) > 4 else ""
                    if user and work_order:
                        now = datetime.utcnow().isoformat()
                        self.db.execute("INSERT OR IGNORE INTO client_parts (created_at, user_id, work_order, comments) VALUES (?, ?, ?, ?)", (now, user, work_order, comments))
                        counts["client_parts"] += 1
            elif include("client_parts"):
                missing_sheets.append("client_parts")

            # agents
            if include("agents") and "agents" in wb.sheetnames:
                sh = wb["agents"]
                for row in sh.iter_rows(min_row=2):
                    agent_name = str(row[0].value or "").strip()
                    a_num = DepotRules.normalize_user_id(row[1].value)
                    tier = 1
                    if len(row) > 2 and row[2].value is not None:
                        try:
                            tier = DepotRules.normalize_agent_tier(row[2].value)
                        except Exception as exc:
                            _runtime_log_event(
                                "depot.import_workbook.agent_tier_parse_failed",
                                severity="warning",
                                summary="Agent tier value in workbook could not be parsed; defaulting to tier 1.",
                                exc=exc,
                                context={"raw_tier": row[2].value},
                            )
                            tier = 1
                    icon_path = str(row[3].value or "").strip() if len(row) > 3 else ""
                    location = str(row[4].value or "").strip() if len(row) > 4 else ""
                    if a_num:
                        self.upsert_agent(a_num, agent_name or a_num, tier, icon_path, location)
                        counts["agents"] += 1
            elif include("agents"):
                missing_sheets.append("agents")

            # QA flags
            if include("qa_flags") and "qa_flags" in wb.sheetnames:
                sh = wb["qa_flags"]
                for row in sh.iter_rows(min_row=2):
                    flag_name = str(row[0].value or "").strip()
                    severity = str(row[1].value or "Medium").strip() if len(row) > 1 else "Medium"
                    icon_path = str(row[2].value or "").strip() if len(row) > 2 else ""
                    if flag_name:
                        self.upsert_qa_flag(flag_name, severity, icon_path)
                        counts["qa_flags"] += 1
            elif include("qa_flags"):
                missing_sheets.append("qa_flags")

            # admin allowlist
            if include("admin_users") and "admin" in wb.sheetnames:
                sh = wb["admin"]
                for row in sh.iter_rows(min_row=2):
                    user_id = DepotRules.normalize_user_id(row[0].value)
                    admin_name = str(row[1].value or "").strip() if len(row) > 1 else ""
                    position = str(row[2].value or "").strip() if len(row) > 2 else ""
                    location = str(row[3].value or "").strip() if len(row) > 3 else ""
                    icon_path = str(row[4].value or "").strip() if len(row) > 4 else ""
                    if user_id:
                        self.add_admin_user(user_id, admin_name, position, location, icon_path)
                        counts["admin_users"] += 1
            elif include("admin_users"):
                missing_sheets.append("admin")
            imported_total = sum(int(v) for v in counts.values())
            label_by_key = {key: label for key, _sheet, label in WORKBOOK_IMPORT_SPECS}
            included_keys = [key for key, _sheet, _label in WORKBOOK_IMPORT_SPECS if include(key)]
            lines: list[str] = [f"Import complete. Rows processed: {imported_total}"]
            if included_keys:
                lines.append("Selected tables:")
                for key in included_keys:
                    lines.append(f"- {label_by_key.get(key, key)}: {counts.get(key, 0)}")
            if missing_sheets:
                lines.append("Missing sheets skipped: " + ", ".join(missing_sheets))
            return True, "\n".join(lines)
        except Exception as exc:
            context = {
                "workbook_path": str(xlsm_path),
                "selected_tables": sorted(str(name) for name in import_tables) if import_tables else [],
            }
            _runtime_log_event(
                "depot.import_workbook_unhandled",
                severity="critical",
                summary="Unhandled workbook import failure.",
                exc=exc,
                context=context,
            )
            _escalate_runtime_issue_once(
                "depot.import_workbook_unhandled",
                "Workbook import failed unexpectedly. Review the runtime log for diagnostics.",
                details=f"{type(exc).__name__}: {exc}",
                context=context,
            )
            return False, str(exc)


class PartNotesDialog(QDialog):
    def __init__(
        self,
        role: str,
        part_data: dict[str, Any],
        tracker: DepotTracker | None = None,
        app_window: "QuickInputsWindow" | None = None,
        current_user: str = "",
        parent: QWidget | None = None,
    ):
        super().__init__(parent)
        self.setObjectName("PartNotesDialog")
        self.role = "qa" if str(role).strip().lower() == "qa" else "agent"
        self.app_window = app_window
        self._style_kind = "qa" if self.role == "qa" else "agent"
        self.part_data = part_data
        self.tracker = tracker
        self.current_user = DepotRules.normalize_user_id(current_user)
        self._working_original_user = DepotRules.normalize_user_id(str(part_data.get("working_user_id", "") or ""))

        work_order = str(part_data.get("work_order", "") or "")
        self.setWindowTitle(f"Notes - {work_order}")
        self.setModal(True)
        self.resize(560, 320)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(8)

        summary = QLabel(
            f"Work Order: {work_order}    Category: {part_data.get('category', '-')}    "
            f"Client: {'Yes' if bool(part_data.get('client_unit', False)) else 'No'}"
        )
        summary.setWordWrap(True)
        layout.addWidget(summary)

        other_label = QLabel("Agent Note (Other User)" if self.role == "qa" else "QA Note (Other User)")
        other_label.setProperty("section", True)
        layout.addWidget(other_label)
        self.other_note_value = QLabel()
        self.other_note_value.setWordWrap(True)
        self.other_note_value.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        other_text = str(part_data.get("agent_comment" if self.role == "qa" else "qa_comment", "") or "").strip()
        self.other_note_value.setText(other_text if other_text else "(none)")
        layout.addWidget(self.other_note_value)

        own_label = QLabel("QA Note (Your Note)" if self.role == "qa" else "Agent Note (Your Note)")
        own_label.setProperty("section", True)
        layout.addWidget(own_label)
        self.own_note_input = QTextEdit()
        self.own_note_input.setFixedHeight(86)
        own_value = str(part_data.get("qa_comment" if self.role == "qa" else "agent_comment", "") or "").strip()
        self.own_note_input.setPlainText(own_value)
        layout.addWidget(self.own_note_input)

        if self.role == "qa":
            qa_flag_row = QHBoxLayout()
            qa_flag_row.addWidget(QLabel("Flag"), 0)
            self.qa_flag_combo = QComboBox()
            flag_options = self.tracker.get_qa_flag_options(include_none=True) if self.tracker is not None else list(QA_FLAG_OPTIONS)
            for option in flag_options:
                self.qa_flag_combo.addItem(option)
            current_flag = str(part_data.get("qa_flag", "") or "").strip()
            current_display = current_flag if current_flag else "None"
            idx = self.qa_flag_combo.findText(current_display)
            self.qa_flag_combo.setCurrentIndex(idx if idx >= 0 else 0)
            qa_flag_row.addWidget(self.qa_flag_combo, 1)
            layout.addLayout(qa_flag_row)
        else:
            self.work_toggle_check = QCheckBox("Working this unit (\U0001F527)")
            if self.current_user:
                if self._working_original_user and self._working_original_user != self.current_user:
                    self.work_toggle_check.setChecked(True)
                    self.work_toggle_check.setEnabled(False)
                    lock_notice = QLabel(f"Currently marked by {self._working_original_user}.")
                    lock_notice.setProperty("muted", True)
                    layout.addWidget(lock_notice)
                else:
                    self.work_toggle_check.setChecked(self._working_original_user == self.current_user)
            else:
                self.work_toggle_check.setEnabled(False)
            layout.addWidget(self.work_toggle_check)

            flag_text = str(part_data.get("qa_flag", "") or "").strip()
            flag_value = flag_text if flag_text else "None"
            layout.addWidget(QLabel(f"Flag: {flag_value}"))

        buttons = QHBoxLayout()
        self.save_btn = QPushButton("Save")
        self.cancel_btn = QPushButton("Cancel")
        self.save_btn.setProperty("actionRole", "save")
        self.cancel_btn.setProperty("actionRole", "reset")
        buttons.addWidget(self.save_btn)
        buttons.addWidget(self.cancel_btn)
        buttons.addStretch(1)
        layout.addLayout(buttons)

        self.save_btn.clicked.connect(self.accept)
        self.cancel_btn.clicked.connect(self.reject)

        if app_window is not None:
            base_css = app_window._popup_theme_stylesheet(self._style_kind, force_opaque_root=True)
            resolved = app_window._resolved_popup_theme(self._style_kind)
            dialog_bg = normalize_hex(
                resolved.get("background", app_window.palette_data.get("surface", DEFAULT_THEME_SURFACE)),
                app_window.palette_data.get("surface", DEFAULT_THEME_SURFACE),
            )
            dialog_text = normalize_hex(
                resolved.get("text", app_window.palette_data.get("label_text", "#FFFFFF")),
                app_window.palette_data.get("label_text", "#FFFFFF"),
            )
            if contrast_ratio(dialog_bg, dialog_text) < 4.0:
                dialog_text = readable_text(dialog_bg)
            field_bg = normalize_hex(
                resolved.get("field_bg", app_window.palette_data.get("input_bg", "#FFFFFF")),
                app_window.palette_data.get("input_bg", "#FFFFFF"),
            )
            field_text = dialog_text if contrast_ratio(field_bg, dialog_text) >= 4.0 else readable_text(field_bg)
            field_border = shift(field_bg, -0.38)
            self.setStyleSheet(
                base_css
                + (
                    "QDialog#PartNotesDialog {"
                    "background-color: transparent;"
                    f"color: {dialog_text};"
                    "}"
                    "QDialog#PartNotesDialog QLabel {"
                    f"color: {dialog_text};"
                    "background-color: transparent;"
                    "font-weight: 700;"
                    "}"
                    "QDialog#PartNotesDialog QTextEdit, QDialog#PartNotesDialog QLineEdit, QDialog#PartNotesDialog QComboBox {"
                    f"background-color: {rgba_css(field_bg, 0.95)};"
                    f"color: {field_text};"
                    f"border: 1px solid {field_border};"
                    "}"
                )
            )

    def paintEvent(self, event) -> None:  # noqa: N802
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform)
        if self.app_window is not None:
            theme = self.app_window._resolved_popup_theme(self._style_kind)
            base_bg = normalize_hex(
                theme.get("background", self.app_window.palette_data.get("surface", DEFAULT_THEME_SURFACE)),
                self.app_window.palette_data.get("surface", DEFAULT_THEME_SURFACE),
            )
            bg_color = QColor(base_bg)
            if bool(theme.get("transparent", False)):
                bg_color.setAlpha(220)
            painter.fillRect(self.rect(), bg_color)
            bg = self.app_window.render_background_pixmap(self.rect().size(), kind=self._style_kind)
            if not bg.isNull():
                painter.drawPixmap(self.rect(), bg)
        super().paintEvent(event)

    def values(self) -> dict[str, str]:
        own_note = self.own_note_input.toPlainText().strip()
        out: dict[str, str] = {
            "own_note": own_note,
            "qa_flag": "",
            "qa_flag_image_path": "",
            "working_user_id": "__UNCHANGED__",
        }
        if self.role == "qa":
            selected_flag = str(self.qa_flag_combo.currentText() if hasattr(self, "qa_flag_combo") else "").strip()
            if selected_flag.lower() == "none":
                selected_flag = ""
            out["qa_flag"] = selected_flag
            out["qa_flag_image_path"] = ""
        elif hasattr(self, "work_toggle_check") and self.work_toggle_check.isEnabled():
            out["working_user_id"] = self.current_user if self.work_toggle_check.isChecked() else ""
        return out


class AlertPulseTabBar(QTabBar):
    """Tab bar with per-tab alert pulse/bloom rendering without changing tab width."""

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._alert_indices: set[int] = set()
        self._ack_indices: set[int] = set()
        self._flash_on = True
        self._paint_error_logged = False
        self._normal_text_color = QColor("#FFFFFF")
        self._alert_color = QColor("#D95A5A")
        self._ack_color = QColor("#D1A91F")
        self.setDrawBase(False)

    def set_alert_visual_state(
        self,
        alert_indices: set[int],
        ack_indices: set[int],
        flash_on: bool,
        normal_text_color: QColor,
    ) -> None:
        self._alert_indices = {int(idx) for idx in alert_indices if idx >= 0}
        self._ack_indices = {int(idx) for idx in ack_indices if idx >= 0}
        self._flash_on = bool(flash_on)
        self._normal_text_color = QColor(normal_text_color)
        self.update()

    def _paint_bloom(self, painter: QPainter, tab_rect: QRect, pulse_on: bool, acknowledged: bool) -> None:
        if tab_rect.isNull():
            return
        base = QColor(self._ack_color if acknowledged else self._alert_color)
        if acknowledged:
            inner_alpha = 126
            mid_alpha = 82
            outer_alpha = 46
        elif pulse_on:
            inner_alpha = 210
            mid_alpha = 148
            outer_alpha = 92
        else:
            inner_alpha = 22
            mid_alpha = 10
            outer_alpha = 0

        outer_rect = QRectF(tab_rect.adjusted(-14, -8, 14, 9))
        mid_rect = QRectF(tab_rect.adjusted(-8, -5, 8, 6))
        inner_rect = QRectF(tab_rect.adjusted(-3, -2, 3, 3))

        painter.save()
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        painter.setPen(Qt.PenStyle.NoPen)

        for rect, alpha in ((outer_rect, outer_alpha), (mid_rect, mid_alpha), (inner_rect, inner_alpha)):
            path = QPainterPath()
            path.addRoundedRect(rect, 9.0, 9.0)
            fill = QColor(base)
            fill.setAlpha(int(max(0, alpha)))
            painter.fillPath(path, fill)
        painter.restore()

    def paintEvent(self, event) -> None:  # noqa: N802
        try:
            painter = QStylePainter(self)
            for idx in range(int(self.count())):
                opt = QStyleOptionTab()
                self.initStyleOption(opt, idx)
                alert_enabled = idx in self._alert_indices
                acknowledged = idx in self._ack_indices
                pulse_on = bool(alert_enabled and not acknowledged and self._flash_on)

                palette = QPalette(opt.palette)
                if alert_enabled:
                    if pulse_on:
                        text_color = QColor("#FFFFFF")
                    elif acknowledged:
                        text_color = QColor("#FFF9E5")
                    else:
                        text_color = QColor("#FFF2F2")
                else:
                    text_color = QColor(self._normal_text_color)
                palette.setColor(QPalette.ColorRole.ButtonText, text_color)
                palette.setColor(QPalette.ColorRole.WindowText, text_color)
                opt.palette = palette

                painter.drawControl(QStyle.ControlElement.CE_TabBarTabShape, opt)
                if alert_enabled:
                    self._paint_bloom(painter, self.tabRect(idx), pulse_on, acknowledged)
                painter.drawControl(QStyle.ControlElement.CE_TabBarTabLabel, opt)
        except Exception as exc:
            if not self._paint_error_logged:
                self._paint_error_logged = True
                _runtime_log_event(
                    "ui.alert_tabbar_paint_failed",
                    severity="error",
                    summary="Alert pulse tab bar paint failed; falling back to default tab rendering.",
                    exc=exc,
                    context={"tab_count": int(self.count())},
                )
            super().paintEvent(event)


class DepotAgentWindow(DepotFramelessToolWindow):
    FLAG_WORK_REVERT_HOURS = 2

    def __init__(self, tracker: DepotTracker, current_user: str, app_window: "QuickInputsWindow" | None = None):
        super().__init__(
            app_window,
            window_title="Agent",
            theme_kind="agent",
            size=(780, 292),
            minimum_size=(740, 278),
        )
        self.tracker = tracker
        self.current_user = DepotRules.normalize_user_id(current_user)
        self._always_on_top_config_key = "agent_window_always_on_top"
        self._window_always_on_top = self._load_window_always_on_top_preference(self._always_on_top_config_key, default=True)
        self.set_window_always_on_top(self._window_always_on_top)
        self._current_agent_tier = self._resolve_current_agent_tier()
        self._is_tech3_user = int(self._current_agent_tier) == 3

        self.agent_tabs = QTabWidget(self)
        self._agent_tab_bar = AlertPulseTabBar(self.agent_tabs)
        self.agent_tabs.setTabBar(self._agent_tab_bar)
        self.root_layout.addWidget(self.agent_tabs)

        self.work_tab = QWidget()
        self.parts_tab = QWidget()
        self.cat_parts_tab = QWidget()
        self.client_tab = QWidget()
        self.team_client_tab: QWidget | None = None

        self.agent_tabs.addTab(self.work_tab, "Work")
        self.agent_tabs.addTab(self.parts_tab, "Parts")
        self.agent_tabs.addTab(self.cat_parts_tab, "Cat Parts")
        self.agent_tabs.addTab(self.client_tab, "Client")
        if self.team_client_tab is not None:
            self.agent_tabs.addTab(self.team_client_tab, "Team Client")

        self._tab_indices: dict[str, int] = {"parts": 1, "cat_parts": 2, "client": 3}
        if self.team_client_tab is not None:
            self._tab_indices["team_client"] = 4
        self._tab_titles: dict[str, str] = {"parts": "Parts", "cat_parts": "Cat Parts", "client": "Client"}
        if self.team_client_tab is not None:
            self._tab_titles["team_client"] = "Team Client"
        self._tab_alert_states: dict[str, bool] = {"parts": False, "cat_parts": False, "client": False}
        if self.team_client_tab is not None:
            self._tab_alert_states["team_client"] = False
        self._tab_alert_ack_states: dict[str, bool] = {"parts": False, "cat_parts": False, "client": False}
        if self.team_client_tab is not None:
            self._tab_alert_ack_states["team_client"] = False
        self._tab_flash_on = True
        self._parts_has_flagged_rows = False
        self._cat_parts_has_flagged_rows = False
        self._parts_has_urgent_flagged_rows = False
        self._parts_has_in_progress_flagged_rows = False
        self._cat_parts_has_urgent_flagged_rows = False
        self._cat_parts_has_in_progress_flagged_rows = False
        self._client_due_ack_ids: set[int] = set()
        self._client_due_active_ids: set[int] = set()
        self._team_client_due_count = 0
        self._tab_flash_timer = QTimer(self)
        self._tab_flash_timer.setInterval(700)
        self._tab_flash_timer.timeout.connect(self._on_tab_alert_flash_tick)
        self._tab_flash_timer.start()
        self._flag_watchdog_timer = QTimer(self)
        self._flag_watchdog_timer.setInterval(60000)
        self._flag_watchdog_timer.timeout.connect(self._refresh_flag_alert_watchdog)
        self._flag_watchdog_timer.start()
        self.agent_tabs.currentChanged.connect(self._on_agent_tab_changed)

        self._build_work_tab()
        self._build_parts_tab()
        self._build_cat_parts_tab()
        self._build_client_tab()
        if self.team_client_tab is not None:
            self._build_team_client_tab()

        self._refresh_recent_submissions_label()

        if self.app_window is not None:
            self.apply_theme_styles()
            self.app_window.active_agent_window = self

    def _build_work_tab(self):
        root = QHBoxLayout(self.work_tab)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(6)

        left_wrap = QWidget(self.work_tab)
        left_col = QVBoxLayout(left_wrap)
        left_col.setContentsMargins(0, 0, 0, 0)
        left_col.setSpacing(4)

        form_wrap = QWidget(left_wrap)
        left_layout = QFormLayout(form_wrap)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setHorizontalSpacing(6)
        left_layout.setVerticalSpacing(3)

        self.work_user_lbl = QLabel(self.current_user)
        self.work_order_input = QLineEdit()
        self.work_status = QComboBox()
        self.work_status.addItems(["Complete", "Junk Out", "Other", "Part Order", "RTV", "Triaged"])
        self.work_client_check = QCheckBox("")
        self.work_comments = QLineEdit()
        self.agent_always_on_top_check = QCheckBox("Always on top")
        self.agent_always_on_top_check.setChecked(bool(self._window_always_on_top))
        self.agent_always_on_top_check.toggled.connect(self._on_agent_always_on_top_toggled)
        self.work_submit_btn = QPushButton("Submit")
        self.work_submit_btn.clicked.connect(self._submit_work_entry)
        self.work_submit_btn.setFixedHeight(24)

        action_row_wrap = QWidget(form_wrap)
        action_row = QHBoxLayout(action_row_wrap)
        action_row.setContentsMargins(0, 0, 0, 0)
        action_row.setSpacing(6)
        action_row.addWidget(self.agent_always_on_top_check, 0)
        action_row.addStretch(1)
        action_row.addWidget(self.work_submit_btn, 0)

        left_layout.addRow("User", self.work_user_lbl)
        left_layout.addRow("Work Order", self.work_order_input)
        left_layout.addRow("Status Update", self.work_status)
        left_layout.addRow("Client", self.work_client_check)
        left_layout.addRow("Comments", self.work_comments)
        left_layout.addRow("", action_row_wrap)
        left_col.addWidget(form_wrap, 0)

        self.recent_submissions_label = QLabel()
        self.recent_submissions_label.setWordWrap(False)
        self.recent_submissions_label.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        self.recent_submissions_label.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed)
        line_height = max(14, int(self.recent_submissions_label.fontMetrics().lineSpacing()))
        self.recent_submissions_label.setFixedHeight((line_height * 4) + 6)
        left_col.addWidget(self.recent_submissions_label, 0)
        left_col.addStretch(1)
        root.addWidget(left_wrap, 3)

        chart_wrap = QWidget(self.work_tab)
        chart_wrap.setMinimumWidth(304)
        chart_layout = QVBoxLayout(chart_wrap)
        chart_layout.setContentsMargins(0, 0, 0, 0)
        chart_layout.setSpacing(1)

        chart_title = QLabel("Touch Summary")
        chart_title.setProperty("section", True)
        chart_layout.addWidget(chart_title)

        self.agent_touch_bars: dict[str, TouchDistributionBar] = {}
        self.agent_touch_legends: dict[str, QLabel] = {}
        self.agent_touch_row_totals: dict[str, QLabel] = {}
        for key, heading in (("today", "Today"), ("last_7", "7 Day"), ("last_30", "30 Day")):
            bar = TouchDistributionBar(chart_wrap)
            bar.setFixedHeight(12)
            self.agent_touch_bars[key] = bar
            row = QHBoxLayout()
            row.setContentsMargins(0, 0, 0, 0)
            row.setSpacing(3)
            lbl = QLabel(heading)
            lbl.setProperty("section", True)
            lbl.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            lbl.setFixedWidth(58)
            row.addWidget(lbl, 0)
            row.addWidget(bar, 1)
            total_lbl = QLabel("0")
            total_lbl.setProperty("section", True)
            total_lbl.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            total_lbl.setFixedWidth(34)
            total_lbl.setStyleSheet("font-size: 15px; font-weight: 800;")
            self.agent_touch_row_totals[key] = total_lbl
            row.addWidget(total_lbl, 0)
            chart_layout.addLayout(row)
            legend = QLabel("No scans.")
            legend.setWordWrap(False)
            legend.setAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
            legend.setMargin(0)
            self.agent_touch_legends[key] = legend
            chart_layout.addWidget(legend)

        root.addWidget(chart_wrap, 2)

        self.work_order_input.returnPressed.connect(self._submit_work_entry)

    def apply_theme_styles(self) -> None:
        super().apply_theme_styles()
        self._apply_tab_alert_visuals()

    def _resolve_current_agent_tier(self) -> int:
        try:
            row = self.tracker.db.fetchone("SELECT tier FROM agents WHERE user_id=? LIMIT 1", (self.current_user,))
            if row is None:
                return 1
            return DepotRules.normalize_agent_tier(row["tier"])
        except Exception as exc:
            _runtime_log_event(
                "ui.agent_tier_lookup_failed",
                severity="warning",
                summary="Failed resolving agent tier for tab-alert behavior.",
                exc=exc,
                context={"user_id": str(self.current_user)},
            )
            return 1

    def _tab_normal_text_color(self) -> QColor:
        if self.app_window is not None:
            return QColor(normalize_hex(self.app_window.palette_data.get("label_text", "#FFFFFF"), "#FFFFFF"))
        return QColor("#FFFFFF")

    def _tab_key_for_index(self, index: int) -> str:
        for key, idx in self._tab_indices.items():
            if int(idx) == int(index):
                return key
        return ""

    def _acknowledge_tab_alert(self, key: str) -> None:
        if key not in self._tab_alert_states:
            return
        if not bool(self._tab_alert_states.get(key, False)):
            return
        if bool(self._tab_alert_ack_states.get(key, False)):
            return
        self._tab_alert_ack_states[key] = True
        self._apply_tab_alert_visuals()

    def _on_agent_tab_changed(self, index: int) -> None:
        key = self._tab_key_for_index(index)
        if key == "client":
            self._acknowledge_tab_alert(key)

    def _set_tab_alert(self, key: str, enabled: bool, acknowledged: bool | None = None) -> None:
        if key not in self._tab_alert_states:
            return
        was_enabled = bool(self._tab_alert_states.get(key, False))
        enabled_now = bool(enabled)
        self._tab_alert_states[key] = enabled_now
        if acknowledged is not None:
            self._tab_alert_ack_states[key] = bool(enabled_now and acknowledged)
        else:
            if enabled_now and not was_enabled:
                self._tab_alert_ack_states[key] = False
            elif not enabled_now:
                self._tab_alert_ack_states[key] = False
            if enabled_now:
                tab_idx = int(self._tab_indices.get(key, -1))
                if tab_idx >= 0 and tab_idx == int(self.agent_tabs.currentIndex()):
                    self._tab_alert_ack_states[key] = True
        self._apply_tab_alert_visuals()

    def _update_tab_alert_states(self) -> None:
        parts_alert = bool(self._parts_has_urgent_flagged_rows or self._parts_has_in_progress_flagged_rows)
        parts_ack = bool(self._parts_has_in_progress_flagged_rows and not self._parts_has_urgent_flagged_rows)
        self._set_tab_alert("parts", parts_alert, acknowledged=parts_ack)

        cat_urgent = bool(self._is_tech3_user and self._cat_parts_has_urgent_flagged_rows)
        cat_in_progress = bool(self._is_tech3_user and self._cat_parts_has_in_progress_flagged_rows)
        cat_alert = bool(cat_urgent or cat_in_progress)
        cat_ack = bool(cat_in_progress and not cat_urgent)
        self._set_tab_alert("cat_parts", cat_alert, acknowledged=cat_ack)

        client_alert = bool(
            any(
                int(part_id) not in self._client_due_ack_ids
                for part_id in getattr(self, "_client_due_active_ids", set())
            )
        )
        self._set_tab_alert("client", client_alert)
        if "team_client" in self._tab_alert_states:
            self._set_tab_alert("team_client", bool(int(self._team_client_due_count) > 0), acknowledged=False)

    def _apply_tab_alert_visuals(self) -> None:
        if not hasattr(self, "agent_tabs"):
            return
        tab_bar = self.agent_tabs.tabBar()
        normal_color = self._tab_normal_text_color()
        alert_indices: set[int] = set()
        ack_indices: set[int] = set()
        for key, idx in self._tab_indices.items():
            if idx < 0 or idx >= int(self.agent_tabs.count()):
                continue
            base_text = self._tab_titles.get(key, self.agent_tabs.tabText(idx))
            self.agent_tabs.setTabText(idx, base_text)
            should_alert = bool(self._tab_alert_states.get(key, False))
            if should_alert:
                alert_indices.add(int(idx))
                if bool(self._tab_alert_ack_states.get(key, False)):
                    ack_indices.add(int(idx))
        if isinstance(tab_bar, AlertPulseTabBar):
            tab_bar.set_alert_visual_state(alert_indices, ack_indices, bool(self._tab_flash_on), normal_color)
        else:
            flashing_color = QColor("#F4BCBC")
            acknowledged_color = QColor("#E6C177")
            for key, idx in self._tab_indices.items():
                if idx < 0 or idx >= int(self.agent_tabs.count()):
                    continue
                if idx not in alert_indices:
                    tab_bar.setTabTextColor(idx, normal_color)
                    continue
                if idx in ack_indices:
                    tab_bar.setTabTextColor(idx, acknowledged_color)
                else:
                    tab_bar.setTabTextColor(idx, flashing_color if self._tab_flash_on else normal_color)

    def _on_tab_alert_flash_tick(self) -> None:
        if not any(bool(value) for value in self._tab_alert_states.values()):
            self._tab_flash_on = True
            self._apply_tab_alert_visuals()
            return
        self._tab_flash_on = not bool(self._tab_flash_on)
        self._apply_tab_alert_visuals()

    def _on_agent_always_on_top_toggled(self, checked: bool) -> None:
        self._window_always_on_top = self._apply_window_always_on_top_preference(self._always_on_top_config_key, checked)

    @staticmethod
    def _agent_touch_color(touch: str) -> str:
        palette = {
            DepotRules.TOUCH_COMPLETE: "#21B46D",
            DepotRules.TOUCH_JUNK: "#D95A5A",
            DepotRules.TOUCH_PART_ORDER: "#D3A327",
            DepotRules.TOUCH_RTV: "#4F86D9",
            "Triaged": "#20AFA8",
            DepotRules.TOUCH_OTHER: "#8E97A8",
        }
        return normalize_hex(palette.get(str(touch or "").strip(), "#6F7C91"), "#6F7C91")

    def _query_agent_touch_counts(self, start_date: str, end_date: str) -> tuple[dict[str, int], int]:
        rows = self.tracker.db.fetchall(
            "SELECT touch, COUNT(*) AS c "
            "FROM submissions "
            "WHERE user_id=? "
            "AND COALESCE(NULLIF(TRIM(entry_date), ''), SUBSTR(created_at, 1, 10))>=? "
            "AND COALESCE(NULLIF(TRIM(entry_date), ''), SUBSTR(created_at, 1, 10))<=? "
            "GROUP BY touch",
            (self.current_user, start_date, end_date),
        )
        counts: dict[str, int] = {}
        total = 0
        for row in rows:
            touch = str(row["touch"] or "").strip()
            if not touch:
                continue
            count = int(max(0, safe_int(row["c"], 0)))
            counts[touch] = count
            total += count
        return counts, total

    def _refresh_work_touch_chart(self) -> None:
        if not hasattr(self, "agent_touch_bars"):
            return

        today = datetime.now().date()
        ranges = {
            "today": (today, today),
            "last_7": (today - timedelta(days=6), today),
            "last_30": (today - timedelta(days=29), today),
        }

        totals: dict[str, int] = {"today": 0, "last_7": 0, "last_30": 0}
        touch_order = (
            DepotRules.TOUCH_COMPLETE,
            DepotRules.TOUCH_JUNK,
            DepotRules.TOUCH_PART_ORDER,
            DepotRules.TOUCH_RTV,
            "Triaged",
            DepotRules.TOUCH_OTHER,
        )

        try:
            for key, (start_dt, end_dt) in ranges.items():
                counts, total = self._query_agent_touch_counts(start_dt.isoformat(), end_dt.isoformat())
                totals[key] = total
                ordered: list[str] = []
                for touch in touch_order:
                    if touch in counts:
                        ordered.append(touch)
                for touch in sorted(counts.keys()):
                    if touch not in ordered:
                        ordered.append(touch)

                segments = [(touch, int(counts.get(touch, 0)), self._agent_touch_color(touch)) for touch in ordered if int(counts.get(touch, 0)) > 0]
                self.agent_touch_bars[key].set_segments(segments)
                if key in self.agent_touch_row_totals:
                    self.agent_touch_row_totals[key].setText(str(int(total)))
                if segments:
                    legend_cells = [
                        (
                            "<td align='center' style='padding:0 4px;'>"
                            f"<span style='color:{color}; font-weight:700'>{DepotRules.chart_touch_label(touch)}</span>: {count}"
                            "</td>"
                        )
                        for touch, count, color in segments
                    ]
                    legend_text = "<table width='100%' cellspacing='0' cellpadding='0'><tr>" + "".join(legend_cells) + "</tr></table>"
                    self.agent_touch_legends[key].setTextFormat(Qt.TextFormat.RichText)
                    self.agent_touch_legends[key].setText(legend_text)
                else:
                    self.agent_touch_legends[key].setTextFormat(Qt.TextFormat.PlainText)
                    self.agent_touch_legends[key].setText("No scans.")
        except Exception as exc:
            _runtime_log_event(
                "ui.agent_touch_chart_query_failed",
                severity="warning",
                summary="Failed refreshing agent touch summary chart.",
                exc=exc,
                context={"user_id": str(self.current_user)},
            )
            for key in ("today", "last_7", "last_30"):
                self.agent_touch_bars[key].set_segments([])
                self.agent_touch_legends[key].setTextFormat(Qt.TextFormat.PlainText)
                self.agent_touch_legends[key].setText("Unavailable")
                if key in self.agent_touch_row_totals:
                    self.agent_touch_row_totals[key].setText("0")
            totals = {"today": 0, "last_7": 0, "last_30": 0}

    def _submit_work_entry(self):
        wo = self.work_order_input.text().strip()
        if not wo:
            self._show_themed_message(QMessageBox.Icon.Warning, "Validation", "Work order is required.")
            return
        touch = self.work_status.currentText()
        client_unit = self.work_client_check.isChecked()
        comments = self.work_comments.text().strip()

        try:
            self.tracker.submit_work(self.current_user, wo, touch, client_unit, comments)
            self.work_order_input.clear()
            self.work_comments.clear()
            self.work_order_input.setFocus()
            self._refresh_recent_submissions_label()
            self._refresh_client_followup()
            self._refresh_agent_parts()
            self._refresh_category_parts()
            if self.team_client_tab is not None:
                self._refresh_team_client_followup()
            if self.app_window is not None:
                qa_window = getattr(self.app_window, "active_qa_window", None)
                if qa_window is not None and qa_window.isVisible():
                    try:
                        qa_window._refresh_assigned_parts()
                        qa_window._refresh_delivered_parts()
                        qa_window._refresh_completed_parts()
                    except Exception as exc:
                        _runtime_log_event(
                            "ui.agent_qa_window_refresh_failed",
                            severity="warning",
                            summary="Agent submit succeeded but QA window refresh failed.",
                            exc=exc,
                            context={"user_id": str(self.current_user)},
                        )
        except Exception as exc:
            self._show_themed_message(QMessageBox.Icon.Critical, "Error", f"Failed to save: {exc}")

    def _refresh_recent_submissions_label(self) -> None:
        try:
            rows = self.tracker.db.fetchall(
                "SELECT work_order, touch, client_unit, created_at FROM submissions WHERE user_id=? ORDER BY created_at DESC LIMIT 3",
                (self.current_user,),
            )
        except Exception as exc:
            _runtime_log_event(
                "ui.agent_recent_submissions_query_failed",
                severity="warning",
                summary="Failed querying recent submissions for agent panel; showing unavailable fallback.",
                exc=exc,
                context={"user_id": str(self.current_user)},
            )
            self.recent_submissions_label.setText("Recent submissions: unavailable")
            self._refresh_work_touch_chart()
            self._refresh_client_followup()
            if self.team_client_tab is not None:
                self._refresh_team_client_followup()
            return

        if not rows:
            self.recent_submissions_label.setText("Latest 3 submissions:\n1. (none)\n2. (none)\n3. (none)")
            self._refresh_work_touch_chart()
            self._refresh_client_followup()
            if self.team_client_tab is not None:
                self._refresh_team_client_followup()
            return

        lines: list[str] = ["Latest 3 submissions:"]
        for index, row in enumerate(rows, start=1):
            wo = str(row["work_order"])
            touch = str(row["touch"])
            client_marker = " \u2713" if int(row["client_unit"] or 0) else ""
            created = str(row["created_at"] or "")
            stamp = created[11:16] if len(created) >= 16 else created
            lines.append(f"{index}. {wo} ({touch}{client_marker}) [{stamp}]")

        for index in range(len(rows) + 1, 4):
            lines.append(f"{index}. (none)")

        self.recent_submissions_label.setText("\n".join(lines))
        self._refresh_work_touch_chart()
        self._refresh_client_followup()
        if self.team_client_tab is not None:
            self._refresh_team_client_followup()

    def _build_parts_tab(self):
        layout = QVBoxLayout(self.parts_tab)
        self.parts_table = QTableWidget()
        configure_standard_table(
            self.parts_table,
            ["Work Order", "Client", "Flag", "Age", "Working", "Installed", "Category", "QA Note"],
            resize_modes={
                0: QHeaderView.ResizeMode.ResizeToContents,
                1: QHeaderView.ResizeMode.ResizeToContents,
                2: QHeaderView.ResizeMode.ResizeToContents,
                3: QHeaderView.ResizeMode.ResizeToContents,
                4: QHeaderView.ResizeMode.ResizeToContents,
                5: QHeaderView.ResizeMode.ResizeToContents,
                6: QHeaderView.ResizeMode.ResizeToContents,
                7: QHeaderView.ResizeMode.Stretch,
            },
            stretch_last=True,
        )
        self.parts_table.horizontalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        self.parts_table.cellClicked.connect(
            lambda row, col: self._on_parts_table_cell_clicked("parts", self.parts_table, row, col)
        )
        self.parts_table.itemDoubleClicked.connect(self._copy_work_order_from_table_item)
        self.parts_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.parts_table.customContextMenuRequested.connect(
            lambda pos: self._open_agent_notes_from_context(self.parts_table, pos)
        )
        self.parts_refresh_btn = QPushButton("Refresh")
        self.parts_refresh_btn.clicked.connect(self._refresh_agent_parts)
        self.parts_open_notes_btn = QPushButton("Open Notes")
        self.parts_open_notes_btn.setProperty("actionRole", "pick")
        self.parts_open_notes_btn.clicked.connect(lambda: self._open_agent_notes_for_table(self.parts_table))
        self.parts_working_btn = QPushButton("Agent Is Working This")
        self.parts_working_btn.setProperty("actionRole", "apply")
        self.parts_working_btn.clicked.connect(lambda: self._toggle_selected_part_working(self.parts_table))
        self.parts_installed_btn = QPushButton("Parts Installed")
        self.parts_installed_btn.setProperty("actionRole", "apply")
        self.parts_installed_btn.clicked.connect(lambda: self._toggle_selected_part_installed(self.parts_table))
        self.parts_workorder_search = QLineEdit()
        self.parts_workorder_search.setPlaceholderText("Search work order...")
        self.parts_workorder_search.setClearButtonEnabled(True)
        self.parts_workorder_search.textChanged.connect(self._refresh_agent_parts)
        controls = QHBoxLayout()
        controls.addWidget(QLabel("Work Order:"))
        controls.addWidget(self.parts_workorder_search, 1)
        controls.addWidget(self.parts_refresh_btn)
        controls.addWidget(self.parts_open_notes_btn)
        controls.addWidget(self.parts_working_btn)
        controls.addWidget(self.parts_installed_btn)
        layout.addLayout(controls)
        layout.addWidget(self.parts_table)
        self._refresh_agent_parts()

    def _copy_work_order_from_table_item(self, item: QTableWidgetItem) -> None:
        table, work_order = _copy_work_order_from_table_item(item)
        if work_order:
            self._show_copy_notice(table, f"Copied Work Order: {work_order}", duration_ms=4200)

    def _open_agent_notes_from_context(self, table: QTableWidget, pos: QPoint) -> None:
        if not _select_table_row_by_context_pos(table, pos):
            return
        self._open_agent_notes_for_table(table)

    @staticmethod
    def _flag_tooltip(flag: str, qa_comment: str, agent_comment: str, has_image: bool) -> str:
        flag_text = flag if flag else "None"
        qa_text = qa_comment if qa_comment else "(none)"
        agent_text = agent_comment if agent_comment else "(none)"
        return (
            "Double-click to copy work order.\n"
            "Right-click to open notes.\n"
            f"Flag: {flag_text}\n"
            f"QA Note: {qa_text}\n"
            f"Agent Note: {agent_text}"
        )

    @staticmethod
    def _part_age_label(created_at: str) -> str:
        raw = str(created_at or "").strip()
        if len(raw) < 10:
            return "-"
        try:
            created_day = datetime.strptime(raw[:10], "%Y-%m-%d").date()
        except Exception:
            return "-"
        age_days = max(0, (datetime.utcnow().date() - created_day).days)
        return f"{age_days}d"

    @staticmethod
    def _note_preview(note: str, max_len: int = 64) -> str:
        cleaned = " ".join(str(note or "").split()).strip()
        if not cleaned:
            return "(none)"
        if len(cleaned) <= max_len:
            return cleaned
        return cleaned[: max(0, max_len - 3)].rstrip() + "..."

    @staticmethod
    def _center_item(item: QTableWidgetItem) -> QTableWidgetItem:
        return _center_table_item(item)

    def _working_status_icon(self) -> QIcon:
        icon = QIcon.fromTheme("applications-engineering")
        if icon.isNull():
            icon = QIcon.fromTheme("tools")
        if icon.isNull():
            icon = QIcon.fromTheme("preferences-system")
        if icon.isNull():
            icon = self.style().standardIcon(QStyle.StandardPixmap.SP_FileDialogDetailedView)
        return icon

    def _working_available_icon(self) -> QIcon:
        icon = QIcon.fromTheme("system-search")
        if icon.isNull():
            icon = QIcon.fromTheme("edit-find")
        if icon.isNull():
            icon = self.style().standardIcon(QStyle.StandardPixmap.SP_DialogOpenButton)
        return icon

    def _client_checked_icon(self) -> QIcon:
        return self.style().standardIcon(QStyle.StandardPixmap.SP_DialogApplyButton)

    def _part_on_hand_icon(self) -> QIcon:
        return self.style().standardIcon(QStyle.StandardPixmap.SP_DialogApplyButton)

    def _part_installed_icon(self) -> QIcon:
        icon = QIcon.fromTheme("applications-engineering")
        if icon.isNull():
            icon = QIcon.fromTheme("tools")
        if icon.isNull():
            icon = QIcon.fromTheme("preferences-system")
        if icon.isNull():
            icon = self.style().standardIcon(QStyle.StandardPixmap.SP_DialogApplyButton)
        return icon

    def _followup_done_icon(self) -> QIcon:
        return self.style().standardIcon(QStyle.StandardPixmap.SP_DialogApplyButton)

    def _followup_clock_icon(self, color_hex: str) -> QIcon:
        color = QColor(normalize_hex(color_hex, "#21B46D"))
        pix = QPixmap(16, 16)
        pix.fill(Qt.GlobalColor.transparent)
        painter = QPainter(pix)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        pen = QPen(color)
        pen.setWidth(2)
        painter.setPen(pen)
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.drawEllipse(2, 2, 12, 12)
        painter.drawLine(8, 8, 8, 5)
        painter.drawLine(8, 8, 11, 9)
        painter.end()
        return QIcon(pix)

    def _followup_wait_icon_by_days(self, days_since_action: int) -> tuple[QIcon, str]:
        days = int(max(0, safe_int(days_since_action, 0)))
        if days <= 0:
            return self._followup_clock_icon("#21B46D"), DepotRules.followup_stage_label(0)
        if days == 1:
            return self._followup_clock_icon("#D1A91F"), DepotRules.followup_stage_label(1)
        return self._followup_clock_icon("#D95A5A"), DepotRules.followup_stage_label(2)

    @staticmethod
    def _parse_iso_date(raw_value: str) -> date | None:
        stamp = str(raw_value or "").strip()
        if not stamp:
            return None
        try:
            return datetime.fromisoformat(stamp.replace("Z", "+00:00")).date()
        except Exception:
            pass
        if len(stamp) >= 10:
            try:
                return datetime.strptime(stamp[:10], "%Y-%m-%d").date()
            except Exception:
                return None
        return None

    @staticmethod
    def _parse_iso_datetime(raw_value: str) -> datetime | None:
        stamp = str(raw_value or "").strip()
        if not stamp:
            return None
        try:
            parsed = datetime.fromisoformat(stamp.replace("Z", "+00:00"))
            if parsed.tzinfo is not None:
                parsed = parsed.astimezone().replace(tzinfo=None)
            return parsed
        except Exception:
            pass
        candidate = stamp.replace("T", " ")
        for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(candidate[: len(datetime.now().strftime(fmt))], fmt)
            except Exception:
                continue
        return None

    @classmethod
    def _is_working_flag_stale(cls, raw_stamp: str) -> bool:
        updated_at = cls._parse_iso_datetime(raw_stamp)
        if updated_at is None:
            return True
        return bool((datetime.now() - updated_at) >= timedelta(hours=int(max(1, cls.FLAG_WORK_REVERT_HOURS))))

    def _flag_alert_counts_from_rows(self, rows: list[sqlite3.Row]) -> tuple[int, int]:
        urgent_count = 0
        in_progress_count = 0
        for row in rows:
            flag = str(row["qa_flag"] or "").strip()
            image_path = str(row["qa_flag_image_path"] or "").strip()
            if not flag and not image_path:
                continue
            working_user = DepotRules.normalize_user_id(str(row["working_user_id"] or ""))
            working_stamp = str(row["working_updated_at"] or "").strip()
            if working_user and not self._is_working_flag_stale(working_stamp):
                in_progress_count += 1
            else:
                urgent_count += 1
        return urgent_count, in_progress_count

    def _refresh_flag_alert_watchdog(self) -> None:
        if not self.isVisible():
            return
        try:
            part_rows = self.tracker.db.fetchall(
                "SELECT COALESCE(qa_flag, '') AS qa_flag, COALESCE(qa_flag_image_path, '') AS qa_flag_image_path, "
                "COALESCE(working_user_id, '') AS working_user_id, COALESCE(working_updated_at, '') AS working_updated_at "
                "FROM parts WHERE assigned_user_id=? AND is_active=1",
                (self.current_user,),
            )
            part_urgent, part_in_progress = self._flag_alert_counts_from_rows(part_rows)
            self._parts_has_urgent_flagged_rows = bool(part_urgent > 0)
            self._parts_has_in_progress_flagged_rows = bool(part_in_progress > 0)
            self._parts_has_flagged_rows = bool((part_urgent + part_in_progress) > 0)

            if self._is_tech3_user:
                cat_rows = self.tracker.db.fetchall(
                    "SELECT COALESCE(qa_flag, '') AS qa_flag, COALESCE(qa_flag_image_path, '') AS qa_flag_image_path, "
                    "COALESCE(working_user_id, '') AS working_user_id, COALESCE(working_updated_at, '') AS working_updated_at "
                    "FROM parts WHERE is_active=1"
                )
                cat_urgent, cat_in_progress = self._flag_alert_counts_from_rows(cat_rows)
                self._cat_parts_has_urgent_flagged_rows = bool(cat_urgent > 0)
                self._cat_parts_has_in_progress_flagged_rows = bool(cat_in_progress > 0)
                self._cat_parts_has_flagged_rows = bool((cat_urgent + cat_in_progress) > 0)
            else:
                self._cat_parts_has_urgent_flagged_rows = False
                self._cat_parts_has_in_progress_flagged_rows = False
                self._cat_parts_has_flagged_rows = False
            if self.team_client_tab is not None:
                self._refresh_team_client_followup()
            self._update_tab_alert_states()
        except Exception as exc:
            _runtime_log_event(
                "ui.agent_flag_alert_watchdog_failed",
                severity="warning",
                summary="Failed refreshing agent flag alert watchdog state.",
                exc=exc,
                context={"user_id": str(self.current_user)},
            )

    @staticmethod
    def _format_working_updated_stamp(raw_stamp: str) -> str:
        stamp = str(raw_stamp or "").strip()
        if not stamp:
            return ""
        try:
            parsed = datetime.fromisoformat(stamp.replace("Z", "+00:00"))
            return parsed.strftime("%Y-%m-%d %I:%M %p")
        except Exception:
            return stamp.replace("T", " ")

    def _toggle_selected_part_working(self, table: QTableWidget) -> None:
        part_id = _selected_part_id_from_table(table)
        if part_id is None:
            self._show_themed_message(QMessageBox.Icon.Warning, "Validation", "Select a row first.")
            return
        row = self.tracker.db.fetchone(
            "SELECT COALESCE(working_user_id, '') AS working_user_id "
            "FROM parts WHERE id=?",
            (int(part_id),),
        )
        if row is None:
            self._show_themed_message(QMessageBox.Icon.Warning, "Missing", "Selected part no longer exists.")
            return
        working_user = DepotRules.normalize_user_id(str(row["working_user_id"] or ""))
        if working_user and working_user != self.current_user:
            self._show_themed_message(
                QMessageBox.Icon.Warning,
                "In Use",
                f"This unit is already marked as being worked by {working_user}.",
            )
            return
        next_user = "" if working_user == self.current_user else self.current_user
        try:
            self.tracker.set_part_working_user(part_id, next_user)
        except Exception as exc:
            _runtime_log_event(
                "ui.agent_set_part_working_failed",
                severity="warning",
                summary="Failed updating working-owner flag for part.",
                exc=exc,
                context={"user_id": str(self.current_user), "part_id": int(part_id), "next_user": str(next_user)},
            )
            self._show_themed_message(QMessageBox.Icon.Critical, "Save Failed", f"Could not update flag:\n{type(exc).__name__}: {exc}")
            return
        self._refresh_agent_parts()
        self._refresh_category_parts()
        if self.app_window is not None:
            qa_window = getattr(self.app_window, "active_qa_window", None)
            if qa_window is not None and qa_window.isVisible():
                qa_window._refresh_assigned_parts()
                qa_window._refresh_delivered_parts()

    def _toggle_selected_part_installed(self, table: QTableWidget) -> None:
        part_id = _selected_part_id_from_table(table)
        if part_id is None:
            self._show_themed_message(QMessageBox.Icon.Warning, "Validation", "Select a row first.")
            return
        row = self.tracker.db.fetchone(
            "SELECT COALESCE(parts_installed, 0) AS parts_installed "
            "FROM parts WHERE id=?",
            (int(part_id),),
        )
        if row is None:
            self._show_themed_message(QMessageBox.Icon.Warning, "Missing", "Selected part no longer exists.")
            return
        next_state = 0 if int(row["parts_installed"] or 0) else 1
        try:
            self.tracker.set_part_installed(part_id, bool(next_state), self.current_user)
        except Exception as exc:
            _runtime_log_event(
                "ui.agent_set_part_installed_failed",
                severity="warning",
                summary="Failed toggling installed state for part.",
                exc=exc,
                context={
                    "user_id": str(self.current_user),
                    "part_id": int(part_id),
                    "next_state": int(next_state),
                },
            )
            self._show_themed_message(
                QMessageBox.Icon.Critical,
                "Save Failed",
                f"Could not update installed state:\n{type(exc).__name__}: {exc}",
            )
            return
        self._refresh_agent_parts()
        self._refresh_category_parts()
        if self.app_window is not None:
            qa_window = getattr(self.app_window, "active_qa_window", None)
            if qa_window is not None and qa_window.isVisible():
                qa_window._refresh_assigned_parts()
                qa_window._refresh_delivered_parts()

    def _on_parts_table_cell_clicked(self, tab_key: str, table: QTableWidget, row: int, col: int) -> None:
        if row < 0 or col < 0:
            return
        hdr = table.horizontalHeaderItem(col)
        header = str(hdr.text() or "").strip() if hdr is not None else ""
        if header == "Working":
            table.selectRow(row)
            self._toggle_selected_part_working(table)
            return
        if header == "Installed":
            table.selectRow(row)
            self._toggle_selected_part_installed(table)
            return

    def eventFilter(self, watched, event) -> bool:  # noqa: N802
        return super().eventFilter(watched, event)

    def _open_agent_notes_for_table(self, table: QTableWidget) -> None:
        part_id = _selected_part_id_from_table(table)
        if part_id is None:
            self._show_themed_message(QMessageBox.Icon.Warning, "Validation", "Select a row first.")
            return

        row = self.tracker.db.fetchone(
            "SELECT id, work_order, category, client_unit, COALESCE(qa_comment, '') AS qa_comment, "
            "COALESCE(agent_comment, '') AS agent_comment, COALESCE(qa_flag, '') AS qa_flag, "
            "COALESCE(qa_flag_image_path, '') AS qa_flag_image_path, COALESCE(comments, '') AS comments, "
            "COALESCE(working_user_id, '') AS working_user_id, COALESCE(working_updated_at, '') AS working_updated_at "
            "FROM parts WHERE id=?",
            (part_id,),
        )
        if row is None:
            self._show_themed_message(QMessageBox.Icon.Warning, "Missing", "Selected part no longer exists.")
            return

        qa_comment = str(row["qa_comment"] or row["comments"] or "").strip()
        image_path = self.tracker.resolve_qa_flag_icon(
            str(row["qa_flag"] or "").strip(),
            str(row["qa_flag_image_path"] or ""),
        )
        part_data = {
            "id": int(row["id"]),
            "work_order": str(row["work_order"] or ""),
            "category": str(row["category"] or ""),
            "client_unit": bool(int(row["client_unit"] or 0)),
            "qa_comment": qa_comment,
            "agent_comment": str(row["agent_comment"] or "").strip(),
            "qa_flag": str(row["qa_flag"] or "").strip(),
            "qa_flag_image_path": image_path,
            "working_user_id": DepotRules.normalize_user_id(str(row["working_user_id"] or "")),
            "working_updated_at": str(row["working_updated_at"] or "").strip(),
        }
        dialog = PartNotesDialog(
            "agent",
            part_data,
            tracker=self.tracker,
            app_window=self.app_window,
            current_user=self.current_user,
            parent=self,
        )
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return
        values = dialog.values()
        self.tracker.update_part_agent_comment(part_id, values.get("own_note", ""))
        working_user_id = str(values.get("working_user_id", "__UNCHANGED__"))
        if working_user_id != "__UNCHANGED__":
            self.tracker.set_part_working_user(part_id, working_user_id)
        self._refresh_agent_parts()
        self._refresh_category_parts()
        if self.app_window is not None:
            qa_window = getattr(self.app_window, "active_qa_window", None)
            if qa_window is not None and qa_window.isVisible():
                qa_window._refresh_assigned_parts()
                qa_window._refresh_delivered_parts()

    def _refresh_agent_parts(self):
        search_text = ""
        if hasattr(self, "parts_workorder_search"):
            search_text = str(self.parts_workorder_search.text() or "").strip()
        query = (
            "SELECT p.id, p.created_at, p.work_order, p.category, p.client_unit, COALESCE(p.qa_comment, '') AS qa_comment, "
            "COALESCE(p.agent_comment, '') AS agent_comment, COALESCE(p.comments, '') AS comments, "
            "COALESCE(p.qa_flag, '') AS qa_flag, COALESCE(p.qa_flag_image_path, '') AS qa_flag_image_path, "
            "COALESCE(p.working_user_id, '') AS working_user_id, COALESCE(p.working_updated_at, '') AS working_updated_at, "
            "COALESCE(p.parts_installed, 0) AS parts_installed, "
            "COALESCE(p.parts_installed_by, '') AS parts_installed_by, COALESCE(p.parts_installed_at, '') AS parts_installed_at "
            "FROM parts p "
            "WHERE p.assigned_user_id=? AND p.is_active=1 "
            "AND p.id=("
            "SELECT MAX(p2.id) FROM parts p2 WHERE p2.is_active=1 AND p2.work_order=p.work_order"
            ")"
        )
        params: list[Any] = [self.current_user]
        if search_text:
            query += " AND p.work_order LIKE ?"
            params.append(f"%{search_text}%")
        query += " ORDER BY p.created_at ASC, p.id ASC LIMIT 300"
        rows = self.tracker.db.fetchall(query, tuple(params))
        self.parts_table.setRowCount(0)
        urgent_flagged_rows = 0
        in_progress_flagged_rows = 0
        for row_idx, r in enumerate(rows):
            self.parts_table.insertRow(row_idx)
            part_id = int(r["id"])
            work_order = str(r["work_order"] or "").strip()
            category = str(r["category"] or "").strip() or "Other"
            is_client = bool(int(r["client_unit"] or 0))
            age_text = self._part_age_label(str(r["created_at"] or ""))
            qa_comment = str(r["qa_comment"] or r["comments"] or "").strip()
            agent_comment = str(r["agent_comment"] or "").strip()
            flag = str(r["qa_flag"] or "").strip()
            working_user = DepotRules.normalize_user_id(str(r["working_user_id"] or ""))
            working_stamp = str(r["working_updated_at"] or "").strip()
            parts_installed = bool(int(r["parts_installed"] or 0))
            parts_installed_by = DepotRules.normalize_user_id(str(r["parts_installed_by"] or ""))
            parts_installed_at = str(r["parts_installed_at"] or "").strip()
            image_abs = self.tracker.resolve_qa_flag_icon(
                str(r["qa_flag"] or "").strip(),
                str(r["qa_flag_image_path"] or ""),
            )
            has_flag = bool(str(flag).strip() or str(image_abs).strip())
            if has_flag:
                if working_user and not self._is_working_flag_stale(working_stamp):
                    in_progress_flagged_rows += 1
                else:
                    urgent_flagged_rows += 1

            client_item = QTableWidgetItem("")
            client_item.setData(Qt.ItemDataRole.UserRole, part_id)
            if is_client:
                client_item.setIcon(self._client_checked_icon())
                client_item.setToolTip("Client unit")
            else:
                client_item.setToolTip("Non-client unit")
            self._center_item(client_item)
            flag_item = QTableWidgetItem("" if image_abs else (flag if flag else ""))
            flag_item.setData(Qt.ItemDataRole.UserRole, part_id)
            flag_item.setToolTip(self._flag_tooltip(flag, qa_comment, agent_comment, bool(image_abs)))
            if image_abs:
                flag_item.setIcon(QIcon(image_abs))
            self._center_item(flag_item)
            working_item = QTableWidgetItem("\U0001F527" if working_user else "")
            working_item.setData(Qt.ItemDataRole.UserRole, part_id)
            if working_user:
                working_tip = f"Agent working this unit: {working_user}"
                friendly_stamp = self._format_working_updated_stamp(working_stamp)
                if friendly_stamp:
                    working_tip += f"\nUpdated: {friendly_stamp}"
                working_item.setToolTip(working_tip)
            else:
                working_item.setToolTip("No agent is marked as working this unit.")
            self._center_item(working_item)
            installed_item = QTableWidgetItem("")
            installed_item.setData(Qt.ItemDataRole.UserRole, part_id)
            if parts_installed:
                installed_item.setIcon(self._part_installed_icon())
                installed_tip = "Parts installed."
                if parts_installed_by:
                    installed_tip += f"\nBy: {parts_installed_by}"
                friendly_installed = self._format_working_updated_stamp(parts_installed_at)
                if friendly_installed:
                    installed_tip += f"\nAt: {friendly_installed}"
                installed_item.setToolTip(installed_tip)
            else:
                installed_item.setToolTip("Click to mark parts installed.")
            self._center_item(installed_item)
            qa_note_item = QTableWidgetItem(self._note_preview(qa_comment))
            qa_note_item.setToolTip(f"QA Note: {qa_comment if qa_comment else '(none)'}")
            self._center_item(qa_note_item)
            age_item = self._center_item(QTableWidgetItem(age_text))
            work_item = self._center_item(QTableWidgetItem(work_order))
            work_item.setData(Qt.ItemDataRole.UserRole, part_id)
            category_item = self._center_item(QTableWidgetItem(category))
            self.parts_table.setItem(row_idx, 0, work_item)
            self.parts_table.setItem(row_idx, 1, client_item)
            self.parts_table.setItem(row_idx, 2, flag_item)
            self.parts_table.setItem(row_idx, 3, age_item)
            self.parts_table.setItem(row_idx, 4, working_item)
            self.parts_table.setItem(row_idx, 5, installed_item)
            self.parts_table.setItem(row_idx, 6, category_item)
            self.parts_table.setItem(row_idx, 7, qa_note_item)
        self._parts_has_flagged_rows = bool((urgent_flagged_rows + in_progress_flagged_rows) > 0)
        self._parts_has_urgent_flagged_rows = bool(urgent_flagged_rows > 0)
        self._parts_has_in_progress_flagged_rows = bool(in_progress_flagged_rows > 0)
        self._update_tab_alert_states()

    def _build_cat_parts_tab(self):
        layout = QVBoxLayout(self.cat_parts_tab)
        filter_layout = QHBoxLayout()
        self.cat_filter = QComboBox()
        self._refresh_category_filter_options()
        self.cat_filter.currentTextChanged.connect(self._refresh_category_parts)
        filter_layout.addWidget(QLabel("Category:"))
        filter_layout.addWidget(self.cat_filter, 1)
        self.cat_workorder_search = QLineEdit()
        self.cat_workorder_search.setPlaceholderText("Search work order...")
        self.cat_workorder_search.setClearButtonEnabled(True)
        self.cat_workorder_search.textChanged.connect(self._refresh_category_parts)
        filter_layout.addWidget(QLabel("Work Order:"))
        filter_layout.addWidget(self.cat_workorder_search, 1)
        self.cat_refresh_btn = QPushButton("Refresh")
        self.cat_refresh_btn.clicked.connect(self._refresh_category_parts)
        filter_layout.addWidget(self.cat_refresh_btn, 0)
        self.cat_open_notes_btn = QPushButton("Open Notes")
        self.cat_open_notes_btn.setProperty("actionRole", "pick")
        self.cat_open_notes_btn.clicked.connect(lambda: self._open_agent_notes_for_table(self.cat_parts_table))
        filter_layout.addWidget(self.cat_open_notes_btn, 0)
        self.cat_working_btn = QPushButton("Agent Is Working This")
        self.cat_working_btn.setProperty("actionRole", "apply")
        self.cat_working_btn.clicked.connect(lambda: self._toggle_selected_part_working(self.cat_parts_table))
        filter_layout.addWidget(self.cat_working_btn, 0)
        self.cat_installed_btn = QPushButton("Parts Installed")
        self.cat_installed_btn.setProperty("actionRole", "apply")
        self.cat_installed_btn.clicked.connect(lambda: self._toggle_selected_part_installed(self.cat_parts_table))
        filter_layout.addWidget(self.cat_installed_btn, 0)
        layout.addLayout(filter_layout)

        headers = ["Work Order", "Client", "Flag", "Age", "Working", "Installed", "Category", "QA Note"]
        resize_modes: dict[int, QHeaderView.ResizeMode] = {
            0: QHeaderView.ResizeMode.ResizeToContents,
            1: QHeaderView.ResizeMode.ResizeToContents,
            2: QHeaderView.ResizeMode.ResizeToContents,
            3: QHeaderView.ResizeMode.ResizeToContents,
            4: QHeaderView.ResizeMode.ResizeToContents,
            5: QHeaderView.ResizeMode.ResizeToContents,
            6: QHeaderView.ResizeMode.ResizeToContents,
            7: QHeaderView.ResizeMode.Stretch,
        }
        if self._is_tech3_user:
            headers = ["Work Order", "Client", "Flag", "Age", "Working", "Installed", "A#", "Category", "QA Note"]
            resize_modes = {
                0: QHeaderView.ResizeMode.ResizeToContents,
                1: QHeaderView.ResizeMode.ResizeToContents,
                2: QHeaderView.ResizeMode.ResizeToContents,
                3: QHeaderView.ResizeMode.ResizeToContents,
                4: QHeaderView.ResizeMode.ResizeToContents,
                5: QHeaderView.ResizeMode.ResizeToContents,
                6: QHeaderView.ResizeMode.ResizeToContents,
                7: QHeaderView.ResizeMode.ResizeToContents,
                8: QHeaderView.ResizeMode.Stretch,
            }

        self.cat_parts_table = QTableWidget()
        configure_standard_table(
            self.cat_parts_table,
            headers,
            resize_modes=resize_modes,
            stretch_last=True,
        )
        self.cat_parts_table.horizontalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        self.cat_parts_table.cellClicked.connect(
            lambda row, col: self._on_parts_table_cell_clicked("cat_parts", self.cat_parts_table, row, col)
        )
        self.cat_parts_table.itemDoubleClicked.connect(self._copy_work_order_from_table_item)
        self.cat_parts_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.cat_parts_table.customContextMenuRequested.connect(
            lambda pos: self._open_agent_notes_from_context(self.cat_parts_table, pos)
        )
        layout.addWidget(self.cat_parts_table)

        self._refresh_category_parts()

    def _refresh_category_filter_options(self) -> None:
        if not hasattr(self, "cat_filter"):
            return
        previous = self.cat_filter.currentText().strip() or "All"
        categories: list[str] = ["All", "Appliance", "Audio", "PC", "TV", "Other"]
        try:
            rows = self.tracker.db.fetchall(
                "SELECT DISTINCT category FROM parts WHERE is_active=1 AND category IS NOT NULL AND TRIM(category) <> '' ORDER BY category ASC"
            )
            for row in rows:
                value = str(row["category"] or "").strip()
                if value and value not in categories:
                    categories.append(value)
        except Exception as exc:
            _runtime_log_event(
                "ui.agent_category_filter_query_failed",
                severity="warning",
                summary="Failed loading category filter options; continuing with defaults.",
                exc=exc,
                context={"user_id": str(self.current_user)},
            )

        self.cat_filter.blockSignals(True)
        self.cat_filter.clear()
        self.cat_filter.addItems(categories)
        if previous in categories:
            self.cat_filter.setCurrentText(previous)
        else:
            self.cat_filter.setCurrentIndex(0)
        self.cat_filter.blockSignals(False)

    def _refresh_category_parts(self):
        self._refresh_category_filter_options()
        agent_name_lookup: dict[str, str] = {}
        if self._is_tech3_user:
            try:
                for agent_row in self.tracker.list_agents():
                    agent_user = DepotRules.normalize_user_id(str(agent_row.get("user_id", "") or ""))
                    if not agent_user:
                        continue
                    agent_name_lookup[agent_user] = str(agent_row.get("agent_name", "") or "").strip()
            except Exception as exc:
                _runtime_log_event(
                    "ui.agent_category_parts_agent_lookup_failed",
                    severity="warning",
                    summary="Failed resolving agent names for Tech 3 category parts view.",
                    exc=exc,
                    context={"user_id": str(self.current_user)},
                )
        cat = self.cat_filter.currentText().strip()
        search_text = ""
        if hasattr(self, "cat_workorder_search"):
            search_text = str(self.cat_workorder_search.text() or "").strip()
        if cat and cat != "All":
            query = (
                "SELECT p.id, p.created_at, p.work_order, COALESCE(p.assigned_user_id, '') AS assigned_user_id, "
                "p.category, p.client_unit, COALESCE(p.qa_comment, '') AS qa_comment, "
                "COALESCE(p.agent_comment, '') AS agent_comment, COALESCE(p.comments, '') AS comments, "
                "COALESCE(p.qa_flag, '') AS qa_flag, COALESCE(p.qa_flag_image_path, '') AS qa_flag_image_path, "
                "COALESCE(p.working_user_id, '') AS working_user_id, COALESCE(p.working_updated_at, '') AS working_updated_at, "
                "COALESCE(p.parts_installed, 0) AS parts_installed, "
                "COALESCE(p.parts_installed_by, '') AS parts_installed_by, COALESCE(p.parts_installed_at, '') AS parts_installed_at "
                "FROM parts p "
                "WHERE p.category=? AND p.is_active=1 "
                "AND p.id=("
                "SELECT MAX(p2.id) FROM parts p2 WHERE p2.is_active=1 AND p2.work_order=p.work_order"
                ")"
            )
            params: list[Any] = [cat]
            if search_text:
                query += " AND p.work_order LIKE ?"
                params.append(f"%{search_text}%")
            query += (
                " ORDER BY p.client_unit DESC, CASE WHEN TRIM(COALESCE(p.qa_flag, '')) <> '' THEN 1 ELSE 0 END DESC, "
                "p.created_at ASC, p.id ASC LIMIT 300"
            )
            rows = self.tracker.db.fetchall(query, tuple(params))
        else:
            query = (
                "SELECT p.id, p.created_at, p.work_order, COALESCE(p.assigned_user_id, '') AS assigned_user_id, "
                "p.category, p.client_unit, COALESCE(p.qa_comment, '') AS qa_comment, "
                "COALESCE(p.agent_comment, '') AS agent_comment, COALESCE(p.comments, '') AS comments, "
                "COALESCE(p.qa_flag, '') AS qa_flag, COALESCE(p.qa_flag_image_path, '') AS qa_flag_image_path, "
                "COALESCE(p.working_user_id, '') AS working_user_id, COALESCE(p.working_updated_at, '') AS working_updated_at, "
                "COALESCE(p.parts_installed, 0) AS parts_installed, "
                "COALESCE(p.parts_installed_by, '') AS parts_installed_by, COALESCE(p.parts_installed_at, '') AS parts_installed_at "
                "FROM parts p "
                "WHERE p.is_active=1 "
                "AND p.id=("
                "SELECT MAX(p2.id) FROM parts p2 WHERE p2.is_active=1 AND p2.work_order=p.work_order"
                ")"
            )
            params = []
            if search_text:
                query += " AND p.work_order LIKE ?"
                params.append(f"%{search_text}%")
            query += (
                " ORDER BY p.client_unit DESC, CASE WHEN TRIM(COALESCE(p.qa_flag, '')) <> '' THEN 1 ELSE 0 END DESC, "
                "p.created_at ASC, p.id ASC LIMIT 300"
            )
            rows = self.tracker.db.fetchall(query, tuple(params))
        self.cat_parts_table.setRowCount(0)
        urgent_flagged_rows = 0
        in_progress_flagged_rows = 0
        for row_idx, r in enumerate(rows):
            self.cat_parts_table.insertRow(row_idx)
            part_id = int(r["id"])
            work_order = str(r["work_order"] or "").strip()
            assigned_user = DepotRules.normalize_user_id(str(r["assigned_user_id"] or ""))
            category = str(r["category"] or "").strip() or "Other"
            is_client = bool(int(r["client_unit"] or 0))
            age_text = self._part_age_label(str(r["created_at"] or ""))
            qa_comment = str(r["qa_comment"] or r["comments"] or "").strip()
            agent_comment = str(r["agent_comment"] or "").strip()
            flag = str(r["qa_flag"] or "").strip()
            working_user = DepotRules.normalize_user_id(str(r["working_user_id"] or ""))
            working_stamp = str(r["working_updated_at"] or "").strip()
            parts_installed = bool(int(r["parts_installed"] or 0))
            parts_installed_by = DepotRules.normalize_user_id(str(r["parts_installed_by"] or ""))
            parts_installed_at = str(r["parts_installed_at"] or "").strip()
            image_abs = self.tracker.resolve_qa_flag_icon(
                str(r["qa_flag"] or "").strip(),
                str(r["qa_flag_image_path"] or ""),
            )
            has_flag = bool(str(flag).strip() or str(image_abs).strip())
            if has_flag:
                if working_user and not self._is_working_flag_stale(working_stamp):
                    in_progress_flagged_rows += 1
                else:
                    urgent_flagged_rows += 1

            client_item = QTableWidgetItem("")
            client_item.setData(Qt.ItemDataRole.UserRole, part_id)
            if is_client:
                client_item.setIcon(self._client_checked_icon())
                client_item.setToolTip("Client unit")
            else:
                client_item.setToolTip("Non-client unit")
            self._center_item(client_item)
            flag_item = QTableWidgetItem("" if image_abs else (flag if flag else ""))
            flag_item.setData(Qt.ItemDataRole.UserRole, part_id)
            flag_item.setToolTip(self._flag_tooltip(flag, qa_comment, agent_comment, bool(image_abs)))
            if image_abs:
                flag_item.setIcon(QIcon(image_abs))
            self._center_item(flag_item)
            working_item = QTableWidgetItem("\U0001F527" if working_user else "")
            working_item.setData(Qt.ItemDataRole.UserRole, part_id)
            if working_user:
                working_tip = f"Agent working this unit: {working_user}"
                friendly_stamp = self._format_working_updated_stamp(working_stamp)
                if friendly_stamp:
                    working_tip += f"\nUpdated: {friendly_stamp}"
                working_item.setToolTip(working_tip)
            else:
                working_item.setToolTip("No agent is marked as working this unit.")
            self._center_item(working_item)
            installed_item = QTableWidgetItem("")
            installed_item.setData(Qt.ItemDataRole.UserRole, part_id)
            if parts_installed:
                installed_item.setIcon(self._part_installed_icon())
                installed_tip = "Parts installed."
                if parts_installed_by:
                    installed_tip += f"\nBy: {parts_installed_by}"
                friendly_installed = self._format_working_updated_stamp(parts_installed_at)
                if friendly_installed:
                    installed_tip += f"\nAt: {friendly_installed}"
                installed_item.setToolTip(installed_tip)
            else:
                installed_item.setToolTip("Click to mark parts installed.")
            self._center_item(installed_item)
            qa_note_item = QTableWidgetItem(self._note_preview(qa_comment))
            qa_note_item.setToolTip(f"QA Note: {qa_comment if qa_comment else '(none)'}")
            self._center_item(qa_note_item)
            age_item = self._center_item(QTableWidgetItem(age_text))
            work_item = self._center_item(QTableWidgetItem(work_order))
            work_item.setData(Qt.ItemDataRole.UserRole, part_id)
            category_item = self._center_item(QTableWidgetItem(category))
            assigned_item = self._center_item(QTableWidgetItem(assigned_user if assigned_user else "-"))
            if assigned_user:
                assigned_name = str(agent_name_lookup.get(assigned_user, "") or "").strip()
                if assigned_name and assigned_name != assigned_user:
                    assigned_item.setToolTip(f"{assigned_name}")
                else:
                    assigned_item.setToolTip(assigned_user)
            else:
                assigned_item.setToolTip("Unassigned")
            self.cat_parts_table.setItem(row_idx, 0, work_item)
            self.cat_parts_table.setItem(row_idx, 1, client_item)
            self.cat_parts_table.setItem(row_idx, 2, flag_item)
            self.cat_parts_table.setItem(row_idx, 3, age_item)
            self.cat_parts_table.setItem(row_idx, 4, working_item)
            if self._is_tech3_user:
                self.cat_parts_table.setItem(row_idx, 5, installed_item)
                self.cat_parts_table.setItem(row_idx, 6, assigned_item)
                self.cat_parts_table.setItem(row_idx, 7, category_item)
                self.cat_parts_table.setItem(row_idx, 8, qa_note_item)
            else:
                self.cat_parts_table.setItem(row_idx, 5, installed_item)
                self.cat_parts_table.setItem(row_idx, 6, category_item)
                self.cat_parts_table.setItem(row_idx, 7, qa_note_item)
        self._cat_parts_has_flagged_rows = bool((urgent_flagged_rows + in_progress_flagged_rows) > 0)
        self._cat_parts_has_urgent_flagged_rows = bool(urgent_flagged_rows > 0)
        self._cat_parts_has_in_progress_flagged_rows = bool(in_progress_flagged_rows > 0)
        self._update_tab_alert_states()

    def _build_client_tab(self):
        layout = QVBoxLayout(self.client_tab)
        summary = QLabel("Client follow-up queue for this agent (Client + Other daily, Part Order after 21 days).")
        summary.setWordWrap(True)
        summary.setProperty("muted", True)
        layout.addWidget(summary)

        controls = QHBoxLayout()
        controls.addWidget(QLabel("Follow-up Queue"), 0)
        controls.addStretch(1)
        self.client_refresh_btn = QPushButton("Refresh")
        self.client_refresh_btn.clicked.connect(self._refresh_client_followup)
        controls.addWidget(self.client_refresh_btn, 0)
        layout.addLayout(controls)

        self.client_due_summary = QLabel("No follow-up alerts.")
        self.client_due_summary.setProperty("section", True)
        layout.addWidget(self.client_due_summary)

        self.client_followup_table = QTableWidget()
        configure_standard_table(
            self.client_followup_table,
            ["Due", "Work Order", "Status", "Last Update", "Age", "Notes"],
            resize_modes={
                0: QHeaderView.ResizeMode.ResizeToContents,
                1: QHeaderView.ResizeMode.ResizeToContents,
                2: QHeaderView.ResizeMode.ResizeToContents,
                3: QHeaderView.ResizeMode.ResizeToContents,
                4: QHeaderView.ResizeMode.ResizeToContents,
                5: QHeaderView.ResizeMode.Stretch,
            },
            stretch_last=True,
        )
        self.client_followup_table.horizontalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        self.client_followup_table.cellClicked.connect(self._on_client_followup_cell_clicked)
        self.client_followup_table.itemDoubleClicked.connect(self._copy_work_order_from_table_item)
        layout.addWidget(self.client_followup_table, 1)

        self._client_due_items: list[QTableWidgetItem] = []
        self._client_due_active_ids = set()
        self._client_due_flash_on = True
        self._client_due_flash_timer = QTimer(self)
        self._client_due_flash_timer.setInterval(700)
        self._client_due_flash_timer.timeout.connect(self._on_client_due_flash_tick)
        self._client_due_flash_timer.start()
        self._refresh_client_followup()

    def _build_team_client_tab(self) -> None:
        if self.team_client_tab is None:
            return
        layout = QVBoxLayout(self.team_client_tab)
        summary = QLabel("Tech 3 view: team client follow-up items due now.")
        summary.setWordWrap(True)
        summary.setProperty("muted", True)
        layout.addWidget(summary)

        controls = QHBoxLayout()
        controls.addWidget(QLabel("Team Follow-up Queue"), 0)
        controls.addStretch(1)
        self.team_client_refresh_btn = QPushButton("Refresh")
        self.team_client_refresh_btn.clicked.connect(self._refresh_team_client_followup)
        controls.addWidget(self.team_client_refresh_btn, 0)
        layout.addLayout(controls)

        self.team_client_due_summary = QLabel("No team follow-up alerts.")
        self.team_client_due_summary.setProperty("section", True)
        layout.addWidget(self.team_client_due_summary)

        self.team_client_followup_table = QTableWidget()
        configure_standard_table(
            self.team_client_followup_table,
            ["Agent", "Due", "Work Order", "Status", "Last Update", "Age", "Notes"],
            resize_modes={
                0: QHeaderView.ResizeMode.ResizeToContents,
                1: QHeaderView.ResizeMode.ResizeToContents,
                2: QHeaderView.ResizeMode.ResizeToContents,
                3: QHeaderView.ResizeMode.ResizeToContents,
                4: QHeaderView.ResizeMode.ResizeToContents,
                5: QHeaderView.ResizeMode.ResizeToContents,
                6: QHeaderView.ResizeMode.Stretch,
            },
            stretch_last=True,
        )
        self.team_client_followup_table.horizontalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        self.team_client_followup_table.itemDoubleClicked.connect(self._copy_work_order_from_table_item)
        layout.addWidget(self.team_client_followup_table, 1)
        self._refresh_team_client_followup()

    def _refresh_team_client_followup(self) -> None:
        if self.team_client_tab is None or not hasattr(self, "team_client_followup_table"):
            return
        query = (
            "SELECT cp.id, COALESCE(cp.user_id, '') AS user_id, cp.work_order, cp.created_at, "
            "COALESCE(cp.comments, '') AS comments, "
            "COALESCE(cp.followup_last_action, '') AS followup_last_action, "
            "COALESCE(cp.followup_last_action_at, '') AS followup_last_action_at, "
            "COALESCE(cp.followup_last_actor, '') AS followup_last_actor, "
            "COALESCE(cp.followup_no_contact_count, 0) AS followup_no_contact_count, "
            "COALESCE(("
            "SELECT s.touch FROM submissions s "
            "WHERE s.user_id=cp.user_id AND s.work_order=cp.work_order "
            "AND s.client_unit=1 AND s.touch IN ('Part Order', 'Other') "
            "ORDER BY s.created_at DESC, s.id DESC LIMIT 1"
            "), '') AS latest_touch, "
            "COALESCE(("
            "SELECT s.entry_date FROM submissions s "
            "WHERE s.user_id=cp.user_id AND s.work_order=cp.work_order "
            "AND s.client_unit=1 AND s.touch IN ('Part Order', 'Other') "
            "ORDER BY s.created_at DESC, s.id DESC LIMIT 1"
            "), '') AS latest_touch_date, "
            "COALESCE(("
            "SELECT MAX(s.entry_date) FROM submissions s "
            "WHERE s.user_id=cp.user_id AND s.work_order=cp.work_order "
            "AND s.client_unit=1 AND s.touch='Part Order'"
            "), '') AS last_part_order_date "
            "FROM client_parts cp "
            "ORDER BY cp.created_at DESC, cp.id DESC LIMIT 600"
        )
        try:
            rows = self.tracker.db.fetchall(query)
        except Exception as exc:
            _runtime_log_event(
                "ui.agent_team_client_followup_query_failed",
                severity="warning",
                summary="Failed loading team client follow-up rows.",
                exc=exc,
                context={"user_id": str(self.current_user)},
            )
            self.team_client_followup_table.setRowCount(0)
            self.team_client_due_summary.setText("Team follow-up unavailable. Details were logged.")
            self._team_client_due_count = 0
            self._update_tab_alert_states()
            return

        today = datetime.now().date()
        today_iso = today.isoformat()
        self.team_client_followup_table.setRowCount(0)
        kept_row = 0
        due_total = 0
        due_active_total = 0
        for r in rows:
            latest_touch = str(r["latest_touch"] or "").strip()
            if latest_touch not in {DepotRules.TOUCH_PART_ORDER, DepotRules.TOUCH_OTHER}:
                continue
            work_order = str(r["work_order"] or "").strip()
            if not work_order:
                continue
            user_id = DepotRules.normalize_user_id(str(r["user_id"] or ""))
            last_action = DepotRules.normalize_followup_action(str(r["followup_last_action"] or ""))
            last_action_at = str(r["followup_last_action_at"] or "").strip()
            last_action_actor = DepotRules.normalize_user_id(str(r["followup_last_actor"] or ""))
            no_contact_count = int(max(0, safe_int(r["followup_no_contact_count"], 0)))
            action_date = self._parse_iso_date(last_action_at)
            days_since_action = max(0, (today - action_date).days) if action_date is not None else -1
            last_update = str(r["latest_touch_date"] or "").strip()
            if not last_update:
                last_update = str(r["created_at"] or "")[:10]
            age_days = -1
            if len(last_update) >= 10:
                try:
                    age_days = max(0, (today - datetime.strptime(last_update[:10], "%Y-%m-%d").date()).days)
                except Exception:
                    age_days = -1
            age_text = f"{age_days}d" if age_days >= 0 else "-"
            notes = str(r["comments"] or "").strip()

            due = False
            due_reason = ""
            if latest_touch == DepotRules.TOUCH_OTHER:
                due = bool(last_update and last_update < today_iso)
                if due:
                    due_reason = "Daily follow-up required for Client + Other."
            elif latest_touch == DepotRules.TOUCH_PART_ORDER:
                part_order_date = str(r["last_part_order_date"] or "").strip() or last_update
                if len(part_order_date) >= 10:
                    try:
                        po_days = max(0, (today - datetime.strptime(part_order_date[:10], "%Y-%m-%d").date()).days)
                    except Exception:
                        po_days = 0
                    if po_days >= 21:
                        due = True
                        due_reason = f"Part Order follow-up due ({po_days} days)."
            if not due:
                continue
            has_action = bool(last_action)
            due_active = bool(due and not has_action)

            self.team_client_followup_table.insertRow(kept_row)
            agent_item = self._center_item(QTableWidgetItem(user_id if user_id else "-"))
            due_item = self._center_item(QTableWidgetItem(""))
            if due_active:
                due_item.setText(DepotRules.followup_stage_label(0))
                due_item.setIcon(self._followup_clock_icon("#21B46D"))
            elif has_action:
                if last_action == DepotRules.CLIENT_FOLLOWUP_WORK_APPROVED:
                    due_item.setText("")
                    due_item.setIcon(self._followup_done_icon())
                else:
                    clock_icon, stage_text = self._followup_wait_icon_by_days(days_since_action)
                    due_item.setText(stage_text)
                    due_item.setIcon(clock_icon)
            action_line = f"\nLast follow-up: {last_action}" if last_action else ""
            actor_line = f" by {last_action_actor}" if last_action_actor else ""
            when_line = f" ({self._format_working_updated_stamp(last_action_at)})" if last_action_at else ""
            attempt_line = f"\nNo-contact follow-ups: {no_contact_count}" if no_contact_count > 0 else ""
            due_item.setToolTip(f"{due_reason}{action_line}{actor_line}{when_line}{attempt_line}")

            work_item = self._center_item(QTableWidgetItem(work_order))
            status_item = self._center_item(QTableWidgetItem(latest_touch))
            update_item = self._center_item(QTableWidgetItem(last_update if last_update else "-"))
            age_item = self._center_item(QTableWidgetItem(age_text))
            note_item = self._center_item(QTableWidgetItem(self._note_preview(notes)))
            note_item.setToolTip(f"Comments: {notes if notes else '(none)'}")

            self.team_client_followup_table.setItem(kept_row, 0, agent_item)
            self.team_client_followup_table.setItem(kept_row, 1, due_item)
            self.team_client_followup_table.setItem(kept_row, 2, work_item)
            self.team_client_followup_table.setItem(kept_row, 3, status_item)
            self.team_client_followup_table.setItem(kept_row, 4, update_item)
            self.team_client_followup_table.setItem(kept_row, 5, age_item)
            self.team_client_followup_table.setItem(kept_row, 6, note_item)

            due_total += 1
            if due_active:
                due_active_total += 1
            kept_row += 1

        self._team_client_due_count = int(due_active_total)
        if due_total > 0:
            self.team_client_due_summary.setText(f"Team follow-up rows: {due_total} | Due now: {due_active_total}")
        else:
            self.team_client_due_summary.setText("No team follow-up alerts.")
        self._update_tab_alert_states()

    def _refresh_client_followup(self):
        if not hasattr(self, "client_followup_table"):
            return
        query = (
            "SELECT cp.id, cp.work_order, cp.created_at, COALESCE(cp.comments, '') AS comments, "
            "COALESCE(cp.followup_last_action, '') AS followup_last_action, "
            "COALESCE(cp.followup_last_action_at, '') AS followup_last_action_at, "
            "COALESCE(cp.followup_last_actor, '') AS followup_last_actor, "
            "COALESCE(cp.followup_no_contact_count, 0) AS followup_no_contact_count, "
            "COALESCE(cp.followup_stage_logged, -1) AS followup_stage_logged, "
            "COALESCE(("
            "SELECT s.touch FROM submissions s "
            "WHERE s.user_id=cp.user_id AND s.work_order=cp.work_order "
            "AND s.client_unit=1 AND s.touch IN ('Part Order', 'Other') "
            "ORDER BY s.created_at DESC, s.id DESC LIMIT 1"
            "), '') AS latest_touch, "
            "COALESCE(("
            "SELECT s.entry_date FROM submissions s "
            "WHERE s.user_id=cp.user_id AND s.work_order=cp.work_order "
            "AND s.client_unit=1 AND s.touch IN ('Part Order', 'Other') "
            "ORDER BY s.created_at DESC, s.id DESC LIMIT 1"
            "), '') AS latest_touch_date, "
            "COALESCE(("
            "SELECT MAX(s.entry_date) FROM submissions s "
            "WHERE s.user_id=cp.user_id AND s.work_order=cp.work_order "
            "AND s.client_unit=1 AND s.touch='Part Order'"
            "), '') AS last_part_order_date "
            "FROM client_parts cp "
            "WHERE cp.user_id=? "
            "ORDER BY cp.created_at DESC, cp.id DESC LIMIT 300"
        )
        try:
            rows = self.tracker.db.fetchall(query, (self.current_user,))
        except Exception as exc:
            _runtime_log_event(
                "ui.agent_client_followup_query_failed",
                severity="warning",
                summary="Failed loading agent client follow-up rows.",
                exc=exc,
                context={"user_id": str(self.current_user)},
            )
            self.client_followup_table.setRowCount(0)
            self.client_due_summary.setText("Client follow-up unavailable. Details were logged.")
            self._client_due_items = []
            self._client_due_active_ids = set()
            self._client_due_ack_ids = set()
            self._update_tab_alert_states()
            return

        today = datetime.now().date()
        today_iso = today.isoformat()
        self.client_followup_table.setRowCount(0)
        previous_active_ids = set(getattr(self, "_client_due_active_ids", set()))
        due_items: list[QTableWidgetItem] = []
        due_count = 0
        due_active_ids: set[int] = set()
        kept_row = 0
        for r in rows:
            client_part_id = int(r["id"])
            latest_touch = str(r["latest_touch"] or "").strip()
            if latest_touch not in {DepotRules.TOUCH_PART_ORDER, DepotRules.TOUCH_OTHER}:
                continue
            work_order = str(r["work_order"] or "").strip()
            if not work_order:
                continue
            last_action = DepotRules.normalize_followup_action(str(r["followup_last_action"] or ""))
            last_action_at = str(r["followup_last_action_at"] or "").strip()
            last_action_actor = DepotRules.normalize_user_id(str(r["followup_last_actor"] or ""))
            no_contact_count = int(max(0, safe_int(r["followup_no_contact_count"], 0)))
            stage_logged = int(safe_int(r["followup_stage_logged"], -1))
            action_date = self._parse_iso_date(last_action_at)
            days_since_action = (
                max(0, (today - action_date).days)
                if action_date is not None
                else -1
            )
            last_update = str(r["latest_touch_date"] or "").strip()
            if not last_update:
                last_update = str(r["created_at"] or "")[:10]
            age_days = -1
            if len(last_update) >= 10:
                try:
                    age_days = max(0, (today - datetime.strptime(last_update[:10], "%Y-%m-%d").date()).days)
                except Exception:
                    age_days = -1
            age_text = f"{age_days}d" if age_days >= 0 else "-"
            notes = str(r["comments"] or "").strip()

            due = False
            due_reason = ""
            if latest_touch == DepotRules.TOUCH_OTHER:
                # "Other" requires daily follow-up after the submission day.
                due = bool(last_update and last_update < today_iso)
                if due:
                    due_reason = "Daily follow-up required for Client + Other."
            elif latest_touch == DepotRules.TOUCH_PART_ORDER:
                part_order_date = str(r["last_part_order_date"] or "").strip() or last_update
                if len(part_order_date) >= 10:
                    try:
                        po_days = max(0, (today - datetime.strptime(part_order_date[:10], "%Y-%m-%d").date()).days)
                    except Exception:
                        po_days = 0
                    if po_days >= 21:
                        due = True
                        due_reason = f"Part Order follow-up due ({po_days} days)."

            has_action = bool(last_action)
            due_active = bool(due and not has_action)

            self.client_followup_table.insertRow(kept_row)
            due_item = self._center_item(QTableWidgetItem(""))
            due_item.setData(Qt.ItemDataRole.UserRole, client_part_id)
            due_item.setData(Qt.ItemDataRole.UserRole + 1, 1 if due else 0)
            due_item.setData(Qt.ItemDataRole.UserRole + 2, 1 if has_action else 0)
            if due_active:
                due_item.setText(DepotRules.followup_stage_label(0))
                due_item.setIcon(self._followup_clock_icon("#21B46D"))
            elif due and has_action:
                if last_action == DepotRules.CLIENT_FOLLOWUP_WORK_APPROVED:
                    due_item.setText("")
                    due_item.setIcon(self._followup_done_icon())
                else:
                    stage_idx = int(clamp(int(max(0, days_since_action)), 0, 2)) if days_since_action >= 0 else 0
                    if stage_idx > stage_logged:
                        try:
                            notes = self.tracker.update_client_followup_stage(client_part_id, stage_idx)
                            stage_logged = stage_idx
                        except Exception as exc:
                            _runtime_log_event(
                                "ui.agent_client_followup_stage_update_failed",
                                severity="warning",
                                summary="Failed recording follow-up day-stage in client comments.",
                                exc=exc,
                                context={
                                    "user_id": str(self.current_user),
                                    "client_part_id": int(client_part_id),
                                    "work_order": str(work_order),
                                    "stage_index": int(stage_idx),
                                },
                            )
                    clock_icon, stage_text = self._followup_wait_icon_by_days(days_since_action)
                    due_item.setText(stage_text)
                    due_item.setIcon(clock_icon)
            if due:
                action_line = f"\nLast follow-up: {last_action}" if last_action else ""
                actor_line = f" by {last_action_actor}" if last_action_actor else ""
                when_line = f" ({self._format_working_updated_stamp(last_action_at)})" if last_action_at else ""
                attempt_line = f"\nNo-contact follow-ups: {no_contact_count}" if no_contact_count > 0 else ""
                if has_action and last_action != DepotRules.CLIENT_FOLLOWUP_WORK_APPROVED and days_since_action >= 0:
                    stage_name = DepotRules.followup_stage_label(0 if days_since_action <= 0 else (1 if days_since_action == 1 else 2))
                    wait_line = f"\nWaiting Stage: {stage_name}"
                else:
                    wait_line = ""
                if has_action and last_action:
                    due_item.setToolTip(
                        f"{due_reason}{wait_line}{action_line}{actor_line}{when_line}{attempt_line}\n"
                        "Click this cell to log another follow-up action."
                    )
                else:
                    due_item.setToolTip(
                        f"{due_reason}{action_line}{actor_line}{when_line}{attempt_line}\n"
                        "Click this cell to log follow-up action."
                    )
            else:
                due_item.setToolTip("No follow-up due.")
            work_item = self._center_item(QTableWidgetItem(work_order))
            status_item = self._center_item(QTableWidgetItem(latest_touch))
            update_item = self._center_item(QTableWidgetItem(last_update if last_update else "-"))
            age_item = self._center_item(QTableWidgetItem(age_text))
            note_item = self._center_item(QTableWidgetItem(self._note_preview(notes)))
            note_item.setToolTip(f"Comments: {notes if notes else '(none)'}")

            self.client_followup_table.setItem(kept_row, 0, due_item)
            self.client_followup_table.setItem(kept_row, 1, work_item)
            self.client_followup_table.setItem(kept_row, 2, status_item)
            self.client_followup_table.setItem(kept_row, 3, update_item)
            self.client_followup_table.setItem(kept_row, 4, age_item)
            self.client_followup_table.setItem(kept_row, 5, note_item)
            if due_active:
                due_active_ids.add(client_part_id)
                if int(client_part_id) not in self._client_due_ack_ids:
                    due_items.append(due_item)
                due_count += 1
            kept_row += 1

        self._client_due_ack_ids.intersection_update(due_active_ids)
        self._client_due_active_ids = due_active_ids
        if due_active_ids.difference(previous_active_ids):
            self._tab_alert_ack_states["client"] = False
        self._client_due_items = due_items
        self._client_due_flash_on = True
        self._apply_client_due_flash_visuals()
        self._update_tab_alert_states()
        if due_count > 0:
            self.client_due_summary.setText(f"Follow-up due now: {due_count}")
        else:
            self.client_due_summary.setText("No follow-up alerts.")

    def _on_client_followup_cell_clicked(self, row: int, col: int) -> None:
        if col != 0:
            return
        if row < 0 or row >= int(self.client_followup_table.rowCount()):
            return
        due_item = self.client_followup_table.item(row, 0)
        work_item = self.client_followup_table.item(row, 1)
        if due_item is None or work_item is None:
            return
        is_due_row = safe_int(due_item.data(Qt.ItemDataRole.UserRole + 1), 0) > 0
        has_logged_action = safe_int(due_item.data(Qt.ItemDataRole.UserRole + 2), 0) > 0
        if not is_due_row and not has_logged_action:
            return
        client_part_id = safe_int(due_item.data(Qt.ItemDataRole.UserRole), 0)
        if client_part_id <= 0:
            return
        if is_due_row and client_part_id in self._client_due_active_ids:
            self._client_due_ack_ids.add(int(client_part_id))
            due_item.setBackground(QColor(0, 0, 0, 0))
            self._client_due_items = [
                item
                for item in self._client_due_items
                if item is not None and safe_int(item.data(Qt.ItemDataRole.UserRole), 0) != int(client_part_id)
            ]
            self._apply_client_due_flash_visuals()
            self._update_tab_alert_states()
        work_order = str(work_item.text() or "").strip() or "(unknown)"
        action, ok = show_flowgrid_themed_input_item(
            self,
            self.app_window,
            "agent",
            "Client Follow-up",
            (
                f"Work Order: {work_order}\n"
                "Select follow-up outcome:"
            ),
            list(DepotRules.CLIENT_FOLLOWUP_ACTIONS),
            0,
            False,
        )
        if not ok:
            return
        action_text = DepotRules.normalize_followup_action(str(action or ""))
        if not action_text:
            return
        try:
            no_contact_count = self.tracker.mark_client_followup_action(client_part_id, action_text, self.current_user)
        except Exception as exc:
            _runtime_log_event(
                "ui.agent_client_followup_mark_failed",
                severity="warning",
                summary="Failed recording client follow-up action.",
                exc=exc,
                context={
                    "user_id": str(self.current_user),
                    "client_part_id": int(client_part_id),
                    "work_order": str(work_order),
                    "action": str(action_text),
                },
            )
            self._show_themed_message(
                QMessageBox.Icon.Warning,
                "Follow-up",
                f"Could not save follow-up update:\n{type(exc).__name__}: {exc}",
            )
            return
        self._refresh_client_followup()
        if self.team_client_tab is not None:
            self._refresh_team_client_followup()
        if action_text in DepotRules.CLIENT_FOLLOWUP_NO_CONTACT_ACTIONS and int(no_contact_count) == 3:
            self._show_themed_message(
                QMessageBox.Icon.Warning,
                "No Contact Alert",
                "Please ship unit back to store due to no contact from client.",
            )

    def _apply_client_due_flash_visuals(self) -> None:
        if not hasattr(self, "_client_due_items"):
            return
        if not self._client_due_items:
            return
        on_color = QColor("#D95A5A")
        on_color.setAlpha(105)
        off_color = QColor(0, 0, 0, 0)
        for item in list(self._client_due_items):
            if item is None:
                continue
            item.setBackground(on_color if self._client_due_flash_on else off_color)

    def _on_client_due_flash_tick(self) -> None:
        if not hasattr(self, "_client_due_items"):
            return
        if not self._client_due_items:
            self._client_due_flash_on = True
            return
        self._client_due_flash_on = not bool(getattr(self, "_client_due_flash_on", False))
        self._apply_client_due_flash_visuals()


class DepotQAWindow(DepotFramelessToolWindow):
    def __init__(self, tracker: DepotTracker, current_user: str, app_window: "QuickInputsWindow" | None = None):
        super().__init__(app_window, window_title="QA/WCS", theme_kind="qa", size=(820, 550))
        self.tracker = tracker
        self.current_user = DepotRules.normalize_user_id(current_user)
        self._always_on_top_config_key = "qa_window_always_on_top"
        self._window_always_on_top = self._load_window_always_on_top_preference(self._always_on_top_config_key, default=True)
        self.set_window_always_on_top(self._window_always_on_top)

        self.qa_tabs = QTabWidget(self)
        self.root_layout.addWidget(self.qa_tabs)

        self.submit_tab = QWidget()
        self.assigned_tab = QWidget()
        self.delivered_tab = QWidget()
        self.completed_tab = QWidget()

        self.qa_tabs.addTab(self.submit_tab, "Submit")
        self.qa_tabs.addTab(self.assigned_tab, "Assigned Parts")
        self.qa_tabs.addTab(self.delivered_tab, "Parts Delivered")
        self.qa_tabs.addTab(self.completed_tab, "Completed")

        self._build_qa_submit_tab()
        self._build_qa_assigned_tab()
        self._build_qa_delivered_tab()
        self._build_qa_completed_tab()

        self.recent_submissions_label = QLabel()
        self.recent_submissions_label.setWordWrap(True)
        self.recent_submissions_label.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        self.root_layout.addWidget(self.recent_submissions_label)
        self._refresh_recent_submissions_label()

        if self.app_window is not None:
            self.apply_theme_styles()
            self.app_window.active_qa_window = self

    def _build_qa_submit_tab(self):
        layout = QFormLayout(self.submit_tab)
        self.qa_work_order = QLineEdit()
        self.qa_assign_to = QComboBox()
        self._populate_agents()
        self.qa_category = QComboBox()
        self.qa_category.addItems(["Appliance", "Audio", "PC", "TV", "Other"])
        self.qa_client_check = QCheckBox("Client")
        self.qa_comments = QLineEdit()
        self.qa_always_on_top_check = QCheckBox("Always on top")
        self.qa_always_on_top_check.setChecked(bool(self._window_always_on_top))
        self.qa_always_on_top_check.toggled.connect(self._on_qa_always_on_top_toggled)
        self.qa_flag_combo = QComboBox()
        self._populate_flags()
        self.qa_bulk_parts_input = QTextEdit()
        self.qa_bulk_parts_input.setAcceptRichText(False)
        self.qa_bulk_parts_input.setMinimumHeight(116)
        self.qa_bulk_parts_input.setPlaceholderText(
            "Paste tab-separated rows here:\n"
            "LPN<TAB>Part#<TAB>Part Description<TAB>Shipping Info\n"
            "Only rows with shipping info containing 'delivered' are imported."
        )
        self.qa_bulk_parts_input.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.qa_bulk_parts_input.customContextMenuRequested.connect(self._show_qa_bulk_parts_context_menu)
        self.qa_bulk_import_btn = QPushButton("Import Delivered Rows")
        self.qa_bulk_import_btn.setProperty("actionRole", "pick")
        self.qa_bulk_import_btn.clicked.connect(self._submit_qa_bulk_parts)
        self.qa_bulk_import_status = QLabel("")
        self.qa_bulk_import_status.setWordWrap(True)
        self.qa_bulk_import_status.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)

        layout.addRow("Work Order", self.qa_work_order)
        layout.addRow("Assign To", self.qa_assign_to)
        layout.addRow("Category", self.qa_category)
        layout.addRow("Client", self.qa_client_check)
        layout.addRow("Flag", self.qa_flag_combo)
        layout.addRow("QA Comment", self.qa_comments)
        layout.addRow("Parts Paste", self.qa_bulk_parts_input)
        layout.addRow("", self.qa_bulk_import_btn)
        layout.addRow("", self.qa_bulk_import_status)
        layout.addRow("", self.qa_always_on_top_check)

        self.qa_work_order.returnPressed.connect(self._submit_qa_part)

    def _populate_flags(self) -> None:
        previous = self.qa_flag_combo.currentText().strip() if hasattr(self, "qa_flag_combo") else "None"
        options = self.tracker.get_qa_flag_options(include_none=True)
        self.qa_flag_combo.blockSignals(True)
        self.qa_flag_combo.clear()
        self.qa_flag_combo.addItems(options)
        idx = self.qa_flag_combo.findText(previous)
        self.qa_flag_combo.setCurrentIndex(idx if idx >= 0 else 0)
        self.qa_flag_combo.blockSignals(False)

    def _populate_agents(self):
        previous = self.qa_assign_to.currentData()
        rows = self.tracker.list_agents()

        self.qa_assign_to.blockSignals(True)
        self.qa_assign_to.clear()
        for r in rows:
            tier_label = DepotRules.agent_tier_label(r.get("tier", 1))
            display = (
                f"{r['user_id']} - {r['agent_name']} ({tier_label})"
                if r["agent_name"]
                else f"{r['user_id']} ({tier_label})"
            )
            self.qa_assign_to.addItem(display, r["user_id"])
            idx = self.qa_assign_to.count() - 1
            icon_path = str(r["icon_path"] or "").strip()
            if icon_path and Path(icon_path).exists():
                self.qa_assign_to.setItemIcon(idx, QIcon(icon_path))

        if self.qa_assign_to.count() == 0:
            self.qa_assign_to.addItem(f"{self.current_user} - Current User", self.current_user)

        if previous:
            found_idx = self.qa_assign_to.findData(previous)
            if found_idx >= 0:
                self.qa_assign_to.setCurrentIndex(found_idx)
        self.qa_assign_to.blockSignals(False)

    def _on_qa_always_on_top_toggled(self, checked: bool) -> None:
        self._window_always_on_top = self._apply_window_always_on_top_preference(self._always_on_top_config_key, checked)

    def _submit_qa_part(self):
        wo = self.qa_work_order.text().strip()
        bulk_text = str(self.qa_bulk_parts_input.toPlainText() or "").strip() if hasattr(self, "qa_bulk_parts_input") else ""
        if not wo:
            if bulk_text:
                self._submit_qa_bulk_parts()
                return
            self._show_themed_message(QMessageBox.Icon.Warning, "Validation", "Work order required.")
            return
        if bulk_text:
            # Enter on submit should include both form submission and delivered import in one action.
            self._submit_qa_bulk_parts()
            return
        assigned = self.qa_assign_to.currentData() or self.current_user
        client = self.qa_client_check.isChecked()
        comments = self.qa_comments.text().strip()
        category = self.qa_category.currentText()
        selected_flag = str(self.qa_flag_combo.currentText() if hasattr(self, "qa_flag_combo") else "").strip()
        if selected_flag.lower() == "none":
            selected_flag = ""

        self.tracker.submit_part(
            self.current_user,
            assigned,
            wo,
            category,
            client,
            comments,
            selected_flag,
        )
        self.qa_work_order.clear()
        self.qa_comments.clear()
        if hasattr(self, "qa_flag_combo"):
            self.qa_flag_combo.setCurrentIndex(0)
        if hasattr(self, "qa_bulk_import_status"):
            self.qa_bulk_import_status.clear()
        self.qa_work_order.setFocus()
        self._refresh_after_qa_submit()

    def _refresh_after_qa_submit(self) -> None:
        self._refresh_recent_submissions_label()
        self._refresh_assigned_parts()
        self._refresh_delivered_parts()
        self._refresh_completed_parts()
        if self.app_window is not None:
            agent_window = getattr(self.app_window, "active_agent_window", None)
            if agent_window is not None and agent_window.isVisible():
                try:
                    agent_window._refresh_agent_parts()
                    agent_window._refresh_category_parts()
                except Exception as exc:
                    _runtime_log_event(
                        "ui.qa_agent_window_refresh_failed",
                        severity="warning",
                        summary="QA submit succeeded but paired agent window refresh failed.",
                        exc=exc,
                        context={"user_id": str(self.current_user)},
                    )

    def _show_qa_bulk_parts_context_menu(self, pos: QPoint) -> None:
        if not hasattr(self, "qa_bulk_parts_input"):
            return
        menu = self.qa_bulk_parts_input.createStandardContextMenu()
        menu.addSeparator()
        paste_import_action = menu.addAction("Paste Clipboard and Import Delivered Rows")
        chosen = menu.exec(self.qa_bulk_parts_input.mapToGlobal(pos))
        if chosen == paste_import_action and self._append_clipboard_to_qa_bulk_parts():
            self._submit_qa_bulk_parts()
        menu.deleteLater()

    def _append_clipboard_to_qa_bulk_parts(self) -> bool:
        if not hasattr(self, "qa_bulk_parts_input"):
            return False
        clipboard = QApplication.clipboard()
        clipboard_text = str(clipboard.text() or "")
        normalized_clipboard = clipboard_text.replace("\r\n", "\n").replace("\r", "\n").strip("\n")
        if not normalized_clipboard.strip():
            self._show_themed_message(QMessageBox.Icon.Warning, "Paste", "Clipboard is empty.")
            return False
        existing_text = str(self.qa_bulk_parts_input.toPlainText() or "")
        if existing_text.strip():
            merged = f"{existing_text.rstrip()}\n{normalized_clipboard}"
        else:
            merged = normalized_clipboard
        self.qa_bulk_parts_input.setPlainText(merged)
        return True

    def _submit_qa_bulk_parts(self) -> None:
        if not hasattr(self, "qa_bulk_parts_input"):
            return
        raw_text = str(self.qa_bulk_parts_input.toPlainText() or "")
        if not raw_text.strip():
            self._show_themed_message(
                QMessageBox.Icon.Warning,
                "Bulk Import",
                "Paste tab-separated rows before importing.",
            )
            return

        form_work_order = DepotRules.normalize_work_order(str(self.qa_work_order.text() or "").strip())
        if not form_work_order:
            self._show_themed_message(
                QMessageBox.Icon.Warning,
                "Bulk Import",
                "Work order required. Set the Work Order field, then import delivered rows.",
            )
            return

        assigned = self.qa_assign_to.currentData() or self.current_user
        client = self.qa_client_check.isChecked()
        category = self.qa_category.currentText()
        selected_flag = str(self.qa_flag_combo.currentText() if hasattr(self, "qa_flag_combo") else "").strip()
        if selected_flag.lower() == "none":
            selected_flag = ""
        base_comment = str(self.qa_comments.text() or "").strip()

        parsed_rows: list[tuple[int, str, str, str, str]] = []
        duplicate_rows_in_paste = 0
        skipped_not_delivered = 0
        skipped_missing_lpn = 0
        skipped_bad_format = 0
        seen_row_signatures: set[tuple[str, str, str, str]] = set()

        for line_no, raw_line in enumerate(raw_text.splitlines(), start=1):
            line = str(raw_line or "").strip()
            if not line:
                continue
            columns = [str(col or "").strip() for col in raw_line.split("\t")]
            if len(columns) < 4:
                columns = [str(col or "").strip() for col in re.split(r"\s{2,}", line) if str(col or "").strip()]
            if len(columns) < 4:
                skipped_bad_format += 1
                continue

            lpn = DepotRules.normalize_work_order(columns[0])
            part_number = str(columns[1] or "").strip()
            part_description = str(columns[2] or "").strip()
            shipping_info = " ".join(str(col or "").strip() for col in columns[3:] if str(col or "").strip())
            shipping_info_normalized = shipping_info.strip().casefold()

            if not lpn:
                skipped_missing_lpn += 1
                continue
            if not re.match(r"^delivered\b", shipping_info_normalized):
                skipped_not_delivered += 1
                continue
            row_sig = (
                lpn,
                part_number.strip().casefold(),
                part_description.strip().casefold(),
                shipping_info_normalized,
            )
            if row_sig in seen_row_signatures:
                duplicate_rows_in_paste += 1
                continue

            seen_row_signatures.add(row_sig)
            parsed_rows.append((line_no, lpn, part_number, part_description, shipping_info))

        inserted_count = 0
        updated_existing = 0
        failed_count = 0
        existing_active = self.tracker.find_active_parts_by_work_orders([form_work_order])

        if parsed_rows:
            try:
                created_part_id = self.tracker.submit_part(
                    self.current_user,
                    assigned,
                    form_work_order,
                    category,
                    client,
                    base_comment,
                    selected_flag,
                    True,
                )
                for _line_no, lpn, part_number, part_description, shipping_info in parsed_rows:
                    self.tracker.upsert_part_detail(
                        created_part_id,
                        lpn,
                        part_number,
                        part_description,
                        shipping_info,
                        True,
                    )
                if form_work_order in existing_active:
                    updated_existing += 1
                else:
                    inserted_count += 1
            except Exception as exc:
                failed_count += 1
                _runtime_log_event(
                    "ui.qa_bulk_part_submit_failed",
                    severity="warning",
                    summary="QA bulk import failed for one row; continued with remaining rows.",
                    exc=exc,
                    context={
                        "user_id": str(self.current_user),
                        "line_no": int(parsed_rows[0][0]) if parsed_rows else -1,
                        "work_order": str(form_work_order),
                        "assigned_user_id": str(assigned),
                    },
                )

        total_lines = len([line for line in raw_text.splitlines() if str(line or "").strip()])
        status_text = (
            f"Bulk import: inserted {inserted_count}, updated {updated_existing}. "
            f"Not delivered {skipped_not_delivered}, duplicate rows in paste {duplicate_rows_in_paste}, missing LPN {skipped_missing_lpn}, "
            f"bad format {skipped_bad_format}, failed {failed_count}, lines {total_lines}."
        )
        if hasattr(self, "qa_bulk_import_status"):
            self.qa_bulk_import_status.setText(status_text)

        if (inserted_count + updated_existing) > 0:
            self.qa_work_order.clear()
            self.qa_comments.clear()
            if hasattr(self, "qa_flag_combo"):
                self.qa_flag_combo.setCurrentIndex(0)
            self.qa_work_order.setFocus()
            self._refresh_after_qa_submit()

    def _refresh_recent_submissions_label(self) -> None:
        try:
            rows = self.tracker.db.fetchall(
                "SELECT work_order, assigned_user_id, category, client_unit, created_at FROM parts WHERE user_id=? ORDER BY created_at DESC LIMIT 3",
                (self.current_user,),
            )
        except Exception as exc:
            _runtime_log_event(
                "ui.qa_recent_submissions_query_failed",
                severity="warning",
                summary="Failed querying recent submissions for QA panel; showing unavailable fallback.",
                exc=exc,
                context={"user_id": str(self.current_user)},
            )
            self.recent_submissions_label.setText("Recent submissions: unavailable")
            return

        if not rows:
            self.recent_submissions_label.setText("Latest 3 submissions:\n1. (none)\n2. (none)\n3. (none)")
            return

        lines: list[str] = ["Latest 3 submissions:"]
        for index, row in enumerate(rows, start=1):
            wo = str(row["work_order"])
            assigned = str(row["assigned_user_id"])
            category = str(row["category"])
            client_marker = " \u2713" if int(row["client_unit"] or 0) else ""
            created = str(row["created_at"] or "")
            stamp = created[11:16] if len(created) >= 16 else created
            lines.append(f"{index}. {wo} -> {assigned} ({category}{client_marker}) [{stamp}]")

        for index in range(len(rows) + 1, 4):
            lines.append(f"{index}. (none)")

        self.recent_submissions_label.setText("\n".join(lines))

    def _build_qa_assigned_tab(self):
        layout = QVBoxLayout(self.assigned_tab)
        self.qa_assigned_table = QTableWidget()
        configure_standard_table(
            self.qa_assigned_table,
            ["Work Order", "Client", "Flag", "Age", "Working", "Assigned Agent", "Category", "QA Note", "Agent Note"],
            resize_modes={
                0: QHeaderView.ResizeMode.ResizeToContents,
                1: QHeaderView.ResizeMode.ResizeToContents,
                2: QHeaderView.ResizeMode.ResizeToContents,
                3: QHeaderView.ResizeMode.ResizeToContents,
                4: QHeaderView.ResizeMode.ResizeToContents,
                5: QHeaderView.ResizeMode.ResizeToContents,
                6: QHeaderView.ResizeMode.ResizeToContents,
                7: QHeaderView.ResizeMode.ResizeToContents,
                8: QHeaderView.ResizeMode.Stretch,
            },
            stretch_last=True,
        )
        self.qa_assigned_table.itemDoubleClicked.connect(
            self._copy_work_order_from_assigned_item
        )
        self.qa_assigned_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.qa_assigned_table.customContextMenuRequested.connect(
            lambda pos: self._open_qa_notes_from_context(self.qa_assigned_table, pos)
        )
        self.qa_assigned_refresh = QPushButton("Refresh")
        self.qa_assigned_refresh.clicked.connect(self._refresh_assigned_parts)
        self.qa_assigned_open_notes_btn = QPushButton("Open Notes / Flag")
        self.qa_assigned_open_notes_btn.setProperty("actionRole", "pick")
        self.qa_assigned_open_notes_btn.clicked.connect(
            lambda: self._open_selected_qa_notes_for_table(self.qa_assigned_table)
        )
        self.qa_assigned_workorder_search = QLineEdit()
        self.qa_assigned_workorder_search.setPlaceholderText("Search work order...")
        self.qa_assigned_workorder_search.setClearButtonEnabled(True)
        self.qa_assigned_workorder_search.textChanged.connect(self._refresh_assigned_parts)
        controls = QHBoxLayout()
        controls.addWidget(QLabel("Work Order:"))
        controls.addWidget(self.qa_assigned_workorder_search, 1)
        controls.addWidget(self.qa_assigned_refresh)
        controls.addWidget(self.qa_assigned_open_notes_btn)
        layout.addLayout(controls)
        layout.addWidget(self.qa_assigned_table)
        self._refresh_assigned_parts()

    def _build_qa_delivered_tab(self):
        layout = QVBoxLayout(self.delivered_tab)
        self.qa_delivered_table = QTableWidget()
        configure_standard_table(
            self.qa_delivered_table,
            [
                "Work Order",
                "Installed",
                "Age",
                "Assigned Agent",
                "Category",
                "LPN",
                "Part #",
                "Part Description",
                "Shipping Info",
                "QA Note",
            ],
            resize_modes={
                0: QHeaderView.ResizeMode.ResizeToContents,
                1: QHeaderView.ResizeMode.ResizeToContents,
                2: QHeaderView.ResizeMode.ResizeToContents,
                3: QHeaderView.ResizeMode.ResizeToContents,
                4: QHeaderView.ResizeMode.ResizeToContents,
                5: QHeaderView.ResizeMode.ResizeToContents,
                6: QHeaderView.ResizeMode.ResizeToContents,
                7: QHeaderView.ResizeMode.ResizeToContents,
                8: QHeaderView.ResizeMode.ResizeToContents,
                9: QHeaderView.ResizeMode.Stretch,
            },
            stretch_last=True,
        )
        self.qa_delivered_table.itemDoubleClicked.connect(self._copy_work_order_from_assigned_item)
        self.qa_delivered_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.qa_delivered_table.customContextMenuRequested.connect(
            lambda pos: self._open_qa_notes_from_context(self.qa_delivered_table, pos)
        )

        self.qa_delivered_refresh = QPushButton("Refresh")
        self.qa_delivered_refresh.clicked.connect(self._refresh_delivered_parts)
        self.qa_delivered_open_notes_btn = QPushButton("Open Notes / Flag")
        self.qa_delivered_open_notes_btn.setProperty("actionRole", "pick")
        self.qa_delivered_open_notes_btn.clicked.connect(
            lambda: self._open_selected_qa_notes_for_table(self.qa_delivered_table)
        )
        self.qa_delivered_export_btn = QPushButton("Export CSV")
        self.qa_delivered_export_btn.clicked.connect(self._export_qa_delivered_csv)
        self.qa_delivered_workorder_search = QLineEdit()
        self.qa_delivered_workorder_search.setPlaceholderText("Search work order...")
        self.qa_delivered_workorder_search.setClearButtonEnabled(True)
        self.qa_delivered_workorder_search.textChanged.connect(self._refresh_delivered_parts)

        controls = QHBoxLayout()
        controls.addWidget(QLabel("Work Order:"))
        controls.addWidget(self.qa_delivered_workorder_search, 1)
        controls.addWidget(self.qa_delivered_refresh)
        controls.addWidget(self.qa_delivered_open_notes_btn)
        controls.addWidget(self.qa_delivered_export_btn)
        layout.addLayout(controls)
        layout.addWidget(self.qa_delivered_table)
        self._refresh_delivered_parts()

    def _build_qa_completed_tab(self):
        layout = QVBoxLayout(self.completed_tab)
        self.qa_completed_table = QTableWidget()
        configure_standard_table(
            self.qa_completed_table,
            ["Client", "Flag", "Age", "Working", "Work Order", "Assigned Agent", "Category", "Outcome", "Closed At", "QA Note", "Agent Note"],
            resize_modes={
                0: QHeaderView.ResizeMode.ResizeToContents,
                1: QHeaderView.ResizeMode.ResizeToContents,
                2: QHeaderView.ResizeMode.ResizeToContents,
                3: QHeaderView.ResizeMode.ResizeToContents,
                4: QHeaderView.ResizeMode.ResizeToContents,
                5: QHeaderView.ResizeMode.ResizeToContents,
                6: QHeaderView.ResizeMode.ResizeToContents,
                7: QHeaderView.ResizeMode.ResizeToContents,
                8: QHeaderView.ResizeMode.ResizeToContents,
                9: QHeaderView.ResizeMode.ResizeToContents,
                10: QHeaderView.ResizeMode.Stretch,
            },
            stretch_last=True,
        )
        self.qa_completed_table.itemDoubleClicked.connect(
            self._copy_work_order_from_assigned_item
        )
        self.qa_completed_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.qa_completed_table.customContextMenuRequested.connect(
            lambda pos: self._open_qa_notes_from_context(self.qa_completed_table, pos)
        )
        self.qa_completed_refresh = QPushButton("Refresh")
        self.qa_completed_refresh.clicked.connect(self._refresh_completed_parts)
        self.qa_completed_open_notes_btn = QPushButton("Open Notes / Flag")
        self.qa_completed_open_notes_btn.setProperty("actionRole", "pick")
        self.qa_completed_open_notes_btn.clicked.connect(
            lambda: self._open_selected_qa_notes_for_table(self.qa_completed_table)
        )
        self.qa_completed_workorder_search = QLineEdit()
        self.qa_completed_workorder_search.setPlaceholderText("Search work order...")
        self.qa_completed_workorder_search.setClearButtonEnabled(True)
        self.qa_completed_workorder_search.textChanged.connect(self._refresh_completed_parts)
        controls = QHBoxLayout()
        controls.addWidget(QLabel("Work Order:"))
        controls.addWidget(self.qa_completed_workorder_search, 1)
        controls.addWidget(self.qa_completed_refresh)
        controls.addWidget(self.qa_completed_open_notes_btn)
        layout.addLayout(controls)
        layout.addWidget(self.qa_completed_table)
        self._refresh_completed_parts()

    def _copy_work_order_from_assigned_item(self, item: QTableWidgetItem) -> None:
        table, work_order = _copy_work_order_from_table_item(item)
        if work_order:
            self._show_copy_notice(table, f"Copied Work Order: {work_order}", duration_ms=4200)

    def _open_qa_notes_from_context(self, table: QTableWidget, pos: QPoint) -> None:
        if not _select_table_row_by_context_pos(table, pos):
            return
        self._open_selected_qa_notes_for_table(table)

    def _open_selected_qa_notes(self) -> None:
        self._open_selected_qa_notes_for_table(self.qa_assigned_table)

    def _open_selected_qa_notes_for_table(self, table: QTableWidget) -> None:
        if table.currentRow() < 0:
            self._show_themed_message(QMessageBox.Icon.Warning, "Validation", "Select a row first.")
            return
        part_id = _selected_part_id_from_table(table)
        if part_id is None:
            return

        row = self.tracker.db.fetchone(
            "SELECT id, work_order, category, client_unit, COALESCE(qa_comment, '') AS qa_comment, "
            "COALESCE(agent_comment, '') AS agent_comment, COALESCE(qa_flag, '') AS qa_flag, "
            "COALESCE(qa_flag_image_path, '') AS qa_flag_image_path, COALESCE(comments, '') AS comments "
            "FROM parts WHERE id=?",
            (part_id,),
        )
        if row is None:
            return
        qa_comment = str(row["qa_comment"] or row["comments"] or "").strip()
        image_path = self.tracker.resolve_qa_flag_icon(
            str(row["qa_flag"] or "").strip(),
            str(row["qa_flag_image_path"] or ""),
        )
        part_data = {
            "id": int(row["id"]),
            "work_order": str(row["work_order"] or ""),
            "category": str(row["category"] or ""),
            "client_unit": bool(int(row["client_unit"] or 0)),
            "qa_comment": qa_comment,
            "agent_comment": str(row["agent_comment"] or "").strip(),
            "qa_flag": str(row["qa_flag"] or "").strip(),
            "qa_flag_image_path": image_path,
        }
        dialog = PartNotesDialog("qa", part_data, tracker=self.tracker, app_window=self.app_window, parent=self)
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return
        values = dialog.values()
        self.tracker.update_part_qa_fields(
            part_id,
            values.get("own_note", ""),
            values.get("qa_flag", ""),
        )
        self._refresh_assigned_parts()
        self._refresh_delivered_parts()
        self._refresh_completed_parts()
        if self.app_window is not None:
            agent_window = getattr(self.app_window, "active_agent_window", None)
            if agent_window is not None and agent_window.isVisible():
                agent_window._refresh_agent_parts()
                agent_window._refresh_category_parts()

    def _qa_agent_meta_lookup(self) -> dict[str, tuple[str, str]]:
        agent_meta: dict[str, tuple[str, str]] = {}
        try:
            for agent_row in self.tracker.list_agents():
                agent_user = DepotRules.normalize_user_id(str(agent_row.get("user_id", "") or ""))
                if not agent_user:
                    continue
                agent_meta[agent_user] = (
                    str(agent_row.get("agent_name", "") or "").strip(),
                    str(agent_row.get("icon_path", "") or "").strip(),
                )
        except Exception as exc:
            _runtime_log_event(
                "ui.qa_agent_meta_query_failed",
                severity="warning",
                summary="QA tabs could not resolve agent metadata.",
                exc=exc,
                context={"user_id": str(self.current_user)},
            )
        return agent_meta

    def _refresh_assigned_parts(self):
        search_text = ""
        if hasattr(self, "qa_assigned_workorder_search"):
            search_text = str(self.qa_assigned_workorder_search.text() or "").strip()
        query = (
            "SELECT p.id, p.created_at, p.work_order, p.assigned_user_id, p.category, p.client_unit, COALESCE(p.qa_comment, '') AS qa_comment, "
            "COALESCE(p.agent_comment, '') AS agent_comment, COALESCE(p.comments, '') AS comments, "
            "COALESCE(p.qa_flag, '') AS qa_flag, COALESCE(p.qa_flag_image_path, '') AS qa_flag_image_path, "
            "COALESCE(p.working_user_id, '') AS working_user_id, COALESCE(p.working_updated_at, '') AS working_updated_at "
            "FROM parts p WHERE p.is_active=1 "
            "AND p.id=("
            "SELECT MAX(p2.id) FROM parts p2 WHERE p2.is_active=1 AND p2.work_order=p.work_order"
            ")"
        )
        params: list[Any] = []
        if search_text:
            query += " AND p.work_order LIKE ?"
            params.append(f"%{search_text}%")
        query += " ORDER BY p.created_at ASC, p.id ASC LIMIT 300"

        rows = self.tracker.db.fetchall(query, tuple(params))
        agent_meta = self._qa_agent_meta_lookup()

        self.qa_assigned_table.setRowCount(0)
        for row_idx, r in enumerate(rows):
            self.qa_assigned_table.insertRow(row_idx)
            part_id = int(r["id"])
            work_order = str(r["work_order"] or "").strip()
            assigned = DepotRules.normalize_user_id(str(r["assigned_user_id"] or ""))
            category = str(r["category"] or "").strip() or "Other"
            age_text = DepotAgentWindow._part_age_label(str(r["created_at"] or ""))
            qa_comment = str(r["qa_comment"] or r["comments"] or "").strip()
            agent_comment = str(r["agent_comment"] or "").strip()
            flag = str(r["qa_flag"] or "").strip()
            working_user = DepotRules.normalize_user_id(str(r["working_user_id"] or ""))
            working_stamp = str(r["working_updated_at"] or "").strip()
            image_abs = self.tracker.resolve_qa_flag_icon(
                str(r["qa_flag"] or "").strip(),
                str(r["qa_flag_image_path"] or ""),
            )
            assigned_name, assigned_icon = agent_meta.get(assigned, ("", ""))

            client_item = QTableWidgetItem("")
            client_item.setData(Qt.ItemDataRole.UserRole, part_id)
            if int(r["client_unit"] or 0):
                client_item.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_DialogApplyButton))

            flag_item = QTableWidgetItem("" if image_abs else (flag if flag else ""))
            flag_item.setData(Qt.ItemDataRole.UserRole, part_id)
            flag_item.setToolTip(DepotAgentWindow._flag_tooltip(flag, qa_comment, agent_comment, bool(image_abs)))
            if image_abs:
                flag_item.setIcon(QIcon(image_abs))

            working_item = QTableWidgetItem("\U0001F527" if working_user else "")
            working_item.setData(Qt.ItemDataRole.UserRole, part_id)
            if working_user:
                working_tip = f"Agent working this unit: {working_user}"
                friendly_stamp = DepotAgentWindow._format_working_updated_stamp(working_stamp)
                if friendly_stamp:
                    working_tip += f"\nUpdated: {friendly_stamp}"
                working_item.setToolTip(working_tip)
            else:
                working_item.setToolTip("No agent is marked as working this unit.")

            assigned_text = "-"
            if assigned:
                assigned_text = f"{assigned} - {assigned_name}" if assigned_name else assigned
            assigned_item = QTableWidgetItem(assigned_text)
            if assigned_icon and Path(assigned_icon).exists():
                assigned_item.setIcon(QIcon(assigned_icon))

            qa_note_item = QTableWidgetItem(DepotAgentWindow._note_preview(qa_comment))
            qa_note_item.setToolTip(f"QA Note: {qa_comment if qa_comment else '(none)'}")
            agent_note_item = QTableWidgetItem(DepotAgentWindow._note_preview(agent_comment))
            agent_note_item.setToolTip(f"Agent Note: {agent_comment if agent_comment else '(none)'}")
            work_item = QTableWidgetItem(work_order)
            work_item.setData(Qt.ItemDataRole.UserRole, part_id)

            self.qa_assigned_table.setItem(row_idx, 0, _center_table_item(work_item))
            self.qa_assigned_table.setItem(row_idx, 1, _center_table_item(client_item))
            self.qa_assigned_table.setItem(row_idx, 2, _center_table_item(flag_item))
            self.qa_assigned_table.setItem(row_idx, 3, _center_table_item(QTableWidgetItem(age_text)))
            self.qa_assigned_table.setItem(row_idx, 4, _center_table_item(working_item))
            self.qa_assigned_table.setItem(row_idx, 5, _center_table_item(assigned_item))
            self.qa_assigned_table.setItem(row_idx, 6, _center_table_item(QTableWidgetItem(category)))
            self.qa_assigned_table.setItem(row_idx, 7, _center_table_item(qa_note_item))
            self.qa_assigned_table.setItem(row_idx, 8, _center_table_item(agent_note_item))

    def _refresh_delivered_parts(self):
        if not hasattr(self, "qa_delivered_table"):
            return

        def _split_merged_values(raw_value: str) -> list[str]:
            values = [
                str(piece or "").strip()
                for piece in str(raw_value or "").split(" | ")
                if str(piece or "").strip()
            ]
            return values if values else [""]

        search_text = ""
        if hasattr(self, "qa_delivered_workorder_search"):
            search_text = str(self.qa_delivered_workorder_search.text() or "").strip()
        query = (
            "SELECT p.id, p.created_at, p.work_order, p.assigned_user_id, p.category, "
            "COALESCE(p.qa_comment, '') AS qa_comment, COALESCE(p.comments, '') AS comments, "
            "COALESCE(p.parts_installed, 0) AS parts_installed, "
            "COALESCE(p.parts_installed_by, '') AS parts_installed_by, COALESCE(p.parts_installed_at, '') AS parts_installed_at "
            "FROM parts p "
            "WHERE p.is_active=1 "
            "AND p.id=("
            "SELECT MAX(p2.id) FROM parts p2 WHERE p2.is_active=1 AND p2.work_order=p.work_order"
            ") "
            "AND EXISTS ("
            "SELECT 1 FROM part_details d2 "
            "JOIN parts p3 ON p3.id=d2.part_id "
            "WHERE p3.is_active=1 AND p3.work_order=p.work_order AND COALESCE(d2.delivered, 0)=1"
            ")"
        )
        params: list[Any] = []
        if search_text:
            query += " AND p.work_order LIKE ?"
            params.append(f"%{search_text}%")
        query += " ORDER BY COALESCE(p.parts_installed, 0) ASC, p.created_at ASC, p.id ASC LIMIT 400"

        rows = self.tracker.db.fetchall(query, tuple(params))
        agent_meta = self._qa_agent_meta_lookup()
        self.qa_delivered_table.setRowCount(0)
        installed_icon = QIcon.fromTheme("applications-engineering")
        if installed_icon.isNull():
            installed_icon = QIcon.fromTheme("tools")
        if installed_icon.isNull():
            installed_icon = QIcon.fromTheme("preferences-system")
        if installed_icon.isNull():
            installed_icon = self.style().standardIcon(QStyle.StandardPixmap.SP_DialogApplyButton)
        for r in rows:
            part_id = int(r["id"])
            work_order = str(r["work_order"] or "").strip()
            assigned = DepotRules.normalize_user_id(str(r["assigned_user_id"] or ""))
            assigned_name, assigned_icon = agent_meta.get(assigned, ("", ""))
            assigned_text = f"{assigned} - {assigned_name}" if assigned and assigned_name else (assigned or "-")
            category = str(r["category"] or "").strip() or "Other"
            age_text = DepotAgentWindow._part_age_label(str(r["created_at"] or ""))
            qa_comment = str(r["qa_comment"] or r["comments"] or "").strip()
            parts_installed = bool(int(r["parts_installed"] or 0))
            parts_installed_by = DepotRules.normalize_user_id(str(r["parts_installed_by"] or ""))
            parts_installed_at = str(r["parts_installed_at"] or "").strip()
            detail_rows = self.tracker.db.fetchall(
                "SELECT COALESCE(d.lpn, '') AS lpn, "
                "COALESCE(d.part_number, '') AS part_number, "
                "COALESCE(d.part_description, '') AS part_description, "
                "COALESCE(d.installed_keys, '') AS installed_keys, "
                "COALESCE(d.shipping_info, '') AS shipping_info "
                "FROM part_details d "
                "JOIN parts p ON p.id=d.part_id "
                "WHERE p.is_active=1 AND p.work_order=? AND COALESCE(d.delivered, 0)=1 "
                "ORDER BY d.id ASC",
                (work_order,),
            )
            detail_line_items: list[tuple[str, str, str, str, bool]] = []
            for detail in detail_rows:
                lpn_values = _split_merged_values(str(detail["lpn"] or ""))
                part_values = _split_merged_values(str(detail["part_number"] or ""))
                desc_values = _split_merged_values(str(detail["part_description"] or ""))
                ship_values = _split_merged_values(str(detail["shipping_info"] or ""))
                max_len = max(len(lpn_values), len(part_values), len(desc_values), len(ship_values), 1)
                installed_keys_raw = str(detail["installed_keys"] or "").strip()
                installed_key_set: set[str] = set()
                if installed_keys_raw:
                    try:
                        parsed = json.loads(installed_keys_raw)
                        if isinstance(parsed, list):
                            for value in parsed:
                                value_text = str(value or "").strip()
                                if value_text:
                                    installed_key_set.add(value_text)
                    except Exception:
                        for value in installed_keys_raw.split(" | "):
                            value_text = str(value or "").strip()
                            if value_text:
                                installed_key_set.add(value_text)

                def _value_for(values: list[str], idx: int) -> str:
                    if idx < len(values):
                        return str(values[idx] or "").strip()
                    if len(values) == 1 and max_len > 1:
                        return str(values[0] or "").strip()
                    return ""

                def _line_key(lpn_value: str, part_value: str, desc_value: str, ship_value: str) -> str:
                    return json.dumps(
                        [
                            str(lpn_value or "").strip().casefold(),
                            str(part_value or "").strip().casefold(),
                            str(desc_value or "").strip().casefold(),
                            str(ship_value or "").strip().casefold(),
                        ],
                        ensure_ascii=True,
                        separators=(",", ":"),
                    )

                for idx in range(max_len):
                    lpn_value = _value_for(lpn_values, idx)
                    part_value = _value_for(part_values, idx)
                    desc_value = _value_for(desc_values, idx)
                    ship_value = _value_for(ship_values, idx)
                    if not (lpn_value or part_value or desc_value or ship_value):
                        continue
                    key = _line_key(lpn_value, part_value, desc_value, ship_value)
                    row_installed = bool(key in installed_key_set)
                    if not installed_key_set and parts_installed:
                        # Legacy fallback before per-line install snapshots existed.
                        row_installed = True
                    detail_line_items.append((lpn_value, part_value, desc_value, ship_value, row_installed))
            if not detail_line_items:
                detail_line_items.append(("", "", "", "", bool(parts_installed)))

            for lpn, part_number, part_description, shipping_info, row_installed in detail_line_items:
                insert_row_idx = self.qa_delivered_table.rowCount()
                self.qa_delivered_table.insertRow(insert_row_idx)

                installed_item = QTableWidgetItem("")
                installed_item.setData(Qt.ItemDataRole.UserRole, part_id)
                if row_installed:
                    installed_item.setIcon(installed_icon)
                    tip = "Parts installed."
                    if parts_installed_by:
                        tip += f"\nBy: {parts_installed_by}"
                    friendly_stamp = DepotAgentWindow._format_working_updated_stamp(parts_installed_at)
                    if friendly_stamp:
                        tip += f"\nAt: {friendly_stamp}"
                    installed_item.setToolTip(tip)
                else:
                    installed_item.setToolTip("Waiting for agent install update.")

                assigned_item = QTableWidgetItem(assigned_text)
                if assigned_icon and Path(assigned_icon).exists():
                    assigned_item.setIcon(QIcon(assigned_icon))
                qa_note_item = QTableWidgetItem(DepotAgentWindow._note_preview(qa_comment))
                qa_note_item.setToolTip(f"QA Note: {qa_comment if qa_comment else '(none)'}")
                work_item = QTableWidgetItem(work_order)
                work_item.setData(Qt.ItemDataRole.UserRole, part_id)

                self.qa_delivered_table.setItem(insert_row_idx, 0, _center_table_item(work_item))
                self.qa_delivered_table.setItem(insert_row_idx, 1, _center_table_item(installed_item))
                self.qa_delivered_table.setItem(insert_row_idx, 2, _center_table_item(QTableWidgetItem(age_text)))
                self.qa_delivered_table.setItem(insert_row_idx, 3, _center_table_item(assigned_item))
                self.qa_delivered_table.setItem(insert_row_idx, 4, _center_table_item(QTableWidgetItem(category)))
                self.qa_delivered_table.setItem(insert_row_idx, 5, _center_table_item(QTableWidgetItem(lpn)))
                self.qa_delivered_table.setItem(insert_row_idx, 6, _center_table_item(QTableWidgetItem(part_number)))
                self.qa_delivered_table.setItem(insert_row_idx, 7, _center_table_item(QTableWidgetItem(part_description)))
                self.qa_delivered_table.setItem(insert_row_idx, 8, _center_table_item(QTableWidgetItem(shipping_info)))
                self.qa_delivered_table.setItem(insert_row_idx, 9, _center_table_item(qa_note_item))

    def _export_qa_delivered_csv(self) -> None:
        if not hasattr(self, "qa_delivered_table"):
            return
        table = self.qa_delivered_table
        if table.rowCount() <= 0 or table.columnCount() <= 0:
            self._show_themed_message(QMessageBox.Icon.Information, "Export CSV", "No delivered rows to export.")
            return

        stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        default_name = f"qa_parts_delivered_{stamp}.csv"
        if self.app_window is not None and self.app_window.config_path.parent.exists():
            start_dir = self.app_window.config_path.parent
        else:
            start_dir = Path.home()

        out_path, _ = show_flowgrid_themed_save_file_name(
            self,
            self.app_window,
            "qa",
            "Export Delivered Parts",
            str(start_dir / default_name),
            "CSV Files (*.csv);;All Files (*.*)",
        )
        if not out_path:
            return

        headers: list[str] = []
        for col in range(table.columnCount()):
            hdr = table.horizontalHeaderItem(col)
            headers.append(hdr.text() if hdr is not None else f"col_{col}")

        try:
            with open(out_path, "w", newline="", encoding="utf-8-sig") as handle:
                writer = csv.writer(handle)
                writer.writerow(headers)
                for row in range(table.rowCount()):
                    values: list[str] = []
                    for col in range(table.columnCount()):
                        item = table.item(row, col)
                        value = item.text() if item is not None else ""
                        if headers[col].strip().casefold() == "installed":
                            value = "Yes" if (item is not None and not item.icon().isNull()) else ""
                        values.append(value)
                    writer.writerow(values)
            self._show_themed_message(QMessageBox.Icon.Information, "Export CSV", f"Exported:\n{out_path}")
        except Exception as exc:
            _runtime_log_event(
                "ui.qa_delivered_export_csv_failed",
                severity="error",
                summary="QA delivered parts CSV export failed.",
                exc=exc,
                context={"path": str(out_path)},
            )
            self._show_themed_message(
                QMessageBox.Icon.Warning,
                "Export CSV",
                f"Failed to export CSV:\n{type(exc).__name__}: {exc}\n\nDetails were logged for support.",
            )

    def _refresh_completed_parts(self):
        if not hasattr(self, "qa_completed_table"):
            return
        search_text = ""
        if hasattr(self, "qa_completed_workorder_search"):
            search_text = str(self.qa_completed_workorder_search.text() or "").strip()
        query = (
            "SELECT p.id, p.created_at, p.work_order, p.assigned_user_id, p.category, p.client_unit, "
            "COALESCE(p.qa_comment, '') AS qa_comment, COALESCE(p.agent_comment, '') AS agent_comment, "
            "COALESCE(p.comments, '') AS comments, COALESCE(p.qa_flag, '') AS qa_flag, COALESCE(p.qa_flag_image_path, '') AS qa_flag_image_path, "
            "COALESCE(p.working_user_id, '') AS working_user_id, COALESCE(p.working_updated_at, '') AS working_updated_at, "
            "COALESCE(ls.touch, '') AS latest_touch, COALESCE(ls.created_at, '') AS latest_touch_at "
            "FROM parts p "
            "LEFT JOIN submissions ls ON ls.id = ("
            "SELECT s2.id FROM submissions s2 WHERE s2.work_order = p.work_order "
            "ORDER BY s2.created_at DESC, s2.id DESC LIMIT 1"
            ") "
            "WHERE p.is_active=0 AND COALESCE(ls.touch, '') IN (?, ?, ?)"
        )
        params: list[Any] = [
            DepotRules.TOUCH_COMPLETE,
            DepotRules.TOUCH_JUNK,
            DepotRules.TOUCH_RTV,
        ]
        if search_text:
            query += " AND p.work_order LIKE ?"
            params.append(f"%{search_text}%")
        query += " ORDER BY COALESCE(NULLIF(TRIM(ls.created_at), ''), p.created_at) DESC, p.id DESC LIMIT 400"

        rows = self.tracker.db.fetchall(query, tuple(params))
        agent_meta = self._qa_agent_meta_lookup()
        self.qa_completed_table.setRowCount(0)
        for row_idx, r in enumerate(rows):
            self.qa_completed_table.insertRow(row_idx)
            part_id = int(r["id"])
            work_order = str(r["work_order"] or "").strip()
            assigned = DepotRules.normalize_user_id(str(r["assigned_user_id"] or ""))
            category = str(r["category"] or "").strip() or "Other"
            age_text = DepotAgentWindow._part_age_label(str(r["created_at"] or ""))
            qa_comment = str(r["qa_comment"] or r["comments"] or "").strip()
            agent_comment = str(r["agent_comment"] or "").strip()
            flag = str(r["qa_flag"] or "").strip()
            working_user = DepotRules.normalize_user_id(str(r["working_user_id"] or ""))
            working_stamp = str(r["working_updated_at"] or "").strip()
            outcome_text = str(r["latest_touch"] or "").strip()
            closed_at_raw = str(r["latest_touch_at"] or "").strip()
            if len(closed_at_raw) >= 19:
                closed_at_text = f"{closed_at_raw[:10]} {closed_at_raw[11:19]}"
            elif len(closed_at_raw) >= 16:
                closed_at_text = f"{closed_at_raw[:10]} {closed_at_raw[11:16]}"
            else:
                closed_at_text = closed_at_raw[:19] if closed_at_raw else "-"

            image_abs = self.tracker.resolve_qa_flag_icon(
                str(r["qa_flag"] or "").strip(),
                str(r["qa_flag_image_path"] or ""),
            )
            assigned_name, assigned_icon = agent_meta.get(assigned, ("", ""))

            client_item = QTableWidgetItem("")
            client_item.setData(Qt.ItemDataRole.UserRole, part_id)
            if int(r["client_unit"] or 0):
                client_item.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_DialogApplyButton))

            flag_item = QTableWidgetItem("" if image_abs else (flag if flag else ""))
            flag_item.setData(Qt.ItemDataRole.UserRole, part_id)
            flag_item.setToolTip(DepotAgentWindow._flag_tooltip(flag, qa_comment, agent_comment, bool(image_abs)))
            if image_abs:
                flag_item.setIcon(QIcon(image_abs))

            working_item = QTableWidgetItem("\U0001F527" if working_user else "")
            working_item.setData(Qt.ItemDataRole.UserRole, part_id)
            if working_user:
                working_tip = f"Agent working this unit: {working_user}"
                friendly_stamp = DepotAgentWindow._format_working_updated_stamp(working_stamp)
                if friendly_stamp:
                    working_tip += f"\nUpdated: {friendly_stamp}"
                working_item.setToolTip(working_tip)
            else:
                working_item.setToolTip("No agent is marked as working this unit.")

            assigned_text = "-"
            if assigned:
                assigned_text = f"{assigned} - {assigned_name}" if assigned_name else assigned
            assigned_item = QTableWidgetItem(assigned_text)
            if assigned_icon and Path(assigned_icon).exists():
                assigned_item.setIcon(QIcon(assigned_icon))

            qa_note_item = QTableWidgetItem(DepotAgentWindow._note_preview(qa_comment))
            qa_note_item.setToolTip(f"QA Note: {qa_comment if qa_comment else '(none)'}")
            agent_note_item = QTableWidgetItem(DepotAgentWindow._note_preview(agent_comment))
            agent_note_item.setToolTip(f"Agent Note: {agent_comment if agent_comment else '(none)'}")

            self.qa_completed_table.setItem(row_idx, 0, _center_table_item(client_item))
            self.qa_completed_table.setItem(row_idx, 1, _center_table_item(flag_item))
            self.qa_completed_table.setItem(row_idx, 2, _center_table_item(QTableWidgetItem(age_text)))
            self.qa_completed_table.setItem(row_idx, 3, _center_table_item(working_item))
            self.qa_completed_table.setItem(row_idx, 4, _center_table_item(QTableWidgetItem(work_order)))
            self.qa_completed_table.setItem(row_idx, 5, _center_table_item(assigned_item))
            self.qa_completed_table.setItem(row_idx, 6, _center_table_item(QTableWidgetItem(category)))
            self.qa_completed_table.setItem(row_idx, 7, _center_table_item(QTableWidgetItem(outcome_text)))
            self.qa_completed_table.setItem(row_idx, 8, _center_table_item(QTableWidgetItem(closed_at_text)))
            self.qa_completed_table.setItem(row_idx, 9, _center_table_item(qa_note_item))
            self.qa_completed_table.setItem(row_idx, 10, _center_table_item(agent_note_item))


class IconCropDialog(FlowgridThemedDialog):
    def __init__(self, image: QImage, app_window: "QuickInputsWindow" | None = None, parent: QWidget | None = None):
        super().__init__(parent, app_window, "admin")
        self._source_image = image.copy()
        self._result_image = QImage()

        self.setWindowTitle("Icon Creator")
        self.setModal(True)
        self.resize(920, 560)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(8)

        hint = QLabel(
            "Drag to move the image inside the icon frame. "
            "Use mouse wheel or slider to zoom. The live preview shows the final icon."
        )
        hint.setWordWrap(True)
        layout.addWidget(hint)

        body = QHBoxLayout()
        body.setSpacing(8)
        self.arrange_preview = IconArrangePreview(self._source_image, self)
        body.addWidget(self.arrange_preview, 1)

        preview_col = QVBoxLayout()
        preview_col.setSpacing(6)
        preview_col.addWidget(QLabel("Final Icon Preview"))
        self.output_preview = QLabel()
        self.output_preview.setFixedSize(130, 130)
        self.output_preview.setFrameShape(QFrame.Shape.Box)
        preview_col.addWidget(self.output_preview, 0, Qt.AlignmentFlag.AlignCenter)
        preview_col.addStretch(1)
        body.addLayout(preview_col, 0)
        layout.addLayout(body, 1)

        zoom_row = QHBoxLayout()
        zoom_row.addWidget(QLabel("Zoom"), 0)
        self.zoom_slider = QSlider(Qt.Orientation.Horizontal)
        self.zoom_slider.setRange(100, 400)
        self.zoom_slider.setSingleStep(5)
        self.zoom_slider.setPageStep(20)
        self.zoom_slider.setValue(100)
        zoom_row.addWidget(self.zoom_slider, 1)
        layout.addLayout(zoom_row)

        actions = QHBoxLayout()
        self.crop_btn = QPushButton("Apply")
        self.cancel_btn = QPushButton("Cancel")
        self.crop_btn.setProperty("actionRole", "save")
        self.cancel_btn.setProperty("actionRole", "reset")
        actions.addWidget(self.crop_btn)
        actions.addWidget(self.cancel_btn)
        layout.addLayout(actions)

        self.zoom_slider.valueChanged.connect(self._on_zoom_slider_changed)
        self.arrange_preview.zoom_changed.connect(self._on_preview_zoom_changed)
        self.arrange_preview.view_changed.connect(self._update_output_preview)
        self.crop_btn.clicked.connect(self._accept_crop)
        self.cancel_btn.clicked.connect(self.reject)

        self.apply_theme_styles(force_opaque_root=True)
        self._update_output_preview()

    @staticmethod
    def _set_preview_label_image(target: QLabel, image: QImage) -> None:
        if image.isNull():
            target.clear()
            return
        pix = QPixmap.fromImage(image)
        if pix.isNull():
            target.clear()
            return
        fitted = pix.scaled(
            target.size(),
            Qt.AspectRatioMode.KeepAspectRatio,
            Qt.TransformationMode.SmoothTransformation,
        )
        target.setPixmap(fitted)

    def _on_zoom_slider_changed(self, value: int) -> None:
        self.arrange_preview.set_zoom(float(value) / 100.0)

    def _on_preview_zoom_changed(self, value: float) -> None:
        slider_value = int(round(float(value) * 100.0))
        if self.zoom_slider.value() == slider_value:
            return
        self.zoom_slider.blockSignals(True)
        self.zoom_slider.setValue(slider_value)
        self.zoom_slider.blockSignals(False)

    def _update_output_preview(self) -> None:
        preview = self.arrange_preview.render_cropped_image(180)
        self._set_preview_label_image(self.output_preview, preview)

    def _accept_crop(self) -> None:
        arranged = self.arrange_preview.render_cropped_image(256)
        if arranged.isNull():
            return
        self._result_image = arranged
        self.accept()

    def result_image(self) -> QImage:
        return self._result_image.copy()


class IconArrangePreview(QWidget):
    zoom_changed = Signal(float)
    view_changed = Signal()

    def __init__(self, image: QImage, parent: QWidget | None = None):
        super().__init__(parent)
        self._image = image.copy()
        self._frame_rect = QRect()
        self._zoom = 1.0
        self._pan = QPointF(0.0, 0.0)
        self._drag_active = False
        self._last_pos = QPoint()
        self.setMinimumSize(360, 250)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.setCursor(Qt.CursorShape.OpenHandCursor)
        self.set_source_image(image, reset_view=True)

    def set_source_image(self, image: QImage, *, reset_view: bool = True) -> None:
        self._image = image.copy()
        if reset_view:
            self._zoom = 1.0
            self._pan = QPointF(0.0, 0.0)
            self.zoom_changed.emit(self._zoom)
        self._clamp_pan()
        self.view_changed.emit()
        self.update()

    def _calc_frame_rect(self) -> QRect:
        bounds = self.rect().adjusted(10, 10, -10, -10)
        if bounds.width() <= 0 or bounds.height() <= 0:
            return QRect()
        side = max(1, min(bounds.width(), bounds.height()))
        x = bounds.left() + max(0, (bounds.width() - side) // 2)
        y = bounds.top() + max(0, (bounds.height() - side) // 2)
        return QRect(x, y, side, side)

    def zoom(self) -> float:
        return float(self._zoom)

    def set_zoom(self, value: float) -> None:
        new_zoom = float(clamp(float(value), 1.0, 4.0))
        if abs(new_zoom - self._zoom) < 0.001:
            return
        self._zoom = new_zoom
        self._clamp_pan()
        self.zoom_changed.emit(self._zoom)
        self.view_changed.emit()
        self.update()

    def _clamp_pan(self, frame: QRectF | None = None) -> None:
        if self._image.isNull():
            self._pan = QPointF(0.0, 0.0)
            return
        draw_frame = QRectF(frame) if frame is not None else QRectF(self._frame_rect if self._frame_rect.isValid() else self._calc_frame_rect())
        if draw_frame.width() <= 0 or draw_frame.height() <= 0:
            self._pan = QPointF(0.0, 0.0)
            return
        base_scale = max(
            draw_frame.width() / max(1, self._image.width()),
            draw_frame.height() / max(1, self._image.height()),
        )
        scale = base_scale * self._zoom
        scaled_w = float(self._image.width()) * scale
        scaled_h = float(self._image.height()) * scale
        max_x = max(0.0, (scaled_w - draw_frame.width()) / 2.0)
        max_y = max(0.0, (scaled_h - draw_frame.height()) / 2.0)
        self._pan.setX(float(clamp(self._pan.x(), -max_x, max_x)))
        self._pan.setY(float(clamp(self._pan.y(), -max_y, max_y)))

    def _draw_image_in_frame(self, painter: QPainter, frame: QRectF) -> None:
        if self._image.isNull() or frame.width() <= 0 or frame.height() <= 0:
            return
        self._clamp_pan(frame)
        base_scale = max(
            frame.width() / max(1, self._image.width()),
            frame.height() / max(1, self._image.height()),
        )
        scale = base_scale * self._zoom
        scaled_w = float(self._image.width()) * scale
        scaled_h = float(self._image.height()) * scale
        center = frame.center()
        target = QRectF(
            center.x() - (scaled_w / 2.0) + self._pan.x(),
            center.y() - (scaled_h / 2.0) + self._pan.y(),
            scaled_w,
            scaled_h,
        )

        clip = QPainterPath()
        clip.addRoundedRect(frame, 6, 6)
        painter.save()
        painter.setClipPath(clip)
        painter.drawImage(target, self._image)
        painter.restore()
        painter.setPen(QPen(QColor(42, 224, 203), 2, Qt.PenStyle.SolidLine))
        painter.drawRoundedRect(frame, 6, 6)

    def render_cropped_image(self, target_size: int = 256) -> QImage:
        if self._image.isNull():
            return QImage()
        side = max(64, int(target_size))
        output = QImage(side, side, QImage.Format.Format_ARGB32_Premultiplied)
        output.fill(Qt.GlobalColor.transparent)
        painter = QPainter(output)
        painter.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform)
        self._draw_image_in_frame(painter, QRectF(0.0, 0.0, float(side), float(side)))
        painter.end()
        return output

    def mousePressEvent(self, event: QMouseEvent) -> None:  # noqa: N802
        if event.button() != Qt.MouseButton.LeftButton:
            return super().mousePressEvent(event)
        point = event.position().toPoint()
        if not self._frame_rect.isValid() or not self._frame_rect.contains(point):
            return super().mousePressEvent(event)
        self._drag_active = True
        self._last_pos = point
        self.setCursor(Qt.CursorShape.ClosedHandCursor)
        event.accept()

    def mouseMoveEvent(self, event: QMouseEvent) -> None:  # noqa: N802
        if not self._drag_active:
            return super().mouseMoveEvent(event)
        point = event.position().toPoint()
        delta = point - self._last_pos
        self._last_pos = point
        self._pan = QPointF(self._pan.x() + delta.x(), self._pan.y() + delta.y())
        self._clamp_pan()
        self.view_changed.emit()
        self.update()
        event.accept()

    def mouseReleaseEvent(self, event: QMouseEvent) -> None:  # noqa: N802
        if not self._drag_active:
            return super().mouseReleaseEvent(event)
        self._drag_active = False
        self.setCursor(Qt.CursorShape.OpenHandCursor)
        event.accept()

    def wheelEvent(self, event) -> None:  # noqa: N802
        point = event.position().toPoint()
        if self._frame_rect.isValid() and self._frame_rect.contains(point):
            delta_y = event.angleDelta().y()
            if delta_y != 0:
                factor = 1.1 if delta_y > 0 else (1.0 / 1.1)
                self.set_zoom(self._zoom * factor)
                event.accept()
                return
        super().wheelEvent(event)

    def paintEvent(self, event) -> None:  # noqa: N802
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform)
        painter.fillRect(self.rect(), QColor(10, 10, 10, 220))
        if self._image.isNull():
            return
        self._frame_rect = self._calc_frame_rect()
        if not self._frame_rect.isValid():
            return
        self._draw_image_in_frame(painter, QRectF(self._frame_rect))


class DepotAdminDialog(DepotFramelessToolWindow):
    def __init__(self, tracker: DepotTracker, current_user: str, app_window: "QuickInputsWindow" | None = None):
        super().__init__(app_window, window_title="Admin Panel", theme_kind="admin", size=(620, 500))
        self.tracker = tracker
        self.current_user = DepotRules.normalize_user_id(current_user)
        self._agent_cache: dict[str, dict[str, Any]] = {}
        self._qa_flag_cache: dict[str, dict[str, Any]] = {}
        self._admin_cache: dict[str, dict[str, Any]] = {}

        self.whoami_label = QLabel(f"Current User: {self.current_user}")
        self.root_layout.addWidget(self.whoami_label)

        self.admin_tabs = QTabWidget(self)
        self.admin_tabs.setObjectName("AdminTabs")
        self.root_layout.addWidget(self.admin_tabs, 1)

        self.agent_tab = QWidget()
        self.qa_tab = QWidget()
        self.admin_tab = QWidget()
        self.agent_tab.setObjectName("AdminAgentsTab")
        self.qa_tab.setObjectName("AdminQaTab")
        self.admin_tab.setObjectName("AdminAdminsTab")
        self.admin_tabs.addTab(self.agent_tab, "Agents")
        self.admin_tabs.addTab(self.qa_tab, "QA")
        self.admin_tabs.addTab(self.admin_tab, "Admins")

        self._build_agents_tab()
        self._build_qa_tab()
        self._build_admins_tab()

        if self.app_window is not None:
            self.apply_theme_styles()

        self.refresh_agents()
        self.refresh_qa_flags()
        self.refresh_admins()

    def apply_theme_styles(self) -> None:
        if self.app_window is None:
            return
        super().apply_theme_styles()

    def _build_agents_tab(self) -> None:
        layout = QVBoxLayout(self.agent_tab)
        layout.setContentsMargins(6, 6, 6, 6)
        layout.setSpacing(6)

        form = QFormLayout()
        self.agent_number_input = QLineEdit()
        self.agent_name_input = QLineEdit()
        self.agent_location_input = QLineEdit()
        self.agent_tier_combo = QComboBox()
        self.agent_tier_combo.addItem("Tech 1", 1)
        self.agent_tier_combo.addItem("Tech 2", 2)
        self.agent_tier_combo.addItem("Tech 3", 3)
        self.agent_tier_combo.addItem("MP", 4)

        self.agent_icon_input = QLineEdit()
        self.agent_icon_browse = QPushButton("Browse")
        icon_row = QHBoxLayout()
        icon_row.setContentsMargins(0, 0, 0, 0)
        icon_row.setSpacing(4)
        icon_row.addWidget(self.agent_icon_input, 1)
        icon_row.addWidget(self.agent_icon_browse, 0)
        icon_wrap = QWidget()
        icon_wrap.setLayout(icon_row)

        form.addRow("Agent Number", self.agent_number_input)
        form.addRow("Worker Name", self.agent_name_input)
        form.addRow("Location", self.agent_location_input)
        form.addRow("Tech", self.agent_tier_combo)
        form.addRow("Icon", icon_wrap)
        layout.addLayout(form)

        btn_row = QHBoxLayout()
        self.agent_save_btn = QPushButton("Add / Update")
        self.agent_remove_btn = QPushButton("Remove")
        self.agent_clear_btn = QPushButton("Clear")
        self.agent_save_btn.setProperty("actionRole", "save")
        self.agent_remove_btn.setProperty("actionRole", "reset")
        self.agent_clear_btn.setProperty("actionRole", "pick")
        btn_row.addWidget(self.agent_save_btn)
        btn_row.addWidget(self.agent_remove_btn)
        btn_row.addWidget(self.agent_clear_btn)
        layout.addLayout(btn_row)

        self.agent_table = QTableWidget()
        configure_standard_table(
            self.agent_table,
            ["Agent Number", "Worker Name", "Location", "Tech", "Icon"],
            resize_modes={
                0: QHeaderView.ResizeMode.ResizeToContents,
                1: QHeaderView.ResizeMode.Stretch,
                2: QHeaderView.ResizeMode.ResizeToContents,
                3: QHeaderView.ResizeMode.ResizeToContents,
                4: QHeaderView.ResizeMode.ResizeToContents,
            },
            stretch_last=True,
        )
        layout.addWidget(self.agent_table, 1)

        self.agent_icon_browse.clicked.connect(self._browse_agent_icon)
        self.agent_save_btn.clicked.connect(self._save_agent)
        self.agent_remove_btn.clicked.connect(self._remove_selected_agent)
        self.agent_clear_btn.clicked.connect(self._clear_agent_form)
        self.agent_table.itemSelectionChanged.connect(self._on_agent_selected)
        self.agent_table.cellDoubleClicked.connect(self._on_agent_double_clicked_row)

    def _build_qa_tab(self) -> None:
        layout = QVBoxLayout(self.qa_tab)
        layout.setContentsMargins(6, 6, 6, 6)
        layout.setSpacing(6)

        form = QFormLayout()
        self.qa_flag_name_input = QLineEdit()
        self.qa_flag_severity_combo = QComboBox()
        for value in QA_FLAG_SEVERITY_OPTIONS:
            self.qa_flag_severity_combo.addItem(value)

        self.qa_flag_icon_input = QLineEdit()
        self.qa_flag_icon_browse = QPushButton("Browse")
        self.qa_flag_icon_browse.setProperty("actionRole", "pick")
        icon_row = QHBoxLayout()
        icon_row.setContentsMargins(0, 0, 0, 0)
        icon_row.setSpacing(4)
        icon_row.addWidget(self.qa_flag_icon_input, 1)
        icon_row.addWidget(self.qa_flag_icon_browse, 0)
        icon_wrap = QWidget()
        icon_wrap.setLayout(icon_row)

        form.addRow("Flag Name", self.qa_flag_name_input)
        form.addRow("Severity", self.qa_flag_severity_combo)
        form.addRow("Icon", icon_wrap)
        layout.addLayout(form)

        btn_row = QHBoxLayout()
        self.qa_flag_save_btn = QPushButton("Add / Update")
        self.qa_flag_remove_btn = QPushButton("Remove")
        self.qa_flag_clear_btn = QPushButton("Clear")
        self.qa_flag_save_btn.setProperty("actionRole", "save")
        self.qa_flag_remove_btn.setProperty("actionRole", "reset")
        self.qa_flag_clear_btn.setProperty("actionRole", "pick")
        btn_row.addWidget(self.qa_flag_save_btn)
        btn_row.addWidget(self.qa_flag_remove_btn)
        btn_row.addWidget(self.qa_flag_clear_btn)
        layout.addLayout(btn_row)

        self.qa_flags_table = QTableWidget()
        configure_standard_table(
            self.qa_flags_table,
            ["Flag Name", "Severity", "Icon"],
            resize_modes={
                0: QHeaderView.ResizeMode.Stretch,
                1: QHeaderView.ResizeMode.ResizeToContents,
                2: QHeaderView.ResizeMode.ResizeToContents,
            },
            stretch_last=True,
        )
        layout.addWidget(self.qa_flags_table, 1)

        self.qa_flag_icon_browse.clicked.connect(self._browse_qa_flag_icon)
        self.qa_flag_save_btn.clicked.connect(self._save_qa_flag)
        self.qa_flag_remove_btn.clicked.connect(self._remove_selected_qa_flag)
        self.qa_flag_clear_btn.clicked.connect(self._clear_qa_flag_form)
        self.qa_flags_table.itemSelectionChanged.connect(self._on_qa_flag_selected)
        self.qa_flags_table.cellDoubleClicked.connect(self._on_qa_flag_double_clicked_row)

    def _build_admins_tab(self) -> None:
        layout = QVBoxLayout(self.admin_tab)
        layout.setContentsMargins(6, 6, 6, 6)
        layout.setSpacing(6)

        form = QFormLayout()
        self.admin_user_input = QLineEdit()
        self.admin_name_input = QLineEdit()
        self.admin_position_input = QLineEdit()
        self.admin_location_input = QLineEdit()

        self.admin_icon_input = QLineEdit()
        self.admin_icon_browse = QPushButton("Browse")
        self.admin_icon_browse.setProperty("actionRole", "pick")
        icon_row = QHBoxLayout()
        icon_row.setContentsMargins(0, 0, 0, 0)
        icon_row.setSpacing(4)
        icon_row.addWidget(self.admin_icon_input, 1)
        icon_row.addWidget(self.admin_icon_browse, 0)
        icon_wrap = QWidget()
        icon_wrap.setLayout(icon_row)

        form.addRow("Admin Number", self.admin_user_input)
        form.addRow("Name", self.admin_name_input)
        form.addRow("Position", self.admin_position_input)
        form.addRow("Location", self.admin_location_input)
        form.addRow("Icon", icon_wrap)
        layout.addLayout(form)

        btn_row = QHBoxLayout()
        self.admin_add_btn = QPushButton("Add / Update")
        self.admin_add_btn.setProperty("actionRole", "save")
        self.admin_remove_btn = QPushButton("Remove")
        self.admin_remove_btn.setProperty("actionRole", "reset")
        self.admin_clear_btn = QPushButton("Clear")
        self.admin_clear_btn.setProperty("actionRole", "pick")
        btn_row.addWidget(self.admin_add_btn)
        btn_row.addWidget(self.admin_remove_btn)
        btn_row.addWidget(self.admin_clear_btn)
        layout.addLayout(btn_row)

        self.admin_table = QTableWidget()
        configure_standard_table(
            self.admin_table,
            ["Admin Number", "Name", "Position", "Location", "Icon"],
            resize_modes={
                0: QHeaderView.ResizeMode.ResizeToContents,
                1: QHeaderView.ResizeMode.Stretch,
                2: QHeaderView.ResizeMode.ResizeToContents,
                3: QHeaderView.ResizeMode.ResizeToContents,
                4: QHeaderView.ResizeMode.ResizeToContents,
            },
            stretch_last=True,
        )
        layout.addWidget(self.admin_table, 1)

        self.admin_icon_browse.clicked.connect(self._browse_admin_icon)
        self.admin_add_btn.clicked.connect(self._add_admin_user)
        self.admin_remove_btn.clicked.connect(self._remove_selected_admin)
        self.admin_clear_btn.clicked.connect(self._clear_admin_form)
        self.admin_table.itemSelectionChanged.connect(self._on_admin_selected)
        self.admin_table.cellDoubleClicked.connect(self._on_admin_double_clicked_row)

    def refresh_agents(self) -> None:
        self._agent_cache.clear()
        self.agent_table.setRowCount(0)
        for row in self.tracker.list_agents():
            user_id = str(row["user_id"])
            agent_name = str(row["agent_name"] or "")
            tier = DepotRules.normalize_agent_tier(row["tier"])
            location = str(row.get("location", "") or "").strip()
            icon_path = str(row["icon_path"] or "").strip()
            row_idx = self.agent_table.rowCount()
            self.agent_table.insertRow(row_idx)
            user_item = QTableWidgetItem(user_id)
            user_item.setData(Qt.ItemDataRole.UserRole, user_id)
            name_item = QTableWidgetItem(agent_name)
            location_item = QTableWidgetItem(location or "-")
            tier_item = QTableWidgetItem(DepotRules.agent_tier_label(tier))
            icon_item = QTableWidgetItem(Path(icon_path).name if icon_path else "-")
            user_item.setTextAlignment(int(Qt.AlignmentFlag.AlignCenter))
            icon_item.setTextAlignment(int(Qt.AlignmentFlag.AlignCenter))
            if icon_path and Path(icon_path).exists():
                user_item.setIcon(QIcon(icon_path))
                icon_item.setIcon(QIcon(icon_path))
            self.agent_table.setItem(row_idx, 0, user_item)
            self.agent_table.setItem(row_idx, 1, name_item)
            self.agent_table.setItem(row_idx, 2, location_item)
            self.agent_table.setItem(row_idx, 3, tier_item)
            self.agent_table.setItem(row_idx, 4, icon_item)
            self._agent_cache[user_id] = {
                "agent_name": agent_name,
                "tier": tier,
                "location": location,
                "icon_path": icon_path,
            }
        self._notify_qa_agent_list_changed()

    def refresh_qa_flags(self) -> None:
        self._qa_flag_cache.clear()
        self.qa_flags_table.setRowCount(0)
        for row in self.tracker.list_qa_flags():
            flag_name = str(row.get("flag_name", "") or "").strip()
            severity = str(row.get("severity", "Medium") or "Medium").strip()
            icon_path = str(row.get("icon_path", "") or "").strip()
            row_idx = self.qa_flags_table.rowCount()
            self.qa_flags_table.insertRow(row_idx)
            flag_item = QTableWidgetItem(flag_name)
            flag_item.setData(Qt.ItemDataRole.UserRole, flag_name)
            severity_item = QTableWidgetItem(severity)
            icon_item = QTableWidgetItem(Path(icon_path).name if icon_path else "-")
            flag_item.setTextAlignment(int(Qt.AlignmentFlag.AlignCenter))
            icon_item.setTextAlignment(int(Qt.AlignmentFlag.AlignCenter))
            if icon_path and Path(icon_path).exists():
                flag_item.setIcon(QIcon(icon_path))
                icon_item.setIcon(QIcon(icon_path))
            self.qa_flags_table.setItem(row_idx, 0, flag_item)
            self.qa_flags_table.setItem(row_idx, 1, severity_item)
            self.qa_flags_table.setItem(row_idx, 2, icon_item)
            self._qa_flag_cache[flag_name] = {
                "severity": severity,
                "icon_path": icon_path,
            }
        self._notify_qa_flag_list_changed()

    def refresh_admins(self) -> None:
        self._admin_cache.clear()
        self.admin_table.setRowCount(0)
        for row in self.tracker.list_admin_users():
            user_id = str(row["user_id"] or "")
            admin_name = str(row["admin_name"] or "").strip()
            position = str(row["position"] or "").strip()
            location = str(row["location"] or "").strip()
            icon_path = str(row.get("icon_path", "") or "").strip()
            row_idx = self.admin_table.rowCount()
            self.admin_table.insertRow(row_idx)
            user_item = QTableWidgetItem(user_id)
            user_item.setData(Qt.ItemDataRole.UserRole, user_id)
            icon_item = QTableWidgetItem(Path(icon_path).name if icon_path else "-")
            user_item.setTextAlignment(int(Qt.AlignmentFlag.AlignCenter))
            icon_item.setTextAlignment(int(Qt.AlignmentFlag.AlignCenter))
            if icon_path and Path(icon_path).exists():
                user_item.setIcon(QIcon(icon_path))
                icon_item.setIcon(QIcon(icon_path))
            self.admin_table.setItem(row_idx, 0, user_item)
            self.admin_table.setItem(row_idx, 1, QTableWidgetItem(admin_name))
            self.admin_table.setItem(row_idx, 2, QTableWidgetItem(position))
            self.admin_table.setItem(row_idx, 3, QTableWidgetItem(location))
            self.admin_table.setItem(row_idx, 4, icon_item)
            self._admin_cache[user_id] = {
                "admin_name": admin_name,
                "position": position,
                "location": location,
                "icon_path": icon_path,
            }

    def _notify_qa_agent_list_changed(self) -> None:
        if self.app_window is None:
            return
        qa_window = getattr(self.app_window, "active_qa_window", None)
        if qa_window is not None and qa_window.isVisible():
            qa_window._populate_agents()
            qa_window._refresh_assigned_parts()
            qa_window._refresh_delivered_parts()
            qa_window._refresh_completed_parts()

    def _notify_qa_flag_list_changed(self) -> None:
        if self.app_window is None:
            return
        qa_window = getattr(self.app_window, "active_qa_window", None)
        if qa_window is not None and qa_window.isVisible():
            qa_window._populate_flags()
            qa_window._refresh_assigned_parts()
            qa_window._refresh_delivered_parts()
            qa_window._refresh_completed_parts()
        agent_window = getattr(self.app_window, "active_agent_window", None)
        if agent_window is not None and agent_window.isVisible():
            agent_window._refresh_agent_parts()
            agent_window._refresh_category_parts()

    def _clear_agent_form(self) -> None:
        self.agent_number_input.clear()
        self.agent_name_input.clear()
        self.agent_location_input.clear()
        self.agent_tier_combo.setCurrentIndex(0)
        self.agent_icon_input.clear()

    def _clear_qa_flag_form(self) -> None:
        self.qa_flag_name_input.clear()
        self.qa_flag_severity_combo.setCurrentIndex(1 if self.qa_flag_severity_combo.count() > 1 else 0)
        self.qa_flag_icon_input.clear()

    def _clear_admin_form(self) -> None:
        self.admin_user_input.clear()
        self.admin_name_input.clear()
        self.admin_position_input.clear()
        self.admin_location_input.clear()
        self.admin_icon_input.clear()

    def _load_agent_into_form(self, user_id: str) -> bool:
        normalized = DepotRules.normalize_user_id(user_id)
        if not normalized:
            return False
        data = self._agent_cache.get(normalized)
        if not data:
            return False
        self.agent_number_input.setText(normalized)
        self.agent_name_input.setText(str(data.get("agent_name", "")))
        self.agent_location_input.setText(str(data.get("location", "")))
        tier = DepotRules.normalize_agent_tier(data.get("tier", 1))
        idx = self.agent_tier_combo.findData(tier)
        self.agent_tier_combo.setCurrentIndex(idx if idx >= 0 else 0)
        self.agent_icon_input.setText(str(data.get("icon_path", "")))
        return True

    def _load_admin_into_form(self, user_id: str) -> bool:
        normalized = DepotRules.normalize_user_id(user_id)
        if not normalized:
            return False
        data = self._admin_cache.get(normalized)
        if not data:
            return False
        self.admin_user_input.setText(normalized)
        self.admin_name_input.setText(str(data.get("admin_name", "")))
        self.admin_position_input.setText(str(data.get("position", "")))
        self.admin_location_input.setText(str(data.get("location", "")))
        self.admin_icon_input.setText(str(data.get("icon_path", "")))
        return True

    def _select_admin_item(self, user_id: str) -> None:
        normalized = DepotRules.normalize_user_id(user_id)
        if not normalized:
            return
        for idx in range(self.admin_table.rowCount()):
            item = self.admin_table.item(idx, 0)
            if item is None:
                continue
            row_user = DepotRules.normalize_user_id(str(item.data(Qt.ItemDataRole.UserRole) or item.text() or ""))
            if row_user == normalized:
                self.admin_table.selectRow(idx)
                return

    def _select_agent_item(self, user_id: str) -> None:
        normalized = DepotRules.normalize_user_id(user_id)
        if not normalized:
            return
        for idx in range(self.agent_table.rowCount()):
            item = self.agent_table.item(idx, 0)
            if item is None:
                continue
            item_user = DepotRules.normalize_user_id(str(item.data(Qt.ItemDataRole.UserRole) or item.text() or ""))
            if item_user == normalized:
                self.agent_table.selectRow(idx)
                return

    def _on_agent_selected(self) -> None:
        row = self.agent_table.currentRow()
        if row < 0:
            return
        item = self.agent_table.item(row, 0)
        if item is None:
            return
        user_id = str(item.data(Qt.ItemDataRole.UserRole) or item.text() or "")
        if not self._load_agent_into_form(user_id):
            return

    def _on_agent_double_clicked_row(self, row: int, _column: int) -> None:
        if row < 0:
            return
        item = self.agent_table.item(row, 0)
        if item is None:
            return
        user_id = str(item.data(Qt.ItemDataRole.UserRole) or item.text() or "")
        if not self._load_agent_into_form(user_id):
            return
        self.agent_name_input.setFocus()
        self.agent_name_input.selectAll()

    def _on_qa_flag_selected(self) -> None:
        row = self.qa_flags_table.currentRow()
        if row < 0:
            return
        item = self.qa_flags_table.item(row, 0)
        if item is None:
            return
        flag_name = str(item.data(Qt.ItemDataRole.UserRole) or item.text() or "").strip()
        if not flag_name:
            return
        details = self._qa_flag_cache.get(flag_name, {})
        self.qa_flag_name_input.setText(flag_name)
        severity = str(details.get("severity", "Medium") or "Medium")
        idx = self.qa_flag_severity_combo.findText(severity)
        self.qa_flag_severity_combo.setCurrentIndex(idx if idx >= 0 else 1)
        self.qa_flag_icon_input.setText(str(details.get("icon_path", "") or ""))

    def _on_qa_flag_double_clicked_row(self, row: int, _column: int) -> None:
        if row < 0:
            return
        self.qa_flags_table.selectRow(row)
        self._on_qa_flag_selected()

    def _on_admin_selected(self) -> None:
        row = self.admin_table.currentRow()
        if row < 0:
            return
        item = self.admin_table.item(row, 0)
        if item is None:
            return
        user_id = str(item.data(Qt.ItemDataRole.UserRole) or item.text() or "")
        self._load_admin_into_form(user_id)

    def _on_admin_double_clicked_row(self, row: int, _column: int) -> None:
        if row < 0:
            return
        item = self.admin_table.item(row, 0)
        if item is None:
            return
        user_id = str(item.data(Qt.ItemDataRole.UserRole) or item.text() or "")
        if self._load_admin_into_form(user_id):
            self.admin_name_input.setFocus()
            self.admin_name_input.selectAll()

    @staticmethod
    def _image_file_dialog_filter() -> str:
        patterns: list[str] = []
        for fmt in QImageReader.supportedImageFormats():
            try:
                ext = bytes(fmt).decode("ascii", errors="ignore").strip().lower()
            except Exception:
                ext = ""
            if ext:
                patterns.append(f"*.{ext}")
        if patterns:
            deduped = " ".join(sorted(set(patterns)))
            return f"Images ({deduped});;All Files (*.*)"
        return "All Files (*.*)"

    @staticmethod
    def _read_icon_image(path: str) -> tuple[QImage, str]:
        reader = QImageReader(path)
        reader.setAutoTransform(True)
        image = reader.read()
        if not image.isNull():
            return image, ""
        error_text = str(reader.errorString() or "").strip()
        fallback_pixmap = QPixmap(path)
        if not fallback_pixmap.isNull():
            return fallback_pixmap.toImage(), ""
        if not error_text:
            error_text = "Unsupported or corrupted image format."
        return QImage(), error_text

    def _browse_qa_flag_icon(self) -> None:
        selected = self._select_icon_path_with_editor(
            "Select QA Flag Icon",
            role_key="qa_flag",
            failure_event_key="ui.admin_qa_flag_icon_open_failed",
            failure_summary="QA flag icon selection failed because the image could not be decoded.",
        )
        if selected:
            self.qa_flag_icon_input.setText(selected)

    def _browse_admin_icon(self) -> None:
        selected = self._select_icon_path_with_editor(
            "Select Admin Icon",
            role_key="admin",
            failure_event_key="ui.admin_admin_icon_open_failed",
            failure_summary="Admin icon selection failed because the image could not be decoded.",
        )
        if selected:
            self.admin_icon_input.setText(selected)

    def _write_cropped_temp_icon(self, image: QImage, role_key: str = "agent") -> str:
        if image.isNull():
            return ""
        safe_key = re.sub(r"[^A-Za-z0-9_-]+", "_", str(role_key or "agent")).strip("_") or "agent"
        temp_dir = self.tracker.db.db_path.parent / ASSETS_DIR_NAME / "_icon_tmp" / safe_key
        try:
            temp_dir.mkdir(parents=True, exist_ok=True)
        except Exception as exc:
            _runtime_log_event(
                "ui.admin_crop_temp_dir_create_failed",
                severity="warning",
                summary="Failed creating temporary crop icon directory.",
                exc=exc,
                context={"temp_dir": str(temp_dir)},
            )
            return ""
        stamp = int(time.time() * 1000)
        temp_path = temp_dir / f"crop_{stamp}.png"
        try:
            if image.save(str(temp_path), "PNG"):
                return str(temp_path)
        except Exception as exc:
            _runtime_log_event(
                "ui.admin_crop_temp_write_failed",
                severity="warning",
                summary="Failed writing cropped icon temp file.",
                exc=exc,
                context={"temp_path": str(temp_path)},
            )
            return ""
        return ""

    def _edit_icon_with_popup(
        self,
        icon_path: str,
        *,
        role_key: str,
        failure_event_key: str,
        failure_summary: str,
    ) -> str:
        image, load_error = self._read_icon_image(icon_path)
        if image.isNull():
            _runtime_log_event(
                failure_event_key,
                severity="warning",
                summary=failure_summary,
                context={
                    "icon_path": str(icon_path),
                    "error": load_error,
                },
            )
            details = f"\n\nDetails: {load_error}" if load_error else ""
            self._show_themed_message(QMessageBox.Icon.Warning, "Invalid Image", f"Could not open that image file.{details}")
            return ""
        dialog = IconCropDialog(image, app_window=self.app_window, parent=self)
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return ""
        result_image = dialog.result_image()
        if result_image.isNull():
            return ""
        written = self._write_cropped_temp_icon(result_image, role_key=role_key)
        if not written:
            self._show_themed_message(QMessageBox.Icon.Warning, "Save Failed", "Failed to prepare cropped icon.")
        return written

    def _select_icon_path_with_editor(
        self,
        dialog_title: str,
        *,
        role_key: str,
        failure_event_key: str,
        failure_summary: str,
    ) -> str:
        icon_path, _ = show_flowgrid_themed_open_file_name(
            self,
            self.app_window,
            "admin",
            dialog_title,
            str(Path.home()),
            self._image_file_dialog_filter(),
        )
        if not icon_path:
            return ""
        return self._edit_icon_with_popup(
            icon_path,
            role_key=role_key,
            failure_event_key=failure_event_key,
            failure_summary=failure_summary,
        )

    def _browse_agent_icon(self) -> None:
        selected = self._select_icon_path_with_editor(
            "Select Agent Icon",
            role_key="agent",
            failure_event_key="ui.admin_agent_icon_open_failed",
            failure_summary="Agent icon selection failed because the image could not be decoded.",
        )
        if selected:
            self.agent_icon_input.setText(selected)

    def _save_agent(self) -> None:
        user_id = DepotRules.normalize_user_id(self.agent_number_input.text())
        name = str(self.agent_name_input.text() or "").strip()
        location = str(self.agent_location_input.text() or "").strip()
        tier = DepotRules.normalize_agent_tier(self.agent_tier_combo.currentData() or 1)
        icon_path = str(self.agent_icon_input.text() or "").strip()
        if not user_id or not name:
            self._show_themed_message(QMessageBox.Icon.Warning, "Validation", "Agent Number and Worker Name are required.")
            return
        if icon_path and not Path(icon_path).exists():
            self._show_themed_message(QMessageBox.Icon.Warning, "Validation", "Icon path does not exist. Browse and select a valid icon.")
            return
        stored_icon = self.tracker.upsert_agent(user_id, name, tier, icon_path, location)
        self.agent_icon_input.setText(stored_icon)
        self.refresh_agents()
        self._select_agent_item(user_id)
        self._show_themed_message(QMessageBox.Icon.Information, "Saved", f"Agent {user_id} updated.")

    def _remove_selected_agent(self) -> None:
        row = self.agent_table.currentRow()
        if row < 0:
            self._show_themed_message(QMessageBox.Icon.Warning, "Validation", "Select an agent to remove.")
            return
        item = self.agent_table.item(row, 0)
        if item is None:
            return
        user_id = str(item.data(Qt.ItemDataRole.UserRole) or item.text() or "")
        if not user_id:
            return
        self.tracker.delete_agent(user_id)
        self.refresh_agents()
        self._clear_agent_form()
        self._show_themed_message(QMessageBox.Icon.Information, "Saved", f"Agent {user_id} removed.")

    def _save_qa_flag(self) -> None:
        flag_name = str(self.qa_flag_name_input.text() or "").strip()
        severity = str(self.qa_flag_severity_combo.currentText() or "Medium").strip()
        icon_path = str(self.qa_flag_icon_input.text() or "").strip()
        if not flag_name:
            self._show_themed_message(QMessageBox.Icon.Warning, "Validation", "Flag name is required.")
            return
        if icon_path and not Path(icon_path).exists():
            self._show_themed_message(QMessageBox.Icon.Warning, "Validation", "Flag icon path does not exist.")
            return
        selected_row = self.qa_flags_table.currentRow()
        selected_name = ""
        if selected_row >= 0:
            selected_item = self.qa_flags_table.item(selected_row, 0)
            if selected_item is not None:
                selected_name = str(selected_item.data(Qt.ItemDataRole.UserRole) or selected_item.text() or "").strip()
        if selected_name and selected_name != flag_name and flag_name in self._qa_flag_cache:
            self._show_themed_message(
                QMessageBox.Icon.Warning,
                "Validation",
                "A QA flag with that name already exists.",
            )
            return
        self.tracker.upsert_qa_flag(flag_name, severity, icon_path)
        if selected_name and selected_name != flag_name:
            self.tracker.delete_qa_flag(selected_name)
        self.refresh_qa_flags()
        for row_idx in range(self.qa_flags_table.rowCount()):
            cell = self.qa_flags_table.item(row_idx, 0)
            if cell is None:
                continue
            name = str(cell.data(Qt.ItemDataRole.UserRole) or cell.text() or "").strip()
            if name == flag_name:
                self.qa_flags_table.selectRow(row_idx)
                break
        self._show_themed_message(QMessageBox.Icon.Information, "Saved", f"QA flag '{flag_name}' updated.")

    def _remove_selected_qa_flag(self) -> None:
        row = self.qa_flags_table.currentRow()
        if row < 0:
            self._show_themed_message(QMessageBox.Icon.Warning, "Validation", "Select a QA flag to remove.")
            return
        item = self.qa_flags_table.item(row, 0)
        if item is None:
            return
        flag_name = str(item.data(Qt.ItemDataRole.UserRole) or item.text() or "").strip()
        if not flag_name:
            return
        self.tracker.delete_qa_flag(flag_name)
        self.refresh_qa_flags()
        self._clear_qa_flag_form()
        self._show_themed_message(QMessageBox.Icon.Information, "Saved", f"QA flag '{flag_name}' removed.")

    def _add_admin_user(self) -> None:
        user_id = DepotRules.normalize_user_id(self.admin_user_input.text())
        admin_name = str(self.admin_name_input.text() or "").strip()
        position = str(self.admin_position_input.text() or "").strip()
        location = str(self.admin_location_input.text() or "").strip()
        icon_path = str(self.admin_icon_input.text() or "").strip()
        if not user_id:
            self._show_themed_message(QMessageBox.Icon.Warning, "Validation", "Enter a user id.")
            return
        if not admin_name or not position or not location:
            self._show_themed_message(
                QMessageBox.Icon.Warning,
                "Validation",
                "Admin Number, Name, Position, and Location are required.",
            )
            return
        if icon_path and not Path(icon_path).exists():
            self._show_themed_message(QMessageBox.Icon.Warning, "Validation", "Admin icon path does not exist.")
            return
        stored_icon = self.tracker.add_admin_user(user_id, admin_name, position, location, icon_path)
        self.admin_icon_input.setText(stored_icon)
        self.refresh_admins()
        self._select_admin_item(user_id)
        self._load_admin_into_form(user_id)
        self._show_themed_message(QMessageBox.Icon.Information, "Saved", f"Admin {user_id} updated.")

    def _remove_selected_admin(self) -> None:
        row = self.admin_table.currentRow()
        if row < 0:
            self._show_themed_message(QMessageBox.Icon.Warning, "Validation", "Select an admin user to remove.")
            return
        item = self.admin_table.item(row, 0)
        if item is None:
            return
        user_id = DepotRules.normalize_user_id(str(item.data(Qt.ItemDataRole.UserRole) or item.text() or ""))
        if not user_id:
            return
        if user_id == self.current_user:
            self._show_themed_message(QMessageBox.Icon.Warning, "Validation", "Cannot remove your own admin access from this panel.")
            return
        self.tracker.remove_admin_user(user_id)
        self.refresh_admins()
        self._clear_admin_form()
        self._show_themed_message(QMessageBox.Icon.Information, "Saved", f"Admin {user_id} removed.")


class TouchDistributionBar(QWidget):
    """Simple stacked bar showing filtered submission touch counts by color."""

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._segments: list[tuple[str, int, str]] = []
        self._total = 0
        self.setMinimumHeight(24)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)

    def set_segments(self, segments: list[tuple[str, int, str]]) -> None:
        cleaned: list[tuple[str, int, str]] = []
        total = 0
        for label, count, color in segments:
            safe_count = int(max(0, int(count)))
            if safe_count <= 0:
                continue
            safe_color = normalize_hex(str(color or "#7BDAEA"), "#7BDAEA")
            cleaned.append((str(label), safe_count, safe_color))
            total += safe_count
        self._segments = cleaned
        self._total = total
        self.update()

    def paintEvent(self, event) -> None:  # noqa: N802
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        bounds = self.rect().adjusted(0, 0, -1, -1)
        if bounds.width() <= 2 or bounds.height() <= 2:
            return

        bar_path = QPainterPath()
        bar_path.addRoundedRect(QRectF(bounds), 6.0, 6.0)

        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QColor(12, 18, 24, 110))
        painter.drawPath(bar_path)

        if self._total > 0 and self._segments:
            painter.save()
            painter.setClipPath(bar_path)
            full_width = float(bounds.width())
            x = float(bounds.left())
            remaining_width = full_width
            remaining_total = int(self._total)
            for idx, (_label, count, color_hex) in enumerate(self._segments):
                if idx == len(self._segments) - 1:
                    segment_width = remaining_width
                else:
                    ratio = float(count) / max(1.0, float(remaining_total))
                    segment_width = max(1.0, round(remaining_width * ratio, 2))
                segment_rect = QRectF(x, float(bounds.top()), segment_width, float(bounds.height()))
                painter.setBrush(QColor(color_hex))
                painter.drawRect(segment_rect)
                x += segment_width
                remaining_width -= segment_width
                remaining_total -= count
            painter.restore()

        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.setPen(QPen(QColor(255, 255, 255, 72), 1))
        painter.drawPath(bar_path)


class DepotDashboardDialog(DepotFramelessToolWindow):
    """Tracker data viewer using shared frameless chrome plus filtered submissions metrics."""

    TIMEFRAME_OPTIONS: tuple[tuple[str, str], ...] = (
        ("current_week", "Current Week"),
        ("today", "Today"),
        ("yesterday", "Yesterday"),
        ("last_7_days", "Last 7 Days"),
        ("last_30_days", "Last 30 Days"),
        ("this_month", "This Month"),
        ("all_time", "All Time"),
        ("custom", "Custom"),
    )

    TOUCH_ORDER: tuple[str, ...] = (
        DepotRules.TOUCH_COMPLETE,
        DepotRules.TOUCH_JUNK,
        DepotRules.TOUCH_PART_ORDER,
        DepotRules.TOUCH_RTV,
        "Triaged",
        DepotRules.TOUCH_OTHER,
    )

    TOUCH_COLORS: dict[str, str] = {
        DepotRules.TOUCH_COMPLETE: "#21B46D",
        DepotRules.TOUCH_JUNK: "#D95A5A",
        DepotRules.TOUCH_PART_ORDER: "#D3A327",
        DepotRules.TOUCH_RTV: "#4F86D9",
        "Triaged": "#20AFA8",
        DepotRules.TOUCH_OTHER: "#8E97A8",
    }

    def __init__(self, app_window: "QuickInputsWindow") -> None:
        super().__init__(
            app_window,
            window_title="Data Dashboard",
            theme_kind="dashboard",
            size=(1040, 620),
            minimum_size=(860, 500),
        )
        self.app_window = app_window
        self._date_sync_in_progress = False
        body = QWidget(self)
        self.root_layout.addWidget(body, 1)
        layout = QVBoxLayout(body)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(UI_SPACING_FRAMELESS_TOOL)

        subtitle = QLabel("Live table view from Tracker Hub data.")
        subtitle.setProperty("muted", True)
        subtitle.setWordWrap(True)
        layout.addWidget(subtitle)

        controls = QHBoxLayout()
        controls.setContentsMargins(0, 0, 0, 0)
        controls.setSpacing(6)
        controls.addWidget(QLabel("Table:"), 0)
        self.table_combo = QComboBox()
        for table_name, label_text in TRACKER_DASHBOARD_TABLES:
            self.table_combo.addItem(label_text, table_name)
        self.table_combo.setMaxVisibleItems(max(8, len(TRACKER_DASHBOARD_TABLES)))
        self.table_combo.setSizeAdjustPolicy(QComboBox.SizeAdjustPolicy.AdjustToContentsOnFirstShow)
        self.table_combo.setMinimumWidth(220)
        self.table_combo.setMaximumWidth(340)
        controls.addWidget(self.table_combo, 0)
        controls.addSpacing(8)
        controls.addWidget(QLabel("Rows:"), 0)
        self.limit_spin = QSpinBox()
        self.limit_spin.setRange(50, 5000)
        self.limit_spin.setSingleStep(50)
        self.limit_spin.setValue(300)
        self.limit_spin.setMinimumWidth(110)
        controls.addWidget(self.limit_spin, 0)
        controls.addStretch(1)
        self.refresh_btn = QPushButton("Refresh")
        self.refresh_btn.setProperty("actionRole", "pick")
        self.export_btn = QPushButton("Export CSV")
        self.export_btn.setProperty("actionRole", "save")
        controls.addWidget(self.refresh_btn, 0)
        controls.addWidget(self.export_btn, 0)
        layout.addLayout(controls)

        self.submission_filters_wrap = QWidget(self)
        submission_filters = QHBoxLayout(self.submission_filters_wrap)
        submission_filters.setContentsMargins(0, 0, 0, 0)
        submission_filters.setSpacing(6)
        submission_filters.addWidget(QLabel("Range:"), 0)
        self.timeframe_combo = QComboBox()
        for key, label in self.TIMEFRAME_OPTIONS:
            self.timeframe_combo.addItem(label, key)
        submission_filters.addWidget(self.timeframe_combo, 0)
        submission_filters.addWidget(QLabel("From:"), 0)
        self.start_date_edit = QDateEdit()
        self.start_date_edit.setCalendarPopup(True)
        self.start_date_edit.setDisplayFormat("yyyy-MM-dd")
        submission_filters.addWidget(self.start_date_edit, 0)
        submission_filters.addWidget(QLabel("To:"), 0)
        self.end_date_edit = QDateEdit()
        self.end_date_edit.setCalendarPopup(True)
        self.end_date_edit.setDisplayFormat("yyyy-MM-dd")
        submission_filters.addWidget(self.end_date_edit, 0)
        submission_filters.addWidget(QLabel("User:"), 0)
        self.user_filter_combo = QComboBox()
        self.user_filter_combo.addItem("All Users", "")
        self.user_filter_combo.setMinimumWidth(160)
        submission_filters.addWidget(self.user_filter_combo, 0)
        submission_filters.addStretch(1)
        layout.addWidget(self.submission_filters_wrap)

        self.touch_summary_label = QLabel("")
        self.touch_summary_label.setProperty("muted", True)
        self.touch_summary_label.setWordWrap(True)
        layout.addWidget(self.touch_summary_label)

        self.touch_bar = TouchDistributionBar(self)
        layout.addWidget(self.touch_bar)

        self.touch_legend_label = QLabel("")
        self.touch_legend_label.setWordWrap(True)
        layout.addWidget(self.touch_legend_label)

        self.empty_hint = QLabel("")
        self.empty_hint.setProperty("muted", True)
        self.empty_hint.setWordWrap(True)
        self.empty_hint.hide()
        layout.addWidget(self.empty_hint)

        self.results_tabs = QTabWidget(self)
        self.list_tab = QWidget(self.results_tabs)
        self.table_tab = QWidget(self.results_tabs)
        self.notes_tab = QWidget(self.results_tabs)
        self.results_tabs.addTab(self.list_tab, "List")
        self.results_tabs.addTab(self.table_tab, "Table")
        self.results_tabs.addTab(self.notes_tab, "Notes")
        layout.addWidget(self.results_tabs, 1)

        list_layout = QVBoxLayout(self.list_tab)
        list_layout.setContentsMargins(0, 0, 0, 0)
        list_layout.setSpacing(4)
        self.table = QTableWidget()
        configure_standard_table(self.table, [])
        list_layout.addWidget(self.table, 1)

        table_layout = QVBoxLayout(self.table_tab)
        table_layout.setContentsMargins(0, 0, 0, 0)
        table_layout.setSpacing(4)
        self.table_placeholder_label = QLabel("Chart placeholder: submissions chart options will appear here.")
        self.table_placeholder_label.setWordWrap(True)
        self.table_placeholder_label.setProperty("muted", True)
        self.table_placeholder_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        table_layout.addWidget(self.table_placeholder_label, 1)

        notes_layout = QVBoxLayout(self.notes_tab)
        notes_layout.setContentsMargins(0, 0, 0, 0)
        notes_layout.setSpacing(4)

        notes_intro = QLabel("Testing utility: edit note fields on existing rows.")
        notes_intro.setWordWrap(True)
        notes_intro.setProperty("muted", True)
        notes_layout.addWidget(notes_intro)

        notes_controls = QHBoxLayout()
        notes_controls.setContentsMargins(0, 0, 0, 0)
        notes_controls.setSpacing(6)
        notes_controls.addWidget(QLabel("Field:"), 0)
        self.notes_target_combo = QComboBox()
        self.notes_target_combo.setMinimumWidth(240)
        notes_controls.addWidget(self.notes_target_combo, 0)
        notes_controls.addWidget(QLabel("Rows:"), 0)
        self.notes_limit_spin = QSpinBox()
        self.notes_limit_spin.setRange(25, 5000)
        self.notes_limit_spin.setSingleStep(25)
        self.notes_limit_spin.setValue(200)
        self.notes_limit_spin.setMinimumWidth(100)
        notes_controls.addWidget(self.notes_limit_spin, 0)
        notes_controls.addWidget(QLabel("Work Order:"), 0)
        self.notes_work_order_filter = QLineEdit()
        self.notes_work_order_filter.setPlaceholderText("Optional filter")
        self.notes_work_order_filter.setMinimumWidth(180)
        notes_controls.addWidget(self.notes_work_order_filter, 0)
        self.notes_refresh_btn = QPushButton("Load")
        self.notes_refresh_btn.setProperty("actionRole", "pick")
        notes_controls.addWidget(self.notes_refresh_btn, 0)
        notes_controls.addStretch(1)
        notes_layout.addLayout(notes_controls)

        self.notes_table = QTableWidget()
        notes_headers = ["id", "created_at", "user_id", "work_order", "note_preview"]
        note_resize_modes = {
            0: QHeaderView.ResizeMode.ResizeToContents,
            1: QHeaderView.ResizeMode.ResizeToContents,
            2: QHeaderView.ResizeMode.ResizeToContents,
            3: QHeaderView.ResizeMode.ResizeToContents,
            4: QHeaderView.ResizeMode.Stretch,
        }
        configure_standard_table(self.notes_table, notes_headers, resize_modes=note_resize_modes, stretch_last=True)
        notes_layout.addWidget(self.notes_table, 1)

        self.notes_selected_label = QLabel("Select a row to edit.")
        self.notes_selected_label.setWordWrap(True)
        self.notes_selected_label.setProperty("muted", True)
        notes_layout.addWidget(self.notes_selected_label)

        self.notes_editor = QTextEdit()
        self.notes_editor.setPlaceholderText("Selected note text...")
        self.notes_editor.setMinimumHeight(110)
        self.notes_editor.setEnabled(False)
        notes_layout.addWidget(self.notes_editor)

        notes_actions = QHBoxLayout()
        notes_actions.setContentsMargins(0, 0, 0, 0)
        notes_actions.setSpacing(6)
        self.notes_save_btn = QPushButton("Save Note")
        self.notes_save_btn.setProperty("actionRole", "save")
        self.notes_save_btn.setEnabled(False)
        notes_actions.addWidget(self.notes_save_btn, 0)
        self.notes_status_label = QLabel("")
        self.notes_status_label.setWordWrap(True)
        self.notes_status_label.setProperty("muted", True)
        notes_actions.addWidget(self.notes_status_label, 1)
        notes_layout.addLayout(notes_actions)

        self._notes_selected_row_id: int | None = None

        self.table_combo.currentIndexChanged.connect(self._on_table_changed)
        self.limit_spin.valueChanged.connect(self.refresh_dashboard)
        self.refresh_btn.clicked.connect(self.refresh_dashboard)
        self.export_btn.clicked.connect(self.export_csv)
        self.timeframe_combo.currentIndexChanged.connect(self._on_timeframe_changed)
        self.start_date_edit.dateChanged.connect(self._on_custom_date_changed)
        self.end_date_edit.dateChanged.connect(self._on_custom_date_changed)
        self.user_filter_combo.currentIndexChanged.connect(self.refresh_dashboard)
        self.notes_target_combo.currentIndexChanged.connect(self.refresh_notes_rows)
        self.notes_limit_spin.valueChanged.connect(self.refresh_notes_rows)
        self.notes_work_order_filter.returnPressed.connect(self.refresh_notes_rows)
        self.notes_refresh_btn.clicked.connect(self.refresh_notes_rows)
        self.notes_table.itemSelectionChanged.connect(self._on_notes_selection_changed)
        self.notes_save_btn.clicked.connect(self._save_selected_note)

        self._set_timeframe_key("current_week")
        self._populate_submission_user_filter()
        self._populate_notes_targets()
        self.apply_theme_styles()
        self.refresh_combo_popup_width()
        self.refresh_dashboard()
        self.refresh_notes_rows()

    def closeEvent(self, event) -> None:  # noqa: N802
        self.app_window.config.setdefault("popup_positions", {})["depot_dashboard"] = {
            "x": int(self.x()),
            "y": int(self.y()),
        }
        self.app_window.queue_save_config()
        super().closeEvent(event)

    def apply_theme_styles(self) -> None:
        if self.app_window is None:
            return
        super().apply_theme_styles()
        muted = rgba_css(self.app_window.palette_data["label_text"], 0.80)
        self.setStyleSheet(
            self.styleSheet()
            + (
                "QLabel[muted='true'] {"
                f"color: {muted};"
                "font-weight: 700;"
                "}"
                "QLabel[section='true'] {"
                "font-size: 14px;"
                "font-weight: 800;"
                "}"
            )
        )

    def _on_table_changed(self) -> None:
        self.refresh_combo_popup_width()
        self.refresh_dashboard()

    def _on_timeframe_changed(self) -> None:
        key = str(self.timeframe_combo.currentData() or "").strip()
        if key:
            self._set_timeframe_key(key)
        self.refresh_dashboard()

    def _on_custom_date_changed(self) -> None:
        if self._date_sync_in_progress:
            return
        self._date_sync_in_progress = True
        try:
            if self.start_date_edit.date() > self.end_date_edit.date():
                self.end_date_edit.setDate(self.start_date_edit.date())
            if str(self.timeframe_combo.currentData() or "") != "custom":
                idx = self.timeframe_combo.findData("custom")
                if idx >= 0:
                    self.timeframe_combo.setCurrentIndex(idx)
        finally:
            self._date_sync_in_progress = False
        self.refresh_dashboard()

    def _set_timeframe_key(self, key: str) -> None:
        today = QDate.currentDate()
        start = today
        end = today
        if key == "current_week":
            start = today.addDays(1 - int(today.dayOfWeek()))
            end = start.addDays(6)
        elif key == "today":
            start = today
            end = today
        elif key == "yesterday":
            start = today.addDays(-1)
            end = start
        elif key == "last_7_days":
            end = today
            start = today.addDays(-6)
        elif key == "last_30_days":
            end = today
            start = today.addDays(-29)
        elif key == "this_month":
            start = QDate(today.year(), today.month(), 1)
            end = start.addMonths(1).addDays(-1)
        elif key == "all_time":
            start = today.addDays(-3650)
            end = today
        elif key == "custom":
            return
        self._date_sync_in_progress = True
        try:
            self.start_date_edit.setDate(start)
            self.end_date_edit.setDate(end)
        finally:
            self._date_sync_in_progress = False

    def _current_submission_filters(self) -> tuple[str | None, str | None, str | None]:
        key = str(self.timeframe_combo.currentData() or "current_week").strip()
        start_date: str | None = None
        end_date: str | None = None
        if key != "all_time":
            start_qdate = self.start_date_edit.date()
            end_qdate = self.end_date_edit.date()
            if start_qdate > end_qdate:
                start_qdate, end_qdate = end_qdate, start_qdate
            start_date = start_qdate.toString("yyyy-MM-dd")
            end_date = end_qdate.toString("yyyy-MM-dd")
        selected_user = str(self.user_filter_combo.currentData() or "").strip()
        return start_date, end_date, (selected_user if selected_user else None)

    def _populate_submission_user_filter(self) -> None:
        selected_user = str(self.user_filter_combo.currentData() or "").strip()
        users: list[str] = []
        try:
            rows = self.app_window.depot_db.fetchall(
                "SELECT DISTINCT user_id FROM submissions WHERE COALESCE(user_id, '') <> '' ORDER BY user_id ASC"
            )
            users = [str(row["user_id"]).strip() for row in rows if str(row["user_id"]).strip()]
        except Exception as exc:
            _runtime_log_event(
                "ui.depot_dashboard_user_filter_query_failed",
                severity="warning",
                summary="Dashboard could not refresh submission user filter options.",
                exc=exc,
            )

        self.user_filter_combo.blockSignals(True)
        self.user_filter_combo.clear()
        self.user_filter_combo.addItem("All Users", "")
        for user_id in users:
            self.user_filter_combo.addItem(user_id, user_id)
        if selected_user:
            idx = self.user_filter_combo.findData(selected_user)
            if idx >= 0:
                self.user_filter_combo.setCurrentIndex(idx)
        self.user_filter_combo.blockSignals(False)

    @staticmethod
    def _note_preview_text(value: str, max_len: int = 120) -> str:
        compact = " ".join(str(value or "").replace("\r", "\n").splitlines()).strip()
        if len(compact) <= int(max_len):
            return compact
        safe_max = max(8, int(max_len))
        return compact[: safe_max - 3].rstrip() + "..."

    def _populate_notes_targets(self) -> None:
        selected_key = str(self.notes_target_combo.currentData() or "").strip()
        options: list[tuple[str, str]] = []
        try:
            options = self.app_window.depot_tracker.dashboard_note_target_options()
        except Exception as exc:
            _runtime_log_event(
                "ui.depot_dashboard_notes_target_load_failed",
                severity="warning",
                summary="Dashboard notes editor failed loading editable target list.",
                exc=exc,
            )

        self.notes_target_combo.blockSignals(True)
        self.notes_target_combo.clear()
        for key, label in options:
            self.notes_target_combo.addItem(str(label), str(key))
        if selected_key:
            idx = self.notes_target_combo.findData(selected_key)
            if idx >= 0:
                self.notes_target_combo.setCurrentIndex(idx)
        self.notes_target_combo.blockSignals(False)
        self.notes_refresh_btn.setEnabled(bool(options))
        self.notes_limit_spin.setEnabled(bool(options))
        self.notes_work_order_filter.setEnabled(bool(options))
        self.notes_save_btn.setEnabled(False)

    def refresh_notes_rows(self) -> None:
        headers = ["id", "created_at", "user_id", "work_order", "note_preview"]
        resize_modes = {
            0: QHeaderView.ResizeMode.ResizeToContents,
            1: QHeaderView.ResizeMode.ResizeToContents,
            2: QHeaderView.ResizeMode.ResizeToContents,
            3: QHeaderView.ResizeMode.ResizeToContents,
            4: QHeaderView.ResizeMode.Stretch,
        }
        configure_standard_table(self.notes_table, headers, resize_modes=resize_modes, stretch_last=True)

        target_key = str(self.notes_target_combo.currentData() or "").strip()
        if not target_key:
            self.notes_table.setRowCount(0)
            self._notes_selected_row_id = None
            self.notes_editor.clear()
            self.notes_editor.setEnabled(False)
            self.notes_save_btn.setEnabled(False)
            self.notes_selected_label.setText("No editable note fields are currently available.")
            self.notes_status_label.setText("")
            return

        rows_limit = int(self.notes_limit_spin.value())
        work_order_filter = str(self.notes_work_order_filter.text() or "").strip()
        app = QApplication.instance()
        if app is not None:
            app.setOverrideCursor(Qt.CursorShape.WaitCursor)
        try:
            try:
                rows = self.app_window.depot_tracker.fetch_dashboard_note_rows(
                    target_key,
                    limit=rows_limit,
                    work_order_filter=work_order_filter,
                )
            except Exception as exc:
                _runtime_log_event(
                    "ui.depot_dashboard_notes_query_failed",
                    severity="error",
                    summary="Dashboard notes editor query failed.",
                    exc=exc,
                    context={"target_key": target_key, "limit": rows_limit, "work_order_filter": work_order_filter},
                )
                self._show_themed_message(
                    QMessageBox.Icon.Warning,
                    "Notes load failed",
                    f"Could not load note rows:\n{type(exc).__name__}: {exc}",
                )
                self.notes_table.setRowCount(0)
                self._notes_selected_row_id = None
                self.notes_editor.clear()
                self.notes_editor.setEnabled(False)
                self.notes_save_btn.setEnabled(False)
                self.notes_selected_label.setText("Could not load note rows. Details were logged for support.")
                self.notes_status_label.setText("")
                return
        finally:
            if app is not None:
                app.restoreOverrideCursor()

        self.notes_table.setRowCount(len(rows))
        for row_idx, row in enumerate(rows):
            row_id = int(max(0, safe_int(row.get("id"), 0)))
            created_text = self._normalize_dashboard_datetime(str(row.get("created_at", "") or ""))
            user_text = str(row.get("user_id", "") or "").strip()
            work_order_text = str(row.get("work_order", "") or "").strip()
            note_text = str(row.get("note_text", "") or "").strip()
            preview_text = self._note_preview_text(note_text)

            id_item = QTableWidgetItem(str(row_id))
            id_item.setTextAlignment(int(Qt.AlignmentFlag.AlignCenter))
            id_item.setData(Qt.ItemDataRole.UserRole, note_text)
            self.notes_table.setItem(row_idx, 0, id_item)

            created_item = QTableWidgetItem(created_text)
            created_item.setTextAlignment(int(Qt.AlignmentFlag.AlignCenter))
            self.notes_table.setItem(row_idx, 1, created_item)

            user_item = QTableWidgetItem(user_text)
            user_item.setTextAlignment(int(Qt.AlignmentFlag.AlignCenter))
            self.notes_table.setItem(row_idx, 2, user_item)

            work_order_item = QTableWidgetItem(work_order_text)
            work_order_item.setTextAlignment(int(Qt.AlignmentFlag.AlignCenter))
            self.notes_table.setItem(row_idx, 3, work_order_item)

            preview_item = QTableWidgetItem(preview_text)
            preview_item.setToolTip(note_text if note_text else "(empty)")
            self.notes_table.setItem(row_idx, 4, preview_item)

        self.notes_table.clearSelection()
        self._notes_selected_row_id = None
        self.notes_editor.clear()
        self.notes_editor.setEnabled(False)
        self.notes_save_btn.setEnabled(False)
        self.notes_selected_label.setText("Select a row to edit.")
        self.notes_status_label.setText(f"Loaded {len(rows)} row(s).")

    def _on_notes_selection_changed(self) -> None:
        selected_rows = self.notes_table.selectionModel().selectedRows() if self.notes_table.selectionModel() is not None else []
        if not selected_rows:
            self._notes_selected_row_id = None
            self.notes_editor.clear()
            self.notes_editor.setEnabled(False)
            self.notes_save_btn.setEnabled(False)
            self.notes_selected_label.setText("Select a row to edit.")
            return

        row_idx = int(selected_rows[0].row())
        id_item = self.notes_table.item(row_idx, 0)
        if id_item is None:
            self._notes_selected_row_id = None
            self.notes_editor.clear()
            self.notes_editor.setEnabled(False)
            self.notes_save_btn.setEnabled(False)
            self.notes_selected_label.setText("Select a row to edit.")
            return

        row_id = int(max(0, safe_int(id_item.text(), 0)))
        note_text = str(id_item.data(Qt.ItemDataRole.UserRole) or "")
        work_order_item = self.notes_table.item(row_idx, 3)
        user_item = self.notes_table.item(row_idx, 2)
        work_order_text = work_order_item.text().strip() if work_order_item is not None else ""
        user_text = user_item.text().strip() if user_item is not None else ""

        self._notes_selected_row_id = row_id if row_id > 0 else None
        self.notes_editor.setEnabled(self._notes_selected_row_id is not None)
        self.notes_editor.setPlainText(note_text)
        self.notes_save_btn.setEnabled(self._notes_selected_row_id is not None)
        if self._notes_selected_row_id is not None:
            self.notes_selected_label.setText(
                f"Editing row #{self._notes_selected_row_id} | Work Order: {work_order_text or '(none)'} | User: {user_text or '(none)'}"
            )
        else:
            self.notes_selected_label.setText("Select a row to edit.")
        self.notes_status_label.setText("")

    def _save_selected_note(self) -> None:
        target_key = str(self.notes_target_combo.currentData() or "").strip()
        row_id = self._notes_selected_row_id
        if not target_key or row_id is None:
            self._show_themed_message(QMessageBox.Icon.Warning, "Validation", "Select a note row first.")
            return

        note_text = str(self.notes_editor.toPlainText() or "").strip()
        try:
            self.app_window.depot_tracker.update_dashboard_note_value(target_key, row_id, note_text)
        except Exception as exc:
            _runtime_log_event(
                "ui.depot_dashboard_notes_save_failed",
                severity="error",
                summary="Dashboard notes editor failed saving a note field.",
                exc=exc,
                context={"target_key": target_key, "row_id": row_id},
            )
            self._show_themed_message(
                QMessageBox.Icon.Warning,
                "Save failed",
                f"Could not save note:\n{type(exc).__name__}: {exc}",
            )
            return

        selected_row = self.notes_table.currentRow()
        if selected_row >= 0:
            id_item = self.notes_table.item(selected_row, 0)
            preview_item = self.notes_table.item(selected_row, 4)
            if id_item is not None:
                id_item.setData(Qt.ItemDataRole.UserRole, note_text)
            if preview_item is not None:
                preview_item.setText(self._note_preview_text(note_text))
                preview_item.setToolTip(note_text if note_text else "(empty)")
        self.notes_status_label.setText(f"Saved row #{row_id} at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.")

    def _touch_color(self, touch: str) -> str:
        normalized = str(touch or "").strip()
        return normalize_hex(self.TOUCH_COLORS.get(normalized, "#6F7C91"), "#6F7C91")

    def _refresh_touch_distribution(self, start_date: str | None, end_date: str | None, user_id: str | None) -> None:
        try:
            metrics = self.app_window.depot_tracker.get_dashboard_metrics(
                start_date=start_date,
                end_date=end_date,
                user_id=user_id,
            )
        except Exception as exc:
            _runtime_log_event(
                "ui.depot_dashboard_metrics_query_failed",
                severity="error",
                summary="Dashboard metrics query failed.",
                exc=exc,
                context={"start_date": start_date, "end_date": end_date, "user_id": user_id},
            )
            self.touch_bar.set_segments([])
            self.touch_summary_label.setText("Touch metrics unavailable. Details were logged for support.")
            self.touch_legend_label.setText("")
            return

        by_touch_raw = metrics.get("by_touch", {})
        by_touch: dict[str, int] = {}
        if isinstance(by_touch_raw, dict):
            for key, value in by_touch_raw.items():
                touch_name = str(key or "").strip()
                if not touch_name:
                    continue
                by_touch[touch_name] = int(max(0, safe_int(value, 0)))

        ordered_keys: list[str] = []
        for key in self.TOUCH_ORDER:
            if key in by_touch:
                ordered_keys.append(key)
        for key in sorted(by_touch.keys()):
            if key not in ordered_keys:
                ordered_keys.append(key)

        segments = [
            (touch, int(by_touch.get(touch, 0)), self._touch_color(touch))
            for touch in ordered_keys
            if int(by_touch.get(touch, 0)) > 0
        ]
        self.touch_bar.set_segments(segments)

        total_submissions = int(max(0, safe_int(metrics.get("total_submissions", 0), 0)))
        date_label = "All Time" if start_date is None or end_date is None else f"{start_date} to {end_date}"
        user_label = user_id if user_id else "All Users"
        self.touch_summary_label.setText(
            f"Filtered submissions: {total_submissions} | Range: {date_label} | User: {user_label}"
        )
        if segments:
            legend_chunks = [
                f"<span style='color:{color}; font-weight:700'>{DepotRules.chart_touch_label(touch)}</span>: {count}"
                for touch, count, color in segments
            ]
            self.touch_legend_label.setText(" | ".join(legend_chunks))
            self.touch_legend_label.setTextFormat(Qt.TextFormat.RichText)
        else:
            self.touch_legend_label.setText("No touch activity in the selected filter range.")
            self.touch_legend_label.setTextFormat(Qt.TextFormat.PlainText)

    @staticmethod
    def _normalize_dashboard_datetime(raw_value: Any) -> str:
        text = str(raw_value or "").strip()
        if not text:
            return ""
        try:
            parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
            if parsed.tzinfo is not None:
                try:
                    parsed = parsed.astimezone()
                except Exception:
                    pass
            return parsed.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            pass
        return text.replace("T", " ")

    def _format_dashboard_cell_text(self, col_name: str, raw_value: Any) -> str:
        if raw_value is None:
            return ""
        text = str(raw_value).strip()
        if not text:
            return ""
        normalized_name = str(col_name or "").strip().lower()
        if normalized_name.endswith("_at"):
            return self._normalize_dashboard_datetime(text)
        if normalized_name.endswith("_date") and "T" in text:
            normalized = self._normalize_dashboard_datetime(text)
            return normalized[:10] if normalized else text
        return text

    def _refresh_table_placeholder(
        self,
        table_name: str,
        row_count: int,
        start_date: str | None,
        end_date: str | None,
        user_id: str | None,
    ) -> None:
        if not hasattr(self, "table_placeholder_label"):
            return
        if table_name != "submissions":
            self.table_placeholder_label.setText(
                f"Chart placeholder: chart view is currently reserved for Submissions.\n"
                f"Current source: {table_name} | Rows loaded: {int(row_count)}"
            )
            return
        range_label = "All Time" if start_date is None or end_date is None else f"{start_date} to {end_date}"
        user_label = user_id if user_id else "All Users"
        self.table_placeholder_label.setText(
            "Chart placeholder: submissions trend view will render here.\n"
            f"Range: {range_label} | User: {user_label} | Rows loaded: {int(row_count)}"
        )

    def refresh_combo_popup_width(self) -> None:
        combos: list[QComboBox] = [self.table_combo, self.timeframe_combo, self.user_filter_combo]
        if hasattr(self, "notes_target_combo"):
            combos.append(self.notes_target_combo)
        for combo in combos:
            item_count = int(combo.count())
            if item_count <= 0:
                continue
            fm = combo.fontMetrics()
            widest_text = 0
            for idx in range(item_count):
                widest_text = max(widest_text, fm.horizontalAdvance(combo.itemText(idx)))
            popup_width = int(max(220, widest_text + 56))
            view = combo.view()
            if view is None:
                continue
            view.setMinimumWidth(popup_width)
            view.setTextElideMode(Qt.TextElideMode.ElideNone)
            view.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)

    def refresh_dashboard(self) -> None:
        table_name = str(self.table_combo.currentData() or "").strip()
        if not table_name:
            return

        allowed_tables = {name for name, _label in TRACKER_DASHBOARD_TABLES}
        if table_name not in allowed_tables:
            return

        submissions_mode = table_name == "submissions"
        self.submission_filters_wrap.setVisible(submissions_mode)
        self.touch_summary_label.setVisible(submissions_mode)
        self.touch_bar.setVisible(submissions_mode)
        self.touch_legend_label.setVisible(submissions_mode)
        if submissions_mode:
            self._populate_submission_user_filter()

        limit = int(self.limit_spin.value())
        query = ""
        params: list[Any] = []
        start_date: str | None = None
        end_date: str | None = None
        user_id: str | None = None

        if submissions_mode:
            start_date, end_date, user_id = self._current_submission_filters()
            where_parts: list[str] = []
            if start_date:
                where_parts.append("entry_date >= ?")
                params.append(start_date)
            if end_date:
                where_parts.append("entry_date <= ?")
                params.append(end_date)
            if user_id:
                where_parts.append("user_id = ?")
                params.append(user_id)
            where_clause = f" WHERE {' AND '.join(where_parts)}" if where_parts else ""
            query = (
                "SELECT id, created_at, user_id, work_order, touch, client_unit, entry_date, "
                "CASE "
                "WHEN touch='Part Order' THEN "
                "SUM(CASE WHEN touch='Part Order' THEN 1 ELSE 0 END) OVER ("
                "PARTITION BY user_id, work_order ORDER BY created_at, id "
                "ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW"
                ") "
                "ELSE 0 "
                "END AS part_order_count "
                f"FROM submissions{where_clause} ORDER BY created_at DESC, id DESC LIMIT ?"
            )
            params.append(limit)
        else:
            order_clause = " ORDER BY id DESC"
            if table_name in {"parts", "rtvs", "client_jo", "client_parts"}:
                order_clause = " ORDER BY created_at DESC"
            query = f"SELECT * FROM {table_name}{order_clause} LIMIT ?"
            params.append(limit)

        app = QApplication.instance()
        if app is not None:
            app.setOverrideCursor(Qt.CursorShape.WaitCursor)
        try:
            try:
                rows = self.app_window.depot_db.fetchall(query, tuple(params))
            except Exception as exc:
                _runtime_log_event(
                    "ui.depot_dashboard_query_failed",
                    severity="error",
                    summary="Dashboard table query failed.",
                    exc=exc,
                    context={"table": table_name, "limit": limit, "start_date": start_date, "end_date": end_date, "user_id": user_id},
                )
                self._show_themed_message(
                    QMessageBox.Icon.Critical,
                    "Data load failed",
                    f"Could not load table data:\n{type(exc).__name__}: {exc}",
                )
                self.empty_hint.setText("Could not load data. Details were logged for support.")
                self.empty_hint.show()
                configure_standard_table(self.table, [], stretch_last=True)
                self.table.setRowCount(0)
                self._refresh_table_placeholder(table_name, 0, start_date, end_date, user_id)
                return

            headers: list[str]
            if rows:
                headers = [str(name) for name in rows[0].keys()]
            else:
                try:
                    info_rows = self.app_window.depot_db.fetchall(f"PRAGMA table_info({table_name})")
                except Exception as exc:
                    _runtime_log_event(
                        "ui.depot_dashboard_schema_introspect_failed",
                        severity="warning",
                        summary="Dashboard could not read table schema for empty result set.",
                        exc=exc,
                        context={"table": table_name},
                    )
                    info_rows = []
                headers = [str(r["name"]) for r in info_rows] if info_rows else []

            resize_modes: dict[int, QHeaderView.ResizeMode] = {}
            if headers:
                for idx in range(len(headers)):
                    resize_modes[idx] = QHeaderView.ResizeMode.ResizeToContents
                resize_modes[len(headers) - 1] = QHeaderView.ResizeMode.Stretch
            configure_standard_table(self.table, headers, resize_modes=resize_modes, stretch_last=True)

            self.table.setRowCount(len(rows))
            for row_idx, row in enumerate(rows):
                for col_idx, col_name in enumerate(headers):
                    raw_value = row[col_name]
                    text = self._format_dashboard_cell_text(col_name, raw_value)
                    item = QTableWidgetItem(text)
                    item.setToolTip(text)
                    self.table.setItem(row_idx, col_idx, item)
            self._refresh_table_placeholder(table_name, len(rows), start_date, end_date, user_id)

            if not rows:
                self.empty_hint.setText("No rows in this table for the current row limit/filter.")
                self.empty_hint.show()
            else:
                self.empty_hint.hide()
        finally:
            if app is not None:
                app.restoreOverrideCursor()

        if submissions_mode:
            self._refresh_touch_distribution(start_date, end_date, user_id)
        else:
            self.touch_bar.set_segments([])
            self.touch_summary_label.setText("")
            self.touch_legend_label.setText("")

    def export_csv(self) -> None:
        table = self.table
        if table.columnCount() <= 0:
            self._show_themed_message(QMessageBox.Icon.Information, "Export CSV", "No data to export.")
            return

        table_name = str(self.table_combo.currentData() or "table").strip()
        stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        default_name = f"{table_name}_{stamp}.csv"
        start_dir = self.app_window.config_path.parent if self.app_window.config_path.parent.exists() else Path.home()
        out_path, _ = show_flowgrid_themed_save_file_name(
            self,
            self.app_window,
            "dashboard",
            "Export Dashboard Table",
            str(start_dir / default_name),
            "CSV Files (*.csv);;All Files (*.*)",
        )
        if not out_path:
            return

        headers: list[str] = []
        for col in range(table.columnCount()):
            hdr = table.horizontalHeaderItem(col)
            headers.append(hdr.text() if hdr is not None else f"col_{col}")

        try:
            with open(out_path, "w", newline="", encoding="utf-8-sig") as handle:
                writer = csv.writer(handle)
                writer.writerow(headers)
                for row in range(table.rowCount()):
                    values: list[str] = []
                    for col in range(table.columnCount()):
                        item = table.item(row, col)
                        values.append(item.text() if item is not None else "")
                    writer.writerow(values)
            self._show_themed_message(QMessageBox.Icon.Information, "Export CSV", f"Exported:\n{out_path}")
        except Exception as exc:
            _runtime_log_event(
                "ui.depot_dashboard_export_csv_failed",
                severity="error",
                summary="Dashboard CSV export failed.",
                exc=exc,
                context={"path": str(out_path), "table": table_name},
            )
            self._show_themed_message(
                QMessageBox.Icon.Warning,
                "Export CSV",
                f"Failed to export CSV:\n{type(exc).__name__}: {exc}\n\nDetails were logged for support.",
            )


# Depot specific page will be added into QuickInputsWindow nav
class QuickInputsWindow(QMainWindow):

    def __init__(self) -> None:
        super().__init__()
        self.config_path = _data_file_path(CONFIG_FILENAME)
        self.config: dict[str, Any] = self.load_config()
        self.current_user = DepotRules.normalize_user_id(self.config.get("current_user", detect_current_user_id()))
        self.config["current_user"] = self.current_user
        self._ensure_ui_icon_assets()

        self.palette_data = compute_palette(self.config.get("theme", {}))
        self._pixmap_cache: dict[str, QPixmap] = {}
        self._background_dirty = True
        self._background_cache: dict[tuple[int, int], QPixmap] = {}
        self.image_dialog: ImageLayersDialog | None = None
        self.quick_layout_dialog: QuickLayoutDialog | None = None
        self.quick_radial_menu: QuickRadialMenu | None = None
        self.quick_tabs_widget: QTabWidget | None = None
        self.quick_tab_scrolls: list[QScrollArea] = []
        self.quick_tab_canvases: list[QuickButtonCanvas] = []
        self.last_external_hwnd: int | None = None
        self._saving_timer = QTimer(self)
        self._saving_timer.setInterval(220)
        self._saving_timer.setSingleShot(True)
        self._saving_timer.timeout.connect(self.save_config)
        self._hover_inside = False
        self._hover_revealed = False
        self._hover_delay_timer = QTimer(self)
        self._hover_delay_timer.setSingleShot(True)
        self._hover_delay_timer.timeout.connect(self._on_hover_delay_elapsed)
        self._popup_leave_timer = QTimer(self)
        self._popup_leave_timer.setSingleShot(True)
        self._popup_leave_timer.timeout.connect(self._on_popup_leave_check)
        self._ui_opacity_current = 1.0
        self._ui_opacity_anim = QVariantAnimation(self)
        self._ui_opacity_anim.setEasingCurve(QEasingCurve.Type.InOutQuad)
        self._ui_opacity_anim.valueChanged.connect(lambda value: self._set_ui_opacity(float(value)))
        self._ui_opacity_effects: list[QGraphicsOpacityEffect] = []
        self._corner_radius = 14
        self._drag_offset: QPoint | None = None

        self.setWindowTitle(APP_TITLE)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground, True)
        self.setWindowFlag(Qt.WindowType.FramelessWindowHint, True)
        self.setWindowFlag(Qt.WindowType.Window, True)
        self.setFixedSize(LAUNCH_WIDTH, LAUNCH_HEIGHT)
        self._apply_window_mask()

        self.surface = BackgroundCanvas(self)
        self.setCentralWidget(self.surface)

        self.main_layout = QVBoxLayout(self.surface)
        self.main_layout.setContentsMargins(0, 0, 0, 0)
        self.main_layout.setSpacing(0)

        self.titlebar = TitleBar(self)
        self.main_layout.addWidget(self.titlebar)

        self.body = QWidget()
        self.body_layout = QHBoxLayout(self.body)
        self.body_layout.setContentsMargins(6, 6, 6, 6)
        self.body_layout.setSpacing(6)
        self.main_layout.addWidget(self.body, 1)

        self.sidebar = QWidget()
        self.sidebar.setFixedWidth(SIDEBAR_WIDTH)
        self.sidebar_layout = QVBoxLayout(self.sidebar)
        self.sidebar_layout.setContentsMargins(4, 4, 4, 4)
        self.sidebar_layout.setSpacing(4)

        self.nav_buttons: dict[str, QToolButton] = {}
        self.nav_buttons["quick"] = self._make_nav_button(
            standard_icon_name="SP_MediaPlay",
            icon_filename="grid.png",
            icon_px=30,
        )
        self.nav_buttons["depot"] = self._make_nav_button(standard_icon_name="SP_DirHomeIcon", icon_px=30)

        self.nav_buttons["quick"].clicked.connect(lambda: self.switch_page("quick"))
        self.nav_buttons["depot"].clicked.connect(lambda: self.switch_page("depot"))
        self.nav_buttons["quick"].setToolTip("Input Grid")
        self.nav_buttons["depot"].setToolTip("Tracker Hub")

        self.sidebar_layout.addWidget(self.nav_buttons["quick"])
        self.sidebar_layout.addWidget(self.nav_buttons["depot"])
        self.sidebar_layout.addStretch(1)

        self.settings_button = self._make_nav_button(
            standard_icon_name="SP_FileDialogDetailedView",
            icon_filename="settings.webp",
            icon_px=31,
        )
        self.settings_button.clicked.connect(lambda: self.switch_page("settings"))
        self.settings_button.setToolTip("Settings")
        self.sidebar_layout.addWidget(self.settings_button)

        self.pages = QStackedWidget()

        self.depot_db = DepotDB(_data_file_path(DEPOT_DB_FILENAME))
        _migrate_legacy_agent_icons(self.depot_db.db_path)
        self.depot_tracker = DepotTracker(self.depot_db)
        self.admin_dialog: DepotAdminDialog | None = None
        self.depot_dashboard_dialog: DepotDashboardDialog | None = None

        self.quick_page = self._build_quick_page()
        self.depot_page = self._build_depot_page()
        self.settings_page = self._build_settings_page()

        self.pages.addWidget(self.quick_page)
        self.pages.addWidget(self.depot_page)
        self.pages.addWidget(self.settings_page)

        self.page_index = {"quick": 0, "depot": 1, "settings": 2}
        self._apply_sidebar_position()
        self.switch_page("quick")

        self._foreground_timer = QTimer(self)
        self._foreground_timer.setInterval(260)
        self._foreground_timer.timeout.connect(self._capture_external_target)
        self._foreground_timer.start()

        self._restore_window_position()
        self._apply_window_flags()
        self._init_ui_opacity_effects()
        self.apply_theme_styles()
        self.refresh_quick_grid()
        self.refresh_theme_controls()
        self.refresh_settings_controls()
        self.apply_window_icon()
        _sync_desktop_shortcut(self.config, create_if_missing=False)
        app = QApplication.instance()
        if app is not None:
            app.installEventFilter(self)

    # ---------------------------- Config ---------------------------- #
    def load_config(self) -> dict[str, Any]:
        data: dict[str, Any] = {}
        if self.config_path.exists():
            try:
                data = json.loads(self.config_path.read_text(encoding="utf-8"))
            except Exception as exc:
                context = {"config_path": str(self.config_path)}
                _runtime_log_event(
                    "runtime.config_load_parse_failed",
                    severity="critical",
                    summary="Config parse failed; loading defaults and continuing.",
                    exc=exc,
                    context=context,
                )
                _escalate_runtime_issue_once(
                    "runtime.config_load_parse_failed",
                    "Flowgrid could not parse its config file and loaded defaults for this session.",
                    details=f"{type(exc).__name__}: {exc}",
                    context=context,
                )
                data = {}

        merged = deep_merge(DEFAULT_CONFIG, data)

        if not merged.get("theme_image_layers"):
            old_path = data.get("theme_image_path")
            if old_path:
                merged["theme_image_layers"] = [
                    safe_layer_defaults(
                        {
                            "image_path": old_path,
                            "image_x": data.get("theme_image_x", 0),
                            "image_y": data.get("theme_image_y", 0),
                            "image_scale_mode": data.get("theme_image_scale_mode", "Fill"),
                            "image_anchor": data.get("theme_image_anchor", "Center"),
                            "image_scale_percent": data.get("theme_image_scale_percent", 100),
                        }
                    )
                ]

        cleaned_layers = []
        for layer in merged.get("theme_image_layers", []):
            cleaned_layers.append(safe_layer_defaults(layer if isinstance(layer, dict) else {}))
        merged["theme_image_layers"] = cleaned_layers

        merged["theme"] = {
            "primary": normalize_hex(merged["theme"].get("primary", DEFAULT_THEME_PRIMARY), DEFAULT_THEME_PRIMARY),
            "accent": normalize_hex(merged["theme"].get("accent", DEFAULT_THEME_ACCENT), DEFAULT_THEME_ACCENT),
            "surface": normalize_hex(merged["theme"].get("surface", DEFAULT_THEME_SURFACE), DEFAULT_THEME_SURFACE),
        }

        merged["theme_presets"] = merged.get("theme_presets") or deep_clone(DEFAULT_THEME_PRESETS)
        for preset_name, preset in list(merged["theme_presets"].items()):
            if str(preset_name or "").strip() == "Legacy Blue":
                merged["theme_presets"].pop(preset_name, None)
                continue
            if not isinstance(preset, dict):
                merged["theme_presets"].pop(preset_name, None)
                continue
            merged["theme_presets"][preset_name] = {
                "primary": normalize_hex(preset.get("primary", DEFAULT_THEME_PRIMARY), DEFAULT_THEME_PRIMARY),
                "accent": normalize_hex(preset.get("accent", DEFAULT_THEME_ACCENT), DEFAULT_THEME_ACCENT),
                "surface": normalize_hex(preset.get("surface", DEFAULT_THEME_SURFACE), DEFAULT_THEME_SURFACE),
            }

        if not merged["theme_presets"]:
            merged["theme_presets"] = deep_clone(DEFAULT_THEME_PRESETS)

        legacy_default_theme = {
            "primary": LEGACY_DEFAULT_THEME_PRIMARY,
            "accent": LEGACY_DEFAULT_THEME_ACCENT,
            "surface": LEGACY_DEFAULT_THEME_SURFACE,
        }
        configured_default = merged["theme_presets"].get("Default")
        if isinstance(configured_default, dict):
            normalized_default = {
                "primary": normalize_hex(configured_default.get("primary", DEFAULT_THEME_PRIMARY), DEFAULT_THEME_PRIMARY),
                "accent": normalize_hex(configured_default.get("accent", DEFAULT_THEME_ACCENT), DEFAULT_THEME_ACCENT),
                "surface": normalize_hex(configured_default.get("surface", DEFAULT_THEME_SURFACE), DEFAULT_THEME_SURFACE),
            }
            if normalized_default == legacy_default_theme:
                merged["theme_presets"]["Default"] = deep_clone(DEFAULT_THEME_PRESETS["Default"])

        selected_preset = str(merged.get("selected_theme_preset", "") or "").strip()
        if selected_preset == "Legacy Blue":
            merged["selected_theme_preset"] = "Default"
            selected_preset = "Default"
        if selected_preset == "Default" and merged.get("theme", {}) == legacy_default_theme:
            merged["theme"] = deep_clone(DEFAULT_THEME_PRESETS["Default"])

        if merged.get("selected_theme_preset") not in merged["theme_presets"]:
            merged["selected_theme_preset"] = next(iter(merged["theme_presets"].keys()))

        for kind in ("agent", "qa", "admin", "dashboard"):
            preset_key = f"{kind}_selected_theme_preset"
            if str(merged.get(preset_key, "") or "").strip() == "Legacy Blue":
                merged[preset_key] = str(merged.get("selected_theme_preset", "Default") or "Default")

        legacy_fade_enabled = bool(merged.get("popup_control_fade_enabled", True))
        merged["popup_control_fade_strength"] = int(clamp(int(merged.get("popup_control_fade_strength", 65)), 0, 100))
        merged["popup_control_opacity"] = int(clamp(int(merged.get("popup_control_opacity", 82)), 0, 100))
        merged["popup_control_tail_opacity"] = int(clamp(int(merged.get("popup_control_tail_opacity", 0)), 0, 100))

        style = str(merged.get("popup_control_style", "") or "").strip()
        valid_styles = {"Solid", "Fade Left to Right", "Fade Right to Left", "Fade Center Out"}
        if style not in valid_styles:
            style = "Fade Left to Right" if legacy_fade_enabled else "Solid"
        merged["popup_control_style"] = style
        merged["popup_control_fade_enabled"] = style != "Solid"
        merged["popup_auto_reinherit_enabled"] = bool(merged.get("popup_auto_reinherit_enabled", True))
        auto_reinherit_enabled = bool(merged.get("popup_auto_reinherit_enabled", True))

        popup_valid_styles = {"Solid", "Fade Left to Right", "Fade Right to Left", "Fade Center Out"}
        for popup_key in ("agent_theme", "qa_theme", "admin_theme", "dashboard_theme"):
            popup_theme = merged.get(popup_key, {})
            if not isinstance(popup_theme, dict):
                popup_theme = {}

            popup_theme["background"] = normalize_hex(popup_theme.get("background", "#FFFFFF"), "#FFFFFF")
            popup_theme["text"] = normalize_hex(popup_theme.get("text", "#000000"), "#000000")
            popup_theme["field_bg"] = normalize_hex(popup_theme.get("field_bg", "#FFFFFF"), "#FFFFFF")
            popup_theme["transparent"] = bool(popup_theme.get("transparent", False))
            style_value = str(popup_theme.get("control_style", "Fade Left to Right") or "").strip()
            popup_theme["control_style"] = (
                style_value if style_value in popup_valid_styles else "Fade Left to Right"
            )
            popup_theme["control_opacity"] = int(clamp(safe_int(popup_theme.get("control_opacity", 82), 82), 0, 100))
            popup_theme["control_tail_opacity"] = int(
                clamp(safe_int(popup_theme.get("control_tail_opacity", 0), 0), 0, 100)
            )
            popup_theme["control_fade_strength"] = int(
                clamp(safe_int(popup_theme.get("control_fade_strength", 65), 65), 0, 100)
            )
            popup_theme["header_color"] = normalize_hex(popup_theme.get("header_color", ""), "")
            popup_theme["row_hover_color"] = normalize_hex(popup_theme.get("row_hover_color", ""), "")
            popup_theme["row_selected_color"] = normalize_hex(popup_theme.get("row_selected_color", ""), "")

            cleaned_popup_layers: list[dict[str, Any]] = []
            raw_layers = popup_theme.get("image_layers", [])
            if isinstance(raw_layers, list):
                for layer in raw_layers:
                    cleaned_popup_layers.append(safe_layer_defaults(layer if isinstance(layer, dict) else {}))
            popup_theme["image_layers"] = cleaned_popup_layers
            inherit_value = popup_theme.get("inherit_main_theme")
            if isinstance(inherit_value, bool):
                popup_theme["inherit_main_theme"] = inherit_value
            else:
                # Legacy configs that never customized popup colors/images should inherit Flowgrid theme.
                popup_theme["inherit_main_theme"] = bool(
                    popup_theme["background"] == "#FFFFFF"
                    and popup_theme["text"] == "#000000"
                    and popup_theme["field_bg"] == "#FFFFFF"
                    and not popup_theme["transparent"]
                    and not cleaned_popup_layers
                )
            has_assigned_popup_theme = bool(
                popup_theme["background"] != "#FFFFFF"
                or popup_theme["text"] != "#000000"
                or popup_theme["field_bg"] != "#FFFFFF"
                or popup_theme["transparent"]
                or cleaned_popup_layers
                or popup_theme["control_style"] != "Fade Left to Right"
                or int(popup_theme["control_opacity"]) != 82
                or int(popup_theme["control_tail_opacity"]) != 0
                or int(popup_theme["control_fade_strength"]) != 65
                or bool(popup_theme["header_color"])
                or bool(popup_theme["row_hover_color"])
                or bool(popup_theme["row_selected_color"])
            )
            # Assigned popup theme data always wins over inherited main theme.
            if has_assigned_popup_theme:
                popup_theme["inherit_main_theme"] = False
            elif auto_reinherit_enabled and self._popup_theme_needs_auto_reinherit(popup_theme):
                # Recovery path for legacy/broken configs that were forced out of inherit mode
                # while still holding untouched default values (white/empty fields).
                popup_theme["inherit_main_theme"] = True
            merged[popup_key] = popup_theme

        def normalize_quick_items(raw_items: Any) -> list[dict[str, Any]]:
            cleaned_items: list[dict[str, Any]] = []
            if not isinstance(raw_items, list):
                return cleaned_items
            for idx, item in enumerate(raw_items):
                if not isinstance(item, dict):
                    continue
                action = str(item.get("action", "paste_text")).strip().lower()
                open_target = str(item.get("open_target", "")).strip()
                app_targets = str(item.get("app_targets", "")).strip()
                urls_text = str(item.get("urls", "")).strip()
                browser_path = str(item.get("browser_path", "")).strip()
                text_payload = str(item.get("text", ""))

                # Migrate older macro mode into supported action types.
                if action == "macro":
                    if not app_targets and open_target:
                        app_targets = open_target
                    if app_targets:
                        action = "open_app"
                    elif urls_text or text_payload.strip():
                        action = "open_url"
                    else:
                        action = "paste_text"
                elif action == "macro_sequence":
                    action = "input_sequence"
                elif action not in {"paste_text", "open_url", "open_app", "input_sequence"}:
                    action = "paste_text"

                # Backward compatibility for entries saved before URL list/browser fields existed.
                if action == "open_url" and not urls_text:
                    fallback = open_target or text_payload.strip()
                    urls_text = fallback
                if action == "open_app" and not app_targets and open_target:
                    app_targets = open_target
                normalized_item: dict[str, Any] = {
                    "title": str(item.get("title", f"Item {idx + 1}"))[:64],
                    "tooltip": str(item.get("tooltip", "")),
                    "text": text_payload,
                    "action": action,
                    "open_target": open_target,
                    "app_targets": app_targets,
                    "urls": urls_text,
                    "browser_path": browser_path,
                }
                if isinstance(item.get("x"), (int, float)) and isinstance(item.get("y"), (int, float)):
                    normalized_item["x"] = int(item.get("x", 0))
                    normalized_item["y"] = int(item.get("y", 0))
                cleaned_items.append(normalized_item)
            return cleaned_items

        cleaned_quick_texts = normalize_quick_items(merged.get("quick_texts", []))
        merged["quick_texts"] = cleaned_quick_texts

        cleaned_quick_tabs: list[dict[str, Any]] = []
        raw_quick_tabs = merged.get("quick_tabs", [])
        if isinstance(raw_quick_tabs, list):
            for tab_idx, tab in enumerate(raw_quick_tabs):
                if not isinstance(tab, dict):
                    continue
                tab_name = str(tab.get("name", "")).strip()[:32]
                if not tab_name:
                    tab_name = "Main" if tab_idx == 0 else f"Task {tab_idx + 1}"
                tab_quick_texts = normalize_quick_items(tab.get("quick_texts", []))
                cleaned_quick_tabs.append({"name": tab_name, "quick_texts": tab_quick_texts})

        if not cleaned_quick_tabs:
            cleaned_quick_tabs = [{"name": "Main", "quick_texts": [dict(item) for item in cleaned_quick_texts]}]
        merged["quick_tabs"] = cleaned_quick_tabs

        active_quick_tab = safe_int(merged.get("active_quick_tab", 0), 0)
        if active_quick_tab < 0 or active_quick_tab >= len(cleaned_quick_tabs):
            active_quick_tab = 0
        merged["active_quick_tab"] = active_quick_tab
        merged["quick_texts"] = [
            dict(item) for item in cleaned_quick_tabs[active_quick_tab].get("quick_texts", []) if isinstance(item, dict)
        ]

        family = str(merged.get("quick_button_font_family", "Segoe UI")).strip()
        merged["quick_button_font_family"] = family or "Segoe UI"
        loaded_opacity = float(clamp(float(merged.get("window_opacity", 1.0)), 0.0, 1.0))
        # Prevent a saved 0.00 opacity from making the app appear like it failed to launch.
        if loaded_opacity <= 0.01:
            loaded_opacity = 0.20
        merged["window_opacity"] = loaded_opacity
        current_user = str(merged.get("current_user", "")).strip()
        merged["current_user"] = DepotRules.normalize_user_id(current_user or detect_current_user_id())
        legacy_theme_transparent = data.get("theme_page_transparent_primary_bg")
        if "background_tint_enabled" in data:
            merged["background_tint_enabled"] = bool(merged.get("background_tint_enabled", True))
        elif isinstance(legacy_theme_transparent, bool):
            merged["background_tint_enabled"] = not legacy_theme_transparent
        else:
            merged["background_tint_enabled"] = bool(merged.get("background_tint_enabled", True))

        return merged

    def save_config(self) -> None:
        payload = json.dumps(self.config, indent=2, ensure_ascii=False)
        target = self.config_path
        temp_path = target.with_name(f"{target.name}.tmp")
        try:
            target.parent.mkdir(parents=True, exist_ok=True)
            temp_path.write_text(payload, encoding="utf-8")
            os.replace(temp_path, target)
        except Exception as exc:
            try:
                if temp_path.exists():
                    temp_path.unlink()
            except Exception as cleanup_exc:
                _runtime_log_event(
                    "runtime.config_temp_cleanup_failed",
                    severity="warning",
                    summary="Config save failed and temporary config cleanup also failed.",
                    exc=cleanup_exc,
                    context={"config_path": str(target), "temp_path": str(temp_path)},
                )
            context = {"config_path": str(target)}
            _runtime_log_event(
                "runtime.config_save_failed",
                severity="critical",
                summary="Config save failed; settings may not persist.",
                exc=exc,
                context=context,
            )
            _escalate_runtime_issue_once(
                "runtime.config_save_failed",
                "Flowgrid could not save its config file. Recent settings may not persist.",
                details=f"{type(exc).__name__}: {exc}",
                context=context,
            )

    def queue_save_config(self) -> None:
        self._saving_timer.start()

    # ------------------------- Background --------------------------- #
    def mark_background_dirty(self) -> None:
        self._background_dirty = True
        self._background_cache.clear()
        self.surface.update()

    def load_layer_pixmap(self, path: str) -> QPixmap:
        if path in self._pixmap_cache:
            return self._pixmap_cache[path]

        pixmap = QPixmap()
        if path and Path(path).exists():
            reader = QImageReader(path)
            reader.setAutoTransform(True)
            image = reader.read()
            if not image.isNull():
                pixmap = QPixmap.fromImage(image)

        self._pixmap_cache[path] = pixmap
        return pixmap

    def compute_layer_render(self, layer: dict[str, Any], size: QSize) -> LayerRenderInfo | None:
        if not layer.get("visible", True):
            return None

        path = layer.get("image_path", "")
        pixmap = self.load_layer_pixmap(path)
        if pixmap.isNull() or size.width() <= 0 or size.height() <= 0:
            return None

        target_rect = QRectF(0, 0, float(size.width()), float(size.height()))
        mode = str(layer.get("image_scale_mode", "Fill"))
        anchor = str(layer.get("image_anchor", "Center"))
        scale_percent = int(clamp(int(layer.get("image_scale_percent", 100)), 10, 400)) / 100.0
        img_w = float(pixmap.width())
        img_h = float(pixmap.height())

        if mode == "Stretch":
            draw_w = target_rect.width() * scale_percent
            draw_h = target_rect.height() * scale_percent
        elif mode == "Fit":
            ratio = min(target_rect.width() / img_w, target_rect.height() / img_h)
            ratio *= scale_percent
            draw_w = img_w * ratio
            draw_h = img_h * ratio
        elif mode == "Place":
            draw_w = img_w * scale_percent
            draw_h = img_h * scale_percent
        else:
            ratio = max(target_rect.width() / img_w, target_rect.height() / img_h)
            ratio *= scale_percent
            draw_w = img_w * ratio
            draw_h = img_h * ratio

        x = target_rect.left()
        y = target_rect.top()

        if "Right" in anchor:
            x = target_rect.right() - draw_w
        elif anchor in {"Top", "Center", "Bottom"}:
            x = target_rect.left() + (target_rect.width() - draw_w) / 2

        if "Bottom" in anchor:
            y = target_rect.bottom() - draw_h
        elif anchor in {"Left", "Center", "Right"}:
            y = target_rect.top() + (target_rect.height() - draw_h) / 2

        x += int(layer.get("image_x", 0))
        y += int(layer.get("image_y", 0))

        rect = QRectF(x, y, draw_w, draw_h)
        return LayerRenderInfo(layer=layer, rect=rect, pixmap=pixmap)

    @staticmethod
    def _popup_theme_needs_auto_reinherit(theme: dict[str, Any]) -> bool:
        """True when a popup theme is custom-mode but still carries untouched default values."""
        if not isinstance(theme, dict):
            return False
        if bool(theme.get("inherit_main_theme", False)):
            return False

        background = normalize_hex(theme.get("background", "#FFFFFF"), "#FFFFFF")
        text = normalize_hex(theme.get("text", "#000000"), "#000000")
        field_bg = normalize_hex(theme.get("field_bg", "#FFFFFF"), "#FFFFFF")
        transparent = bool(theme.get("transparent", False))
        control_style = str(theme.get("control_style", "Fade Left to Right") or "").strip()
        control_opacity = int(clamp(safe_int(theme.get("control_opacity", 82), 82), 0, 100))
        control_tail_opacity = int(clamp(safe_int(theme.get("control_tail_opacity", 0), 0), 0, 100))
        control_fade_strength = int(clamp(safe_int(theme.get("control_fade_strength", 65), 65), 0, 100))
        header_color = normalize_hex(theme.get("header_color", ""), "")
        row_hover_color = normalize_hex(theme.get("row_hover_color", ""), "")
        row_selected_color = normalize_hex(theme.get("row_selected_color", ""), "")
        raw_layers = theme.get("image_layers", [])
        has_layers = isinstance(raw_layers, list) and any(isinstance(layer, dict) for layer in raw_layers)

        has_custom_data = bool(
            background != "#FFFFFF"
            or text != "#000000"
            or field_bg != "#FFFFFF"
            or transparent
            or has_layers
            or control_style != "Fade Left to Right"
            or control_opacity != 82
            or control_tail_opacity != 0
            or control_fade_strength != 65
            or bool(header_color)
            or bool(row_hover_color)
            or bool(row_selected_color)
        )
        return not has_custom_data

    @staticmethod
    def _looks_like_unconfigured_popup_theme(theme: dict[str, Any]) -> bool:
        background = normalize_hex(theme.get("background", "#FFFFFF"), "#FFFFFF")
        text = normalize_hex(theme.get("text", "#000000"), "#000000")
        field_bg = normalize_hex(theme.get("field_bg", "#FFFFFF"), "#FFFFFF")
        transparent = bool(theme.get("transparent", False))
        control_style = str(theme.get("control_style", "Fade Left to Right") or "").strip()
        control_opacity = int(clamp(safe_int(theme.get("control_opacity", 82), 82), 0, 100))
        control_tail_opacity = int(clamp(safe_int(theme.get("control_tail_opacity", 0), 0), 0, 100))
        control_fade_strength = int(clamp(safe_int(theme.get("control_fade_strength", 65), 65), 0, 100))
        header_color = normalize_hex(theme.get("header_color", ""), "")
        row_hover_color = normalize_hex(theme.get("row_hover_color", ""), "")
        row_selected_color = normalize_hex(theme.get("row_selected_color", ""), "")
        raw_layers = theme.get("image_layers", [])
        has_layers = isinstance(raw_layers, list) and any(isinstance(layer, dict) for layer in raw_layers)
        has_assigned_data = bool(
            background != "#FFFFFF"
            or text != "#000000"
            or field_bg != "#FFFFFF"
            or transparent
            or has_layers
            or control_style != "Fade Left to Right"
            or control_opacity != 82
            or control_tail_opacity != 0
            or control_fade_strength != 65
            or bool(header_color)
            or bool(row_hover_color)
            or bool(row_selected_color)
        )
        if has_assigned_data:
            return False
        if "inherit_main_theme" in theme:
            return bool(theme.get("inherit_main_theme", False))
        return True

    def _auto_reinherit_popup_defaults(self) -> bool:
        """Recover popup themes that are custom-mode but still effectively default/unconfigured."""
        if not bool(self.config.get("popup_auto_reinherit_enabled", True)):
            return False
        changed = False
        for kind in ("agent", "qa", "admin", "dashboard"):
            key = f"{kind}_theme"
            theme = self.config.get(key, {})
            if not isinstance(theme, dict):
                theme = {}
                self.config[key] = theme
            if self._popup_theme_needs_auto_reinherit(theme):
                theme["inherit_main_theme"] = True
                changed = True
        return changed

    def _default_popup_theme_from_main(self) -> dict[str, Any]:
        layers: list[dict[str, Any]] = []
        for layer in self.config.get("theme_image_layers", []):
            if isinstance(layer, dict):
                layers.append(safe_layer_defaults(layer))
        return {
            "background": normalize_hex(
                self.palette_data.get("control_bg", self.palette_data.get("surface", DEFAULT_THEME_SURFACE)),
                DEFAULT_THEME_SURFACE,
            ),
            "text": normalize_hex(self.palette_data.get("label_text", "#000000"), "#000000"),
            "field_bg": normalize_hex(self.palette_data.get("input_bg", "#FFFFFF"), "#FFFFFF"),
            "transparent": False,
            "inherit_main_theme": True,
            "image_layers": layers,
            "control_style": str(self.config.get("popup_control_style", "Fade Left to Right") or "Fade Left to Right"),
            "control_opacity": int(clamp(safe_int(self.config.get("popup_control_opacity", 82), 82), 0, 100)),
            "control_tail_opacity": int(
                clamp(safe_int(self.config.get("popup_control_tail_opacity", 0), 0), 0, 100)
            ),
            "control_fade_strength": int(
                clamp(safe_int(self.config.get("popup_control_fade_strength", 65), 65), 0, 100)
            ),
            "header_color": normalize_hex(self.config.get("popup_header_color", ""), ""),
            "row_hover_color": normalize_hex(self.config.get("popup_row_hover_color", ""), ""),
            "row_selected_color": normalize_hex(self.config.get("popup_row_selected_color", ""), ""),
        }

    def _resolved_popup_theme(self, kind: str) -> dict[str, Any]:
        base = self._default_popup_theme_from_main()
        raw_theme = self.config.get(f"{kind}_theme", {})
        if not isinstance(raw_theme, dict) or self._looks_like_unconfigured_popup_theme(raw_theme):
            return base

        resolved: dict[str, Any] = {
            "background": normalize_hex(raw_theme.get("background", base["background"]), base["background"]),
            "text": normalize_hex(raw_theme.get("text", base["text"]), base["text"]),
            "field_bg": normalize_hex(raw_theme.get("field_bg", base["field_bg"]), base["field_bg"]),
            "transparent": bool(raw_theme.get("transparent", False)),
            "inherit_main_theme": False,
            "image_layers": [],
            "control_style": str(raw_theme.get("control_style", base["control_style"]) or base["control_style"]),
            "control_opacity": int(
                clamp(safe_int(raw_theme.get("control_opacity", base["control_opacity"]), base["control_opacity"]), 0, 100)
            ),
            "control_tail_opacity": int(
                clamp(
                    safe_int(raw_theme.get("control_tail_opacity", base["control_tail_opacity"]), base["control_tail_opacity"]),
                    0,
                    100,
                )
            ),
            "control_fade_strength": int(
                clamp(
                    safe_int(raw_theme.get("control_fade_strength", base["control_fade_strength"]), base["control_fade_strength"]),
                    0,
                    100,
                )
            ),
            "header_color": normalize_hex(raw_theme.get("header_color", base["header_color"]), base["header_color"]),
            "row_hover_color": normalize_hex(
                raw_theme.get("row_hover_color", base["row_hover_color"]), base["row_hover_color"]
            ),
            "row_selected_color": normalize_hex(
                raw_theme.get("row_selected_color", base["row_selected_color"]), base["row_selected_color"]
            ),
        }
        valid_styles = {"Solid", "Fade Left to Right", "Fade Right to Left", "Fade Center Out"}
        if resolved["control_style"] not in valid_styles:
            resolved["control_style"] = base["control_style"] if base["control_style"] in valid_styles else "Fade Left to Right"
        raw_layers = raw_theme.get("image_layers", [])
        if isinstance(raw_layers, list):
            cleaned: list[dict[str, Any]] = []
            for layer in raw_layers:
                if isinstance(layer, dict):
                    cleaned.append(safe_layer_defaults(layer))
            resolved["image_layers"] = cleaned
        return resolved

    def render_background_pixmap(self, size: QSize, kind: str = "main") -> QPixmap:
        key = (kind, size.width(), size.height())
        cached = self._background_cache.get(key)
        if cached is not None and not self._background_dirty:
            return cached

        pixmap = QPixmap(size)
        pixmap.fill(Qt.GlobalColor.transparent)

        painter = QPainter(pixmap)
        painter.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform)

        if kind == "main":
            layers = self.config.get("theme_image_layers", [])
        else:
            layers = self._resolved_popup_theme(kind).get("image_layers", [])

        for layer in layers:
            info = self.compute_layer_render(layer, size)
            if not info:
                continue
            opacity = float(clamp(float(layer.get("image_opacity", 1.0)), 0.0, 1.0))
            painter.setOpacity(opacity)
            painter.drawPixmap(info.rect.toRect(), info.pixmap)
            painter.setOpacity(1.0)
        painter.end()

        self._background_cache[key] = pixmap
        self._background_dirty = False
        return pixmap

    def _background_tint_enabled(self) -> bool:
        return bool(self.config.get("background_tint_enabled", True))

    def _has_visible_background_layers(self) -> bool:
        for layer in self.config.get("theme_image_layers", []):
            if not isinstance(layer, dict):
                continue
            if not bool(layer.get("visible", True)):
                continue
            path = str(layer.get("image_path", "")).strip()
            if path and Path(path).exists():
                return True
        return False

    def paint_background(self, painter: QPainter, rect: QRect) -> None:
        shell_opacity = float(clamp(getattr(self, "_ui_opacity_current", 1.0), 0.0, 1.0))
        tint_enabled = self._background_tint_enabled()
        if tint_enabled or not self._has_visible_background_layers():
            surface_color = QColor(self.palette_data["surface"])
            surface_color.setAlpha(int(255 * shell_opacity))
            painter.fillRect(rect, surface_color)

        bg = self.render_background_pixmap(rect.size())
        painter.drawPixmap(rect, bg)

        if tint_enabled:
            overlay_color = QColor(self.palette_data["shell_overlay"])
            overlay_color.setAlpha(int(50 * shell_opacity))
            painter.fillRect(rect, overlay_color)

    # -------------------------- UI Build ---------------------------- #
    def _resolve_standard_icon(self, icon_name: str, fallback_name: str = "SP_FileIcon") -> QIcon:
        style = self.style() if self.style() is not None else QApplication.style()
        if style is None:
            return QIcon()
        fallback_enum = getattr(QStyle.StandardPixmap, fallback_name, QStyle.StandardPixmap.SP_FileIcon)
        icon_enum = getattr(QStyle.StandardPixmap, str(icon_name or "").strip(), fallback_enum)
        return style.standardIcon(icon_enum)

    def _ui_icon_dir(self) -> Path:
        icon_dir = _resolve_data_root() / ASSETS_DIR_NAME / FLOWGRID_ICON_PACK_DIR_NAME
        try:
            icon_dir.mkdir(parents=True, exist_ok=True)
        except Exception as exc:
            _runtime_log_event(
                "ui.main_icon_dir_create_failed",
                severity="warning",
                summary="Failed creating local UI icon directory.",
                exc=exc,
                context={"icon_dir": str(icon_dir)},
            )
        return icon_dir

    def _ensure_ui_icon_assets(self) -> None:
        # Keep icon handling file-based: migrate legacy icon files into Assets if present.
        icon_dir = self._ui_icon_dir()
        legacy_candidates = [
            _resolve_data_root() / "ui_icons",
            _resolve_data_root() / ASSETS_DIR_NAME / ASSET_UI_ICON_COMPAT_DIR_NAME,
        ]
        for source_dir in legacy_candidates:
            if not source_dir.exists() or not source_dir.is_dir():
                continue
            try:
                for source_file in source_dir.rglob("*"):
                    if not source_file.is_file():
                        continue
                    target_path = icon_dir / source_file.name
                    if target_path.exists():
                        continue
                    shutil.copy2(source_file, target_path)
            except Exception as exc:
                _runtime_log_event(
                    "ui.main_icon_asset_migrate_failed",
                    severity="warning",
                    summary="Failed migrating legacy UI icon assets into Assets.",
                    exc=exc,
                    context={"source_dir": str(source_dir), "target_dir": str(icon_dir)},
                )

    def _load_ui_icon(self, filename: str, fallback_standard: str = "SP_FileIcon") -> QIcon:
        clean = str(filename or "").strip()
        if clean:
            search_paths = [
                self._ui_icon_dir() / clean,
                _resolve_data_root() / ASSETS_DIR_NAME / ASSET_UI_ICON_COMPAT_DIR_NAME / clean,
                _resolve_data_root() / "ui_icons" / clean,
            ]
            for path in search_paths:
                if not path.exists() or not path.is_file():
                    continue
                icon = QIcon(str(path))
                if not icon.isNull():
                    return icon
        return self._resolve_standard_icon(fallback_standard, "SP_FileIcon")

    def _make_nav_button(
        self,
        icon_text: str = "",
        *,
        standard_icon_name: str = "",
        icon_filename: str = "",
        icon_px: int = 30,
    ) -> QToolButton:
        btn = QToolButton()
        btn.setText(str(icon_text or ""))
        btn.setFixedSize(42, 42)
        btn.setCursor(Qt.CursorShape.PointingHandCursor)
        font = QFont("Segoe UI Symbol", 15, QFont.Weight.Bold)
        btn.setFont(font)
        icon = QIcon()
        if icon_filename:
            fallback = standard_icon_name if standard_icon_name else "SP_FileIcon"
            icon = self._load_ui_icon(icon_filename, fallback)
        elif standard_icon_name:
            icon = self._resolve_standard_icon(standard_icon_name, "SP_FileIcon")
        if not icon.isNull():
            px = int(clamp(safe_int(icon_px, 30), 16, 40))
            btn.setIcon(icon)
            btn.setIconSize(QSize(px, px))
        return btn

    def _build_quick_page(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(8, 6, 8, 8)
        layout.setSpacing(6)

        head_row = QHBoxLayout()
        self.quick_actions_button = QPushButton("")
        self.quick_actions_button.setObjectName("QuickActionsTrigger")
        self.quick_actions_button.setProperty("actionRole", "add")
        self.quick_actions_button.setToolTip("Quick actions")
        self.quick_actions_button.setFixedSize(36, 36)
        self.quick_actions_button.setIcon(self._resolve_standard_icon("SP_DialogOpenButton", "SP_ArrowRight"))
        self.quick_actions_button.setIconSize(QSize(20, 20))
        self.quick_actions_button.clicked.connect(self.toggle_quick_radial_menu)
        head_row.addWidget(self.quick_actions_button, 0)
        head_row.addStretch(1)
        layout.addLayout(head_row)

        self._build_quick_radial_menu()

        self.quick_tabs_widget = QTabWidget()
        self.quick_tabs_widget.setMovable(False)
        self.quick_tabs_widget.setTabPosition(QTabWidget.TabPosition.North)
        self.quick_tabs_widget.currentChanged.connect(self._on_quick_tab_changed)
        layout.addWidget(self.quick_tabs_widget, 1)
        self._rebuild_quick_tab_widgets()

        self._build_quick_editor_dialog()
        return page

    def _build_depot_page(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(8, 6, 8, 8)
        layout.setSpacing(8)

        title = QLabel("Tracker Hub")
        title.setProperty("section", True)
        subtitle = QLabel("Launch windows and maintenance actions.")
        subtitle.setProperty("muted", True)
        layout.addWidget(title)
        layout.addWidget(subtitle)

        self.depot_agent_button = QPushButton("Open Agent")
        self.depot_agent_button.setProperty("actionRole", "launch")
        self.depot_agent_button.setIcon(self._load_ui_icon("wrench.png", "SP_ComputerIcon"))
        self.depot_agent_button.clicked.connect(self._open_depot_agent)
        self.depot_agent_button.setToolTip("Open the Agent popup window.")

        self.depot_qa_button = QPushButton("Open QA/WCS")
        self.depot_qa_button.setProperty("actionRole", "launch")
        self.depot_qa_button.setIcon(self._load_ui_icon("qa.png", "SP_DialogApplyButton"))
        self.depot_qa_button.clicked.connect(self._open_depot_qa)
        self.depot_qa_button.setToolTip("Open the QA/WCS popup window.")

        self.depot_import_button = QPushButton("Import Workbook")
        self.depot_import_button.setProperty("actionRole", "pick")
        self.depot_import_button.setIcon(self._resolve_standard_icon("SP_DialogOpenButton", "SP_FileIcon"))
        self.depot_import_button.clicked.connect(self._import_depot_workbook)
        self.depot_import_button.setToolTip("Import depot workbook data.")

        self.depot_admin_button = QPushButton("Admin")
        self.depot_admin_button.setProperty("actionRole", "pick")
        self.depot_admin_button.setIcon(self._load_ui_icon("user-admin.svg", "SP_FileDialogDetailedView"))
        self.depot_admin_button.clicked.connect(self._open_depot_admin)
        self.depot_admin_button.setToolTip("Open admin panel.")
        self.depot_dashboard_button = QPushButton("Open Data Dashboard")
        self.depot_dashboard_button.setProperty("actionRole", "pick")
        self.depot_dashboard_button.setIcon(self._load_ui_icon("dash.webp", "SP_FileDialogContentsView"))
        self.depot_dashboard_button.clicked.connect(self._open_depot_dashboard)
        self.depot_dashboard_button.setToolTip("Open data dashboard in a separate window.")

        for btn in (
            self.depot_agent_button,
            self.depot_qa_button,
            self.depot_import_button,
            self.depot_admin_button,
            self.depot_dashboard_button,
        ):
            btn.setMinimumHeight(44)
            btn.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
            btn.setIconSize(QSize(24, 24))

        actions_grid = QGridLayout()
        actions_grid.setHorizontalSpacing(8)
        actions_grid.setVerticalSpacing(8)
        actions_grid.setColumnStretch(0, 1)
        actions_grid.setColumnStretch(1, 1)
        actions_grid.addWidget(self.depot_agent_button, 0, 0)
        actions_grid.addWidget(self.depot_qa_button, 0, 1)
        actions_grid.addWidget(self.depot_import_button, 1, 0)
        actions_grid.addWidget(self.depot_admin_button, 1, 1)
        actions_grid.addWidget(self.depot_dashboard_button, 2, 0, 1, 2)
        layout.addLayout(actions_grid)
        layout.addStretch(1)

        return page

    def _show_shell_message(
        self,
        icon: QMessageBox.Icon,
        title: str,
        text: str,
        *,
        theme_kind: str = "main",
    ) -> None:
        """Non-native QMessageBox from the main window, styled like configured popup themes."""
        show_flowgrid_themed_message(self, self, theme_kind, icon, title, text)

    def _refresh_depot_dashboard_combo_popup_width(self) -> None:
        if self.depot_dashboard_dialog is not None:
            self.depot_dashboard_dialog.refresh_combo_popup_width()

    def _refresh_depot_dashboard(self) -> None:
        if self.depot_dashboard_dialog is not None and self.depot_dashboard_dialog.isVisible():
            self.depot_dashboard_dialog.refresh_dashboard()

    def _export_depot_dashboard(self) -> None:
        if self.depot_dashboard_dialog is None:
            self._open_depot_dashboard()
        if self.depot_dashboard_dialog is not None:
            self.depot_dashboard_dialog.export_csv()

    def _open_depot_dashboard(self) -> None:
        self._reveal_immediately()
        if not self.depot_tracker.is_admin_user(self.current_user):
            self._show_shell_message(
                QMessageBox.Icon.Warning,
                "Access Denied",
                "Only administrators can access the dashboard. Please contact an administrator to grant access.",
                theme_kind="dashboard",
            )
            return
        if self.depot_dashboard_dialog is None:
            self.depot_dashboard_dialog = DepotDashboardDialog(self)

        popup_pos = self.config.get("popup_positions", {}).get("depot_dashboard")
        if isinstance(popup_pos, dict) and "x" in popup_pos and "y" in popup_pos and not self.depot_dashboard_dialog.isVisible():
            self.depot_dashboard_dialog.move(int(popup_pos["x"]), int(popup_pos["y"]))

        self.depot_dashboard_dialog.apply_theme_styles()
        self.depot_dashboard_dialog.refresh_combo_popup_width()
        self.depot_dashboard_dialog.refresh_dashboard()
        self.depot_dashboard_dialog.show()
        self.depot_dashboard_dialog.raise_()
        self.depot_dashboard_dialog.activateWindow()


    def _open_depot_agent(self) -> None:
        self._reveal_immediately()
        dlg = getattr(self, "active_agent_window", None)
        if dlg is None:
            dlg = DepotAgentWindow(self.depot_tracker, self.current_user, app_window=self)
            self.active_agent_window = dlg
        dlg.show()
        dlg.raise_()
        dlg.activateWindow()

    def _open_depot_qa(self) -> None:
        self._reveal_immediately()
        dlg = getattr(self, "active_qa_window", None)
        if dlg is None:
            dlg = DepotQAWindow(self.depot_tracker, self.current_user, app_window=self)
            self.active_qa_window = dlg
        dlg.show()
        dlg.raise_()
        dlg.activateWindow()

    def _open_depot_admin(self) -> None:
        self._reveal_immediately()
        if not self.depot_tracker.is_admin_user(self.current_user):
            self._show_shell_message(
                QMessageBox.Icon.Warning,
                "Access Denied",
                "Only administrators can access the admin panel. Please contact an administrator to grant access.",
                theme_kind="admin",
            )
            return

        if self.admin_dialog is None:
            self.admin_dialog = DepotAdminDialog(self.depot_tracker, self.current_user, app_window=self)
        self.admin_dialog.apply_theme_styles()
        self.admin_dialog.refresh_agents()
        self.admin_dialog.refresh_admins()
        self.admin_dialog.show()
        self.admin_dialog.raise_()
        self.admin_dialog.activateWindow()

    def _import_depot_workbook(self) -> None:
        if openpyxl is None:
            self._show_shell_message(
                QMessageBox.Icon.Warning,
                "Depot Workbook Import",
                "openpyxl is not installed.",
                theme_kind="admin",
            )
            return

        start_dir = self.config_path.parent if self.config_path.parent.exists() else Path.home()
        selected_path, _ = show_flowgrid_themed_open_file_name(
            self,
            self,
            "admin",
            "Select Workbook to Import",
            str(start_dir),
            "Excel Workbook (*.xlsm *.xlsx *.xls);;All Files (*.*)",
        )
        if not selected_path:
            return

        import_tables = self._prompt_workbook_import_tables()
        if import_tables is None:
            return

        okay, msg = self.depot_tracker.import_workbook(Path(selected_path), import_tables)
        if okay:
            self._show_shell_message(
                QMessageBox.Icon.Information,
                "Depot Workbook Import",
                msg,
                theme_kind="admin",
            )
            self._refresh_depot_dashboard()
            if hasattr(self, "active_agent_window") and self.active_agent_window is not None and self.active_agent_window.isVisible():
                self.active_agent_window._refresh_agent_parts()
                self.active_agent_window._refresh_category_parts()
            if hasattr(self, "active_qa_window") and self.active_qa_window is not None and self.active_qa_window.isVisible():
                self.active_qa_window._populate_agents()
                self.active_qa_window._populate_flags()
                self.active_qa_window._refresh_assigned_parts()
                self.active_qa_window._refresh_delivered_parts()
                self.active_qa_window._refresh_completed_parts()
            if self.admin_dialog is not None and self.admin_dialog.isVisible():
                self.admin_dialog.refresh_agents()
                self.admin_dialog.refresh_qa_flags()
                self.admin_dialog.refresh_admins()
        else:
            self._show_shell_message(
                QMessageBox.Icon.Warning,
                "Depot Workbook Import",
                msg,
                theme_kind="admin",
            )

    def _prompt_workbook_import_tables(self) -> set[str] | None:
        dialog = FlowgridThemedDialog(self, self, "admin")
        dialog.setWindowTitle("Select Tables to Import")
        dialog.setModal(True)
        dialog.resize(420, 420)
        dialog.apply_theme_styles(force_opaque_root=True)
        dialog_layout = QVBoxLayout(dialog)
        dialog_layout.setContentsMargins(10, 10, 10, 10)
        dialog_layout.setSpacing(8)

        title = QLabel("Choose which workbook tables to import:")
        title.setProperty("section", True)
        dialog_layout.addWidget(title)

        checks: dict[str, QCheckBox] = {}
        for key, sheet, label in WORKBOOK_IMPORT_SPECS:
            box = QCheckBox(f"{label} ({sheet})")
            box.setChecked(True)
            checks[key] = box
            dialog_layout.addWidget(box)

        action_row = QHBoxLayout()
        action_row.setContentsMargins(0, 0, 0, 0)
        action_row.setSpacing(6)
        select_all_btn = QPushButton("Select All")
        select_none_btn = QPushButton("Select None")
        select_all_btn.setProperty("actionRole", "pick")
        select_none_btn.setProperty("actionRole", "pick")
        action_row.addWidget(select_all_btn, 0)
        action_row.addWidget(select_none_btn, 0)
        action_row.addStretch(1)
        dialog_layout.addLayout(action_row)

        footer = QHBoxLayout()
        footer.setContentsMargins(0, 0, 0, 0)
        footer.setSpacing(6)
        import_btn = QPushButton("Import Selected")
        import_btn.setProperty("actionRole", "save")
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setProperty("actionRole", "reset")
        footer.addWidget(import_btn, 0)
        footer.addWidget(cancel_btn, 0)
        footer.addStretch(1)
        dialog_layout.addLayout(footer)

        select_all_btn.clicked.connect(lambda: [box.setChecked(True) for box in checks.values()])
        select_none_btn.clicked.connect(lambda: [box.setChecked(False) for box in checks.values()])
        import_btn.clicked.connect(dialog.accept)
        cancel_btn.clicked.connect(dialog.reject)

        if dialog.exec() != QDialog.DialogCode.Accepted:
            return None

        selected = {key for key, box in checks.items() if box.isChecked()}
        if not selected:
            self._show_shell_message(
                QMessageBox.Icon.Warning,
                "Depot Workbook Import",
                "Select at least one table to import.",
                theme_kind="admin",
            )
            return None
        return selected

    def _build_quick_editor_dialog(self) -> None:
        self.quick_editor_dialog = FlowgridThemedDialog(self, self, "main")
        self.quick_editor_dialog.setObjectName("QuickEditorDialog")
        self.quick_editor_dialog.setWindowTitle("Edit Quick Button")
        self.quick_editor_dialog.setModal(False)
        self.quick_editor_dialog.resize(430, 580)
        self.quick_editor_dialog.apply_theme_styles(force_opaque_root=True)

        editor_layout = QVBoxLayout(self.quick_editor_dialog)
        editor_layout.setContentsMargins(10, 10, 10, 10)
        editor_layout.setSpacing(6)

        self.editor_title = QLineEdit()
        self.editor_tooltip = QLineEdit()
        self.editor_action_combo = QComboBox()
        self.editor_action_combo.addItem("Paste Text", "paste_text")
        self.editor_action_combo.addItem("Open URL(s)", "open_url")
        self.editor_action_combo.addItem("Open App/File", "open_app")
        self.editor_action_combo.addItem("Input Sequence", "input_sequence")

        self.editor_text = QTextEdit()
        self.editor_text.setFixedHeight(82)
        
        self.editor_macro = QTextEdit()
        self.editor_macro.setFixedHeight(82)
        self.editor_macro.setPlaceholderText("Format: Bin location [enter]")
        
        # Macro helper buttons
        macro_buttons_layout = QHBoxLayout()
        macro_buttons_layout.setContentsMargins(0, 0, 0, 0)
        macro_buttons_layout.setSpacing(4)
        
        macro_btn_tab = QPushButton("[Tab]")
        macro_btn_tab.setFixedHeight(28)
        macro_btn_tab.setMaximumWidth(70)
        macro_btn_tab.clicked.connect(lambda: self._insert_macro_simple("[tab]"))
        
        macro_btn_enter = QPushButton("[Enter]")
        macro_btn_enter.setFixedHeight(28)
        macro_btn_enter.setMaximumWidth(70)
        macro_btn_enter.clicked.connect(lambda: self._insert_macro_simple("[enter]"))
        
        macro_btn_delay = QPushButton("Add Delay")
        macro_btn_delay.setFixedHeight(28)
        macro_btn_delay.clicked.connect(self._insert_macro_delay)

        macro_buttons_layout.addWidget(macro_btn_tab, 0)
        macro_buttons_layout.addWidget(macro_btn_enter, 0)
        macro_buttons_layout.addWidget(macro_btn_delay, 1)
        
        self.editor_macro_wrap = QWidget()
        macro_wrap_layout = QVBoxLayout(self.editor_macro_wrap)
        macro_wrap_layout.setContentsMargins(0, 0, 0, 0)
        macro_wrap_layout.setSpacing(4)
        macro_wrap_layout.addWidget(self.editor_macro)
        macro_wrap_layout.addLayout(macro_buttons_layout)
        
        self.editor_apps = QTextEdit()
        self.editor_apps.setFixedHeight(82)
        self.editor_apps_browse = QPushButton("Browse Files/Apps...")
        self.editor_apps_browse.setToolTip("Open file explorer to select one or more app/file targets.")
        self.editor_apps_browse.setProperty("actionRole", "pick")
        self.editor_apps_browse.setFixedHeight(30)
        self.editor_apps_browse_folder = QPushButton("Browse Folder...")
        self.editor_apps_browse_folder.setToolTip("Open folder picker and add a folder target.")
        self.editor_apps_browse_folder.setProperty("actionRole", "pick")
        self.editor_apps_browse_folder.setFixedHeight(30)
        apps_action_row = QHBoxLayout()
        apps_action_row.setContentsMargins(0, 0, 0, 0)
        apps_action_row.addWidget(self.editor_apps_browse, 0)
        apps_action_row.addWidget(self.editor_apps_browse_folder, 0)
        apps_action_row.addStretch(1)
        self.editor_apps_wrap = QWidget()
        apps_wrap_layout = QVBoxLayout(self.editor_apps_wrap)
        apps_wrap_layout.setContentsMargins(0, 0, 0, 0)
        apps_wrap_layout.setSpacing(4)
        apps_wrap_layout.addWidget(self.editor_apps)
        apps_wrap_layout.addLayout(apps_action_row)

        self.editor_urls = QTextEdit()
        self.editor_urls.setFixedHeight(82)

        self.editor_browser_combo = QComboBox()
        self.editor_refresh_browsers_button = QPushButton("Detect")
        self.editor_refresh_browsers_button.setProperty("actionRole", "pick")
        self.editor_refresh_browsers_button.setFixedHeight(30)
        browser_row = QHBoxLayout()
        browser_row.setContentsMargins(0, 0, 0, 0)
        browser_row.setSpacing(4)
        browser_row.addWidget(self.editor_browser_combo, 1)
        browser_row.addWidget(self.editor_refresh_browsers_button, 0)
        self.editor_browser_wrap = QWidget()
        self.editor_browser_wrap.setLayout(browser_row)

        form = QFormLayout()
        form.setSpacing(4)
        form.setContentsMargins(0, 0, 0, 0)
        form.addRow("Title", self.editor_title)
        form.addRow("Context", self.editor_tooltip)
        form.addRow("Action", self.editor_action_combo)
        form.addRow("Text", self.editor_text)
        form.addRow("Input Sequence", self.editor_macro_wrap)
        form.addRow("Apps/Files", self.editor_apps_wrap)
        form.addRow("URLs", self.editor_urls)
        form.addRow("Browser", self.editor_browser_wrap)
        self.editor_form = form
        self.editor_text_label = form.labelForField(self.editor_text)
        self.editor_macro_label = form.labelForField(self.editor_macro_wrap)
        self.editor_apps_label = form.labelForField(self.editor_apps_wrap)
        self.editor_urls_label = form.labelForField(self.editor_urls)
        self.editor_browser_label = form.labelForField(self.editor_browser_wrap)
        editor_layout.addLayout(form)

        action_row = QHBoxLayout()
        self.editor_save_btn = QPushButton("Save")
        self.editor_save_btn.setProperty("actionRole", "save")
        self.editor_delete_btn = QPushButton("Delete")
        self.editor_delete_btn.setProperty("actionRole", "reset")
        self.editor_cancel_btn = QPushButton("Cancel")
        self.editor_cancel_btn.setProperty("actionRole", "pick")
        action_row.addWidget(self.editor_save_btn)
        action_row.addWidget(self.editor_delete_btn)
        action_row.addWidget(self.editor_cancel_btn)
        editor_layout.addLayout(action_row)

        self.editor_save_btn.clicked.connect(self.save_quick_editor)
        self.editor_delete_btn.clicked.connect(self.delete_quick_editor)
        self.editor_cancel_btn.clicked.connect(self.close_quick_editor)
        self.editor_action_combo.currentIndexChanged.connect(self._update_quick_editor_action_ui)
        self.editor_apps_browse.clicked.connect(self.browse_quick_apps)
        self.editor_apps_browse_folder.clicked.connect(self.browse_quick_app_folder)
        self.editor_refresh_browsers_button.clicked.connect(self._refresh_available_browsers)

        self._editing_index: int | None = None
        self._refresh_available_browsers()
        self._update_quick_editor_action_ui()

    def _build_theme_page(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        self.theme_tabs = QTabWidget()
        self.theme_tabs.setUsesScrollButtons(True)
        self.theme_tabs.tabBar().setElideMode(Qt.TextElideMode.ElideNone)
        
        # Main (Flowgrid) theme tab
        main_theme_tab = self._build_main_theme_tab()
        
        # Agent theme tab
        agent_theme_tab = self._build_agent_theme_tab()
        
        # QA theme tab
        qa_theme_tab = self._build_qa_theme_tab()
        
        # Admin theme tab
        admin_theme_tab = self._build_admin_theme_tab()
        
        # Dashboard theme tab
        dashboard_theme_tab = self._build_dashboard_theme_tab()
        
        self.theme_tabs.addTab(self._wrap_scrollable_page(main_theme_tab), "Flowgrid")
        self.theme_tabs.addTab(self._wrap_scrollable_page(agent_theme_tab), "Agent")
        self.theme_tabs.addTab(self._wrap_scrollable_page(qa_theme_tab), "QA")
        self.theme_tabs.addTab(self._wrap_scrollable_page(admin_theme_tab), "Admin")
        self.theme_tabs.addTab(self._wrap_scrollable_page(dashboard_theme_tab), "Dashboard")
        self.theme_tabs.setTabToolTip(0, "Base Flowgrid colors and background layers used as popup defaults.")
        self.theme_tabs.setTabToolTip(1, "Agent popup theme overrides.")
        self.theme_tabs.setTabToolTip(2, "QA/WCS popup theme overrides.")
        self.theme_tabs.setTabToolTip(3, "Admin popup theme overrides.")
        self.theme_tabs.setTabToolTip(4, "Dashboard popup theme overrides.")
        
        layout.addWidget(self.theme_tabs, 1)
        return page

    @staticmethod
    def _wrap_scrollable_page(content: QWidget) -> QScrollArea:
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll.setWidget(content)
        return scroll

    def _build_main_theme_tab(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(8, 6, 8, 8)
        layout.setSpacing(8)

        colors_title = QLabel("Colors")
        colors_title.setProperty("section", True)
        layout.addWidget(colors_title)

        preset_row = QHBoxLayout()
        preset_row.setSpacing(4)
        self.theme_preset_combo = QComboBox()
        self.theme_preset_new = QPushButton("New")
        self.theme_preset_save = QPushButton("Save")
        self.theme_preset_combo.setToolTip("Select a saved Flowgrid color preset.")
        self.theme_preset_new.setToolTip("Create a new preset from the current Flowgrid colors.")
        self.theme_preset_save.setToolTip("Save current Flowgrid colors into the selected preset.")
        self.theme_preset_new.setProperty("actionRole", "new")
        self.theme_preset_save.setProperty("actionRole", "save")
        self.theme_preset_combo.setFixedHeight(28)
        self.theme_preset_new.setFixedSize(54, 28)
        self.theme_preset_save.setFixedSize(54, 28)
        preset_row.addWidget(QLabel("Preset"))
        preset_row.addWidget(self.theme_preset_combo, 1)
        preset_row.addWidget(self.theme_preset_new)
        preset_row.addWidget(self.theme_preset_save)
        layout.addLayout(preset_row)

        self.color_swatches: dict[str, QPushButton] = {}
        colors_grid = QGridLayout()
        colors_grid.setHorizontalSpacing(8)
        colors_grid.setVerticalSpacing(8)
        colors_grid.setColumnStretch(1, 1)

        for row_index, (key, label) in enumerate((("primary", "Primary"), ("accent", "Accent"), ("surface", "Surface"))):
            text = QLabel(label)
            text.setFixedWidth(52)
            swatch = QPushButton()
            swatch.setProperty("actionRole", "pick")
            swatch.setFixedSize(180, 32)
            swatch.setToolTip(f"Pick the {label.lower()} color for the Flowgrid theme.")

            colors_grid.addWidget(text, row_index, 0)
            colors_grid.addWidget(swatch, row_index, 1)

            self.color_swatches[key] = swatch
            swatch.clicked.connect(lambda _=False, c=key: self.pick_theme_color(c))

        layout.addLayout(colors_grid)

        color_actions = QHBoxLayout()
        color_actions.setSpacing(6)
        self.reset_theme_btn = QPushButton("Reset")
        self.image_layers_btn = QPushButton("Background Images")
        self.reset_theme_btn.setToolTip("Reset Flowgrid colors to the selected preset values.")
        self.image_layers_btn.setToolTip("Edit Flowgrid background image layers (position, scale, blend, visibility).")
        self.reset_theme_btn.setProperty("actionRole", "reset")
        self.image_layers_btn.setProperty("actionRole", "pick")
        for btn in (self.reset_theme_btn, self.image_layers_btn):
            btn.setFixedHeight(32)
            color_actions.addWidget(btn, 1)
        layout.addLayout(color_actions)

        self.theme_transparent_bg_check = QCheckBox("Transparent Background")
        self.theme_transparent_bg_check.setToolTip(
            "Disable the Flowgrid tint layer so background images render directly behind controls."
        )
        layout.addWidget(self.theme_transparent_bg_check)
        self.popup_auto_reinherit_check = QCheckBox("Auto-Reinherit Popup Defaults")
        self.popup_auto_reinherit_check.setToolTip(
            "When enabled, popups stuck in an unconfigured custom state are automatically repaired to inherit Flowgrid defaults."
        )
        layout.addWidget(self.popup_auto_reinherit_check)

        layout.addStretch(1)

        self.theme_preset_combo.currentTextChanged.connect(self.on_theme_preset_selected)
        self.theme_preset_new.clicked.connect(self.create_theme_preset)
        self.theme_preset_save.clicked.connect(self.save_theme_preset)
        self.reset_theme_btn.clicked.connect(self.reset_theme)
        self.image_layers_btn.clicked.connect(lambda _checked=False: self.open_image_layers_dialog("main"))
        self.theme_transparent_bg_check.toggled.connect(self.on_theme_page_background_option_changed)
        self.popup_auto_reinherit_check.toggled.connect(self.on_popup_auto_reinherit_changed)
        return page

    def _build_agent_theme_tab(self) -> QWidget:
        return self._build_popup_theme_tab("agent", "Agent Window")

    def _build_qa_theme_tab(self) -> QWidget:
        return self._build_popup_theme_tab("qa", "QA/WCS Window")

    def _build_admin_theme_tab(self) -> QWidget:
        return self._build_popup_theme_tab("admin", "Admin Window")

    def _build_dashboard_theme_tab(self) -> QWidget:
        return self._build_popup_theme_tab("dashboard", "Dashboard Window")

    def _build_popup_theme_tab(self, kind: str, label: str) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(8, 6, 8, 8)
        layout.setSpacing(8)

        title = QLabel(f"{label} Theme")
        title.setProperty("section", True)
        layout.addWidget(title)

        preset_combo_key = f"{kind}_theme_preset_combo"
        preset_row = QHBoxLayout()
        preset_row.setSpacing(4)
        preset_row.addWidget(QLabel("Preset"), 0)
        preset_combo = QComboBox()
        preset_combo.setToolTip(f"Select a preset to seed {label.lower()} colors.")
        setattr(self, preset_combo_key, preset_combo)
        preset_row.addWidget(preset_combo, 1)
        layout.addLayout(preset_row)

        colors_grid = QGridLayout()
        colors_grid.setHorizontalSpacing(8)
        colors_grid.setVerticalSpacing(8)
        colors_grid.setColumnStretch(1, 1)

        swatches_key = f"{kind}_color_swatches"
        if not hasattr(self, swatches_key):
            setattr(self, swatches_key, {})
        color_swatches = getattr(self, swatches_key)

        for row_index, (fld, fld_label) in enumerate((("background", "Background"), ("text", "Text"), ("field_bg", "Controls"))):
            text = QLabel(fld_label)
            text.setMinimumWidth(96)
            swatch = QPushButton()
            swatch.setProperty("actionRole", "pick")
            swatch.setFixedSize(180, 32)
            swatch.setToolTip(f"Pick the {fld_label.lower()} color for this popup.")

            colors_grid.addWidget(text, row_index, 0)
            colors_grid.addWidget(swatch, row_index, 1)

            color_swatches[fld] = swatch
            swatch.clicked.connect(lambda _=False, k=kind, f=fld: self._pick_popup_theme_color(k, f))

        layout.addLayout(colors_grid)

        transparent_check_key = f"{kind}_transparent_bg_check"
        transparent_check = QCheckBox("Transparent Background")
        transparent_check.setToolTip("Make container/frame surfaces transparent while keeping controls and lists readable.")
        setattr(self, transparent_check_key, transparent_check)
        layout.addWidget(transparent_check)

        transparent_check.toggled.connect(lambda checked: self.on_popup_background_option_changed(kind, checked))

        control_form = QFormLayout()
        control_form.setContentsMargins(0, 0, 0, 0)
        control_form.setSpacing(6)

        style_combo_key = f"{kind}_control_style_combo"
        style_combo = QComboBox()
        style_combo.addItems(["Solid", "Fade Left to Right", "Fade Right to Left", "Fade Center Out"])
        style_combo.setToolTip("Choose how input/table control backgrounds are filled.")
        setattr(self, style_combo_key, style_combo)
        control_form.addRow("Fill Style", style_combo)

        fade_slider_key = f"{kind}_control_fade_slider"
        fade_value_key = f"{kind}_control_fade_value"
        fade_slider = QSlider(Qt.Orientation.Horizontal)
        fade_slider.setRange(0, 100)
        fade_slider.setToolTip("How strongly the control fill gradient fades.")
        fade_value = QLabel("65%")
        fade_row = QHBoxLayout()
        fade_row.setContentsMargins(0, 0, 0, 0)
        fade_row.setSpacing(4)
        fade_row.addWidget(fade_slider, 1)
        fade_row.addWidget(fade_value, 0)
        fade_wrap = QWidget()
        fade_wrap.setLayout(fade_row)
        setattr(self, fade_slider_key, fade_slider)
        setattr(self, fade_value_key, fade_value)
        control_form.addRow("Fade", fade_wrap)

        opacity_slider_key = f"{kind}_control_opacity_slider"
        opacity_value_key = f"{kind}_control_opacity_value"
        opacity_slider = QSlider(Qt.Orientation.Horizontal)
        opacity_slider.setRange(0, 100)
        opacity_slider.setToolTip("Primary opacity of controls (inputs, lists, table cells).")
        opacity_value = QLabel("82%")
        opacity_row = QHBoxLayout()
        opacity_row.setContentsMargins(0, 0, 0, 0)
        opacity_row.setSpacing(4)
        opacity_row.addWidget(opacity_slider, 1)
        opacity_row.addWidget(opacity_value, 0)
        opacity_wrap = QWidget()
        opacity_wrap.setLayout(opacity_row)
        setattr(self, opacity_slider_key, opacity_slider)
        setattr(self, opacity_value_key, opacity_value)
        control_form.addRow("Opacity", opacity_wrap)

        tail_slider_key = f"{kind}_control_tail_opacity_slider"
        tail_value_key = f"{kind}_control_tail_opacity_value"
        tail_slider = QSlider(Qt.Orientation.Horizontal)
        tail_slider.setRange(0, 100)
        tail_slider.setToolTip("Opacity at the end of the gradient fill (higher means less fade-out).")
        tail_value = QLabel("0%")
        tail_row = QHBoxLayout()
        tail_row.setContentsMargins(0, 0, 0, 0)
        tail_row.setSpacing(4)
        tail_row.addWidget(tail_slider, 1)
        tail_row.addWidget(tail_value, 0)
        tail_wrap = QWidget()
        tail_wrap.setLayout(tail_row)
        setattr(self, tail_slider_key, tail_slider)
        setattr(self, tail_value_key, tail_value)
        control_form.addRow("End Opacity", tail_wrap)

        optional_swatches_key = f"{kind}_optional_color_swatches"
        setattr(self, optional_swatches_key, {})
        optional_swatches = getattr(self, optional_swatches_key)
        for field, row_label in (
            ("header_color", "List Header"),
            ("row_hover_color", "Row Hover"),
            ("row_selected_color", "Row Selected"),
        ):
            swatch = QPushButton()
            swatch.setProperty("actionRole", "pick")
            swatch.setFixedHeight(28)
            swatch.setToolTip(f"Override {row_label.lower()} color for this popup.")
            clear_btn = QPushButton("Auto")
            clear_btn.setProperty("actionRole", "reset")
            clear_btn.setFixedHeight(28)
            clear_btn.setToolTip(f"Use automatic {row_label.lower()} color based on current theme.")
            color_row = QHBoxLayout()
            color_row.setContentsMargins(0, 0, 0, 0)
            color_row.setSpacing(4)
            color_row.addWidget(swatch, 1)
            color_row.addWidget(clear_btn, 0)
            color_wrap = QWidget()
            color_wrap.setLayout(color_row)
            control_form.addRow(row_label, color_wrap)
            optional_swatches[field] = swatch
            swatch.clicked.connect(lambda _=False, k=kind, f=field: self._pick_popup_optional_color(k, f))
            clear_btn.clicked.connect(lambda _=False, k=kind, f=field: self._clear_popup_optional_color(k, f))

        layout.addLayout(control_form)

        image_row = QHBoxLayout()
        image_btn = QPushButton("Background Images")
        image_btn.setProperty("actionRole", "pick")
        image_btn.setToolTip(f"Edit background image layers for the {label.lower()}.")
        image_row.addWidget(image_btn)
        image_row.addStretch(1)
        layout.addLayout(image_row)

        image_btn.clicked.connect(lambda: self.open_image_layers_dialog(kind))
        preset_combo.currentTextChanged.connect(lambda name, k=kind: self.on_popup_theme_preset_selected(k, name))
        style_combo.currentTextChanged.connect(lambda _value, k=kind: self.on_popup_theme_control_changed(k))
        fade_slider.valueChanged.connect(lambda _value, k=kind: self.on_popup_theme_control_changed(k))
        opacity_slider.valueChanged.connect(lambda _value, k=kind: self.on_popup_theme_control_changed(k))
        tail_slider.valueChanged.connect(lambda _value, k=kind: self.on_popup_theme_control_changed(k))
        
        layout.addStretch(1)
        return page

    def _build_app_settings_tab(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(8, 6, 8, 8)
        layout.setSpacing(8)

        form = QFormLayout()
        form.setContentsMargins(0, 0, 0, 0)
        form.setSpacing(8)

        self.opacity_slider = QSlider(Qt.Orientation.Horizontal)
        self.opacity_slider.setRange(0, 100)
        opacity_tip = "Base opacity for the Flowgrid shell while idle. Lower values make the shell more transparent."
        self.opacity_slider.setToolTip(opacity_tip)
        self.opacity_value = QLabel("1.00")
        self.opacity_value.setToolTip(opacity_tip)
        opacity_row = QHBoxLayout()
        opacity_row.addWidget(self.opacity_slider, 1)
        opacity_row.addWidget(self.opacity_value)
        opacity_wrap = QWidget()
        opacity_wrap.setLayout(opacity_row)

        self.hover_delay_slider = QSlider(Qt.Orientation.Horizontal)
        self.hover_delay_slider.setRange(0, 10)
        hover_delay_tip = "Seconds to wait before auto-revealing full opacity when your cursor enters the window."
        self.hover_delay_slider.setToolTip(hover_delay_tip)
        self.hover_delay_value = QLabel("5s")
        self.hover_delay_value.setToolTip(hover_delay_tip)
        hover_delay_row = QHBoxLayout()
        hover_delay_row.addWidget(self.hover_delay_slider, 1)
        hover_delay_row.addWidget(self.hover_delay_value)
        hover_delay_wrap = QWidget()
        hover_delay_wrap.setLayout(hover_delay_row)

        self.hover_fade_in_slider = QSlider(Qt.Orientation.Horizontal)
        self.hover_fade_in_slider.setRange(0, 10)
        hover_in_tip = "How quickly the shell fades up to full opacity on hover."
        self.hover_fade_in_slider.setToolTip(hover_in_tip)
        self.hover_fade_in_value = QLabel("5s")
        self.hover_fade_in_value.setToolTip(hover_in_tip)
        hover_fade_in_row = QHBoxLayout()
        hover_fade_in_row.addWidget(self.hover_fade_in_slider, 1)
        hover_fade_in_row.addWidget(self.hover_fade_in_value)
        hover_fade_in_wrap = QWidget()
        hover_fade_in_wrap.setLayout(hover_fade_in_row)

        self.hover_fade_out_slider = QSlider(Qt.Orientation.Horizontal)
        self.hover_fade_out_slider.setRange(0, 10)
        hover_out_tip = "How quickly the shell returns to idle opacity after hover ends."
        self.hover_fade_out_slider.setToolTip(hover_out_tip)
        self.hover_fade_out_value = QLabel("5s")
        self.hover_fade_out_value.setToolTip(hover_out_tip)
        hover_fade_out_row = QHBoxLayout()
        hover_fade_out_row.addWidget(self.hover_fade_out_slider, 1)
        hover_fade_out_row.addWidget(self.hover_fade_out_value)
        hover_fade_out_wrap = QWidget()
        hover_fade_out_wrap.setLayout(hover_fade_out_row)

        form.addRow("Idle Opacity", opacity_wrap)
        form.addRow("Hover Delay", hover_delay_wrap)
        form.addRow("Fade In", hover_fade_in_wrap)
        form.addRow("Fade Out", hover_fade_out_wrap)
        layout.addLayout(form)

        self.always_on_top_check = QCheckBox("Always on top")
        self.always_on_top_check.setToolTip("Keep the main Flowgrid shell above normal windows.")
        self.compact_mode_check = QCheckBox("Compact mode")
        self.compact_mode_check.setToolTip(
            "Use tighter spacing and smaller control padding across Flowgrid and popup windows."
        )
        self.sidebar_right_switch = QCheckBox()
        self.sidebar_right_switch.setProperty("switch", True)
        self.sidebar_right_switch.setTristate(False)
        self.sidebar_right_switch.setToolTip("Move the sidebar to the right side of the main Flowgrid window.")
        self.sidebar_switch_status = QLabel()
        self.sidebar_switch_status.setProperty("muted", True)
        self.sidebar_switch_status.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        self.sidebar_switch_status.setMinimumWidth(170)
        layout.addWidget(self.always_on_top_check)
        layout.addWidget(self.compact_mode_check)
        sidebar_row = QHBoxLayout()
        sidebar_row.setContentsMargins(0, 0, 0, 0)
        sidebar_row.setSpacing(6)
        self.sidebar_side_caption = QLabel("Sidebar Side")
        self.sidebar_side_caption.setMinimumWidth(90)
        self.sidebar_left_label = QLabel("Left")
        self.sidebar_left_label.setProperty("muted", True)
        self.sidebar_right_label = QLabel("Right")
        self.sidebar_right_label.setProperty("muted", True)
        self.sidebar_switch_status.setToolTip("Shows the active sidebar side.")
        sidebar_row.addWidget(self.sidebar_side_caption, 0)
        sidebar_row.addWidget(self.sidebar_left_label, 0)
        sidebar_row.addWidget(self.sidebar_right_switch, 0)
        sidebar_row.addWidget(self.sidebar_right_label, 0)
        sidebar_row.addStretch(1)
        layout.addLayout(sidebar_row)
        layout.addWidget(self.sidebar_switch_status)

        icon_row = QHBoxLayout()
        self.pick_icon_button = QPushButton("Set Icon")
        self.clear_icon_button = QPushButton("Clear Icon")
        self.pick_icon_button.setToolTip("Pick a custom icon for the main Flowgrid title bar and desktop shortcut.")
        self.clear_icon_button.setToolTip("Reset the Flowgrid title bar and desktop shortcut back to the default wrench icon.")
        self.pick_icon_button.setProperty("actionRole", "pick")
        self.clear_icon_button.setProperty("actionRole", "reset")
        icon_row.addWidget(self.pick_icon_button)
        icon_row.addWidget(self.clear_icon_button)
        icon_row.addStretch(1)
        layout.addLayout(icon_row)
        layout.addStretch(1)

        self.opacity_slider.valueChanged.connect(self.on_opacity_changed)
        self.hover_delay_slider.valueChanged.connect(self.on_hover_settings_changed)
        self.hover_fade_in_slider.valueChanged.connect(self.on_hover_settings_changed)
        self.hover_fade_out_slider.valueChanged.connect(self.on_hover_settings_changed)
        self.always_on_top_check.toggled.connect(self.on_settings_changed)
        self.compact_mode_check.toggled.connect(self.on_settings_changed)
        self.sidebar_right_switch.toggled.connect(self.on_sidebar_position_changed)
        self.pick_icon_button.clicked.connect(self.pick_custom_icon)
        self.clear_icon_button.clicked.connect(self.clear_custom_icon)
        return tab

    def _build_settings_page(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(8, 6, 8, 8)
        layout.setSpacing(8)

        self.settings_tabs = QTabWidget()
        self.settings_tabs.setObjectName("SettingsTabs")
        self.settings_tabs.setUsesScrollButtons(True)
        main_tab_bar = self.settings_tabs.tabBar()
        main_tab_bar.setObjectName("SettingsMainTabBar")
        main_tab_bar.setElideMode(Qt.TextElideMode.ElideNone)
        self.settings_tabs.addTab(self._build_app_settings_tab(), "App Settings")
        self.settings_tabs.addTab(self._build_theme_page(), "Themes")
        self.settings_tabs.setTabIcon(0, self._resolve_standard_icon("SP_FileDialogDetailedView", "SP_FileIcon"))
        self.settings_tabs.setTabIcon(1, self._resolve_standard_icon("SP_DirOpenIcon", "SP_FileIcon"))
        self.settings_tabs.setTabToolTip(0, "General app behavior and window controls.")
        self.settings_tabs.setTabToolTip(1, "Theme colors, backgrounds, and popup styling.")
        layout.addWidget(self.settings_tabs, 1)
        return page

    def _quick_button_style_tokens(self, font_size: int, button_opacity: float, shape: str, action_type: str = "paste_text") -> tuple[int, int, str, str, str, str]:
        font_size = int(clamp(font_size, 8, 20))
        button_opacity = float(clamp(button_opacity, 0.15, 1.0))
        palette = self.palette_data

        # Adjust base color based on action type for subtle distinction
        base_bg = palette["button_bg"]
        if action_type == "open_url":
            # Shift toward accent color for web links
            base_bg = blend(palette['button_bg'], palette['accent'], 0.25)
        elif action_type == "open_app":
            # Shift toward primary color for apps
            base_bg = blend(palette['button_bg'], palette['primary'], 0.25)
        elif action_type in {"input_sequence", "macro_sequence"}:
            # Shift toward surface color for input sequences.
            base_bg = blend(palette['button_bg'], palette['surface'], 0.35)

        radius = 11
        border = 1
        bg_hex = base_bg
        bg_alpha = button_opacity
        border_color = shift(bg_hex, -0.35)
        text_color = readable_text(bg_hex)
        hover_alpha = min(1.0, bg_alpha + 0.08)

        if shape == "Bordered":
            radius = 4
            border = 2
            bg_hex = base_bg
            bg_alpha = min(1.0, button_opacity * 0.92)
            border_color = shift(bg_hex, -0.45)
        elif shape == "Block":
            radius = 0
            border = 1
            bg_hex = shift(base_bg, -0.08)
            bg_alpha = button_opacity
            border_color = shift(bg_hex, -0.45)
        elif shape == "Pill":
            radius = 999
            border = 1
            bg_hex = base_bg
            bg_alpha = button_opacity
            border_color = shift(bg_hex, -0.42)
        elif shape == "Ghost":
            radius = 10
            border = 2
            bg_hex = base_bg
            bg_alpha = max(0.15, button_opacity * 0.32)
            border_color = shift(bg_hex, -0.62)
            hover_alpha = max(0.45, min(1.0, bg_alpha + 0.22))
        elif shape == "Glass":
            radius = 12
            border = 1
            bg_hex = blend(palette["surface"], palette["button_bg"], 0.48)
            bg_alpha = max(0.35, button_opacity * 0.75)
            border_color = shift(bg_hex, -0.45)
        elif shape == "Outline":
            radius = 9
            border = 2
            bg_hex = palette["surface"]
            bg_alpha = max(0.10, button_opacity * 0.12)
            border_color = shift(palette["button_bg"], -0.68)
            text_color = readable_text(shift(palette["button_bg"], -0.15))
            hover_alpha = max(0.42, min(1.0, button_opacity * 0.55))
        elif shape == "Inset":
            radius = 8
            border = 2
            bg_hex = shift(palette["button_bg"], -0.18)
            bg_alpha = button_opacity
            border_color = shift(bg_hex, -0.42)
        elif shape == "Flat":
            radius = 2
            border = 0
            bg_hex = blend(palette["button_bg"], palette["surface"], 0.18)
            bg_alpha = button_opacity
            border_color = bg_hex
            text_color = readable_text(bg_hex)
        elif shape == "Raised3D":
            radius = 10
            border = 2
            top = shift(base_bg, 0.20)
            bottom = shift(base_bg, -0.20)
            bg_hex = base_bg
            shape_bg = (
                "qlineargradient(x1:0,y1:0,x2:0,y2:1,"
                f" stop:0 {rgba_css(top, min(1.0, button_opacity))},"
                f" stop:1 {rgba_css(bottom, min(1.0, button_opacity))})"
            )
            border_color = shift(bottom, -0.35)
            text_color = readable_text(bottom)
            hover_bg = (
                "qlineargradient(x1:0,y1:0,x2:0,y2:1,"
                f" stop:0 {rgba_css(shift(top, 0.06), min(1.0, button_opacity))},"
                f" stop:1 {rgba_css(shift(bottom, 0.06), min(1.0, button_opacity))})"
            )
            return radius, border, shape_bg, border_color, hover_bg, text_color
        elif shape == "Bevel3D":
            radius = 8
            border = 2
            top = shift(base_bg, 0.12)
            bottom = shift(base_bg, -0.30)
            bg_hex = blend(top, bottom, 0.55)
            shape_bg = (
                "qlineargradient(x1:0,y1:0,x2:0,y2:1,"
                f" stop:0 {rgba_css(top, button_opacity)},"
                f" stop:0.52 {rgba_css(bg_hex, button_opacity)},"
                f" stop:1 {rgba_css(bottom, button_opacity)})"
            )
            border_color = shift(bottom, -0.40)
            text_color = readable_text(bg_hex)
            hover_bg = (
                "qlineargradient(x1:0,y1:0,x2:0,y2:1,"
                f" stop:0 {rgba_css(shift(top, 0.06), button_opacity)},"
                f" stop:0.52 {rgba_css(shift(bg_hex, 0.05), button_opacity)},"
                f" stop:1 {rgba_css(shift(bottom, 0.05), button_opacity)})"
            )
            return radius, border, shape_bg, border_color, hover_bg, text_color
        elif shape == "Ridge3D":
            radius = 6
            border = 2
            c1 = shift(base_bg, 0.20)
            c2 = shift(base_bg, -0.08)
            c3 = shift(base_bg, 0.08)
            c4 = shift(base_bg, -0.28)
            bg_hex = c2
            shape_bg = (
                "qlineargradient(x1:0,y1:0,x2:0,y2:1,"
                f" stop:0 {rgba_css(c1, button_opacity)},"
                f" stop:0.33 {rgba_css(c2, button_opacity)},"
                f" stop:0.66 {rgba_css(c3, button_opacity)},"
                f" stop:1 {rgba_css(c4, button_opacity)})"
            )
            border_color = shift(c4, -0.35)
            text_color = readable_text(c2)
            hover_bg = (
                "qlineargradient(x1:0,y1:0,x2:0,y2:1,"
                f" stop:0 {rgba_css(shift(c1, 0.05), button_opacity)},"
                f" stop:0.33 {rgba_css(shift(c2, 0.05), button_opacity)},"
                f" stop:0.66 {rgba_css(shift(c3, 0.05), button_opacity)},"
                f" stop:1 {rgba_css(shift(c4, 0.05), button_opacity)})"
            )
            return radius, border, shape_bg, border_color, hover_bg, text_color
        elif shape == "Neumorph":
            radius = 14
            border = 1
            bg_hex = blend(palette["surface"], palette["button_bg"], 0.30)
            bg_alpha = max(0.55, button_opacity * 0.85)
            border_color = shift(bg_hex, -0.22)
            text_color = readable_text(bg_hex)
            hover_alpha = min(1.0, bg_alpha + 0.08)
        elif shape == "Retro3D":
            radius = 4
            border = 2
            top = shift(palette["button_bg"], 0.22)
            bottom = shift(palette["button_bg"], -0.30)
            bg_hex = blend(top, bottom, 0.5)
            shape_bg = (
                "qlineargradient(x1:0,y1:0,x2:0,y2:1,"
                f" stop:0 {rgba_css(top, button_opacity)},"
                f" stop:1 {rgba_css(bottom, button_opacity)})"
            )
            border_color = shift(bottom, -0.38)
            text_color = readable_text(bottom)
            hover_bg = (
                "qlineargradient(x1:0,y1:0,x2:0,y2:1,"
                f" stop:0 {rgba_css(shift(top, 0.06), button_opacity)},"
                f" stop:1 {rgba_css(shift(bottom, 0.06), button_opacity)})"
            )
            return radius, border, shape_bg, border_color, hover_bg, text_color
        elif shape == "Neon3D":
            radius = 10
            border = 2
            top = blend(palette["accent"], base_bg, 0.30)
            bottom = blend(palette["primary"], shift(base_bg, -0.14), 0.50)
            bg_hex = blend(top, bottom, 0.5)
            shape_bg = (
                "qlineargradient(x1:0,y1:0,x2:0,y2:1,"
                f" stop:0 {rgba_css(top, min(1.0, button_opacity * 0.95))},"
                f" stop:1 {rgba_css(bottom, min(1.0, button_opacity * 0.95))})"
            )
            border_color = shift(bottom, -0.45)
            text_color = readable_text(bottom)
            hover_bg = (
                "qlineargradient(x1:0,y1:0,x2:0,y2:1,"
                f" stop:0 {rgba_css(shift(top, 0.08), min(1.0, button_opacity))},"
                f" stop:1 {rgba_css(shift(bottom, 0.08), min(1.0, button_opacity))})"
            )
            return radius, border, shape_bg, border_color, hover_bg, text_color
        # Backward compatibility with older saved style names.
        elif shape in {"Neon", "Ocean"}:
            shape = "Neon3D"
            return self._quick_button_style_tokens(font_size, button_opacity, shape)
        elif shape in {"Retro", "Ember"}:
            shape = "Retro3D"
            return self._quick_button_style_tokens(font_size, button_opacity, shape)
        elif shape in {"Royal", "Slate", "Danger", "Forest", "Candy", "Frost"}:
            shape = "Raised3D"
            return self._quick_button_style_tokens(font_size, button_opacity, shape)

        shape_bg = rgba_css(bg_hex, bg_alpha)
        hover_bg = rgba_css(shift(bg_hex, 0.08), hover_alpha)
        return radius, border, shape_bg, border_color, hover_bg, text_color

    def _quick_button_stylesheet(
        self,
        font_size: int,
        button_opacity: float,
        shape: str,
        font_family: str | None = None,
        padding: str = "2px 20px 2px 8px",
        action_type: str = "paste_text",
    ) -> str:
        radius, border, shape_bg, border_color, hover_bg, text_color = self._quick_button_style_tokens(font_size, button_opacity, shape, action_type)
        family_value = str(font_family or self.config.get("quick_button_font_family", "Segoe UI"))
        family_value = family_value.replace("'", "\\'")
        return (
            "QPushButton {"
            f"background-color: {shape_bg};"
            f"color: {text_color};"
            f"border: {border}px solid {border_color};"
            f"border-radius: {radius}px;"
            f"font-size: {int(clamp(font_size, 8, 20))}px;"
            f"font-family: '{family_value}';"
            "font-weight: 700;"
            f"padding: {padding};"
            "text-align: center;"
            "}"
            f"QPushButton:hover {{ background-color: {hover_bg}; }}"
            f"QPushButton:pressed {{ background-color: {hover_bg}; }}"
        )

    def _refresh_theme_preview_buttons(self) -> None:
        if not hasattr(self, "theme_preset_new") or not hasattr(self, "theme_preset_save"):
            return
        font_size = int(clamp(int(self.config.get("quick_button_font_size", 11)), 8, 14))
        font_family = str(self.config.get("quick_button_font_family", "Segoe UI"))
        button_opacity = float(clamp(float(self.config.get("quick_button_opacity", 1.0)), 0.2, 1.0))
        shape = self.config.get("quick_button_shape", "Soft")
        css = self._quick_button_stylesheet(font_size, button_opacity, shape, font_family=font_family, padding="2px 8px 2px 8px")
        for preview_btn in (self.theme_preset_new, self.theme_preset_save):
            preview_btn.setStyleSheet(css)
            poly = build_quick_shape_polygon(shape, preview_btn.width(), preview_btn.height())
            if poly is None:
                preview_btn.clearMask()
            else:
                preview_btn.setMask(QRegion(poly))

    def _apply_sidebar_position(self) -> None:
        while self.body_layout.count() > 0:
            item = self.body_layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.setParent(self.body)

        sidebar_on_right = bool(self.config.get("sidebar_on_right", False))
        if sidebar_on_right:
            self.body_layout.addWidget(self.pages, 1)
            self.body_layout.addWidget(self.sidebar, 0)
        else:
            self.body_layout.addWidget(self.sidebar, 0)
            self.body_layout.addWidget(self.pages, 1)

    # --------------------------- Styling ---------------------------- #
    def apply_theme_styles(self) -> None:
        p = self.palette_data
        compact_mode = bool(self.config.get("compact_mode", True))
        base_font_px = 12 if compact_mode else 13
        section_font_px = 13 if compact_mode else 14
        field_padding = "1px 5px" if compact_mode else "2px 6px"
        text_edit_padding = "3px 5px" if compact_mode else "4px 6px"
        checkbox_spacing = 6 if compact_mode else 8
        checkbox_indicator_px = 12 if compact_mode else 14
        button_padding = "2px 8px" if compact_mode else "4px 10px"
        button_min_height = 24 if compact_mode else 28
        button_font_px = 10 if compact_mode else 11
        scrollbar_width = 8 if compact_mode else 10
        scrollbar_handle_min_height = 14 if compact_mode else 18
        slider_groove_height = 5 if compact_mode else 6
        slider_handle_width = 10 if compact_mode else 12
        slider_handle_margin = "-3px 0px" if compact_mode else "-4px 0px"
        sidebar_color = QColor(p["sidebar_overlay"])
        sidebar_color.setAlpha(125)
        shared_button_bg = rgba_css(p["button_bg"], 0.75)  # 25% transparent for non-quick buttons.
        shared_button_hover = rgba_css(shift(p["button_bg"], 0.08), 0.82)
        add_base = p["primary"]
        apply_base = p["accent"]
        pick_base = blend(p["primary"], p["surface"], 0.45)
        save_base = blend(p["primary"], p["accent"], 0.22)
        new_base = blend(p["surface"], p["primary"], 0.35)
        reset_base = shift(p["surface"], -0.45)
        title_min_bg = rgba_css(blend(p["button_bg"], p["surface"], 0.20), 0.62)
        title_min_border = rgba_css(shift(p["accent"], -0.42), 0.84)
        title_min_hover = rgba_css(shift(blend(p["button_bg"], p["surface"], 0.20), 0.10), 0.76)

        self.surface.setStyleSheet("background: transparent;")
        self.body.setStyleSheet("background: transparent;")
        self.pages.setStyleSheet("background: transparent;")

        self.sidebar.setStyleSheet(
            "QWidget {"
            f"background: rgba({sidebar_color.red()}, {sidebar_color.green()}, {sidebar_color.blue()}, {sidebar_color.alpha()});"
            "border-radius: 0px;"
            "}"
        )

        self.titlebar.setStyleSheet(
            "QWidget {"
            f"background: rgba({QColor(p['shell_overlay']).red()}, {QColor(p['shell_overlay']).green()}, {QColor(p['shell_overlay']).blue()}, 135);"
            f"color: {p['label_text']};"
            "}"
            "QLabel#TitleText { font-size: 13px; font-weight: 600; background: transparent; }"
            "QToolButton {"
            "background: transparent;"
            "border: none;"
            "font-size: 13px;"
            "font-weight: 600;"
            f"color: {p['label_text']};"
            "}"
            "QToolButton#TitleMinButton {"
            f"background: {title_min_bg};"
            f"border: 1px solid {title_min_border};"
            "border-radius: 11px;"
            "font-size: 12px;"
            "font-weight: 800;"
            "}"
            f"QToolButton#TitleMinButton:hover {{ background: {title_min_hover}; }}"
            "QToolButton#TitleCloseButton {"
            "background: rgba(225,80,80,110);"
            "border: 1px solid rgba(255,135,135,175);"
            "border-radius: 11px;"
            "font-size: 12px;"
            "font-weight: 800;"
            "}"
            "QToolButton#TitleCloseButton:hover { background: rgba(235,95,95,155); }"
        )

        global_style = (
            "QWidget {"
            f"color: {p['label_text']};"
            "font-family: 'Segoe UI';"
            f"font-size: {base_font_px}px;"
            "background: transparent;"
            "}"
            "QLabel {"
            "font-weight: 700;"
            "background: transparent;"
            "}"
            "QLabel[muted='true'] {"
            "font-weight: 600;"
            f"color: {rgba_css(p['label_text'], 0.82)};"
            "}"
            "QLabel[section='true'] {"
            f"font-size: {section_font_px}px;"
            "font-weight: 700;"
            "padding-bottom: 2px;"
            "}"
            "QLineEdit, QTextEdit, QSpinBox, QComboBox {"
            f"background: rgba({QColor(p['input_bg']).red()}, {QColor(p['input_bg']).green()}, {QColor(p['input_bg']).blue()}, 228);"
            f"border: 1px solid {shift(p['input_bg'], -0.38)};"
            "border-radius: 3px;"
            f"padding: {field_padding};"
            f"selection-background-color: {p['primary']};"
            f"selection-color: {readable_text(p['primary'])};"
            "}"
            f"QTextEdit {{ padding: {text_edit_padding}; }}"
            f"QCheckBox {{ background: transparent; spacing: {checkbox_spacing}px; font-weight: 700; }}"
            f"QCheckBox::indicator {{ width: {checkbox_indicator_px}px; height: {checkbox_indicator_px}px; }}"
            "QCheckBox[switch='true']::indicator {"
            "width: 44px;"
            "height: 20px;"
            "border-radius: 10px;"
            f"background: {rgba_css(p['button_bg'], 0.55)};"
            f"border: 1px solid {shift(p['button_bg'], -0.40)};"
            "}"
            "QCheckBox[switch='true']::indicator:unchecked {"
            f"background: qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 {rgba_css(p['accent'], 0.62)}, stop:0.45 {rgba_css(p['button_bg'], 0.35)}, stop:1 {rgba_css(p['button_bg'], 0.25)});"
            "}"
            "QCheckBox[switch='true']::indicator:checked {"
            f"background: qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 {rgba_css(p['button_bg'], 0.25)}, stop:0.55 {rgba_css(p['button_bg'], 0.35)}, stop:1 {rgba_css(p['accent'], 0.70)});"
            f"border: 1px solid {shift(p['accent'], -0.45)};"
            "}"
            "QTabBar#SettingsMainTabBar::tab {"
            "font-weight: 800;"
            "}"
            "QScrollArea { background: transparent; border: none; }"
            "QScrollArea > QWidget > QWidget { background: transparent; }"
            "QScrollBar:vertical {"
            "background: rgba(0,0,0,40);"
            f"width: {scrollbar_width}px;"
            "margin: 0px;"
            "border: none;"
            "}"
            "QScrollBar::handle:vertical {"
            "background: rgba(255,255,255,95);"
            f"min-height: {scrollbar_handle_min_height}px;"
            "border-radius: 5px;"
            "}"
            "QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {"
            "height: 0px;"
            "}"
            "QSlider::groove:horizontal {"
            "background: rgba(0,0,0,90);"
            f"height: {slider_groove_height}px;"
            "border-radius: 3px;"
            "}"
            "QSlider::handle:horizontal {"
            f"background: {p['accent']};"
            f"width: {slider_handle_width}px;"
            f"margin: {slider_handle_margin};"
            "border-radius: 3px;"
            "}"
            "QPushButton {"
            f"background-color: {shared_button_bg};"
            f"color: {p['button_text']};"
            f"border: 1px solid {shift(p['button_bg'], -0.40)};"
            "border-radius: 4px;"
            f"padding: {button_padding};"
            f"min-height: {button_min_height}px;"
            f"font-size: {button_font_px}px;"
            "font-weight: 700;"
            "}"
            f"QPushButton:hover {{ background-color: {shared_button_hover}; }}"
            f"QPushButton:pressed {{ background-color: {shared_button_hover}; }}"
            f"QPushButton:checked {{ background-color: {rgba_css(p['accent'], 0.78)}; color: {readable_text(p['accent'])}; border: 1px solid {shift(p['accent'], -0.42)}; }}"
            "QPushButton[actionRole='add'] {"
            f"background-color: {rgba_css(add_base, 0.75)};"
            f"color: {readable_text(add_base)};"
            f"border: 1px solid {shift(add_base, -0.42)};"
            "}"
            "QPushButton[actionRole='apply'] {"
            f"background-color: {rgba_css(apply_base, 0.75)};"
            f"color: {readable_text(apply_base)};"
            f"border: 1px solid {shift(apply_base, -0.42)};"
            "}"
            "QPushButton[actionRole='pick'] {"
            f"background-color: {rgba_css(pick_base, 0.75)};"
            f"color: {readable_text(pick_base)};"
            f"border: 1px solid {shift(pick_base, -0.42)};"
            "}"
            "QPushButton[actionRole='new'] {"
            f"background-color: {rgba_css(new_base, 0.75)};"
            f"color: {readable_text(new_base)};"
            f"border: 1px solid {shift(new_base, -0.42)};"
            "}"
            "QPushButton[actionRole='save'] {"
            f"background-color: {rgba_css(save_base, 0.75)};"
            f"color: {readable_text(save_base)};"
            f"border: 1px solid {shift(save_base, -0.42)};"
            "}"
            "QPushButton[actionRole='reset'] {"
            f"background-color: {rgba_css(reset_base, 0.75)};"
            f"color: {readable_text(reset_base)};"
            f"border: 1px solid {shift(reset_base, -0.42)};"
            "}"
        )
        self.setStyleSheet(global_style)
        if hasattr(self, "quick_actions_button"):
            plus_bg = blend(p["accent"], p["primary"], 0.35)
            plus_hover = shift(plus_bg, 0.08)
            plus_pressed = shift(plus_bg, -0.06)
            plus_border = shift(plus_bg, -0.64)
            self.quick_actions_button.setStyleSheet(
                "QPushButton#QuickActionsTrigger {"
                f"background-color: {rgba_css(plus_bg, 0.94)};"
                f"color: {readable_text(plus_bg)};"
                f"border: 2px solid {plus_border};"
                "border-radius: 18px;"
                "font-size: 22px;"
                "font-weight: 900;"
                "padding: 0px;"
                "}"
                "QPushButton#QuickActionsTrigger:hover {"
                f"background-color: {rgba_css(plus_hover, 0.96)};"
                f"border: 2px solid {shift(plus_border, -0.08)};"
                "}"
                "QPushButton#QuickActionsTrigger:pressed {"
                f"background-color: {rgba_css(plus_pressed, 0.98)};"
                f"border: 2px solid {shift(plus_border, -0.12)};"
                "}"
            )
        if self.quick_radial_menu is not None:
            self.quick_radial_menu.apply_theme_styles(p)

        # Depot tab buttons should be fully opaque (no transparency) as requested
        if hasattr(self, "depot_page"):
            opaque_button_bg = rgba_css(p["button_bg"], 1.0)
            opaque_button_hover = rgba_css(shift(p["button_bg"], 0.08), 1.0)
            opaque_button_pressed = rgba_css(shift(p["button_bg"], -0.06), 1.0)
            opaque_border = shift(p["button_bg"], -0.62)
            depot_button_css = (
                "QPushButton {"
                f"background-color: {opaque_button_bg};"
                f"color: {p['button_text']};"
                f"border: 2px solid {opaque_border};"
                "border-radius: 4px;"
                f"padding: {button_padding};"
                f"min-height: {button_min_height}px;"
                f"font-size: {button_font_px}px;"
                "font-weight: 800;"
                "}"
                f"QPushButton:hover {{ background-color: {opaque_button_hover}; border: 2px solid {shift(opaque_border, -0.08)}; }}"
                f"QPushButton:pressed {{ background-color: {opaque_button_pressed}; border: 2px solid {shift(opaque_border, -0.12)}; }}"
            )
            self.depot_page.setStyleSheet(depot_button_css)
            self._refresh_depot_dashboard_combo_popup_width()

        editor_border = shift(p["control_bg"], -0.30)
        editor_field_border = shift(p["input_bg"], -0.38)
        editor_button_border = shift(p["button_bg"], -0.40)
        editor_button_hover = rgba_css(p["button_bg"], 0.18)
        if hasattr(self, "quick_editor_dialog"):
            base_popup_css = self._popup_theme_stylesheet("main", force_opaque_root=True)
            self.quick_editor_dialog.setStyleSheet(
                base_popup_css
                + (
                    "QDialog#QuickEditorDialog {"
                    f"background: {rgba_css(p['shell_overlay'], 0.90)};"
                    f"color: {p['label_text']};"
                    f"border: 1px solid {editor_border};"
                    "border-radius: 8px;"
                    "}"
                    "QDialog#QuickEditorDialog QLabel {"
                    f"color: {p['label_text']};"
                    "background: transparent;"
                    "font-weight: 700;"
                    "}"
                    "QDialog#QuickEditorDialog QLineEdit, QDialog#QuickEditorDialog QTextEdit, QDialog#QuickEditorDialog QComboBox {"
                    "background: transparent;"
                    f"border: 1px solid {editor_field_border};"
                    "border-radius: 3px;"
                    f"color: {p['label_text']};"
                    "padding: 2px 6px;"
                    "}"
                    "QDialog#QuickEditorDialog QLineEdit:focus, QDialog#QuickEditorDialog QTextEdit:focus, QDialog#QuickEditorDialog QComboBox:focus {"
                    f"border: 1px solid {editor_border};"
                    "background: transparent;"
                    "}"
                    "QDialog#QuickEditorDialog QPushButton {"
                    "background: transparent;"
                    f"border: 1px solid {editor_button_border};"
                    "border-radius: 6px;"
                    f"color: {p['label_text']};"
                    "font-weight: 700;"
                    "}"
                    "QDialog#QuickEditorDialog QPushButton:hover {"
                    f"background: {editor_button_hover};"
                    "}"
                )
            )
        if self.image_dialog is not None:
            self.image_dialog.apply_theme_styles()
        if self.quick_layout_dialog is not None:
            self.quick_layout_dialog.apply_theme_styles()
        if self.depot_dashboard_dialog is not None:
            self.depot_dashboard_dialog.apply_theme_styles()
            self.depot_dashboard_dialog.refresh_combo_popup_width()

        for key, button in self.nav_buttons.items():
            active = self.pages.currentIndex() == self.page_index.get(key)
            self._style_nav_button(button, active)
        self._style_nav_button(self.settings_button, self.pages.currentIndex() == self.page_index["settings"])
        self._refresh_theme_preview_buttons()
        self.refresh_quick_grid()

    def _style_nav_button(self, button: QToolButton, active: bool) -> None:
        p = self.palette_data
        if active:
            bg = rgba_css(p["nav_active"], 0.75)
            fg = readable_text(p["nav_active"])
            border = shift(p["nav_active"], -0.45)
        else:
            bg = rgba_css(p["button_bg"], 0.75)
            fg = p["label_text"]
            border = shift(p["button_bg"], -0.40)
        hover_bg = rgba_css(blend(p["accent"], p["surface"], 0.25), 0.46)

        button.setStyleSheet(
            "QToolButton {"
            f"background: {bg};"
            f"color: {fg};"
            f"border: 1px solid {border};"
            "border-radius: 0px;"
            "padding: 0px;"
            "}"
            f"QToolButton:hover {{ background: {hover_bg}; }}"
        )

    # ------------------------- Page Actions ------------------------- #
    def switch_page(self, page: str) -> None:
        if page != "quick" and self.quick_radial_menu is not None and self.quick_radial_menu.isVisible():
            self.quick_radial_menu.hide()
        index = self.page_index.get(page, 0)
        self.pages.setCurrentIndex(index)
        for key, btn in self.nav_buttons.items():
            self._style_nav_button(btn, key == page)
        self._style_nav_button(self.settings_button, page == "settings")
        self.refresh_all_views()

    def refresh_all_views(self) -> None:
        self.surface.update()
        self.quick_page.update()
        self.settings_page.update()
        
        # Update popup windows if they exist
        if hasattr(self, "active_agent_window") and self.active_agent_window is not None and self.active_agent_window.isVisible():
            self.active_agent_window.update()
        if hasattr(self, "active_qa_window") and self.active_qa_window is not None and self.active_qa_window.isVisible():
            self.active_qa_window.update()
        if self.admin_dialog is not None and self.admin_dialog.isVisible():
            self.admin_dialog.update()
        
        # Update image dialogs if they exist
        if self.image_dialog is not None and self.image_dialog.isVisible():
            self.image_dialog.update()
        if hasattr(self, "agent_image_dialog") and self.agent_image_dialog is not None and self.agent_image_dialog.isVisible():
            self.agent_image_dialog.update()
        if hasattr(self, "qa_image_dialog") and self.qa_image_dialog is not None and self.qa_image_dialog.isVisible():
            self.qa_image_dialog.update()
        if hasattr(self, "admin_image_dialog") and self.admin_image_dialog is not None and self.admin_image_dialog.isVisible():
            self.admin_image_dialog.update()
        if hasattr(self, "dashboard_image_dialog") and self.dashboard_image_dialog is not None and self.dashboard_image_dialog.isVisible():
            self.dashboard_image_dialog.update()
        if self.depot_dashboard_dialog is not None and self.depot_dashboard_dialog.isVisible():
            self.depot_dashboard_dialog.update()

    def _init_ui_opacity_effects(self) -> None:
        self._ui_opacity_effects.clear()
        targets = [
            self.titlebar,
            self.sidebar,
            self.settings_page,
            self.quick_actions_button,
        ]
        for widget in targets:
            effect = QGraphicsOpacityEffect(widget)
            widget.setGraphicsEffect(effect)
            self._ui_opacity_effects.append(effect)
        self._set_ui_opacity(self._base_opacity())

    def _set_ui_opacity(self, value: float) -> None:
        opacity = float(clamp(value, 0.0, 1.0))
        self._ui_opacity_current = opacity
        for effect in self._ui_opacity_effects:
            effect.setOpacity(opacity)
        self.surface.update()

    def _has_active_popup(self) -> bool:
        app = QApplication.instance()
        return bool(app and app.activePopupWidget() is not None)

    def _has_active_internal_dialog(self) -> bool:
        dialogs = (self.image_dialog, self.quick_layout_dialog, self.depot_dashboard_dialog)
        for dialog in dialogs:
            if dialog is not None and dialog.isVisible() and dialog.isActiveWindow():
                return True
        return False

    def _cursor_inside_window(self) -> bool:
        pos = self.mapFromGlobal(QCursor.pos())
        return self.rect().contains(pos)

    def _begin_fade_out(self) -> None:
        self._hover_inside = False
        self._hover_revealed = False
        self._hover_delay_timer.stop()
        self._start_opacity_animation(self._base_opacity(), self._hover_fade_out_ms())

    def _on_popup_leave_check(self) -> None:
        if self._has_active_popup():
            self._popup_leave_timer.start(120)
            return
        if self._has_active_internal_dialog():
            return
        if not self._cursor_inside_window():
            self._begin_fade_out()

    def _reveal_immediately(self) -> None:
        self._hover_inside = True
        self._hover_revealed = True
        self._hover_delay_timer.stop()
        self._popup_leave_timer.stop()
        self._ui_opacity_anim.stop()
        self._set_ui_opacity(1.0)

    # ----------------------- Quick Text Screen ---------------------- #
    def _build_quick_radial_menu(self) -> None:
        self.quick_radial_menu = QuickRadialMenu(self)
        self.quick_radial_menu.action_requested.connect(self._handle_quick_radial_action)
        self.quick_radial_menu.apply_theme_styles(self.palette_data)
        self._sync_quick_tab_actions()

    def toggle_quick_radial_menu(self) -> None:
        if self.quick_radial_menu is None:
            self._build_quick_radial_menu()
        if self.quick_radial_menu is None:
            return
        if self.quick_radial_menu.isVisible():
            self.quick_radial_menu.hide()
            return
        self._reveal_immediately()
        self._sync_quick_tab_actions()
        self.quick_radial_menu.open_anchored_to(self.quick_actions_button)

    def _handle_quick_radial_action(self, action_key: str) -> None:
        if action_key == "add":
            self.open_quick_editor(None)
        elif action_key == "layout":
            self.open_quick_layout_dialog()
        elif action_key == "new_tab":
            self.add_quick_task_tab()
        elif action_key == "rename":
            self.rename_quick_task_tab()
        elif action_key == "remove":
            self.remove_quick_task_tab()

    def _quick_tabs(self) -> list[dict[str, Any]]:
        raw_tabs = self.config.get("quick_tabs")
        if not isinstance(raw_tabs, list):
            raw_tabs = []
            self.config["quick_tabs"] = raw_tabs

        if not raw_tabs:
            legacy_items = self.config.get("quick_texts", [])
            if not isinstance(legacy_items, list):
                legacy_items = []
            raw_tabs.append({"name": "Main", "quick_texts": legacy_items})

        changed = False
        for idx, tab in enumerate(raw_tabs):
            if not isinstance(tab, dict):
                raw_tabs[idx] = {"name": "Main" if idx == 0 else f"Task {idx + 1}", "quick_texts": []}
                changed = True
                continue
            name = str(tab.get("name", "")).strip()[:32]
            if not name:
                name = "Main" if idx == 0 else f"Task {idx + 1}"
            if str(tab.get("name", "")) != name:
                tab["name"] = name
                changed = True
            if not isinstance(tab.get("quick_texts"), list):
                tab["quick_texts"] = []
                changed = True

        if changed:
            self.queue_save_config()
        return raw_tabs

    def _active_quick_tab_index(self) -> int:
        tabs = self._quick_tabs()
        idx = safe_int(self.config.get("active_quick_tab", 0), 0)
        if idx < 0 or idx >= len(tabs):
            idx = 0
            self.config["active_quick_tab"] = idx
        return idx

    def _sync_legacy_quick_texts(self, tab_index: int | None = None) -> None:
        tabs = self._quick_tabs()
        if not tabs:
            self.config["quick_texts"] = []
            return
        idx = self._active_quick_tab_index() if tab_index is None else safe_int(tab_index, 0)
        if idx < 0 or idx >= len(tabs):
            idx = 0
        tab_items = tabs[idx].get("quick_texts", [])
        self.config["quick_texts"] = tab_items if isinstance(tab_items, list) else []

    def _active_quick_texts(self) -> list[dict[str, Any]]:
        tabs = self._quick_tabs()
        idx = self._active_quick_tab_index()
        tab = tabs[idx]
        entries = tab.get("quick_texts", [])
        if not isinstance(entries, list):
            entries = []
            tab["quick_texts"] = entries
        self.config["active_quick_tab"] = idx
        self._sync_legacy_quick_texts(idx)
        return entries

    def _sync_quick_tab_actions(self) -> None:
        has_tabs = bool(self._quick_tabs())
        can_remove = len(self._quick_tabs()) > 1
        if self.quick_radial_menu is not None:
            self.quick_radial_menu.set_action_enabled("rename", has_tabs)
            self.quick_radial_menu.set_action_enabled("remove", can_remove)

    def _make_quick_tab_canvas_page(self) -> tuple[QWidget, QScrollArea, QuickButtonCanvas]:
        page = QWidget()
        wrapper = QVBoxLayout(page)
        wrapper.setContentsMargins(0, 0, 0, 0)
        wrapper.setSpacing(0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(False)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll.viewport().setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)

        canvas = QuickButtonCanvas()
        scroll.setWidget(canvas)
        wrapper.addWidget(scroll, 1)
        return page, scroll, canvas

    def _rebuild_quick_tab_widgets(self) -> None:
        if self.quick_tabs_widget is None:
            return

        active_index = self._active_quick_tab_index()
        for scroll in self.quick_tab_scrolls:
            try:
                scroll.viewport().removeEventFilter(self)
            except Exception as exc:
                _runtime_log_event(
                    "ui.quick_tab_event_filter_remove_failed",
                    severity="warning",
                    summary="Failed removing viewport event filter while rebuilding quick tab widgets.",
                    exc=exc,
                    context={"scroll_object": repr(scroll)},
                )
        self.quick_tab_scrolls.clear()
        self.quick_tab_canvases.clear()

        tabs = self._quick_tabs()
        self.quick_tabs_widget.blockSignals(True)
        self.quick_tabs_widget.clear()
        for idx, tab in enumerate(tabs):
            page, scroll, canvas = self._make_quick_tab_canvas_page()
            self.quick_tabs_widget.addTab(page, str(tab.get("name", f"Task {idx + 1}")))
            self.quick_tab_scrolls.append(scroll)
            self.quick_tab_canvases.append(canvas)
            scroll.viewport().installEventFilter(self)

        if self.quick_tab_scrolls:
            if active_index < 0 or active_index >= len(self.quick_tab_scrolls):
                active_index = 0
            self.quick_tabs_widget.setCurrentIndex(active_index)
            self.quick_scroll = self.quick_tab_scrolls[active_index]
            self.quick_canvas = self.quick_tab_canvases[active_index]
            self.config["active_quick_tab"] = active_index
            self._sync_legacy_quick_texts(active_index)
        self.quick_tabs_widget.blockSignals(False)
        self._sync_quick_tab_actions()

    def _on_quick_tab_changed(self, index: int) -> None:
        if index < 0 or index >= len(self.quick_tab_scrolls):
            return
        self.quick_scroll = self.quick_tab_scrolls[index]
        self.quick_canvas = self.quick_tab_canvases[index]
        self.config["active_quick_tab"] = int(index)
        self._sync_legacy_quick_texts(index)
        self.refresh_quick_grid()
        if self.quick_layout_dialog is not None and self.quick_layout_dialog.isVisible():
            self.quick_layout_dialog.refresh_cards()
        self.queue_save_config()

    def add_quick_task_tab(self) -> None:
        tabs = self._quick_tabs()
        default_name = f"Task {len(tabs) + 1}"
        name, ok = show_flowgrid_themed_input_text(
            self,
            self,
            "main",
            "New Input Grid Tab",
            "Tab name:",
            default_name,
        )
        if not ok:
            return
        tab_name = str(name).strip()[:32] or default_name
        tabs.append({"name": tab_name, "quick_texts": []})
        new_index = len(tabs) - 1
        self.config["active_quick_tab"] = new_index
        self._sync_legacy_quick_texts(new_index)
        self._rebuild_quick_tab_widgets()
        if self.quick_tabs_widget is not None:
            self.quick_tabs_widget.setCurrentIndex(new_index)
        self.refresh_quick_grid()
        if self.quick_layout_dialog is not None and self.quick_layout_dialog.isVisible():
            self.quick_layout_dialog.refresh_cards()
        self.queue_save_config()

    def rename_quick_task_tab(self) -> None:
        tabs = self._quick_tabs()
        if not tabs:
            return
        idx = self._active_quick_tab_index()
        current_name = str(tabs[idx].get("name", f"Task {idx + 1}"))
        name, ok = show_flowgrid_themed_input_text(
            self,
            self,
            "main",
            "Rename Input Grid Tab",
            "Tab name:",
            current_name,
        )
        if not ok:
            return
        updated_name = str(name).strip()[:32] or current_name
        tabs[idx]["name"] = updated_name
        if self.quick_tabs_widget is not None and 0 <= idx < self.quick_tabs_widget.count():
            self.quick_tabs_widget.setTabText(idx, updated_name)
        self.queue_save_config()

    def remove_quick_task_tab(self) -> None:
        tabs = self._quick_tabs()
        if len(tabs) <= 1:
            self._show_shell_message(
                QMessageBox.Icon.Information,
                "Input Grid",
                "At least one quick-input tab is required.",
                theme_kind="main",
            )
            return
        idx = self._active_quick_tab_index()
        tabs.pop(idx)
        next_index = min(idx, len(tabs) - 1)
        self.config["active_quick_tab"] = max(0, next_index)
        self._sync_legacy_quick_texts(self.config["active_quick_tab"])
        self.close_quick_editor()
        self._rebuild_quick_tab_widgets()
        if self.quick_tabs_widget is not None:
            self.quick_tabs_widget.setCurrentIndex(self.config["active_quick_tab"])
        self.refresh_quick_grid()
        if self.quick_layout_dialog is not None and self.quick_layout_dialog.isVisible():
            self.quick_layout_dialog.refresh_cards()
        self.queue_save_config()

    def _default_quick_position(self, index: int, width: int, height: int) -> tuple[int, int]:
        gap_x = 10
        gap_y = 10
        if hasattr(self, "quick_scroll"):
            available_width = max(120, int(self.quick_scroll.viewport().width()))
        else:
            available_width = max(120, LAUNCH_WIDTH - SIDEBAR_WIDTH - 24)
        columns = max(1, available_width // max(1, width + gap_x))
        col = index % columns
        row = index // columns
        return col * (width + gap_x), row * (height + gap_y)

    def _quick_viewport_width(self) -> int:
        try:
            return max(0, int(self.quick_scroll.viewport().width()))
        except Exception:
            return 0

    def _quick_positions_can_persist(self) -> bool:
        # Avoid persisting clamped placeholder geometry during early startup before viewport layout settles.
        return bool(self.isVisible() and self._quick_viewport_width() > 120)

    def refresh_quick_grid(self) -> None:
        self.quick_canvas.clear_cards()
        self.quick_canvas.clear_alignment_guides()

        quick_texts = self._active_quick_texts()
        self.quick_canvas.configure_grid(show_grid=False, snap_enabled=False)
        self.quick_canvas.set_viewport_width(max(120, self._quick_viewport_width()))
        can_persist_positions = self._quick_positions_can_persist()

        if not quick_texts:
            self.quick_canvas.set_placeholder("No quick text buttons yet.", self.palette_data["muted_text"])
            return

        width = int(self.config.get("quick_button_width", 140))
        height = int(self.config.get("quick_button_height", 40))
        font_size = int(self.config.get("quick_button_font_size", 11))
        font_family = str(self.config.get("quick_button_font_family", "Segoe UI"))
        shape = self.config.get("quick_button_shape", "Soft")
        button_opacity = float(clamp(float(self.config.get("quick_button_opacity", 1.0)), 0.15, 1.0))
        updated_positions = False

        for idx, item in enumerate(quick_texts):
            card = QuickButtonCard(
                idx,
                str(item.get("title", "Untitled"))[:28],
                str(item.get("tooltip", "")),
                self.quick_canvas,
            )
            action_type = self._quick_action_kind(item)
            card.apply_visual_style(width, height, font_size, font_family, shape, button_opacity, self.palette_data, action_type)
            card.set_layout_mode(False)
            card.insert_requested.connect(self.insert_quick_text)
            card.edit_requested.connect(self.open_quick_editor)

            raw_x = item.get("x")
            raw_y = item.get("y")
            if isinstance(raw_x, (int, float)) and isinstance(raw_y, (int, float)):
                pos_x, pos_y = int(raw_x), int(raw_y)
            else:
                pos_x, pos_y = self._default_quick_position(idx, width, height)
                if can_persist_positions:
                    item["x"] = int(pos_x)
                    item["y"] = int(pos_y)
                    updated_positions = True

            snapped_x, snapped_y = self.quick_canvas.place_card(card, pos_x, pos_y, snap=False)
            if (
                can_persist_positions
                and (
                    safe_int(item.get("x", -99999), -99999) != snapped_x
                    or safe_int(item.get("y", -99999), -99999) != snapped_y
                )
            ):
                item["x"] = int(snapped_x)
                item["y"] = int(snapped_y)
                updated_positions = True

        if updated_positions:
            self.queue_save_config()

    def open_quick_editor(self, index: int | None) -> None:
        quick_texts = self._active_quick_texts()
        self._editing_index = index
        self._refresh_available_browsers()

        if index is None:
            self.editor_title.setText("")
            self.editor_tooltip.setText("")
            self.editor_action_combo.setCurrentIndex(0)
            self.editor_text.setPlainText("")
            self.editor_macro.setPlainText("")
            self.editor_apps.setPlainText("")
            self.editor_urls.setPlainText("")
            self.editor_browser_combo.setCurrentIndex(0)
            self.editor_delete_btn.setEnabled(False)
        else:
            if index < 0 or index >= len(quick_texts):
                return
            entry = quick_texts[index]
            self.editor_title.setText(str(entry.get("title", "")))
            self.editor_tooltip.setText(str(entry.get("tooltip", "")))
            action = str(entry.get("action", "paste_text")).strip().lower()
            if action == "macro_sequence":
                action = "input_sequence"
            action_index = self.editor_action_combo.findData(action)
            if action_index < 0:
                action_index = self.editor_action_combo.findData("paste_text")
            self.editor_action_combo.setCurrentIndex(max(0, action_index))
            
            # Handle input sequence vs regular text.
            if action == "input_sequence":
                self.editor_macro.setPlainText(str(entry.get("text", "")))
                self.editor_text.setPlainText("")
            else:
                self.editor_text.setPlainText(str(entry.get("text", "")))
                self.editor_macro.setPlainText("")
            
            app_targets = str(entry.get("app_targets", "")).strip()
            if not app_targets:
                app_targets = str(entry.get("open_target", "")).strip()
            self.editor_apps.setPlainText(app_targets)
            self.editor_urls.setPlainText(str(entry.get("urls", "")))
            browser_path = str(entry.get("browser_path", "")).strip()
            browser_index = self.editor_browser_combo.findData(browser_path)
            if browser_index < 0:
                browser_index = 0
            self.editor_browser_combo.setCurrentIndex(browser_index)
            self.editor_delete_btn.setEnabled(True)

        # Opening the editor is an explicit interaction; reveal full shell opacity immediately.
        self._reveal_immediately()
        if not self.quick_editor_dialog.isVisible():
            px = int(self.x() + max(8, (self.width() - self.quick_editor_dialog.width()) / 2))
            py = int(self.y() + max(36, (self.height() - self.quick_editor_dialog.height()) / 2))
            self.quick_editor_dialog.move(px, py)
        self.quick_editor_dialog.show()
        self.quick_editor_dialog.raise_()
        self.quick_editor_dialog.activateWindow()

    @staticmethod
    def _input_sequence_contains_blocked_credentials(sequence_text: str) -> bool:
        text = str(sequence_text or "")
        if not text.strip():
            return False
        lowered = text.lower()
        if re.search(r"\b(user(name)?|user[_\s-]?id|login|sign[\s-]?in|pass(word)?|passwd|pwd)\b", lowered):
            return True
        if re.search(r"\b(user(name)?|user[_\s-]?id|pass(word)?|passwd|pwd)\s*[:=]", lowered):
            return True
        # Block common "email/username [tab] password" style sequences.
        if re.search(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}.*\[(tab|enter)\]", text, re.IGNORECASE):
            return True
        return False

    def save_quick_editor(self) -> None:
        title = self.editor_title.text().strip() or "Untitled"
        tooltip = self.editor_tooltip.text().strip()
        action = str(self.editor_action_combo.currentData() or "paste_text")
        text = self.editor_text.toPlainText()
        macro = self.editor_macro.toPlainText()
        app_targets = self.editor_apps.toPlainText().strip()
        urls = self.editor_urls.toPlainText().strip()
        browser_path = str(self.editor_browser_combo.currentData() or "").strip()

        if action == "paste_text":
            macro = ""
            app_targets = ""
            urls = ""
            browser_path = ""
        elif action == "input_sequence":
            text = macro
            app_targets = ""
            urls = ""
            browser_path = ""
            macro = ""
            if self._input_sequence_contains_blocked_credentials(text):
                _runtime_log_event(
                    "ui.quick_input_sequence_sensitive_content_blocked",
                    severity="warning",
                    summary="Blocked saving input sequence containing credential-like content.",
                    context={"title": title, "user_id": str(self.current_user)},
                )
                self._show_shell_message(
                    QMessageBox.Icon.Warning,
                    "Input Sequence Blocked",
                    "Input Sequences cannot store username/password or login credentials.\n"
                    "Please remove sensitive fields and save again.",
                )
                return
        elif action == "open_url":
            text = ""
            macro = ""
            app_targets = ""
        elif action == "open_app":
            text = ""
            macro = ""
            urls = ""
            browser_path = ""

        quick_texts = self._active_quick_texts()

        if self._editing_index is None:
            pos_x, pos_y = self._default_quick_position(
                len(quick_texts),
                int(self.config.get("quick_button_width", 140)),
                int(self.config.get("quick_button_height", 40)),
            )
            entry = {
                "title": title,
                "tooltip": tooltip,
                "text": text,
                "action": action,
                "open_target": "",
                "app_targets": app_targets,
                "urls": urls,
                "browser_path": browser_path,
                "x": int(pos_x),
                "y": int(pos_y),
            }
            quick_texts.append(entry)
        elif 0 <= self._editing_index < len(quick_texts):
            existing = quick_texts[self._editing_index]
            entry = {
                "title": title,
                "tooltip": tooltip,
                "text": text,
                "action": action,
                "open_target": "",
                "app_targets": app_targets,
                "urls": urls,
                "browser_path": browser_path,
                "x": safe_int(existing.get("x", 0), 0) if isinstance(existing, dict) else 0,
                "y": safe_int(existing.get("y", 0), 0) if isinstance(existing, dict) else 0,
            }
            quick_texts[self._editing_index] = entry

        self.queue_save_config()
        self.close_quick_editor()
        self.refresh_quick_grid()
        if self.quick_layout_dialog is not None and self.quick_layout_dialog.isVisible():
            self.quick_layout_dialog.refresh_cards()

    def delete_quick_editor(self) -> None:
        if self._editing_index is None:
            return

        quick_texts = self._active_quick_texts()
        if 0 <= self._editing_index < len(quick_texts):
            quick_texts.pop(self._editing_index)
            self.queue_save_config()
            self.close_quick_editor()
            self.refresh_quick_grid()
            if self.quick_layout_dialog is not None and self.quick_layout_dialog.isVisible():
                self.quick_layout_dialog.refresh_cards()

    def close_quick_editor(self) -> None:
        self._editing_index = None
        self.quick_editor_dialog.hide()

    def _capture_external_target(self) -> None:
        if os.name != "nt":
            return

        hwnd = user32.GetForegroundWindow()
        if not hwnd:
            return

        pid = ctypes.c_ulong(0)
        user32.GetWindowThreadProcessId(hwnd, ctypes.byref(pid))
        if pid.value != os.getpid():
            self.last_external_hwnd = int(hwnd)

    def _is_shift_pressed(self) -> bool:
        return bool(QApplication.keyboardModifiers() & Qt.KeyboardModifier.ShiftModifier)

    @staticmethod
    def _resolve_context_script_path(script_name: str) -> Path | None:
        candidates = [
            Path(__file__).with_name(script_name),
            Path.cwd() / script_name,
        ]
        for candidate in candidates:
            if candidate.exists():
                return candidate
        return None

    @staticmethod
    def _entry_context_text(entry: dict[str, Any]) -> str:
        context = str(entry.get("context", "")).strip()
        if context:
            return context
        return str(entry.get("tooltip", "")).strip()

    def _launch_shift_context_script_for_entry(self, entry: dict[str, Any]) -> bool:
        if not self._is_shift_pressed():
            return False

        context_text = self._entry_context_text(entry).lower()
        if not context_text:
            return False

        for context_keyword, script_name in SHIFT_CONTEXT_SCRIPT_LAUNCHERS.items():
            if context_keyword.lower() not in context_text:
                continue
            script_path = self._resolve_context_script_path(script_name)
            if script_path is None:
                return False
            return self._open_app_target(str(script_path))
        return False

    def _send_ctrl_v(self) -> None:
        if os.name != "nt":
            return
        user32.keybd_event(VK_CONTROL, 0, 0, 0)
        user32.keybd_event(VK_V, 0, 0, 0)
        user32.keybd_event(VK_V, 0, KEYEVENTF_KEYUP, 0)
        user32.keybd_event(VK_CONTROL, 0, KEYEVENTF_KEYUP, 0)

    def _send_key(self, key_code: int, shift_held: bool = False, alt_held: bool = False) -> None:
        """Send a single key press with optional modifiers."""
        if os.name != "nt":
            return
        
        modifiers: list[int] = []
        if shift_held:
            modifiers.append(VK_SHIFT)
        if alt_held:
            modifiers.append(VK_ALT)
        
        # Press modifiers
        for mod in modifiers:
            user32.keybd_event(mod, 0, 0, 0)
        
        # If key is return, set as return for consistent behavior
        if key_code == VK_ENTER:
            key_code = VK_RETURN

        # Press and release the key
        user32.keybd_event(key_code, 0, 0, 0)
        user32.keybd_event(key_code, 0, KEYEVENTF_KEYUP, 0)
        
        # Release modifiers
        for mod in reversed(modifiers):
            user32.keybd_event(mod, 0, KEYEVENTF_KEYUP, 0)

    def _parse_macro_sequence(self, sequence: str) -> list[dict[str, Any]]:
        """Parse an input sequence into supported commands.

        Supported syntax:
        - plain text
        - [tab]
        - [enter]
        - [delay: ms]
        """
        import re
        
        commands: list[dict[str, Any]] = []
        
        # Find all [commands] and plain text segments
        pattern = r'(\[([^\]]+)\]|[^\[\]]+)'
        matches = re.finditer(pattern, sequence)
        
        for match in matches:
            part = match.group(0).strip()
            if not part:
                continue
                
            # Check if this part is a [command]
            if part.startswith('[') and part.endswith(']'):
                content = part[1:-1].strip()
                
                # Parse the command
                if ':' in content:
                    cmd_type, cmd_value = content.split(':', 1)
                    cmd_type = cmd_type.strip().lower()
                    cmd_value = cmd_value.strip()

                    if cmd_type == 'delay':
                        try:
                            delay_ms = int(cmd_value)
                            if 0 <= delay_ms <= 60000:  # Max 60 seconds
                                commands.append({'action': 'delay', 'ms': delay_ms})
                        except ValueError as exc:
                            _runtime_log_event(
                                "runtime.macro_delay_parse_failed",
                                severity="warning",
                                summary="Invalid [delay: ms] command value in input sequence; command skipped.",
                                exc=exc,
                                context={"raw_value": cmd_value, "sequence_preview": sequence[:240]},
                            )
                else:
                    # Single-word commands: [tab], [enter].
                    cmd_name = content.lower()
                    if cmd_name == 'tab':
                        commands.append({'action': 'key', 'key': 'tab'})
                    elif cmd_name in {'enter', 'return'}:
                        commands.append({'action': 'key', 'key': 'enter'})
            else:
                # Plain text - treat as type command
                text = part.strip()
                if text:
                    commands.append({'action': 'type', 'text': text})
        
        return commands

    def _execute_macro_sequence(self, sequence: str) -> None:
        """Execute a parsed macro sequence."""
        if os.name != "nt":
            return
        
        # Restore focus to the external window before executing macro
        if self.last_external_hwnd:
            hwnd = int(self.last_external_hwnd)
            try:
                user32.ShowWindow(hwnd, SW_RESTORE)
                user32.SetForegroundWindow(hwnd)
                time.sleep(0.05)
            except Exception as exc:
                _runtime_log_event(
                    "runtime.macro_restore_foreground_failed",
                    severity="warning",
                    summary="Failed restoring foreground window before executing macro sequence.",
                    exc=exc,
                    context={"hwnd": hwnd},
                )
        
        commands = self._parse_macro_sequence(sequence)
        
        for cmd in commands:
            action = cmd.get('action', '')
            
            if action == 'type':
                text = cmd.get('text', '')
                if text:
                    QGuiApplication.clipboard().setText(text)
                    time.sleep(0.05)  # Small delay before pasting
                    self._send_ctrl_v()
                    time.sleep(0.05)  # Small delay after pasting
            
            elif action == 'key':
                key_name = cmd.get('key', '')
                shift_held = cmd.get('shift', False)
                alt_held = cmd.get('alt', False)
                
                if key_name in KEY_CODES:
                    key_code = KEY_CODES[key_name]
                    self._send_key(key_code, shift_held=shift_held, alt_held=alt_held)
                    time.sleep(0.06)  # Small delay between keys
            
            elif action == 'delay':
                delay_ms = cmd.get('ms', 0)
                time.sleep(delay_ms / 1000.0)

    def _quick_action_kind(self, entry: dict[str, Any]) -> str:
        action = str(entry.get("action", "paste_text")).strip().lower()
        if action == "macro_sequence":
            action = "input_sequence"
        if action not in {"paste_text", "open_url", "open_app", "input_sequence"}:
            return "paste_text"
        return action

    def _detect_installed_browsers(self) -> list[tuple[str, str]]:
        found: list[tuple[str, str]] = []
        seen: set[str] = set()

        def add_browser(label: str, path: str) -> None:
            if not path:
                return
            path_obj = Path(path)
            if not path_obj.exists():
                return
            key = str(path_obj).lower()
            if key in seen:
                return
            seen.add(key)
            found.append((label, str(path_obj)))

        checks = [
            (
                "Microsoft Edge",
                ["msedge.exe"],
                [
                    ("PROGRAMFILES(X86)", r"Microsoft\Edge\Application\msedge.exe"),
                    ("PROGRAMFILES", r"Microsoft\Edge\Application\msedge.exe"),
                ],
            ),
            (
                "Google Chrome",
                ["chrome.exe"],
                [
                    ("PROGRAMFILES", r"Google\Chrome\Application\chrome.exe"),
                    ("PROGRAMFILES(X86)", r"Google\Chrome\Application\chrome.exe"),
                    ("LOCALAPPDATA", r"Google\Chrome\Application\chrome.exe"),
                ],
            ),
            (
                "Mozilla Firefox",
                ["firefox.exe"],
                [
                    ("PROGRAMFILES", r"Mozilla Firefox\firefox.exe"),
                    ("PROGRAMFILES(X86)", r"Mozilla Firefox\firefox.exe"),
                ],
            ),
            (
                "Brave",
                ["brave.exe"],
                [
                    ("PROGRAMFILES", r"BraveSoftware\Brave-Browser\Application\brave.exe"),
                    ("PROGRAMFILES(X86)", r"BraveSoftware\Brave-Browser\Application\brave.exe"),
                    ("LOCALAPPDATA", r"BraveSoftware\Brave-Browser\Application\brave.exe"),
                ],
            ),
            (
                "Opera",
                ["opera.exe", "launcher.exe"],
                [
                    ("LOCALAPPDATA", r"Programs\Opera\opera.exe"),
                    ("PROGRAMFILES", r"Opera\launcher.exe"),
                    ("PROGRAMFILES(X86)", r"Opera\launcher.exe"),
                ],
            ),
            (
                "Vivaldi",
                ["vivaldi.exe"],
                [
                    ("LOCALAPPDATA", r"Vivaldi\Application\vivaldi.exe"),
                    ("PROGRAMFILES", r"Vivaldi\Application\vivaldi.exe"),
                ],
            ),
        ]

        for label, exe_names, rel_paths in checks:
            for exe_name in exe_names:
                located = shutil.which(exe_name)
                if located:
                    add_browser(label, located)
            for env_key, rel_path in rel_paths:
                base = os.environ.get(env_key, "")
                if not base:
                    continue
                add_browser(label, str(Path(base) / rel_path))

        return found

    def _refresh_available_browsers(self, preferred_path: str | None = None) -> None:
        if not hasattr(self, "editor_browser_combo"):
            return
        current_path = preferred_path if preferred_path is not None else str(self.editor_browser_combo.currentData() or "").strip()
        self.editor_browser_combo.blockSignals(True)
        self.editor_browser_combo.clear()
        self.editor_browser_combo.addItem("Default Browser", "")
        for label, path in self._detect_installed_browsers():
            self.editor_browser_combo.addItem(label, path)
        target_index = self.editor_browser_combo.findData(current_path)
        if target_index < 0:
            target_index = 0
        self.editor_browser_combo.setCurrentIndex(target_index)
        self.editor_browser_combo.blockSignals(False)

    def _set_editor_row_visible(self, label: QWidget | None, field: QWidget, visible: bool) -> None:
        if label is not None:
            label.setVisible(visible)
        field.setVisible(visible)

    def _update_quick_editor_action_ui(self) -> None:
        action = str(self.editor_action_combo.currentData() or "paste_text")
        text_mode = action == "paste_text"
        input_mode = action in {"input_sequence", "macro_sequence"}
        url_mode = action == "open_url"
        app_mode = action == "open_app"

        self._set_editor_row_visible(self.editor_text_label, self.editor_text, text_mode)
        self._set_editor_row_visible(self.editor_macro_label, self.editor_macro_wrap, input_mode)
        self._set_editor_row_visible(self.editor_apps_label, self.editor_apps_wrap, app_mode)
        self._set_editor_row_visible(self.editor_urls_label, self.editor_urls, url_mode)
        self._set_editor_row_visible(self.editor_browser_label, self.editor_browser_wrap, url_mode)

        if action == "paste_text":
            self.editor_text.setPlaceholderText("Text to copy + paste into the last selected app.")
            self.editor_macro.setPlaceholderText("")
            self.editor_macro.setEnabled(False)
            self.editor_apps.setEnabled(False)
            self.editor_apps_browse.setEnabled(False)
            self.editor_apps_browse_folder.setEnabled(False)
            self.editor_urls.setEnabled(False)
            self.editor_browser_combo.setEnabled(False)
            self.editor_refresh_browsers_button.setEnabled(False)
        elif action in {"input_sequence", "macro_sequence"}:
            self.editor_macro.setPlaceholderText("Format: Bin location [enter]  (credentials are blocked)")
            self.editor_macro.setEnabled(True)
            self.editor_text.setPlaceholderText("")
            self.editor_text.setEnabled(False)
            self.editor_apps.setEnabled(False)
            self.editor_apps_browse.setEnabled(False)
            self.editor_apps_browse_folder.setEnabled(False)
            self.editor_urls.setEnabled(False)
            self.editor_browser_combo.setEnabled(False)
            self.editor_refresh_browsers_button.setEnabled(False)
        elif action == "open_url":
            self.editor_urls.setPlaceholderText("One URL per line.")
            self.editor_macro.setPlaceholderText("")
            self.editor_macro.setEnabled(False)
            self.editor_apps.setEnabled(False)
            self.editor_apps_browse.setEnabled(False)
            self.editor_apps_browse_folder.setEnabled(False)
            self.editor_urls.setEnabled(True)
            self.editor_browser_combo.setEnabled(True)
            self.editor_refresh_browsers_button.setEnabled(True)
        else:
            self.editor_apps.setPlaceholderText("One app/file/folder path per line. Use Browse buttons to pick targets.")
            self.editor_macro.setPlaceholderText("")
            self.editor_macro.setEnabled(False)
            self.editor_apps.setEnabled(True)
            self.editor_apps_browse.setEnabled(True)
            self.editor_apps_browse_folder.setEnabled(True)
            self.editor_urls.setEnabled(False)
            self.editor_browser_combo.setEnabled(False)
            self.editor_refresh_browsers_button.setEnabled(False)

    def browse_quick_apps(self) -> None:
        action = str(self.editor_action_combo.currentData() or "paste_text")
        if action != "open_app":
            return
        current_lines = [line.strip() for line in self.editor_apps.toPlainText().splitlines() if line.strip()]
        current = current_lines[-1] if current_lines else ""
        start_dir = str(Path(current).parent) if current else str(Path.home())
        if not Path(start_dir).exists():
            start_dir = str(Path.home())
        selected_paths, _ = show_flowgrid_themed_open_file_names(
            self,
            self,
            "main",
            "Select App/File Targets",
            start_dir,
            "All Files (*.*);;Programs (*.exe *.lnk *.bat *.cmd *.ps1);;Executables (*.exe *.bat *.cmd);;Shortcuts (*.lnk)",
        )
        if not selected_paths:
            return
        lines = [line.strip() for line in self.editor_apps.toPlainText().splitlines() if line.strip()]
        for selected in selected_paths:
            normalized = str(selected or "").strip()
            if normalized and normalized not in lines:
                lines.append(normalized)
        self.editor_apps.setPlainText("\n".join(lines))

    def browse_quick_app_folder(self) -> None:
        action = str(self.editor_action_combo.currentData() or "paste_text")
        if action != "open_app":
            return
        current_lines = [line.strip() for line in self.editor_apps.toPlainText().splitlines() if line.strip()]
        current = current_lines[-1] if current_lines else ""
        start_dir = str(Path(current).parent) if current else str(Path.home())
        if not Path(start_dir).exists():
            start_dir = str(Path.home())
        selected_dir = show_flowgrid_themed_existing_directory(self, self, "main", "Select Folder Target", start_dir)
        selected_dir = str(selected_dir or "").strip()
        if not selected_dir:
            return
        lines = [line.strip() for line in self.editor_apps.toPlainText().splitlines() if line.strip()]
        if selected_dir not in lines:
            lines.append(selected_dir)
        self.editor_apps.setPlainText("\n".join(lines))

    def _insert_macro_command(self, command: str) -> None:
        """Insert an input-sequence token into the sequence text box."""
        cursor = self.editor_macro.textCursor()
        cursor.movePosition(QTextCursor.End)
        text = self.editor_macro.toPlainText()
        if text and not text[-1].isspace():
            cursor.insertText(" ")
        cursor.insertText(command)
        self.editor_macro.setTextCursor(cursor)

    def _insert_macro_delay(self) -> None:
        """Open dialog to insert [delay: ms] command."""
        seconds, ok = show_flowgrid_themed_input_int(
            self,
            self,
            "main",
            "Insert Delay Command",
            "Enter delay in seconds (0-60):",
            1,
            0,
            60,
            1,
        )
        if ok:
            milliseconds = seconds * 1000
            self._insert_macro_command(f"[delay: {milliseconds}]")

    def _insert_macro_simple(self, command: str) -> None:
        """Insert a simple predefined command."""
        self._insert_macro_command(command)

    def _insert_text_payload(self, text: str) -> None:
        if not text:
            return
        QGuiApplication.clipboard().setText(text)
        if os.name == "nt" and self.last_external_hwnd:
            hwnd = int(self.last_external_hwnd)
            try:
                user32.ShowWindow(hwnd, SW_RESTORE)
                user32.SetForegroundWindow(hwnd)
                time.sleep(0.05)
                self._send_ctrl_v()
            except Exception as exc:
                _runtime_log_event(
                    "runtime.insert_text_foreground_restore_failed",
                    severity="warning",
                    summary="Failed restoring target window before paste; using direct Ctrl+V fallback.",
                    exc=exc,
                    context={"hwnd": hwnd, "text_length": len(text)},
                )
                self._send_ctrl_v()
        else:
            self._send_ctrl_v()

    def _open_url_target(self, target: str, browser_path: str = "") -> bool:
        target = target.strip()
        if not target:
            return False
        url = QUrl.fromUserInput(target)
        if not url.isValid():
            return False
        browser_path = browser_path.strip()
        if browser_path and Path(browser_path).exists():
            detached = QProcess.startDetached(browser_path, [url.toString()])
            if isinstance(detached, tuple):
                return bool(detached[0])
            return bool(detached)
        return bool(QDesktopServices.openUrl(url))

    def _parse_urls(self, urls_text: str) -> list[str]:
        urls: list[str] = []
        for part in urls_text.replace(";", "\n").splitlines():
            value = part.strip()
            if value:
                urls.append(value)
        return urls

    def _parse_targets(self, targets_text: str) -> list[str]:
        targets: list[str] = []
        for part in targets_text.replace(";", "\n").splitlines():
            value = part.strip()
            if value:
                targets.append(value)
        return targets

    def _open_url_targets(self, urls_text: str, browser_path: str = "", fallback_target: str = "", fallback_text: str = "") -> bool:
        urls = self._parse_urls(urls_text)
        if not urls:
            if fallback_target.strip():
                urls = [fallback_target.strip()]
            elif fallback_text.strip():
                urls = [fallback_text.strip()]
        browser_path = browser_path.strip()
        if browser_path and Path(browser_path).exists() and len(urls) > 1:
            args: list[str] = []
            for target in urls:
                url = QUrl.fromUserInput(target)
                if url.isValid():
                    args.append(url.toString())
            if args:
                detached = QProcess.startDetached(browser_path, args)
                if isinstance(detached, tuple):
                    return bool(detached[0])
                return bool(detached)
        opened = False
        for url in urls:
            opened = self._open_url_target(url, browser_path=browser_path) or opened
        return opened

    def _open_app_target(self, target: str) -> bool:
        target = target.strip()
        if not target:
            return False
        if os.name == "nt":
            try:
                os.startfile(target)  # type: ignore[attr-defined]
                return True
            except Exception as exc:
                _runtime_log_event(
                    "runtime.quick_open_app_startfile_failed",
                    severity="warning",
                    summary="os.startfile failed for quick action target; trying alternate open methods.",
                    exc=exc,
                    context={"target": target},
                )
        if Path(target).exists():
            return bool(QDesktopServices.openUrl(QUrl.fromLocalFile(str(Path(target).resolve()))))
        return bool(QDesktopServices.openUrl(QUrl.fromUserInput(target)))

    def _open_app_targets(self, targets_text: str, fallback_target: str = "", fallback_text: str = "") -> bool:
        targets = self._parse_targets(targets_text)
        if not targets:
            if fallback_target.strip():
                targets = [fallback_target.strip()]
            elif fallback_text.strip():
                targets = [fallback_text.strip()]
        opened = False
        for target in targets:
            opened = self._open_app_target(target) or opened
        return opened

    def insert_quick_text(self, index: int) -> None:
        quick_texts = self._active_quick_texts()
        if index < 0 or index >= len(quick_texts):
            return
        entry = quick_texts[index]
        if not isinstance(entry, dict):
            return

        if self._launch_shift_context_script_for_entry(entry):
            return

        action = self._quick_action_kind(entry)
        text = str(entry.get("text", ""))
        target = str(entry.get("open_target", "")).strip()
        app_targets = str(entry.get("app_targets", "")).strip()
        urls = str(entry.get("urls", ""))
        browser_path = str(entry.get("browser_path", "")).strip()

        if action == "open_url":
            opened = self._open_url_targets(urls, browser_path=browser_path, fallback_target=target, fallback_text=text)
            if not opened:
                context = {
                    "index": int(index),
                    "title": str(entry.get("title", "")),
                    "urls": urls,
                    "browser_path": browser_path,
                    "fallback_target": target,
                }
                _runtime_log_event(
                    "runtime.quick_action_open_url_failed",
                    severity="critical",
                    summary="Quick action failed to open any URL target.",
                    context=context,
                )
                _escalate_runtime_issue_once(
                    "runtime.quick_action_open_url_failed",
                    "Quick action could not open the configured URL target(s).",
                    details="Review the URLs and browser path configured for this quick action.",
                    context=context,
                )
        elif action == "open_app":
            opened = self._open_app_targets(app_targets, fallback_target=target, fallback_text=text)
            if not opened:
                context = {
                    "index": int(index),
                    "title": str(entry.get("title", "")),
                    "app_targets": app_targets,
                    "fallback_target": target,
                }
                _runtime_log_event(
                    "runtime.quick_action_open_app_failed",
                    severity="critical",
                    summary="Quick action failed to open any application or file target.",
                    context=context,
                )
                _escalate_runtime_issue_once(
                    "runtime.quick_action_open_app_failed",
                    "Quick action could not open the configured app/file target(s).",
                    details="Verify the path(s) and launch permissions for this quick action.",
                    context=context,
                )
        elif action in {"input_sequence", "macro_sequence"}:
            self._execute_macro_sequence(text)
        else:
            self._insert_text_payload(text)

    # ------------------------- Theme Screen ------------------------- #
    def refresh_theme_controls(self) -> None:
        theme = self.config.get("theme", {})
        for key in ("primary", "accent", "surface"):
            value = normalize_hex(theme.get(key, "#FFFFFF"), "#FFFFFF")
            self.color_swatches[key].setText(value)
            self.color_swatches[key].setStyleSheet(
                "QPushButton {"
                f"background-color: {rgba_css(value, 0.75)};"
                f"color: {readable_text(value)};"
                f"border: 1px solid {shift(value, -0.45)};"
                "font-weight: 700;"
                "}"
            )

        presets = self.config.get("theme_presets", {})
        current = self.config.get("selected_theme_preset")
        self.theme_preset_combo.blockSignals(True)
        self.theme_preset_combo.clear()
        self.theme_preset_combo.addItems(list(presets.keys()))
        if current in presets:
            self.theme_preset_combo.setCurrentText(current)
        self.theme_preset_combo.blockSignals(False)
        self.theme_transparent_bg_check.blockSignals(True)
        self.theme_transparent_bg_check.setChecked(not bool(self.config.get("background_tint_enabled", True)))
        self.theme_transparent_bg_check.blockSignals(False)
        self.popup_auto_reinherit_check.blockSignals(True)
        self.popup_auto_reinherit_check.setChecked(bool(self.config.get("popup_auto_reinherit_enabled", True)))
        self.popup_auto_reinherit_check.blockSignals(False)

        self._refresh_popup_theme_tab("agent")
        self._refresh_popup_theme_tab("qa")
        self._refresh_popup_theme_tab("admin")
        self._refresh_popup_theme_tab("dashboard")

        self._refresh_theme_preview_buttons()

    def pick_theme_color(self, key: str) -> None:
        current = QColor(self.config.get("theme", {}).get(key, "#FFFFFF"))
        color = show_flowgrid_themed_color(self, self, "main", f"Pick {key.title()} Color", current)
        if not color.isValid():
            return
        self.config["theme"][key] = normalize_hex(color.name().upper(), self.config["theme"].get(key, "#FFFFFF"))
        self._theme_updated()

    def _theme_updated(self) -> None:
        self.palette_data = compute_palette(self.config.get("theme", {}))
        self.mark_background_dirty()
        self.apply_theme_styles()
        self.refresh_theme_controls()
        self.refresh_all_views()
        self.queue_save_config()
        self._refresh_popup_themes()

    def _popup_control_fill_css(self, popup_theme: dict[str, Any], field_bg: str) -> str:
        style = str(popup_theme.get("control_style", "Fade Left to Right") or "Fade Left to Right").strip()
        if style not in {"Solid", "Fade Left to Right", "Fade Right to Left", "Fade Center Out"}:
            style = "Fade Left to Right"
        fade_strength = int(clamp(safe_int(popup_theme.get("control_fade_strength", 65), 65), 0, 100))
        base_alpha = float(clamp(safe_int(popup_theme.get("control_opacity", 82), 82), 0, 100)) / 100.0
        tail_alpha = float(clamp(safe_int(popup_theme.get("control_tail_opacity", 0), 0), 0, 100)) / 100.0
        if style == "Solid" or fade_strength <= 0:
            return rgba_css(field_bg, base_alpha)

        computed_end = base_alpha * max(0.0, 1.0 - (fade_strength / 100.0))
        end_alpha = max(tail_alpha, computed_end)
        mid_alpha = max(end_alpha, min(1.0, (base_alpha + end_alpha) / 2.0 + 0.04))
        if style == "Fade Right to Left":
            return (
                "qlineargradient(x1:1,y1:0,x2:0,y2:0,"
                f"stop:0 {rgba_css(field_bg, base_alpha)},"
                f"stop:0.58 {rgba_css(field_bg, mid_alpha)},"
                f"stop:1 {rgba_css(field_bg, end_alpha)})"
            )
        if style == "Fade Center Out":
            return (
                "qlineargradient(x1:0,y1:0,x2:1,y2:0,"
                f"stop:0 {rgba_css(field_bg, end_alpha)},"
                f"stop:0.5 {rgba_css(field_bg, base_alpha)},"
                f"stop:1 {rgba_css(field_bg, end_alpha)})"
            )
        return (
            "qlineargradient(x1:0,y1:0,x2:1,y2:0,"
            f"stop:0 {rgba_css(field_bg, base_alpha)},"
            f"stop:0.58 {rgba_css(field_bg, mid_alpha)},"
            f"stop:1 {rgba_css(field_bg, end_alpha)})"
        )

    def _materialize_popup_theme_for_edit(self, kind: str) -> dict[str, Any]:
        """Freeze inherited popup theme values into the kind-specific config before first customization."""
        key = f"{kind}_theme"
        theme = self.config.setdefault(key, {})
        if not isinstance(theme, dict):
            theme = {}
            self.config[key] = theme

        base = self._resolved_popup_theme(kind)
        inherited_mode = bool(theme.get("inherit_main_theme", False) or self._looks_like_unconfigured_popup_theme(theme))
        if inherited_mode:
            theme["background"] = normalize_hex(base.get("background", "#FFFFFF"), "#FFFFFF")
            theme["text"] = normalize_hex(base.get("text", "#000000"), "#000000")
            theme["field_bg"] = normalize_hex(base.get("field_bg", "#FFFFFF"), "#FFFFFF")
            inherited_layers = base.get("image_layers", [])
            theme["image_layers"] = [
                safe_layer_defaults(layer) for layer in inherited_layers if isinstance(layer, dict)
            ]
        else:
            theme["background"] = normalize_hex(theme.get("background", base["background"]), base["background"])
            theme["text"] = normalize_hex(theme.get("text", base["text"]), base["text"])
            theme["field_bg"] = normalize_hex(theme.get("field_bg", base["field_bg"]), base["field_bg"])

            raw_layers = theme.get("image_layers")
            if isinstance(raw_layers, list):
                cleaned_layers: list[dict[str, Any]] = []
                for layer in raw_layers:
                    if isinstance(layer, dict):
                        cleaned_layers.append(safe_layer_defaults(layer))
                theme["image_layers"] = cleaned_layers
            else:
                inherited_layers = base.get("image_layers", [])
                theme["image_layers"] = [
                    safe_layer_defaults(layer) for layer in inherited_layers if isinstance(layer, dict)
                ]
        return theme

    def _popup_theme_stylesheet(self, kind: str, force_opaque_root: bool = False) -> str:
        theme = self._resolved_popup_theme(kind)
        compact_mode = bool(self.config.get("compact_mode", True))
        tab_padding = "2px 8px" if compact_mode else "4px 10px"
        field_padding = "1px 5px" if compact_mode else "2px 6px"
        header_padding = "3px 5px" if compact_mode else "4px 6px"
        button_padding = "1px 7px" if compact_mode else "2px 8px"
        check_indicator_px = 12 if compact_mode else 14
        transparent = bool(theme.get("transparent", False))
        bg = normalize_hex(theme.get("background", self.palette_data["surface"]), self.palette_data["surface"])
        text = normalize_hex(theme.get("text", self.palette_data["label_text"]), self.palette_data["label_text"])
        field_bg = normalize_hex(theme.get("field_bg", self.palette_data["input_bg"]), self.palette_data["input_bg"])
        field_fill = self._popup_control_fill_css(theme, field_bg)
        field_border = shift(field_bg, -0.38)
        selection_bg = self.palette_data["accent"]
        selection_text = readable_text(selection_bg)
        row_selected_bg = normalize_hex(theme.get("row_selected_color", selection_bg), selection_bg)
        row_selected_text = readable_text(row_selected_bg)
        hover_bg = normalize_hex(theme.get("row_hover_color", ""), "")
        if not hover_bg:
            hover_bg = blend(row_selected_bg, field_bg, 0.55)
        hover_text = readable_text(hover_bg)
        header_bg = normalize_hex(theme.get("header_color", ""), "")
        if not header_bg:
            header_bg = blend(field_bg, bg, 0.25)
        button_bg = blend(self.palette_data["button_bg"], field_bg, 0.30)
        button_hover = shift(button_bg, 0.08)
        button_pressed = shift(button_bg, -0.08)
        button_text = readable_text(button_bg)
        button_border = shift(button_bg, -0.40)
        save_bg = blend(selection_bg, button_bg, 0.50)
        save_border = shift(save_bg, -0.40)
        save_text = readable_text(save_bg)
        pick_bg = blend(button_bg, field_bg, 0.25)
        pick_border = shift(pick_bg, -0.38)
        pick_text = readable_text(pick_bg)
        reset_bg = blend("#BE4E4E", button_bg, 0.45)
        reset_border = shift(reset_bg, -0.42)
        reset_text = readable_text(reset_bg)
        new_bg = blend(self.palette_data["primary"], button_bg, 0.45)
        new_border = shift(new_bg, -0.40)
        new_text = readable_text(new_bg)
        button_bg_css = rgba_css(button_bg, 0.78)
        button_hover_css = rgba_css(button_hover, 0.90)
        button_pressed_css = rgba_css(button_pressed, 0.92)
        save_bg_css = rgba_css(save_bg, 0.82)
        new_bg_css = rgba_css(new_bg, 0.82)
        pick_bg_css = rgba_css(pick_bg, 0.82)
        reset_bg_css = rgba_css(reset_bg, 0.82)
        root_bg = bg if (force_opaque_root or not transparent) else "transparent"
        tab_bg = "transparent"
        tab_hover = rgba_css(selection_bg, 0.26)
        tab_selected_bg = rgba_css(selection_bg, 0.34)
        tab_pane_bg = "transparent"
        root_selector = "QDialog"
        container_transparent_css = (
            "QWidget, QFrame, QGroupBox, QScrollArea, QTabWidget, QStackedWidget {"
            "background-color: transparent;"
            "}"
        )
        return (
            f"{root_selector} {{"
            f"background-color: {root_bg};"
            f"color: {text};"
            "}"
            "QLabel {"
            "background-color: transparent;"
            f"color: {text};"
            "font-weight: 700;"
            "}"
            + container_transparent_css
            + (
            "QTabWidget::pane {"
            f"background-color: {tab_pane_bg};"
            f"border: 1px solid {field_border};"
            "border-radius: 4px;"
            "}"
            "QTabBar::tab {"
            f"background-color: {tab_bg};"
            f"color: {text};"
            f"border: 1px solid {field_border};"
            f"padding: {tab_padding};"
            "border-top-left-radius: 4px;"
            "border-top-right-radius: 4px;"
            "margin-right: 2px;"
            "}"
            "QTabBar::tab:selected {"
            f"background-color: {tab_selected_bg};"
            f"color: {selection_text};"
            "}"
            "QTabBar::tab:hover {"
            f"background-color: {tab_hover};"
            f"color: {selection_text};"
            "}"
            "QLineEdit, QTextEdit, QPlainTextEdit, QSpinBox, QDateEdit, QComboBox, QListWidget, QTableWidget, QListView, QTreeView {"
            f"background: {field_fill};"
            f"color: {text};"
            f"border: 1px solid {field_border};"
            "border-radius: 4px;"
            f"padding: {field_padding};"
            f"selection-background-color: {selection_bg};"
            f"selection-color: {selection_text};"
            "}"
            f"QCheckBox::indicator {{ width: {check_indicator_px}px; height: {check_indicator_px}px; }}"
            "QTableWidget, QListView, QTreeView {"
            "gridline-color: "
            f"{shift(field_border, -0.10)};"
            "alternate-background-color: transparent;"
            "}"
            "QHeaderView::section {"
            f"background-color: {header_bg};"
            f"color: {text};"
            f"border: 1px solid {field_border};"
            f"padding: {header_padding};"
            "font-weight: 700;"
            "}"
            "QComboBox QAbstractItemView {"
            f"background: {field_fill};"
            f"color: {text};"
            f"border: 1px solid {field_border};"
            f"selection-background-color: {row_selected_bg};"
            f"selection-color: {row_selected_text};"
            "}"
            "QListWidget::item, QTableWidget::item, QListView::item, QTreeView::item, QComboBox QAbstractItemView::item {"
            "background-color: transparent;"
            "}"
            "QListWidget::item:hover, QTableWidget::item:hover, QListView::item:hover, QTreeView::item:hover, QComboBox QAbstractItemView::item:hover {"
            f"background-color: {hover_bg};"
            f"color: {hover_text};"
            "}"
            "QListWidget::item:selected, QTableWidget::item:selected, QListView::item:selected, QTreeView::item:selected, QComboBox QAbstractItemView::item:selected {"
            f"background-color: {row_selected_bg};"
            f"color: {row_selected_text};"
            "}"
            "QPushButton {"
            f"background-color: {button_bg_css};"
            f"color: {button_text};"
            f"border: 1px solid {button_border};"
            "border-radius: 4px;"
            f"padding: {button_padding};"
            "}"
            "QPushButton#DepotFramelessCloseButton {"
            "background-color: rgba(225,80,80,110);"
            "border: 1px solid rgba(255,135,135,175);"
            "border-radius: 11px;"
            f"color: {readable_text('#E15050')};"
            "font-size: 12px;"
            "font-weight: 800;"
            "padding: 0px;"
            "}"
            "QPushButton#DepotFramelessCloseButton:hover {"
            "background-color: rgba(235,95,95,155);"
            "}"
            "QPushButton#DepotFramelessCloseButton:pressed {"
            "background-color: rgba(198,74,74,178);"
            "}"
            "QPushButton:hover {"
            f"background-color: {button_hover_css};"
            "}"
            "QPushButton:pressed {"
            f"background-color: {button_pressed_css};"
            "}"
            "QPushButton[actionRole='save'] {"
            f"background-color: {save_bg_css};"
            f"color: {save_text};"
            f"border: 1px solid {save_border};"
            "}"
            "QPushButton[actionRole='new'] {"
            f"background-color: {new_bg_css};"
            f"color: {new_text};"
            f"border: 1px solid {new_border};"
            "}"
            "QPushButton[actionRole='pick'] {"
            f"background-color: {pick_bg_css};"
            f"color: {pick_text};"
            f"border: 1px solid {pick_border};"
            "}"
            "QPushButton[actionRole='reset'] {"
            f"background-color: {reset_bg_css};"
            f"color: {reset_text};"
            f"border: 1px solid {reset_border};"
            "}"
            )
        )

    def _select_popup_color(self, kind: str, field: str) -> None:
        current = QColor(self._resolved_popup_theme(kind).get(field, "#FFFFFF"))
        color = show_flowgrid_themed_color(self, self, kind, f"Pick {kind.title()} {field.title()} Color", current)
        if not color.isValid():
            return
        theme = self._materialize_popup_theme_for_edit(kind)
        theme[field] = normalize_hex(color.name().upper(), theme.get(field, "#FFFFFF"))
        theme["inherit_main_theme"] = False
        self.queue_save_config()
        self._refresh_popup_themes()

    def _pick_popup_theme_color(self, kind: str, field: str) -> None:
        current = QColor(self._resolved_popup_theme(kind).get(field, "#FFFFFF"))
        color = show_flowgrid_themed_color(self, self, kind, f"Pick {kind.title()} {field.title()} Color", current)
        if not color.isValid():
            return
        theme = self._materialize_popup_theme_for_edit(kind)
        theme[field] = normalize_hex(color.name().upper(), theme.get(field, "#FFFFFF"))
        theme["inherit_main_theme"] = False
        self._refresh_popup_theme_tab(kind)
        self._refresh_popup_themes()
        self.queue_save_config()

    def _popup_optional_default_color(self, kind: str, field: str) -> str:
        theme = self._resolved_popup_theme(kind)
        field_bg = normalize_hex(theme.get("field_bg", self.palette_data.get("input_bg", "#FFFFFF")), "#FFFFFF")
        selected = normalize_hex(
            theme.get("row_selected_color", self.palette_data.get("accent", DEFAULT_THEME_ACCENT)),
            DEFAULT_THEME_ACCENT,
        )
        if field == "header_color":
            return blend(
                field_bg,
                normalize_hex(theme.get("background", self.palette_data.get("surface", DEFAULT_THEME_SURFACE)), DEFAULT_THEME_SURFACE),
                0.25,
            )
        if field == "row_hover_color":
            return blend(selected, field_bg, 0.55)
        if field == "row_selected_color":
            return selected
        return field_bg

    def _pick_popup_optional_color(self, kind: str, field: str) -> None:
        theme = self._materialize_popup_theme_for_edit(kind)
        current_hex = normalize_hex(theme.get(field, ""), "")
        if not current_hex:
            current_hex = self._popup_optional_default_color(kind, field)
        color = show_flowgrid_themed_color(
            self,
            self,
            kind,
            f"Pick {kind.title()} {field.replace('_', ' ').title()} Color",
            QColor(current_hex),
        )
        if not color.isValid():
            return
        theme[field] = normalize_hex(color.name().upper(), "")
        theme["inherit_main_theme"] = False
        self._refresh_popup_theme_tab(kind)
        self._refresh_popup_themes()
        self.queue_save_config()

    def _clear_popup_optional_color(self, kind: str, field: str) -> None:
        theme = self._materialize_popup_theme_for_edit(kind)
        theme[field] = ""
        theme["inherit_main_theme"] = False
        self._refresh_popup_theme_tab(kind)
        self._refresh_popup_themes()
        self.queue_save_config()

    def on_popup_theme_control_changed(self, kind: str) -> None:
        theme = self._materialize_popup_theme_for_edit(kind)
        style_combo = getattr(self, f"{kind}_control_style_combo", None)
        fade_slider = getattr(self, f"{kind}_control_fade_slider", None)
        opacity_slider = getattr(self, f"{kind}_control_opacity_slider", None)
        tail_slider = getattr(self, f"{kind}_control_tail_opacity_slider", None)
        if style_combo is None or fade_slider is None or opacity_slider is None or tail_slider is None:
            return
        style = str(style_combo.currentText() or "Fade Left to Right").strip()
        if style not in {"Solid", "Fade Left to Right", "Fade Right to Left", "Fade Center Out"}:
            style = "Fade Left to Right"
        theme["control_style"] = style
        theme["control_fade_strength"] = int(clamp(int(fade_slider.value()), 0, 100))
        theme["control_opacity"] = int(clamp(int(opacity_slider.value()), 0, 100))
        theme["control_tail_opacity"] = int(clamp(int(tail_slider.value()), 0, 100))
        theme["inherit_main_theme"] = False

        fade_value = getattr(self, f"{kind}_control_fade_value", None)
        if fade_value is not None:
            fade_value.setText(f"{int(theme['control_fade_strength'])}%")
        opacity_value = getattr(self, f"{kind}_control_opacity_value", None)
        if opacity_value is not None:
            opacity_value.setText(f"{int(theme['control_opacity'])}%")
        tail_value = getattr(self, f"{kind}_control_tail_opacity_value", None)
        if tail_value is not None:
            tail_value.setText(f"{int(theme['control_tail_opacity'])}%")

        self._refresh_popup_themes()
        self.queue_save_config()

    def _refresh_popup_themes(self) -> None:
        if hasattr(self, "active_agent_window") and self.active_agent_window is not None:
            self.active_agent_window.apply_theme_styles()
        if hasattr(self, "active_qa_window") and self.active_qa_window is not None:
            self.active_qa_window.apply_theme_styles()
        if hasattr(self, "admin_dialog") and self.admin_dialog is not None:
            self.admin_dialog.apply_theme_styles()
        if hasattr(self, "depot_dashboard_dialog") and self.depot_dashboard_dialog is not None:
            self.depot_dashboard_dialog.apply_theme_styles()

    def on_popup_theme_preset_selected(self, kind: str, name: str) -> None:
        preset_name = str(name or "").strip()
        if not preset_name:
            return
        presets = self.config.get("theme_presets", {})
        preset = presets.get(preset_name)
        if not isinstance(preset, dict):
            return
        primary = normalize_hex(str(preset.get("primary", DEFAULT_THEME_PRIMARY)), DEFAULT_THEME_PRIMARY)
        accent = normalize_hex(str(preset.get("accent", DEFAULT_THEME_ACCENT)), DEFAULT_THEME_ACCENT)
        surface = normalize_hex(str(preset.get("surface", DEFAULT_THEME_SURFACE)), DEFAULT_THEME_SURFACE)
        popup_palette = compute_palette({"primary": primary, "accent": accent, "surface": surface})
        theme = self.config.setdefault(f"{kind}_theme", {})
        theme["background"] = normalize_hex(popup_palette.get("control_bg", surface), surface)
        theme["text"] = normalize_hex(popup_palette.get("label_text", readable_text(surface)), readable_text(surface))
        theme["field_bg"] = normalize_hex(
            popup_palette.get("input_bg", blend(surface, primary, 0.18)),
            blend(surface, primary, 0.18),
        )
        theme["inherit_main_theme"] = False
        self.config[f"{kind}_selected_theme_preset"] = preset_name
        self._refresh_popup_theme_tab(kind)
        self._refresh_popup_themes()
        self.queue_save_config()

    def _refresh_popup_theme_tab(self, kind: str) -> None:
        swatches_key = f"{kind}_color_swatches"
        if not hasattr(self, swatches_key):
            return
        color_swatches = getattr(self, swatches_key)
        theme = self._resolved_popup_theme(kind)

        preset_combo = getattr(self, f"{kind}_theme_preset_combo", None)
        if preset_combo is not None:
            presets = self.config.get("theme_presets", {})
            selected = str(self.config.get(f"{kind}_selected_theme_preset", "") or "").strip()
            if not selected:
                selected = str(self.config.get("selected_theme_preset", "") or "").strip()
            preset_combo.blockSignals(True)
            preset_combo.clear()
            preset_combo.addItems(list(presets.keys()))
            if selected in presets:
                preset_combo.setCurrentText(selected)
            elif preset_combo.count() > 0:
                preset_combo.setCurrentIndex(0)
            preset_combo.blockSignals(False)

        for fld, swatch in color_swatches.items():
            value = normalize_hex(theme.get(fld, "#FFFFFF"), "#FFFFFF")
            swatch.setText(value)
            swatch.setStyleSheet(
                "QPushButton {"
                f"background-color: {rgba_css(value, 0.75)};"
                f"color: {readable_text(value)};"
                f"border: 1px solid {shift(value, -0.45)};"
                "font-weight: 700;"
                "}"
            )

        transparent_check_key = f"{kind}_transparent_bg_check"
        if hasattr(self, transparent_check_key):
            transparent_check = getattr(self, transparent_check_key)
            transparent_check.blockSignals(True)
            transparent_check.setChecked(bool(theme.get("transparent", False)))
            transparent_check.blockSignals(False)

        style_combo = getattr(self, f"{kind}_control_style_combo", None)
        if style_combo is not None:
            style_combo.blockSignals(True)
            style_value = str(theme.get("control_style", "Fade Left to Right") or "Fade Left to Right")
            style_idx = style_combo.findText(style_value)
            style_combo.setCurrentIndex(style_idx if style_idx >= 0 else 1)
            style_combo.blockSignals(False)

        fade_slider = getattr(self, f"{kind}_control_fade_slider", None)
        fade_value = getattr(self, f"{kind}_control_fade_value", None)
        if fade_slider is not None:
            fade_slider.blockSignals(True)
            fade_num = int(clamp(int(theme.get("control_fade_strength", 65)), 0, 100))
            fade_slider.setValue(fade_num)
            fade_slider.blockSignals(False)
            if fade_value is not None:
                fade_value.setText(f"{fade_num}%")

        opacity_slider = getattr(self, f"{kind}_control_opacity_slider", None)
        opacity_value = getattr(self, f"{kind}_control_opacity_value", None)
        if opacity_slider is not None:
            opacity_slider.blockSignals(True)
            opacity_num = int(clamp(int(theme.get("control_opacity", 82)), 0, 100))
            opacity_slider.setValue(opacity_num)
            opacity_slider.blockSignals(False)
            if opacity_value is not None:
                opacity_value.setText(f"{opacity_num}%")

        tail_slider = getattr(self, f"{kind}_control_tail_opacity_slider", None)
        tail_value = getattr(self, f"{kind}_control_tail_opacity_value", None)
        if tail_slider is not None:
            tail_slider.blockSignals(True)
            tail_num = int(clamp(int(theme.get("control_tail_opacity", 0)), 0, 100))
            tail_slider.setValue(tail_num)
            tail_slider.blockSignals(False)
            if tail_value is not None:
                tail_value.setText(f"{tail_num}%")

        optional_swatches_key = f"{kind}_optional_color_swatches"
        if hasattr(self, optional_swatches_key):
            optional_swatches = getattr(self, optional_swatches_key)
            for field, swatch in optional_swatches.items():
                raw_value = normalize_hex(theme.get(field, ""), "")
                color_value = raw_value or self._popup_optional_default_color(kind, field)
                label = raw_value if raw_value else f"Auto ({color_value})"
                swatch.setText(label)
                swatch.setStyleSheet(
                    "QPushButton {"
                    f"background-color: {rgba_css(color_value, 0.75)};"
                    f"color: {readable_text(color_value)};"
                    f"border: 1px solid {shift(color_value, -0.45)};"
                    "font-weight: 700;"
                    "}"
                )

    def on_theme_page_background_option_changed(self, checked: bool) -> None:
        self.config["background_tint_enabled"] = not bool(checked)
        self.refresh_settings_controls()
        self.refresh_all_views()
        self.queue_save_config()

    def on_popup_auto_reinherit_changed(self, checked: bool) -> None:
        enabled = bool(checked)
        self.config["popup_auto_reinherit_enabled"] = enabled
        if enabled:
            repaired = self._auto_reinherit_popup_defaults()
            if repaired:
                self.mark_background_dirty()
                self.refresh_theme_controls()
                self._refresh_popup_themes()
        self.queue_save_config()

    def on_popup_background_option_changed(self, kind: str, checked: bool) -> None:
        theme = self._materialize_popup_theme_for_edit(kind)
        theme["inherit_main_theme"] = False
        theme["transparent"] = bool(checked)
        self._refresh_popup_themes()
        self.queue_save_config()

    def reset_theme(self) -> None:
        selected = self.config.get("selected_theme_preset")
        presets = self.config.get("theme_presets", {})
        fallback = presets.get(selected) or next(iter(presets.values()))
        self.config["theme"] = deep_clone(fallback)
        self._theme_updated()

    def on_theme_preset_selected(self, name: str) -> None:
        if not name:
            return
        preset = self.config.get("theme_presets", {}).get(name)
        if not preset:
            return
        self.config["selected_theme_preset"] = name
        self.config["theme"] = deep_clone(preset)
        self._theme_updated()

    def create_theme_preset(self) -> None:
        name = f"Preset {len(self.config.get('theme_presets', {})) + 1}"
        base = name
        suffix = 1
        while name in self.config["theme_presets"]:
            suffix += 1
            name = f"{base} {suffix}"
        self.config["theme_presets"][name] = deep_clone(self.config["theme"])
        self.config["selected_theme_preset"] = name
        self.refresh_theme_controls()
        self.queue_save_config()

    def save_theme_preset(self) -> None:
        name = self.theme_preset_combo.currentText().strip()
        if not name:
            return
        self.config["theme_presets"][name] = deep_clone(self.config["theme"])
        self.config["selected_theme_preset"] = name
        self.refresh_theme_controls()
        self.queue_save_config()

    def open_image_layers_dialog(self, kind: str | bool = "main") -> None:
        if isinstance(kind, bool):
            kind = "main"
        if not isinstance(kind, str):
            kind = "main"
        kind = kind.strip().lower()
        if kind not in {"main", "agent", "qa", "admin", "dashboard"}:
            kind = "main"

        if kind == "main":
            if self.image_dialog is None:
                self.image_dialog = ImageLayersDialog(self, kind="main")

            popup_pos = self.config.get("popup_positions", {}).get("image_layers")
            if isinstance(popup_pos, dict) and "x" in popup_pos and "y" in popup_pos and not self.image_dialog.isVisible():
                self.image_dialog.move(int(popup_pos["x"]), int(popup_pos["y"]))

            self.image_dialog.refresh_list()
            self.image_dialog.apply_theme_styles()
            self.image_dialog.show()
            self.image_dialog.raise_()
            self.image_dialog.activateWindow()
        else:
            # Popup dialogs (agent, qa, admin, dashboard)
            dialog_key = f"{kind}_image_dialog"
            if not hasattr(self, dialog_key):
                setattr(self, dialog_key, ImageLayersDialog(self, kind=kind))
            dialog = getattr(self, dialog_key)

            popup_pos = self.config.get("popup_positions", {}).get(f"image_layers_{kind}")
            if isinstance(popup_pos, dict) and "x" in popup_pos and "y" in popup_pos and not dialog.isVisible():
                dialog.move(int(popup_pos["x"]), int(popup_pos["y"]))

            dialog.refresh_list()
            dialog.apply_theme_styles()
            dialog.show()
            dialog.raise_()
            dialog.activateWindow()

    def open_quick_layout_dialog(self) -> None:
        self._reveal_immediately()
        if self.quick_layout_dialog is None:
            self.quick_layout_dialog = QuickLayoutDialog(self)

        popup_pos = self.config.get("popup_positions", {}).get("quick_layout")
        if isinstance(popup_pos, dict) and "x" in popup_pos and "y" in popup_pos and not self.quick_layout_dialog.isVisible():
            self.quick_layout_dialog.move(int(popup_pos["x"]), int(popup_pos["y"]))

        self.quick_layout_dialog.apply_theme_styles()
        self.quick_layout_dialog.refresh_cards()
        self.quick_layout_dialog.show()
        self.quick_layout_dialog.raise_()
        self.quick_layout_dialog.activateWindow()

    # ----------------------- Settings Screen ------------------------ #
    def refresh_settings_controls(self) -> None:
        self.opacity_slider.blockSignals(True)
        self.hover_delay_slider.blockSignals(True)
        self.hover_fade_in_slider.blockSignals(True)
        self.hover_fade_out_slider.blockSignals(True)
        self.always_on_top_check.blockSignals(True)
        self.compact_mode_check.blockSignals(True)
        self.sidebar_right_switch.blockSignals(True)

        opacity = float(clamp(float(self.config.get("window_opacity", 1.0)), 0.0, 1.0))
        self.opacity_slider.setValue(int(opacity * 100))
        self.opacity_value.setText(f"{opacity:.2f}")
        delay_s = int(clamp(int(self.config.get("hover_reveal_delay_s", 5)), 0, 10))
        fade_in_s = int(clamp(int(self.config.get("hover_fade_in_s", 5)), 0, 10))
        fade_out_s = int(clamp(int(self.config.get("hover_fade_out_s", 5)), 0, 10))
        self.hover_delay_slider.setValue(delay_s)
        self.hover_fade_in_slider.setValue(fade_in_s)
        self.hover_fade_out_slider.setValue(fade_out_s)
        self.hover_delay_value.setText(f"{delay_s}s")
        self.hover_fade_in_value.setText(f"{fade_in_s}s")
        self.hover_fade_out_value.setText(f"{fade_out_s}s")
        self.always_on_top_check.setChecked(bool(self.config.get("always_on_top", False)))
        self.compact_mode_check.setChecked(bool(self.config.get("compact_mode", True)))
        self.sidebar_right_switch.setChecked(bool(self.config.get("sidebar_on_right", False)))

        self.opacity_slider.blockSignals(False)
        self.hover_delay_slider.blockSignals(False)
        self.hover_fade_in_slider.blockSignals(False)
        self.hover_fade_out_slider.blockSignals(False)
        self.always_on_top_check.blockSignals(False)
        self.compact_mode_check.blockSignals(False)
        self.sidebar_right_switch.blockSignals(False)
        self._refresh_sidebar_switch_caption()

    def on_settings_changed(self) -> None:
        self.config["always_on_top"] = bool(self.always_on_top_check.isChecked())
        self.config["compact_mode"] = bool(self.compact_mode_check.isChecked())
        self._apply_window_flags()
        self.apply_theme_styles()
        self.refresh_all_views()
        self.queue_save_config()

    def on_sidebar_position_changed(self, checked: bool) -> None:
        self.config["sidebar_on_right"] = bool(checked)
        self._refresh_sidebar_switch_caption()
        self._apply_sidebar_position()
        self.refresh_all_views()
        self.queue_save_config()

    def _refresh_sidebar_switch_caption(self) -> None:
        is_right = bool(self.sidebar_right_switch.isChecked())
        if hasattr(self, "sidebar_switch_status"):
            self.sidebar_switch_status.setText("Sidebar position: Right" if is_right else "Sidebar position: Left")
        if hasattr(self, "sidebar_left_label"):
            self.sidebar_left_label.setStyleSheet("font-weight: 800;" if not is_right else "font-weight: 500;")
        if hasattr(self, "sidebar_right_label"):
            self.sidebar_right_label.setStyleSheet("font-weight: 800;" if is_right else "font-weight: 500;")

    def on_hover_settings_changed(self) -> None:
        self.config["hover_reveal_delay_s"] = int(self.hover_delay_slider.value())
        self.config["hover_fade_in_s"] = int(self.hover_fade_in_slider.value())
        self.config["hover_fade_out_s"] = int(self.hover_fade_out_slider.value())
        self.hover_delay_value.setText(f"{self.config['hover_reveal_delay_s']}s")
        self.hover_fade_in_value.setText(f"{self.config['hover_fade_in_s']}s")
        self.hover_fade_out_value.setText(f"{self.config['hover_fade_out_s']}s")
        if not self._hover_inside:
            self._set_ui_opacity(self._base_opacity())
        self.queue_save_config()

    def _base_opacity(self) -> float:
        return float(clamp(float(self.config.get("window_opacity", 1.0)), 0.0, 1.0))

    def _hover_delay_ms(self) -> int:
        return int(clamp(int(self.config.get("hover_reveal_delay_s", 5)), 0, 10) * 1000)

    def _hover_fade_in_ms(self) -> int:
        return int(clamp(int(self.config.get("hover_fade_in_s", 5)), 0, 10) * 1000)

    def _hover_fade_out_ms(self) -> int:
        return int(clamp(int(self.config.get("hover_fade_out_s", 5)), 0, 10) * 1000)

    def _start_opacity_animation(self, target_opacity: float, duration_ms: int) -> None:
        target = float(clamp(target_opacity, 0.0, 1.0))
        self._ui_opacity_anim.stop()
        if duration_ms <= 0:
            self._set_ui_opacity(target)
            return
        self._ui_opacity_anim.setDuration(duration_ms)
        self._ui_opacity_anim.setStartValue(self._ui_opacity_current)
        self._ui_opacity_anim.setEndValue(target)
        self._ui_opacity_anim.start()

    def _on_hover_delay_elapsed(self) -> None:
        if not self._hover_inside:
            return
        if self._base_opacity() >= 0.999:
            return
        self._hover_revealed = True
        self._start_opacity_animation(1.0, self._hover_fade_in_ms())

    def on_opacity_changed(self, slider_value: int) -> None:
        opacity = clamp(slider_value / 100.0, 0.0, 1.0)
        self.config["window_opacity"] = opacity
        self.opacity_value.setText(f"{opacity:.2f}")
        if self._hover_revealed:
            self._set_ui_opacity(1.0)
        else:
            self._set_ui_opacity(opacity)
        self.queue_save_config()

    def _apply_window_flags(self) -> None:
        self.setWindowOpacity(1.0)
        self.setWindowFlag(Qt.WindowType.WindowStaysOnTopHint, bool(self.config.get("always_on_top", False)))
        self.show()
        if self._hover_revealed:
            self._set_ui_opacity(1.0)
        else:
            self._set_ui_opacity(self._base_opacity())

    def eventFilter(self, watched, event) -> bool:  # noqa: N802
        if event.type() == QEvent.Type.Resize:
            try:
                if watched in [scroll.viewport() for scroll in self.quick_tab_scrolls]:
                    self.refresh_quick_grid()
                    return False
            except Exception as exc:
                _runtime_log_event(
                    "ui.event_filter_resize_handler_failed",
                    severity="warning",
                    summary="Resize event filter handling failed; continuing with default event processing.",
                    exc=exc,
                    context={"watched": repr(watched)},
                )
        if event.type() == QEvent.Type.MouseButtonPress and isinstance(watched, QWidget):
            if watched is self or self.isAncestorOf(watched):
                self._reveal_immediately()
        return super().eventFilter(watched, event)

    # ---------------------- Icon / Titlebar ------------------------- #
    def pick_custom_icon(self) -> None:
        icon_path, _ = show_flowgrid_themed_open_file_name(
            self,
            self,
            "main",
            "Select Window Icon",
            str(Path.home()),
            "Images (*.png *.ico *.jpg *.jpeg *.bmp *.webp);;All Files (*.*)",
        )
        if not icon_path:
            return
        self.config["app_icon_path"] = icon_path
        self.apply_window_icon()
        self._sync_desktop_shortcut_after_icon_change()
        self.queue_save_config()

    def clear_custom_icon(self) -> None:
        self.config["app_icon_path"] = ""
        self.apply_window_icon()
        self._sync_desktop_shortcut_after_icon_change()
        self.queue_save_config()

    def _load_icon_image(self, icon_path: str) -> QImage:
        return _load_icon_image_file(icon_path)

    def _is_mostly_opaque(self, image: QImage) -> bool:
        return _is_image_mostly_opaque(image)

    def _estimate_corner_matte(self, image: QImage) -> QColor:
        return _estimate_icon_corner_matte(image)

    def _cleanup_icon_transparency(self, image: QImage) -> QImage:
        return _cleanup_icon_transparency_image(image)

    def _build_smoothed_icon(self, icon_path: str) -> QIcon:
        return _build_smoothed_qicon(icon_path)

    def _sync_desktop_shortcut_after_icon_change(self) -> None:
        status, detail = _sync_desktop_shortcut(self.config, create_if_missing=False)
        if status == "failed":
            self._show_themed_message(
                QMessageBox.Icon.Warning,
                "Shortcut Update Failed",
                f"{detail}\n\nSee the runtime log for additional diagnostics.",
            )

    def apply_window_icon(self) -> None:
        icon_source = _resolve_active_app_icon_path(self.config)
        icon = self._build_smoothed_icon(str(icon_source)) if icon_source is not None else QIcon()
        if icon.isNull():
            icon = QApplication.style().standardIcon(QApplication.style().StandardPixmap.SP_DesktopIcon)
        self.setWindowIcon(icon)
        app = QApplication.instance()
        if app is not None:
            app.setWindowIcon(icon)
        self.titlebar.update_icon(icon)

    # ------------------------- Window events ------------------------ #
    def _apply_window_mask(self) -> None:
        if self.width() <= 0 or self.height() <= 0:
            return
        path = QPainterPath()
        path.addRoundedRect(QRectF(self.rect()), self._corner_radius, self._corner_radius)
        self.setMask(QRegion(path.toFillPolygon().toPolygon()))

    def _restore_window_position(self) -> None:
        pos = self.config.get("window_position")
        if isinstance(pos, dict) and "x" in pos and "y" in pos:
            x = safe_int(pos.get("x", 0), 0)
            y = safe_int(pos.get("y", 0), 0)
            win_w = max(120, int(self.width()))
            win_h = max(120, int(self.height()))
            target_rect = QRect(int(x), int(y), win_w, win_h)

            screens = QGuiApplication.screens()
            visible_geometry: QRect | None = None
            for screen in screens:
                try:
                    geometry = screen.availableGeometry()
                except Exception as exc:
                    _runtime_log_event(
                        "ui.restore_window_screen_geometry_failed",
                        severity="warning",
                        summary="Failed reading screen geometry while restoring window position; checking next screen.",
                        exc=exc,
                    )
                    continue
                if geometry.intersects(target_rect):
                    visible_geometry = geometry
                    break

            if visible_geometry is None:
                primary = QGuiApplication.primaryScreen()
                if primary is not None:
                    geometry = primary.availableGeometry()
                    x = int(geometry.left() + max(0, (geometry.width() - win_w) / 2))
                    y = int(geometry.top() + max(0, (geometry.height() - win_h) / 2))
                    self.config["window_position"] = {"x": x, "y": y}
                    self.queue_save_config()
                self.move(int(x), int(y))
                return

            max_x = int(visible_geometry.right() - win_w + 1)
            max_y = int(visible_geometry.bottom() - win_h + 1)
            clamped_x = int(clamp(x, visible_geometry.left(), max_x))
            clamped_y = int(clamp(y, visible_geometry.top(), max_y))
            if clamped_x != x or clamped_y != y:
                self.config["window_position"] = {"x": clamped_x, "y": clamped_y}
                self.queue_save_config()
            self.move(clamped_x, clamped_y)

    def mousePressEvent(self, event: QMouseEvent) -> None:  # noqa: N802
        if event.button() == Qt.MouseButton.LeftButton and _can_start_window_drag(self, event.position().toPoint()):
            self._drag_offset = event.globalPosition().toPoint() - self.frameGeometry().topLeft()
            event.accept()
            return
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event: QMouseEvent) -> None:  # noqa: N802
        if self._drag_offset and event.buttons() & Qt.MouseButton.LeftButton:
            self.move(event.globalPosition().toPoint() - self._drag_offset)
            event.accept()
            return
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event: QMouseEvent) -> None:  # noqa: N802
        self._drag_offset = None
        super().mouseReleaseEvent(event)

    def resizeEvent(self, event) -> None:  # noqa: N802
        self._apply_window_mask()
        super().resizeEvent(event)

    def enterEvent(self, event) -> None:  # noqa: N802
        self._hover_inside = True
        self._hover_delay_timer.stop()
        self._popup_leave_timer.stop()
        self._ui_opacity_anim.stop()

        # If we were already at full opacity, keep it stable when re-entering quickly.
        if self._ui_opacity_current >= 0.985:
            self._hover_revealed = True
            self._set_ui_opacity(1.0)
        elif self._base_opacity() < 0.999:
            self._hover_revealed = False
            self._hover_delay_timer.start(self._hover_delay_ms())
        else:
            self._hover_revealed = True
            self._set_ui_opacity(1.0)
        super().enterEvent(event)

    def leaveEvent(self, event) -> None:  # noqa: N802
        # Start fade-out check immediately when pointer exits the app.
        self._popup_leave_timer.stop()
        self._on_popup_leave_check()
        super().leaveEvent(event)

    def moveEvent(self, event) -> None:  # noqa: N802
        self.config["window_position"] = {"x": int(self.x()), "y": int(self.y())}
        self.queue_save_config()
        super().moveEvent(event)

    def changeEvent(self, event) -> None:  # noqa: N802
        if event.type() == QEvent.Type.ActivationChange and not self.isActiveWindow():
            if not self._has_active_popup() and not self._has_active_internal_dialog():
                self._popup_leave_timer.stop()
                self._begin_fade_out()
        super().changeEvent(event)

    def closeEvent(self, event) -> None:  # noqa: N802
        app = QApplication.instance()
        if app is not None:
            app.removeEventFilter(self)
        for scroll in self.quick_tab_scrolls:
            try:
                scroll.viewport().removeEventFilter(self)
            except Exception as exc:
                _runtime_log_event(
                    "ui.close_remove_event_filter_failed",
                    severity="warning",
                    summary="Failed removing viewport event filter during app close.",
                    exc=exc,
                    context={"scroll_object": repr(scroll)},
                )
        self.config["window_position"] = {"x": int(self.x()), "y": int(self.y())}
        if self.image_dialog is not None:
            self.config.setdefault("popup_positions", {})["image_layers"] = {
                "x": int(self.image_dialog.x()),
                "y": int(self.image_dialog.y()),
            }
            self.image_dialog.close()
        if self.quick_layout_dialog is not None:
            self.config.setdefault("popup_positions", {})["quick_layout"] = {
                "x": int(self.quick_layout_dialog.x()),
                "y": int(self.quick_layout_dialog.y()),
            }
            self.quick_layout_dialog.close()
        if self.quick_radial_menu is not None:
            self.quick_radial_menu.close()
        if self.depot_dashboard_dialog is not None:
            self.depot_dashboard_dialog.close()
        if self.admin_dialog is not None:
            self.admin_dialog.close()
        if hasattr(self, "quick_editor_dialog") and self.quick_editor_dialog is not None:
            self.quick_editor_dialog.close()
        self.save_config()
        super().closeEvent(event)


def main() -> int:
    if os.name == "nt":
        try:
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("QuickInputs.QtApp")
        except Exception as exc:
            _runtime_log_event(
                "bootstrap.app_user_model_id_set_failed",
                severity="warning",
                summary="Failed setting explicit AppUserModelID for Windows shell integration.",
                exc=exc,
            )

    app = QApplication(sys.argv)
    app.setApplicationName(APP_TITLE)
    app.setQuitOnLastWindowClosed(True)
    window = QuickInputsWindow()
    window.show()
    return app.exec()


def _run_command_line_mode() -> int | None:
    if "--install" in _CLI_FLAGS:
        return _run_installer_mode(launch_after_install="--no-launch" not in _CLI_FLAGS)
    if "--create-shortcut" in _CLI_FLAGS:
        return _run_installer_mode(launch_after_install=False)
    return None


if __name__ == "__main__":
    try:
        cli_result = _run_command_line_mode()
        raise SystemExit(main() if cli_result is None else cli_result)
    except SystemExit:
        raise
    except Exception as exc:
        details = "".join(traceback.format_exception(type(exc), exc, exc.__traceback__))
        _notify_launch_error("TH-9000", "Fatal runtime crash during launch.", details)
        raise SystemExit(1)

